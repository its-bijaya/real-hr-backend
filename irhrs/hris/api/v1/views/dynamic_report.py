from django.contrib.auth import get_user_model
from django.db.models.fields.reverse_related import ForeignObjectRel
from django.forms.utils import pretty_name
from django.utils.functional import cached_property
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.filters import SearchFilter
from rest_framework.response import Response
from django.utils.html import strip_tags
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import get_column_letter

from config import settings
from irhrs.core.mixins.serializers import create_dummy_serializer, add_fields_to_serializer_class
from irhrs.core.mixins.viewset_mixins import ListViewSetMixin, OrganizationMixin, PastUserFilterMixin
from irhrs.core.utils import nested_getattr
from irhrs.core.utils.filters import FilterMapBackend, OrderingFilterMap
from irhrs.export.mixins.export import BackgroundTableExportMixin
from irhrs.hris.api.v1.permissions import HRISReportPermission
from irhrs.hris.api.v1.serializers.dynamic_report import DynamicHRISReportSerializer
from irhrs.permission.constants.permissions import HRIS_REPORTS_PERMISSION

USER = get_user_model()


class DynamicHRISReportViewSet(BackgroundTableExportMixin, OrganizationMixin, PastUserFilterMixin, ListViewSetMixin):
    """
    HRIS report for selected fields by user

    selecting fields and categories

    -------------------------------------------------------------------
    filters

        search=full_name
        division=division_slug
        user_status=past/current

    --------------------------------------------------------------------
    selecting categories

    All fields and categories are listed at `dynamic-report/fields` API

    send category names as

    ```url
        ?categories=category_name1,category_name2
    ```

    for multiple values append `-m` to category name


    ```url
        ?categories=cateoory_name1-m,,category_name2-m
    ```

    send field names as

    ```url
        ?fields=field_name1,field_name2
    ```

    Example:

    ```url
        ?categories=g,m,e-m&fields=g_gender,g_code,m_blood_group,e_start_date,e_end_date,e_job_title
    ```


    """
    queryset = USER.objects.all()
    filter_backends = (FilterMapBackend, SearchFilter, OrderingFilterMap)
    search_fields = ('first_name', 'middle_name', 'last_name', )
    filter_map = {
        'division': 'detail__division__slug',
        'user': 'id'
    }
    ordering_fields_map = {
        'full_name': ('first_name', 'middle_name', 'last_name', ),
        'username': 'username'
    }
    permission_classes = [HRISReportPermission]
    notification_permissions = [HRIS_REPORTS_PERMISSION]

    category_relation_map = {
        "g": 'detail',
        "m": 'medical_info',
        "e": 'user_experiences',
        "ed": "user_education",
        "pe": "past_experiences",
        "li": "legal_info",
        "c": "contacts",
        "a": "addresses",
        "b": "userbank"
    }
    category_name_map = {
        "g": "General",
        "m": "Medical Info",
        "e": "Experience",
        "ed": "Education",
        "pe": "Past Experiences",
        "li": "Legal Info",
        "c": "Contacts",
        "a": "Addresses",
        "b": "Bank Detail"
    }

    export_type = 'Dynamic HRIS Report'

    @staticmethod
    def has_user_permission():
        return False

    def get_frontend_redirect_url(self):
        return f'/admin/{self.organization.slug}/hris/reports/generate-report'

    def get_export_fields(self):
        fields_map = [
            {
                "name": "user",
                "title": "User Details",
                "fields": (
                    {"name": "id", "title": "Id"},
                    {"name": "full_name", "title": "Full Name"},
                )
            }
        ]
        for category in self.categories:
            many_in = f"{category['name']}-m" in self.selected_categories
            single_in = category['name'] in self.selected_categories
            if single_in or many_in:
                fields_map.append({
                    "name": f"{category['name']}{'-m' if many_in else ''}",
                    "title": category.get('text'),
                    "fields": [
                        {"name": sub_item.get("field_name"), "title": sub_item.get("text"), "fields": None}
                        for sub_item in category.get('subItems', []) if sub_item.get('field_name') in self.selected_fields
                    ]
                })

        return fields_map

    def get_export_data(self):
        if not self.selected_categories_fields:
            return []
        return super().get_export_data()

    def get_extra_export_data(self):
        extra = super().get_extra_export_data()
        extra["serializer_kwargs"] = {
            "user_fields": [
                'id', 'full_name', 'profile_picture', 'cover_picture', 'job_title'
            ]
        }
        return extra

    def get_serializer_class_params(self):
        return {
            "args": [self.selected_categories_fields, self.category_relation_map, self.cat_fields_map],
            "kwargs": dict(from_export=True)
        }

    @staticmethod
    def get_value(obj, field=None, html_strip=False):
        """get_{field} for serializer used in this view"""
        if field is None:
            attr = obj
        else:
            attr = nested_getattr(obj, field, call=False)

        if attr is None:
            return attr

        if hasattr(attr, 'all'):
            attr = list(attr.all())

        if isinstance(attr, list):
            return ', '.join([DynamicHRISReportViewSet.get_value(atr) for atr in attr])

        if isinstance(attr, str):
            if html_strip:
                return strip_tags(attr).replace("&nbsp;", " ")
            return attr

        if hasattr(attr, 'title'):
            return attr.title

        if hasattr(attr, 'name'):
            return attr.name

        return str(attr)

    @staticmethod
    def get_serializer_class_for_export(selected_categories_fields, category_relation_map, cat_fields_map,
                                        from_export=False):
        category_serializers = dict()
        for category, fields in selected_categories_fields.items():

            cleaned_category = category.replace("-m", "")

            # make all fields SerializerMethodField
            serializer_fields = {f: serializers.SerializerMethodField(allow_null=True) for f in fields}

            # get methods for above fields
            serializer_get_methods = {
                # copy variables are used to prevent all functions referencing to same variable
                f"get_{f}": (lambda s, obj, category_copy=cleaned_category, f_copy=f: DynamicHRISReportViewSet.get_value(
                    obj, f_copy.replace(f"{category_copy}_", "", 1), html_strip=from_export)) for f in fields
            }

            # create serializer with above fields
            serializer = create_dummy_serializer(fields=dict(**serializer_fields, **serializer_get_methods))

            # for many=true fields and if many is not selected use method defined in wrap as field_one
            source = category_relation_map[cleaned_category]
            if "-m" not in category and cat_fields_map[cleaned_category]['allow_many']:
                source = f"{source}_one"

            category_serializers.update({category: serializer(
                source=source,
                many=cat_fields_map[cleaned_category]['allow_many'])}
            )

        # finally create and return serializer
        return add_fields_to_serializer_class(serializer_class=DynamicHRISReportSerializer,
                                              fields=category_serializers)

    @classmethod
    def get_workbook_to_export_file_content(cls, data, title, columns, extra_content,
                                            description=None, **kwargs):
        # overwrite this method for signature(sign) of user
        wb = super().get_workbook_to_export_file_content(
            data, title, columns, extra_content, **kwargs
        )

        ws = wb.active
        rows = next(ws.iter_rows(min_row=5, max_row=5))
        signature_col = [row.column for row in rows if row.value == "Signature"]

        if signature_col:
            signature_col_index = signature_col[0]
            col_letter = get_column_letter(signature_col_index)

            for index in range(6, ws.max_row):
                # set height, width of the row
                ws.row_dimensions[index].height = 50
                ws.column_dimensions[col_letter].width = 11
                signature = ws.cell(row=index, column=signature_col_index)
                signature_val = signature.value

                if not signature_val:
                    continue

                # previously set signature (which was absolute path) is replaced by actual Image
                # from absolute path
                signature.value = None
                try:
                    image_obj = Image(settings.MEDIA_ROOT + signature_val)
                    image_obj.height = image_obj.height * 0.25
                    image_obj.width = image_obj.width * 0.25
                except FileNotFoundError:
                    signature.value = None
                    continue
                ws.add_image(image_obj, anchor=f"{col_letter}{index}")

        return wb

    def get_serializer_class(self, export=False):
        """build serializer"""

        if self.action == 'export' and not export:
            return super().get_serializer_class()

        return self.get_serializer_class_for_export(
            self.selected_categories_fields, self.category_relation_map, self.cat_fields_map
        )

    def get_serializer(self, *args, **kwargs):
        if not self.action == 'export':
            kwargs['user_fields'] = [
                'id', 'full_name', 'profile_picture', 'cover_picture', 'job_title'
            ]
        return super().get_serializer(*args, **kwargs)

    @cached_property
    def categories(self):
        """Generate categories and fields from Model class"""
        excludes = ['id', 'created_at', 'modified_at','completeness_percent', 'created_by', 'modified_by', 'user', 'slug', 'skill']
        additional_fields = {
            # category: fields
            'g': [
                ('email', 'user.email'),
                ('username', 'user.username'),
                ('first_level_supervisor', 'user.first_level_supervisor.full_name'),
                ('signature', 'user.signature')
            ],
            'm': [('chronic_diseases', 'user.chronicdisease_set')]
        }
        categorywise_excludes = {
            'g': ['organization', 'branch', 'division',
                  'job_title', 'employment_level', 'employment_status']
        }
        field_name_map = {
            'employment_status': 'employment_type'
        }

        cats = list()

        for category, relation in self.category_relation_map.items():
            # get field from USER
            field = USER._meta.get_field(relation)
            category_excludes = categorywise_excludes.get(category, [])

            # all fields from related model to that field
            fields = field.related_model._meta.get_fields()

            sub_items = list()

            for f in fields:
                # Exclude reverse relation
                if not isinstance(f, ForeignObjectRel) and f.name not in excludes and f.name not in category_excludes:
                    sub_items.append({
                        "text": pretty_name(field_name_map.get(f.name, f.name)),
                        "value": False,
                        "field_name": f"{category.lower().replace(' ', '_')}_{f.name}"
                    })
            if category in additional_fields:
                additional_set = additional_fields[category]

                for text, field_name in additional_set:
                    sub_items.append({
                        "text": pretty_name(text),
                        "value": False,
                        "field_name": f"{category.lower().replace(' ', '_')}_{field_name}"
                    })


            # Extra keys are for frontend to build the form
            cat = {
                "text": self.category_name_map[category],
                "name": category,
                "value": False,
                "multipleSelected": False,
                "multiple": field.multiple,
                "subItems": sub_items
            }
            cats.append(cat)

        return cats

    @cached_property
    def cat_fields_map(self):
        """
        Actual details required for backend
        """
        cat_fields = dict()
        for cat in self.categories:
            key = cat['name']
            val = {
                "allow_many": cat["multiple"],
                "fields": [f["field_name"] for f in cat["subItems"]]
            }
            cat_fields.update({key: val})

        return cat_fields

    @cached_property
    def available_categories(self):
        """
        List of all available categories including many version
        """
        cats = list()
        for k, v in self.cat_fields_map.items():
            cats.append(k)
            if v["allow_many"]:
                cats.append(f"{k}-m")
        return cats

    @cached_property
    def selected_categories(self):
        """List of selected categories after validating query params"""
        categories = self.request.query_params.get('categories', '').split(',')
        return [cat for cat in categories if cat in self.available_categories]

    @cached_property
    def all_fields_list(self):
        """List of all available fields"""
        return [x for v in self.cat_fields_map.values() for x in v["fields"]]

    @cached_property
    def available_field_names(self):
        """Available fields for selected categories"""
        return [x for k, v in self.cat_fields_map.items()
                for x in v["fields"] if {k, f"{k}-m"}.intersection(self.selected_categories)]

    def _check_fields_validity(self):
        """
        method that checks available field names are unique

        it is made for dev use only as categories are constant
        """
        assert len(self.all_fields_list) == len(set(self.all_fields_list))

    @cached_property
    def selected_fields(self):
        """Final list of selected fields after validating query params"""
        fields = self.request.query_params.get('fields', '').split(',')
        return [field for field in fields if field in self.available_field_names]

    @cached_property
    def selected_categories_fields(self):
        """category to fields map used for serializer"""
        cats = dict()
        for category in self.selected_categories:
            cleaned = category.replace('-m', '')
            avail_fields = self.cat_fields_map.get(cleaned, {}).get('fields', [])

            selected_fields = []
            for field in avail_fields:
                if field in self.selected_fields:
                    selected_fields.append(field)

            cats.update({category: selected_fields})
        return cats

    def get_queryset(self):
        past_employee = self.request.query_params.get('user_status', None)
        qs = super().get_queryset().filter(detail__organization=self.organization)
        return qs.past() if past_employee else qs.current()

    @action(methods=['GET'], detail=False, url_path='fields')
    def get_list_of_fields(self, request, **kwargs):
        return Response(self.categories)
