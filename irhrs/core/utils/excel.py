from collections.abc import Mapping, Sequence
from openpyxl import load_workbook, Workbook
from io import BytesIO


class ExcelDict(Mapping):
    def __init__(self, file):
        wb = self._get_workbook(file)
        ws = wb.active

        # creating error book by initially copying uploaded
        # workbook, deepcopying caused error while loading
        # with openpyxl
        self.error_wb = Workbook()
        self.error_ws = self.error_wb.active
        for value in ws.values:
            self.error_ws.append(value)

        rows = ws.values
        self.header = next(rows)
        self.data = {
            row[0]: dict(zip(self.header[1:], row[1:])) for row in rows if row[0]
        }
        self.to_index = dict(zip(self.data.keys(), range(2, len(self.data) + 2)))
        self.error_field = len(self.header) + 1
        self.errors = {}
        self.error_ws.cell(row=1, column=self.error_field, value="Errors")

    def _get_workbook(self, file):
        if isinstance(file, Workbook):
            return file

        file = BytesIO(file.read())
        return load_workbook(filename=file)

    def __iter__(self):
        return iter(self.data)

    def __len__(self):
        return len(self.data)

    def __getitem__(self, key):
        return self.data[key]

    def __setitem__(self, key, value):
        if not value:
            return
        self.errors[key] = value
        self.error_ws.cell(
            row=self.to_index[key], column=self.error_field, value=str(value)
        )


class ExcelList(Sequence):
    def __init__(self, file):
        wb = self._get_workbook(file)
        ws = wb.active
        self._list = [list(row) for row in ws.values if row[0]]

    def _get_workbook(self, file):
        if isinstance(file, Workbook):
            return file
        file = BytesIO(file.read())
        return load_workbook(filename=file)

    def __getitem__(self, index):
        return self._list[index]

    def __setitem__(self, index, value):
        self._list[index] = value

    def __delitem__(self, index):
        del self._list[index]

    def __len__(self):
        return len(self._list)

    def __insert__(self, index, value):
        return self._list.insert(index, value)

    def generate_workbook(self):
        wb = Workbook()
        ws = wb.active
        for value in self._list:
            ws.append(value)
        return wb
