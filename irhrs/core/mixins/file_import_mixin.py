import uuid
from zipfile import BadZipFile

from django.conf import settings
from django.contrib.auth.models import AnonymousUser
from django.core.cache import cache
from django.http import HttpResponse
from django.utils import timezone
from django_q.tasks import async_task
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from irhrs.core.utils.common import DummyObject
from irhrs.core.utils.filters import inverse_mapping
from irhrs.export.utils.helpers import save_workbook
from irhrs.notification.utils import notify_organization
from irhrs.organization.api.v1.serializers.knowledge_skill_ability import (FileImportSerializer)


class BaseForFileImport:
    queryset_fields_map = None
    failed_url = None
    success_url = None
    import_serializer_class = None

    def get_import_serializer_class(self):
        """
        Serializer class used to import
        default to cls.serializer_class
        """
        return (
            self.import_serializer_class
            or self.__class__.serializer_class
        )

    def __init__(self, *args, **kwargs):
        # TODO: @shital create get_import_fields() and use it

        super().__init__(*args, **kwargs)

        assert hasattr(self, 'import_fields'), "You must specify import " \
                                               "field to generate sample file."
        # if model field is given it sets model_fields_map value else
        # it generates its default model_fields_map values from import_fields
        self.model_fields_map = getattr(
            self,
            'model_fields_map',
            None
        ) or dict(
            zip(
                self.import_fields,
                list(
                    map(
                        lambda fields: fields.lower().replace(' ', '_'),
                        self.import_fields
                    )
                )
            )
        )
        self.validate_model_fields_map(self.model_fields_map)
        self.serializer_field_map = inverse_mapping(self.model_fields_map)

    def get_failed_url(self):
        if not self.failed_url:
            raise AssertionError(
                "'{}' should either include a `failed_url` attribute, "
                "or override the `get_failed_url()` method.".format(self.__class__.__name__)
            )
        return self.failed_url

    def get_success_url(self):
        if not self.success_url:
            raise AssertionError(
                "'{}' should either include a `success_url` attribute, "
                "or override the `get_success_url()` method.".format(self.__class__.__name__)
            )
        return self.success_url

    def get_background_task_name(self):
        task_name = getattr(self, 'background_task_name', None) or self.__class__.__name__
        return task_name

    def get_queryset_fields_map(self):
        return self.queryset_fields_map

    def get_many_to_many_fields(self):
        if hasattr(self, 'many_to_many_fields'):
            return getattr(self, 'many_to_many_fields')
        return []

    @classmethod
    def add_validators(cls, workbook, worksheet, queryset_fields_map, serializer_field_map):
        if queryset_fields_map:
            key_qs = queryset_fields_map
            for key in key_qs:
                worksheet.add_data_validation(
                    cls.get_dv(serializer_field_map.get(key), key_qs.get(key), workbook)
                )
            return worksheet

    @classmethod
    def get_dv(cls, fieldname, qs, workbook):
        export_fields = cls.import_fields
        index = cls.get_index(fieldname, export_fields)
        # TODO @shital use map to return this
        slug_field = getattr(cls, 'slug_field_for_sample', 'slug')
        return cls.get_dv_for_qs(qs, index, fieldname, workbook, field=slug_field)

    @staticmethod
    def get_index(field, l):
        return l.index(field) + 1

    @staticmethod
    def get_dv_for_qs(qs, index, fieldname, workbook, field='slug'):
        fieldname = fieldname.lower().replace(' ', '_')
        if isinstance(qs, list):
            autocomplete = qs
        else:
            autocomplete = list(qs.values_list(field, flat=True))
        # v = ",".join(autocomplete)
        sheet = workbook.create_sheet(fieldname)
        max_row = len(autocomplete)
        for value in autocomplete:
            sheet.append([value])

        cl = get_column_letter(index)

        # dv = DataValidation(type="list", formula1=f'"{v}"', allow_blank=False)

        dv = DataValidation(type="list", formula1=f'{fieldname}!$A$1:$A${max_row}')
        dv.add(f'{cl}2:{cl}1048576')

        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'

        sheet.sheet_state = 'hidden'
        sheet.protection.sheet = True
        sheet.protection.password = f"{fieldname}{max_row}"
        return dv

    @staticmethod
    def validate_model_fields_map(field_map):
        keys = field_map.keys()
        values = field_map.values()
        assert len(keys) == len(set(values)), "Values must be unique for model_fields_map"


class SampleDownloadForImportMixin:
    sample_file_name = 'import'

    def generate_sample_for_import(self):
        values = getattr(self, 'values', None)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(self.import_fields)

        for i in range(1, len(self.import_fields) + 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 30

        if values:
            worksheet.append(values)

        self.add_validators(
            workbook, worksheet,
            self.get_queryset_fields_map(), self.serializer_field_map
        )

        return save_virtual_workbook(workbook)

    def get_sample_file_name(self):
        return self.sample_file_name

    @action(
        detail=False,
        methods=['GET'],
        url_path='import/sample',
        serializer_class=FileImportSerializer
    )
    def download_sample(self, request, *args, **kwargs):
        response = HttpResponse(
            content=self.generate_sample_for_import(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; ' \
                                          f'filename={self.get_sample_file_name()}_sample.xlsx'
        return response


class NormalFileImportBaseMixin(BaseForFileImport):
    def start_import(self, file):
        try:
            workbook = load_workbook(file, read_only=True)
        except (OSError, BadZipFile):
            raise ValidationError(
                {
                    'file': ['The file seems to be corrupted']
                }
            )

        # checking whether the file is empty
        if len(workbook.sheetnames) > 0:
            sheet_name = workbook.sheetnames[0]
        else:
            raise ValidationError(
                {
                    'file': ['The file is empty.']
                }
            )

        worksheet = workbook[sheet_name]
        return list(self.extract_data_for_import(worksheet))

    def extract_data_for_import(self, worksheet):
        headers = None
        self.validate_data(worksheet)
        for index, row in enumerate(worksheet):
            row_data = [cell.value for cell in row]
            if not any(row_data):
                break
            if index == 0:
                headers = [self.model_fields_map.get(header) for header in row_data]
                self.validate_header(headers)
            else:
                yield dict(
                    zip(
                        headers,
                        row_data
                    )
                )

    def reformat_errors(self, errors):
        """
        Convert nested errors to flat errors
        :param errors: Error list
        :return: Flat Error list
        """
        new_errors = list()
        for error in errors:
            error_dict = dict()
            for key, value in error.items():
                if isinstance(value, dict):
                    for k, v in value.items():
                        error_dict.update({self.serializer_field_map.get(k, k): v})
                else:
                    error_dict.update({self.serializer_field_map.get(key, key): value})
            new_errors.append(error_dict)
        return new_errors

    def validate_header(self, headers):
        fields = set(self.import_fields)
        if fields.issubset(set(headers)):
            raise ValidationError(
                {
                    'non_field_errors': ['Could not import holidays. ']
                }
            )

    @staticmethod
    def validate_data(worksheet):
        if worksheet.max_row <= 1:
            raise ValidationError(
                {
                    'non_field_errors': ["No data found for the headers."]
                }
            )


class BackgroundFileImportBaseMixin(BaseForFileImport):
    @staticmethod
    def start_import(cls, path, organization, context, serializer_class,
                     permission, failed_url, success_url, model_fields_map,
                     task_name, queryset_fields_map, non_mandatory_field_value,
                     serializer_field_map, many_to_many_fields=None):
        cache.set(f'{task_name}_import_status_{organization.id}', 'processing', None)
        BackgroundFileImportBaseMixin.save_records(
            cls, path, organization, context, serializer_class,
            permission, failed_url, success_url, model_fields_map,
            task_name, queryset_fields_map, non_mandatory_field_value,
            serializer_field_map, many_to_many_fields
        )
        cache.set(f'block_{task_name}_import_{organization.id}', False, None)

    @staticmethod
    def get_updated_data(data, mtm_fields, separator):
        update_data = {}
        for key, value in data.items():
            if key in mtm_fields and not isinstance(value, list):
                update_data[key] = [value] if separator not in value else value.split(separator)
        data.update(update_data)
        return data

    @staticmethod
    def save_records(
        cls, path, organization, context, serializer_class,
        permission, failed_url, success_url, model_fields_map,
        task_name, queryset_fields_map, non_mandatory_field_value,
        serializer_field_map, many_to_many_fields, separator=','):
        """
         Takes rows of the excel file and process it and save it.

        :param cls:
        :param path:
        :param organization:
        :param context:
        :param serializer_class:
        :param permission:
        :param failed_url:
        :param success_url:
        :param model_fields_map:
        :param task_name:
        :param queryset_fields_map:
        :param non_mandatory_field_value:
        :param serializer_field_map:
        :param many_to_many_fields:
        :param separator:
        :return:
        """

        validation_errors = []

        worksheet = cls.get_sheet(path, organization, permission, failed_url, task_name)
        if worksheet is None:
            cache.set(f'{task_name}_import_status_{organization.id}', 'failed', None)
            cache.set(f'failed_{task_name}_import_{organization.id}', None, None)
            return
        total_data = 0
        imported_data = 0
        _data = cls.extract_data_from_worksheet(
            worksheet,
            model_fields_map,
            permission,
            failed_url,
            organization,
            task_name
        )
        for data in _data:
            if non_mandatory_field_value:
                non_mandatory_field_value_lower = {
                    k.lower().replace(" ", "_"): v for k,v in non_mandatory_field_value.items()
                }
                none_fields = [key for key, value in data.items() if value is None]
                update_data = {
                    key: value for key, value in non_mandatory_field_value_lower.items()
                    if key in none_fields
                }
                data.update(update_data)

            if many_to_many_fields:
                data = cls.get_updated_data(data, many_to_many_fields, separator)

            serializer = serializer_class(
                data=cls.get_data_for_serializer(data, serializer_class),
                context=context
            )
            if serializer.is_valid():
                serializer.save()
                imported_data += 1
            else:
                validation_errors.append(
                    {
                        **data,
                        'errors': cls.stringify_errors(
                            serializer,
                            inverse_mapping(model_fields_map)
                        )
                    }
                )
            total_data += 1

        filename = None
        if validation_errors:
            filename = cls.generate_failed_exports(validation_errors, task_name,
                                                   queryset_fields_map, serializer_field_map)
        if len(validation_errors) > 0:
            if imported_data > 0:
                message_data = {
                    "message": f"{task_name.replace('_', ' ').title()} Import Failed."
                               f" Imported {imported_data} "
                               f"out of {total_data} {task_name.replace('_', ' ')}.",
                    "imported": "partial",
                    "filename": settings.BACKEND_URL + settings.MEDIA_URL + filename,
                    "url": failed_url
                }
            else:
                message_data = {
                    "message": f"Could not import any of {total_data} {task_name.replace('_', ' ')}.",
                    "imported": "partial",
                    "filename": settings.BACKEND_URL + settings.MEDIA_URL + filename,
                    "url": failed_url
                }
        else:
            if imported_data == 0 and total_data == 0:
                message_data = {
                    "message": f"{task_name.replace('_', ' ').title()} Import Failed."
                               f" No {task_name.replace('_', ' ')} were found in the file.",
                    "imported": "failed",
                    "filename": filename,
                    "url": failed_url
                }
            else:
                message_data = {
                    "message": f"{task_name.replace('_', ' ').title()} Import complete."
                               f" Successfully imported {total_data} {task_name.replace('_', ' ')}.",
                    "imported": "success",
                    "filename": None,
                    "url": success_url
                }

        cache.set(f'{task_name}_import_status_{organization.id}', message_data['imported'], None)
        cache.set(f'failed_{task_name}_import_file_{organization.id}', message_data['filename'],
                  None)
        notify_organization(
            text=message_data['message'],
            action=organization,
            organization=organization,
            url=message_data["url"],
            permissions=permission
        )

    @classmethod
    def get_data_for_serializer(cls, data, serializer_class):
        serializer_class_fields = serializer_class._declared_fields
        for key, value in data.items():
            if isinstance(serializer_class_fields.get(key), serializers.DateField) and \
               isinstance(value, timezone.datetime):
                data[key] = value.date()

        nested_data = dict()
        nested_serializer_fields = getattr(cls, 'nested_serializer_fields', {})
        if nested_serializer_fields:
            _data = data.copy()
            for key, value in nested_serializer_fields.items():
                nested_data.update({
                    key: {field: _data.pop(field, None) for field in value}
                })
            nested_data.update(_data)
            return nested_data
        return data

    @classmethod
    def extract_data_from_worksheet(cls, worksheet, model_fields_map, permission,
                                    failed_url, organization, task_name):
        headers = None
        for index, row in enumerate(worksheet):
            row_data = [cell.value for cell in row]
            if not any(row_data):
                break
            if index == 0:
                headers = [model_fields_map.get(header) for header in row_data]
                cls.validate_header(
                    row_data, permission,
                    failed_url, organization, task_name
                )
            else:
                yield dict(
                    zip(
                        headers,
                        row_data
                    )
                )

    @classmethod
    def get_sheet(cls, file, organization, permission, failed_url, task_name):
        """
        :param file: File opened in binary read mode
        :type file: File
        :param organization: Organization Instance
        :type organization: Organization
        :param permission: Permissions to view notification
        :type permission: list
        :param failed_url: frontend url to be redirect when file import fails
        :type failed_url: str
        :param task_name:
        """
        try:
            workbook = load_workbook(filename=file, read_only=True)
        except (OSError, BadZipFile):
            notify_organization(
                text=f'{task_name} Import Failed. Corrupt file was sent.',
                action=organization,
                organization=organization,
                permissions=permission,
                url=failed_url
            )
            return None

        if len(workbook.sheetnames) > 0:
            sheet_name = workbook.sheetnames[0]
        else:
            notify_organization(
                text=f'{task_name} Import Failed. File had no data in it.',
                action=organization,
                organization=organization,
                url=failed_url,
                permissions=permission
            )
            return None

        return workbook[sheet_name]

    @classmethod
    def stringify_errors(cls, serializer, field_name_map):
        """Build string from dictionary of errors"""
        error_dict = serializer.errors
        error_str = ''
        for key, val in error_dict.items():
            if isinstance(val, dict):
                for k, v in val.items():
                    error_str += f"{field_name_map.get(k, k)}: {','.join(v) if isinstance(v, list) else v}\n"
            else:
                error_str += f"{field_name_map.get(key, key)}: {','.join(val) if isinstance(val, list) else val}\n"
        return error_str.strip()

    @classmethod
    def generate_failed_exports(cls, failed, task_name, queryset_fields_map, serializer_field_map):
        fields = cls.import_fields + ['errors']
        wb = Workbook()
        ws = wb.active
        ws.append(fields)

        cls.add_validators(wb, ws, queryset_fields_map, serializer_field_map)
        for fail in failed:
            row = list(fail.values())
            row = list(
                map(
                    lambda x: ', '.join(x) if isinstance(x, list) else x,
                    row
                )
            )
            ws.append(row)

        filename = f"failed_exports/{task_name}/{uuid.uuid4().hex}.xlsx"
        save_workbook(wb, filename)

        return filename

    @classmethod
    def validate_header(cls, headers, permission, failed_url, organization, task_name):
        import_fields = set(cls.import_fields)
        if not import_fields.issubset(set(headers)):
            # notify_organization(
            #     text=f'Could not import {task_name.replace("-", " ")}.',
            #     action=organization,
            #     organization=organization,
            #     url=failed_url,fault
            #     permissions=permission
            # )
            cache.set(f'{task_name}_import_status_{organization.id}', 'failed', None)
            cache.set(f'failed_{task_name}_import_{organization.id}', None, None)


class NormalFileImportMixin(NormalFileImportBaseMixin, SampleDownloadForImportMixin):
    """
        This mixin is used for generating sample files for user as reference
        as well as helps to import data from file to database.

        For proper functioning of this mixin, we need to specify import_fields
        and model_fields_map as class variables. Here import_fields is mandatory
        where as model_fields_map is optional. We need to specify these fields
        within class where we extend this class.

        Example:
        self.import_fields = [
            'NAME',
            'DESCRIPTION'
        ]

        self.model_fields_map = {
            'NAME': 'name',
            'DESCRIPTION': 'description'
        }

        model_fields_map is dict containing mapping between import_fields and
        model_fields. Key for model_fields_map dict represent import_fields where as
        Value for model_fields_map dict represents model_field for that module.

        Final output file as a sample.
        :```````````````````:````````````````````````:
        : NAME              : DESCRIPTION           :
        :```````````````````:``````````````````````:
        :..................:.......................:

        """

    @action(
        detail=False,
        methods=['POST'],
        url_path='import',
        url_name='import',
        serializer_class=FileImportSerializer
    )
    def import_file(self, request, *args, **kwargs):
        file = None
        file_serializer = FileImportSerializer(data=request.data)
        if file_serializer.is_valid(raise_exception=True):
            file = file_serializer.validated_data['file']
        data = self.start_import(file)
        serializer = self.get_import_serializer_class()(
            data=data,
            many=True,
            context=self.get_serializer_context()
        )
        if serializer.is_valid():
            serializer.save()
            return Response({'detail': f'Successfully imported data from file.'})
        return Response(self.reformat_errors(serializer.errors), 400)


class BackgroundFileImportMixin(BackgroundFileImportBaseMixin, SampleDownloadForImportMixin):
    """
        This mixin is used for generating sample files for user as reference
        as well as helps to import data from file to database asynchronously.
        When we upload file for imports, corresponding file will be sent to
        background and get processed. When file gets successfully imports or
        gets failed, user gets notification about the status.

        For proper functioning of this mixin, we need to specify import_fields,
        model_fields_map, values, background_task_name, sample_file_name,
        non_mandatory_field_value, nested_serializer_fields as class variables.
        Here import_fields, failed_url and success_url are mandatory where as
        other fields are optional. We need to specify these fields within class
        where we extend this class.

        For proper explanation, please check how this mixin has been implemented
        for OrganizationHoliday for file import.

        * Note: If there are any many to many field that need to be exported it need's
        to be specified on many_to_many_fields. And on non_mandatory_field_value
        you must provide empty as default value.

        :cvar import_fields:
            This field has all the fields needed to be export
            :type import_fields: list
            example:
            import_fields = [
                'NAME',
                'CATEGORY',
                'START DATE',
                'END DATE',
                'GENDER',
                'DIVISION',
                'RELIGION',
                'ETHNICITY',
                'LOWER_AGE',
                'UPPER_AGE',
                'DESCRIPTION'
            ]
        :cvar values:
            This field holds sample value for sample file. This field in non mandatory field
            :type values: list
            example:
            values = [
                'Dashain',
                'Public',
                '2019-01-01',
                '2019-01-01',
                'All',
                'IT',
                'Hinduism',
                'Tamang',
                16,
                99,
                'Holiday_description'
            ]
        :cvar model_fields_map: Non Mandatory field
            This field holds map between import_fields and model fields
            :type model_fields_map: dict
            example:
                model_fields_map = {
                    'NAME': 'name',
                    'CATEGORY': 'category',
                    'START DATE': 'start_date',
                    'END DATE': 'end_date',
                    'GENDER': 'gender',
                    'DIVISION': 'division',
                    'RELIGION': 'religion',
                    'ETHNICITY': 'ethnicity',
                    'LOWER_AGE': 'lower_age',
                    'UPPER_AGE': 'upper_age',
                    'DESCRIPTION': 'description',
                }

        :cvar failed_url: (Mandatory Field)
            This field is used within notification to redirect user in frontend side.
            :type failed_url: str
            example:
                failed_url='/admin/org-slug/organization/settings/holiday/?status=failed'
        :cvar success_url: (Mandatory Field)
            This field is used within notification to redirect user in frontend side.
            :type failed_url: str
            example:
                failed_url='/admin/org-slug/organization/settings/holiday/?status=success'
        :cvar background_task_name: (Non mandatory field)
            This fields holds data of task name that is run in background.
            :type background_task_name: str
            example:
                background_task_name = 'holiday'
        :cvar sample_file_name: (Non mandatory field)
            This field value is used for naming sample file used as import
            :type sample_file_name: str
            example:
                sample_file_name = 'holiday'
        :cvar non_mandatory_field_value: (Non mandatory field)
            This field is used to specify default value for non mandatory fields of the model.
            :type non_mandatory_field_value: dict
            example:
                non_mandatory_field_value = {
                    'description': '',
                    'gender': 'All',
                    'lower_age': 16,
                    'upper_age': 99
                }
        :cvar nested_serializer_fields: (Non Mandatory field)
            This field is used when ever there is nested serializer used within
            :type nested_serializer_fields: dict
            example:
                 nested_serializer_fields = {
                    'rule': [
                        'gender', 'division', 'religion', 'ethnicity',
                        'category', 'lower_age', 'upper_age'
                    ]
                }
        :cvar import_serializer_Class: (Non Mandatory Field)
            Serializer class used to validate and save imported data
            Defaults to view.serializer_class

        :cvar slug_field_for_sample: (Non Mandatory Field)
            While generating datavalidator (choices field) which field to display
            in choices, It must be unique. Defaults to `slug`
        """

    @property
    def permissions_description_for_notification(self):
        permissions_description_for_notification = []
        for permission in self.permission_classes:
            if hasattr(permission, 'description'):
                for permission_description in permission.description.values():
                    if isinstance(permission_description, list):
                        permissions_description_for_notification += permission_description
        return permissions_description_for_notification

    @action(
        detail=False,
        methods=['POST'],
        url_path='import',
        serializer_class=FileImportSerializer
    )
    def import_file(self, request, *args, **kwargs):
        organization = getattr(self, 'organization', None) or request.user.detail.organization
        task_name = self.get_background_task_name()
        if cache.get(f'block_{task_name}_import_{organization.id}', False):
            status = cache.get(f'{task_name}_import_status_{organization.id}', 'processing')
            raise ValidationError(
                {"non_field_errors": f"Could not start import. Previous import is {status}"})

        # save uploaded file to memory and process it
        serializer = FileImportSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            file = serializer.validated_data['file']
        # path = self._save_file(file)
        context = self.get_serializer_context()
        context['request'] = DummyObject(
            method="POST",
            user=getattr(
                request,
                'user',
                AnonymousUser
            )
        )
        context["view"] = DummyObject()
        cache.set(f'{task_name}_import_status_{organization.id}', 'queued', None)
        cache.set(f'block_{task_name}_import_{organization.id}', True, None)
        async_task(
            self.start_import,
            self.__class__,
            file,
            organization,
            context,
            self.get_import_serializer_class(),
            self.permissions_description_for_notification,
            self.get_failed_url(),
            self.get_success_url(),
            self.model_fields_map,
            task_name,
            self.get_queryset_fields_map(),
            getattr(self, 'non_mandatory_field_value', None),
            self.serializer_field_map,
            self.get_many_to_many_fields()
        )
        return Response(
            dict(
                imported="success",
                message="File sent for background processing."
            )
        )

    @import_file.mapping.get
    def import_file_status(self, request, *args, **kwargs):
        organization = getattr(self, 'organization', None) or request.user.detail.organization
        task_name = self.get_background_task_name()

        lock = cache.get(f'block_{task_name}_import_{organization.id}', False)
        status = cache.get(f'{task_name}_import_status_{organization.id}',
                           'processing' if lock else 'successful')
        failed_import_file = cache.get(f'failed_{task_name}_import_file_{organization.id}', None)
        return Response({
            'lock': lock,
            'status': status,
            'failed_import_file': failed_import_file
        })
