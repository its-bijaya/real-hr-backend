"""@irhrs_docs"""
from io import BytesIO
from django.core.files.storage import default_storage
from openpyxl import load_workbook as xl_load_workbook

from irhrs.export.constants import ADMIN, EXPORTED_AS_CHOICES, QUEUED, PROCESSING, COMPLETED
from ..models import Export


def validate_export_kwargs(func):
    def function(*args, **kwargs):
        exported_as = kwargs.get('exported_as')
        user = kwargs.get('user')
        assert exported_as == ADMIN or user, "`User` must be set if `exported_as` is not Admin"
        assert not exported_as or (
            exported_as and exported_as in EXPORTED_AS_CHOICES, "Exported as must be either `Admin` or `Normal User` ")
        return func(*args, **kwargs)
    return function


def get_user_exported_as_filter(user, exported_as, organization):
    fil = dict()
    if user:
        fil.update({'user': user})
    if exported_as:
        fil.update({'exported_as': exported_as})
    if organization:
        fil.update({'organization': organization})
    return fil


@validate_export_kwargs
def has_pending_export(export_type, user=None, exported_as=None, organization=None):
    """check whether user has pending request"""
    fil = dict(export_type=export_type, status__in=[QUEUED, PROCESSING])
    fil.update(**get_user_exported_as_filter(user, exported_as, organization))
    return Export.objects.filter(**fil).only('id').exists()


@validate_export_kwargs
def get_latest_export(export_type, user=None, exported_as=None, organization=None):
    fil = dict(export_type=export_type, status=COMPLETED)
    fil.update(**get_user_exported_as_filter(user, exported_as, organization))
    return Export.objects.filter(**fil).order_by('-created_at').first()


def save_workbook(workbook, filename):
    """
    Save workbook using default storage.

    Note: Use this instead of workbook.save()

    :param workbook: Workbook to save
    :param filename: filename (don't include media path)
    :return: path saved
    """
    bytes_stream = BytesIO()
    workbook.save(bytes_stream)
    return default_storage.save(filename, bytes_stream)


def load_workbook(filename, mode='rb', *args, **kwargs):
    """
        Load workbook using default storage.

        Note: Use this instead of load_workbook by openpyxl

        :param mode: mode to open file
        :param filename: filename (don't include media path)
        :return: path saved
        """
    input_file = default_storage.open(filename, mode)
    return xl_load_workbook(filename=BytesIO(input_file.read()), *args, **kwargs)


def save_virtual_workbook(workbook):
    """
    use this save_virtual_workbook method instead of openpyxl's
    openpyxl's save_virtual_workbook has been depracated.
    """
    from tempfile import NamedTemporaryFile

    with NamedTemporaryFile() as tf:
        workbook.save(tf.name)
        from io import BytesIO

        in_memory = BytesIO(tf.read())
        return in_memory.getvalue()

