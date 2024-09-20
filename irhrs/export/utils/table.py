"""@irhrs_docs

Table Util
==========

It used openpyxl to create nested tables

This  util has three classes `Column`, `Row` and `Cell` that composes a table.


Eg.

.. code-block :: python

    #  |  Address                    |
    #  |*****************************|
    #  | Street | City | Coordinates |
    #  |        |      |*************|
    #  |        |      | lat | long  |
    #  |*****************************|


    # registering columns

    addressCol = Column(name="Address", title="Address", index=1, accept_cells=True)

    # setting a parent sets nested relation
    streetCol = Column(name="Street", title="Address", index=1, parent=addressCol)
    cityCol = Column(name="City", title="City", index=2, parent=addressCol)
    coordinatesCol = Column(name="Coordinates", title="Lat Long Details", index=3, parent=addressCol)

    latCol = Column(name="lat", title="Latitude", index=1, parent=coordinatesCol)
    longCol = Column(name="long", title="Longitude", index=2, parent=coordinatesCol)

    # registering rows
    r1 = Row(index=1)
    r2 = Row(index=2)

    # creating cells
    c1 = Cell({
            "Street": "Shantinagar",
            "City": "Kathmandu",
            "Coordinates": {
                "lat": "123",
                "long": "456"
            }
        },
        row=r1,
        column=addressCol
    )


    # registering two cells in same row and columns makes multiple values so it splits horizontally as of now
    c2 = Cell({
            "Street": "Tinkune",
            "City": "Kathmandu",
            "Coordinates": {
                "lat": "123",
                "long": "456"
            },
        },
        row=r1,
        column=addressCol
    )



    c11 = Cell({
            "Street": "Shantinagar",
            "City": "Kathmandu",
            "Coordinates": {
                "lat": "123",
                "long": "456"
            }
        },
        row=r2,
        column=addressCol
    )



    c12 = Cell({
            "Street": "Tinkune",
            "City": "Kathmandu",
            "Coordinates": {
                "lat": "123",
                "long": "456"
            },
        },
        row=r2,
        column=addressCol
    )

    c13 = Cell({
            "Street": "Tinkune",
            "City": "Kathmandu",
            "Coordinates": {
                "lat": "123",
                "long": "456"
            },

        },
        row=r2,
        column=addressCol
    )

    # create a table
    table = Table()

    # table optionally takes a description argument which is a list of lines describing the table
    # table = Table(description=["This is example table", "Yout can add descriptions here."])
    # each lines will take a row each with column spanned to 6

    table.add_column(addressCol)

    table.add_row(r1)
    table.add_row(r2)

    table.load_workbook()
    table.save_workbook("addresses.xlsx")

"""

import functools
import itertools
from collections import defaultdict
from collections.abc import Mapping
from datetime import datetime as dt, date
from numbers import Number

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill as Fill, Border, Side, Alignment

from irhrs.core.utils import nested_get
from irhrs.export.utils.helpers import save_workbook


class Column:
    """
    Column of the table.

    :ivar title: Title of the column written the excel
    :ivar index: Index of column, used to order the columns, it is independent of spreadsheet column letter
    :ivar name: Name of column, it looks for this name while searching for values in cell,
        If not set it is default as title
    :ivar accept_cells: Flag that determines whether cells registered to this column hold values
    :ivar parent: Parent column of current column
    :ivar children: set of child columns
    :ivar cells: set of cells registered to this column
    """

    def __init__(self, title, index, name=None, accept_cells=False, parent=None):
        """

        :param title: Title of the column written the excel
        :param index: Index of column, used to order the columns, it is independent of spreadsheet column letter
        :param name: Name of column, it looks for this name while searching for values in cell,
            If not set it is default as title
        :param accept_cells: Flag that determines whether cells registered to this column hold values
        :param parent: Parent column of current column
        """
        self.title = title
        self.name = name or self.title
        self.index = index  # order of column with respect to its siblings
        self.parent = parent
        self.accept_cells = accept_cells  # This column can accept cells or not
        self.children = set()

        self.column_number = None  # start column number in actual work sheet

        self.cells = set()

        if self.parent:
            # update parent's children list
            self.parent.add_child(self)

    def __str__(self):
        return self.title

    def __repr__(self):
        return self.__str__()

    def __hash__(self):
        return hash(self.index)

    def __eq__(self, other):
        return self.index == other.index

    def __lt__(self, other):
        return self.index < other.index

    def add_child(self, child):
        """
        Add child column to this column
        :param child: Child column
        :type child: Column

        :return: None
        """
        return self.children.add(child)

    def add_cell(self, cell):
        """
        Add child column to this column
        :param cell: Add a cell to the column
        :type cell: Cell

        :return: None
        """
        assert self.accept_cells, "Can not add cells when accept_cells is set to false"
        cell = self.validate_cell(cell)
        return self.cells.add(cell)

    def validate_cell(self, cell):
        """
        Validate whether cell is allowed to be added in the column
        :param cell: Cell to be validated
        :type cell: Cell
        :return: None
        """
        if self.children and cell.value:
            assert isinstance(cell.value, Mapping), f'Each data must be dictionary, it is {type(cell.value)}'
        return cell

    @property
    def fields(self):
        """
        Expected field for this column, includes children as well

        :return: dictionary of fields

        Eg. {
            "name": "ParentName",
            "fields": [
                { "name": "ChildName1", "fields": None }
                { "name": "ChildName2", "fields": None }
            ]
        }
        """
        f = dict(name=self.name)
        if not self.children:
            f.update({'fields': None})
        else:
            f.update({'fields': [child.fields for child in self.children]})
        return f

    @property
    def width(self):
        """
        Width of column, including its children
        """

        max_val = 0

        for child in self.children:
            max_val += child.width

        return max(max_val, 1)

    @property
    def count(self):
        """
        Maximum number of multiple children recorded in same row and column; if none defaults to 1
        """
        sorted_cells = sorted(self.cells, key=lambda i: i.row)
        grouped_cells = itertools.groupby(sorted_cells, lambda i: i.row)

        max_val = functools.reduce(
            lambda max_v, grouped_dict: max(max_v, len(list(grouped_dict[1]))),
            itertools.chain([1], grouped_cells)
        )

        return max(max_val, 1)


class Row:
    """
    Row of the table
    """
    def __init__(self, index):
        self.index = index
        self.cells = set()

    def __hash__(self):
        return hash(self.index)

    def __eq__(self, other):
        return self.index == other.index

    def __lt__(self, other):
        return self.index < other.index

    def __str__(self):
        return str(self.index)

    def __repr__(self):
        return self.__str__()

    def add_cell(self, cell):
        """
        Add child column to this column
        :param cell: Add a cell to the row
        :type cell: Cell

        :return: None
        """
        return self.cells.add(cell)

    @property
    def row_span(self):
        sorted_cells = sorted(self.cells, key=lambda i: i.column)
        column_groups = itertools.groupby(sorted_cells, lambda i: i.column)

        return functools.reduce(
            lambda max_v, grouped_dict: max(max_v, len(list(grouped_dict[1]))),
            itertools.chain([1], column_groups))


class Cell:
    """
    Cell of the table

    :ivar row: Row the cell belongs to
    :ivar column: Column that the cell belongs to
    """
    row = None
    column = None

    def __init__(self, value, row=None, column=None):
        self.value = value
        self.set_row(row)
        self.set_column(column)

    def set_row(self, row):
        """
        Set row
        :param row: row of cell
        :type row: Row
        :return: None
        """
        self.row = row
        if isinstance(row, Row):
            row.add_cell(self)

    def set_column(self, column):
        """
        Set column
        :param column: row of cell
        :type column: Row
        :return: None
        """
        self.column = column

        if isinstance(column, Column):
            column.add_cell(self)

    def __str__(self):
        return str(self.value)

    def __repr__(self):
        return self.__str__()


class Table:
    """
    Table class

    :ivar columns: Set of columns of the table
    :ivar rows: Set of rows of the table
    :ivar wb: workbook associated with the table
        filled when `load_workbook` is called.
    """

    ATTRIBUTE_SEPARATOR = '--'  # dot is not used here because it conflicts with defaults used by serializer source

    def __init__(self, description=None, organization=None):
        self.columns = set()  # only top level columns
        self.rows = set()
        self.wb = None
        self.organization = organization

        # a list of description lines, first line will be bold and other will be in normal font
        if description:
            assert isinstance(description, (list, tuple))
        self.description = description

    def add_column(self, col):
        return self.columns.add(col)

    def add_row(self, row):
        assert isinstance(row, Row), "This must be a row"
        return self.rows.add(row)

    def load_workbook(self):
        """Fill data in workbook"""
        wb = Workbook()
        ws = wb.active
        lines_used = self.write_description(ws)
        row_offset = self.write_columns(ws, lines_used=(lines_used + 1)if lines_used > 0 else lines_used)
        self.freeze_headers(ws, row_offset)
        self.write_cells(ws, row_offset)
        self.wb = wb

    def save_workbook(self, name):
        assert self.wb is not None, "You must call `load_workbook` before calling save_workbook"
        return save_workbook(workbook=self.wb, filename=name)

    def write_description(self, ws):
        """Write description of table"""

        lines_used = 0

        if self.organization:
            from irhrs.export.utils.export import ExportBase
            lines_used = ExportBase.insert_org_info(ws, self.organization, lines_used)

        if self.description:
            for line in self.description:
                lines_used += 1
                ws.merge_cells(start_row=lines_used, start_column=1, end_row=lines_used, end_column=7)
                cell = ws.cell(row=lines_used, column=1, value=line)
                if lines_used == 1:
                    # bold the first line
                    bold_font = Font(bold=True)
                    cell.font = bold_font

        return lines_used

    def write_columns(self, ws, lines_used=0):
        """Write column headers"""
        colors = ['CFD8DC', '90A4AE', '78909C']

        def set_headers(columns, headings, depth=0, offset=0, selected_color=None):
            """
            Set offset and width of columns
            """
            fill = itertools.cycle(colors)

            for column in sorted(columns):
                column.column_number = offset
                headings[depth].append((column.title, offset, column.width, selected_color or next(fill)))

                if column.children:
                    set_headers(
                        column.children,
                        headings,
                        depth=depth + 1,
                        offset=offset,
                        selected_color=next(fill)
                    )

                offset += (column.width)

        headings = defaultdict(list)

        set_headers(columns=self.columns, headings=headings, depth=lines_used)

        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        row = 0

        for row_, heading_list in headings.items():

            row = row_ + 1

            for heading in heading_list:
                title, offset, width, color = heading

                offset += 1
                width -= 1

                if width > 0:
                    ws.merge_cells(start_row=row, start_column=offset, end_row=row, end_column=offset + width)

                cell = ws.cell(row=row, column=offset, value=title)
                cell.font = bold_font
                cell.fill = Fill(start_color=color, end_color=color, fill_type='solid')
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        return row

    def write_cells(self, ws, row_offset):
        """Write cells to tha table"""
        def get_field_name(fields, output, prefix=None):
            prefix = prefix + self.ATTRIBUTE_SEPARATOR if prefix else ""
            if fields.get('fields') is None:
                output.append(f"{prefix}fields")
            else:
                for field in fields.get('fields'):
                    get_field_name(field, prefix=f'{prefix}{field.get("name")}', output=output)

        occupied_rows = row_offset
        for row in sorted(self.rows):

            sorted_cells = sorted(row.cells, key=lambda i: i.column)
            column_groups = itertools.groupby(sorted_cells, lambda i: i.column)

            for column, cells_ in column_groups:
                fields = column.fields
                cells_, cells = itertools.tee(cells_, 2)
                len_cells = len(list(cells_))

                start_index = occupied_rows
                field_names = []
                get_field_name(fields, field_names)

                for row_index, cell in enumerate(cells, start=1):

                    for col_index, field_name in enumerate(field_names, start=1):

                        value = nested_get(cell.value, field_name, separator=self.ATTRIBUTE_SEPARATOR)

                        if isinstance(value, list):
                            value = ", ".join(str(item) for item in value)
                        else:
                            if not isinstance(value, (dt, date, Number)):
                                value = str(value) if value is not None else value

                        _cell = ws.cell(row=start_index + row_index,
                                        column=column.column_number + col_index, value=value)
                        _cell.alignment = Alignment(vertical="center")

                        if row_index == len_cells:
                            if row_index == 1:
                                bottom_border = Border(bottom=Side(style='thin'), top=Side(style='thin'))
                            else:
                                bottom_border = Border(bottom=Side(style='thin'))
                            _cell.border = bottom_border
                            if row_index != row.row_span:
                                ws.merge_cells(
                                    start_row=start_index + row_index,
                                    start_column=column.column_number + col_index,
                                    end_row=start_index + row.row_span,
                                    end_column=column.column_number + col_index
                                )
                        elif row_index == 1:
                            top_border = Border(top=Side(style='thin'))
                            _cell.border = top_border
            occupied_rows += row.row_span

    @staticmethod
    def freeze_headers(ws, row_offset):
        ws.freeze_panes = f"A{row_offset + 1}"
