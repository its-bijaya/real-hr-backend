"""@irhrs_docs"""
import itertools
import os
import re
from collections.abc import Mapping
from logging import getLogger

from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.forms.utils import pretty_name
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils import timezone
from fast_map import fast_map_async
from openpyxl import Workbook
from openpyxl.styles import PatternFill as Fill, Font, Alignment, Color, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from xhtml2pdf import pisa

from irhrs.core.utils import nested_getattr, nested_get
from irhrs.core.utils.common import get_today
from irhrs.core.utils.filters import inverse_mapping
from irhrs.export.utils.table import Column, Table, Row, Cell

export_logger = getLogger(__name__)


class ExportBase:
    __val_to_key = None

    @classmethod
    def get_column_head(cls, name, mapping=None):

        if mapping:
            return mapping[name]

        name = name.rsplit('.', 1)[-1]
        return pretty_name(name)

    @classmethod
    def get_column_from_map(cls, mapping=None):
        if not mapping:
            mapping = {}

        value = mapping.get('display')
        col_span = mapping.get('col_span')
        row_span = mapping.get('row_span')
        color = mapping.get('fill')
        child = mapping.get('child')
        return (value, col_span, row_span, color), child

    @staticmethod
    def get_column_cell(obj, name):
        if callable(name):
            attr = name(obj)
        else:
            if isinstance(obj, dict):
                attr = nested_get(obj, name)
            else:
                try:
                    attr = nested_getattr(obj, name)
                except ObjectDoesNotExist:
                    return ''
                if hasattr(attr, '_meta'):
                    return str(attr) if attr else ''
                elif hasattr(attr, 'all'):
                    return ', '.join(str(x) for x in attr.all())

        if isinstance(attr, list):
            return ', '.join(map(str, attr))

        elif type(attr) in [int, float]:
            return attr

        try:
            attr = attr.__str__() if attr is not None else ''
        except Exception as e:
            export_logger.error("Exception in export", exc_info=True)
            return ''
        return attr

    @staticmethod
    def get_column_mapping(columns):
        mapping = None
        if isinstance(columns, dict):
            mapping, columns = inverse_mapping(columns), columns.values()
        return columns, mapping

    @staticmethod
    def insert_org_info(ws, organization, lines_used=0):

        # logo = nested_getattr(organization, 'appearance.logo')
        # if logo and os.path.exists(logo.path):
        #     ws.merge_cells(
        #         start_row=lines_used + 1,
        #         start_column=1, end_row=lines_used + 1,
        #         end_column=7
        #     )
        #     image_obj = Image(logo)
        #     ws.add_image(image_obj, anchor="A1")
        #     dimension = ws.row_dimensions[1]
        #     dimension.height = image_obj.height * 0.75
        #     lines_used += 1

        ws.merge_cells(start_row=lines_used+1, start_column=1, end_row=lines_used+1, end_column=7)
        ws.cell(row=lines_used+1, column=1, value=organization.name)
        ws.merge_cells(start_row=lines_used+2, start_column=1, end_row=lines_used+2, end_column=7)
        ws.cell(row=lines_used+2, column=1,
                value=f"Generated at: ({get_today(with_time=True).strftime('%Y-%m-%d %H:%M:%S')})")
        lines_used += 2
        return lines_used


class ExcelExport(ExportBase):

    @classmethod
    def process(cls, qs, title='Sheet1', columns=None,
                description=None, prepare_export_object=None,
                prepare_export_object_context=None, freeze_first_column=False,
                organization=None):
        """
        :param qs : Iterable of export objects
        :param title :  Title of the worksheet
        :param columns: columns to fetch from the qs ,
                can be iterable of field names or mapping of title and field name

                eg:
                Iterable of fields: ('id','user.name')
                defaults to qs.model._meta.fields

                eg.
                Mapping of title and field name
                {
                    "Name": "user.name"
                }

                set value to #SN to include serial number

        :param description: list of lines describing the export
        :param prepare_export_object: function that will be called before accessing export data
        :param prepare_export_object_context: dict to be unpacked while calling
            prepare_export_object
        :param freeze_first_column: True if freeze first column of export
        :param organization: Organization instance to include organization name and logo
        :return openpyxl , Workbook object

        """

        # if passed custom header to column mapping use
        columns, mapping = cls.get_column_mapping(columns)
        colors = ['CFD8DC', '90A4AE', '78909C']
        color = itertools.cycle(colors)

        prepare_export_object_context = prepare_export_object_context or dict()

        wb = Workbook()
        ws = wb.active
        ws.title = title
        lines_used = 0

        if organization:
            lines_used = cls.insert_org_info(ws, organization, lines_used)

        if description and isinstance(description, (list, tuple)):
            for line in description:
                lines_used += 1
                ws.merge_cells(start_row=lines_used, start_column=1, end_row=lines_used, end_column=7)
                cell = ws.cell(row=lines_used, column=1, value=line)
                if lines_used == 1:
                    # bold the first line
                    bold_font = Font(bold=True)
                    cell.font = bold_font

        columns = columns or [i.__str__().split('.')[-1] for i in qs.model._meta.fields]

        lines_used = lines_used + 2 if lines_used != 0 else 1

        cls.fill_headings(columns, lines_used, ws, mapping, color)

        ws.freeze_panes = f"{'B' if freeze_first_column else 'A'}{lines_used+1}"

        cls.fill_data(columns, lines_used + 1, qs, ws,
                      prepare_export_object, prepare_export_object_context)

        return wb

    @classmethod
    def fill_data(cls, columns, row_offset, qs, ws,
                  prepare_export_object=None,
                  prepare_export_object_context=None):
        """
        Fill data rows in given worksheet
        :param columns: list field names to extract data from object, eg. ('id','user.name')
        :param row_offset: row offset to start filling data
        :param qs: Iterable of export objects
        :param ws: WorkSheet
        :param prepare_export_object: function that will be called before accessing export data
        :param prepare_export_object_context: dict to be unpacked while calling
            prepare_export_object
        :return:
            It returns nothing. It writes into ws
        """

        prepare_export_object_context = prepare_export_object_context or dict()
        for row, obj in enumerate(qs, start=row_offset):
            if prepare_export_object is not None and callable(prepare_export_object):
                obj = prepare_export_object(obj, **prepare_export_object_context)
            for col, column in enumerate(columns, start=1):
                if column == '#SN':
                    # Insert Serial Number if needed
                    value = (row - row_offset) + 1
                else:
                    value = cls.get_column_cell(obj, column)
                _ = ws.cell(row=row, column=col, value=value)

    @classmethod
    def fill_headings(cls, columns, row_offset, ws, mapping=None, color=None):
        """
        Fill headings in given worksheet
        :param columns: list field names to extract data from object, eg. ('id','user.name')
        :param row_offset: row offset to leave before headers
        :param ws: worksheet instance
        :param mapping: [optional] Mapping of title and field name
            {
                "Name": "user.name"
            }
        :param color: [optional] cycle object of color codes to fill into headers
        :return:
            It returns nothing. It writes into ws
        """

        bold = Font(bold=True)
        center_aligned = Alignment(horizontal='center')

        for x, column in enumerate(columns, start=1):
            value = cls.get_column_head(column, mapping)
            cell = ws.cell(row=row_offset if row_offset > 0 else 1, column=x, value=value)

            cell.font = bold
            cell.alignment = center_aligned

            if color:
                color_ = next(color)
                cell.fill = Fill(start_color=color_, end_color=color_, fill_type='solid')

    @classmethod
    def get_response(cls,  qs, title='Sheet1', columns=None, filename="export.xlsx",
                     organization=None):
        wb = ExcelExport.process(qs=qs, columns=columns, title=title, organization=organization)
        return cls.get_response_for_workbook(wb, filename)

    @staticmethod
    def get_response_for_workbook(wb, filename='export.xlsx'):
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="' + \
                                          filename + '"'
        return response


class PDFExport(ExportBase):
    @staticmethod
    def link_callback(uri, rel):
        """
        Convert HTML URIs to absolute system paths so xhtml2pdf can access those
        resources
        """
        static_url = settings.STATIC_URL
        static_root = settings.STATIC_ROOT
        media_url = settings.MEDIA_URL
        media_root = settings.MEDIA_ROOT

        # Let the xhtmltopdf use static files (e.g. css)
        if uri.startswith(media_url):
            path = os.path.join(media_root, uri.replace(media_url, ""))
        elif uri.startswith(static_url):
            path = os.path.join(static_root, uri.replace(static_url, ""))
        else:
            return uri

        # make sure that file exists
        if not os.path.isfile(path):
            raise Exception(
                'media URI must start with %s or %s' % (static_url, media_url)
            )
        return path

    @classmethod
    def get_response(cls, qs, columns=None, template_src='pdf_template.html',
                     filename="export.pdf"):

        # if passed custom header to column mapping use
        columns, mapping = cls.get_column_mapping(columns)

        columns = columns or [i.__str__().split('.')[-1] for i in qs.model._meta.fields]

        template = get_template(template_src)
        context = {
            "queryset": qs,
            "columns": columns,
        }

        # Create a Django response object, and specify content_type as pdf
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; ' \
                                          'filename="{}"'.format(filename)
        # Render the template.
        html = template.render(context)

        # PDF Creation
        pisa_status = pisa.CreatePDF(
            html, dest=response, link_callback=cls.link_callback)
        if pisa_status.err:
            return HttpResponse(
                'We had some errors <pre>' + html + '</pre>')
        return response


class TableExport(ExportBase):
    """
    Sample export format

    .. code-block:: python

        {
            {"name": "user", "fields": ("id", "full_name")},
            {"name": "leave_types", "fields": ("name", "count")}
        }

    """

    @staticmethod
    def get_columns(fields, level=0, parent=None):

        columns = dict()
        for index, field in enumerate(fields):
            if isinstance(field, dict):
                title = field.get('title')
                name = field.get('name') or title
                fields = field.get('fields')
            else:
                title = field
                name = title
                fields = None

            allow_cells = level == 0
            col = Column(title=title, name=name, index=index, accept_cells=allow_cells,
                         parent=parent)

            if fields:
                TableExport.get_columns(fields, level=1, parent=col)

            columns.update({col.name: col})

        return columns

    @staticmethod
    def validate_export_format(export_format):
        assert isinstance(format, dict)
        assert export_format.get('name') is not None
        assert export_format.get('fields') is not None
        return export_format

    @classmethod
    def process(cls, data, export_format, description=None, organization=None):
        """
        :param data: data to export
        :param export_format: export fields
            eg.
            {
                {"name": "user", "fields": ("id", "full_name")},
                {"name": "leave_types", "fields": ("name", "count")}
            }
        :param description:
        :param organization: Organization instance to include logo and name
        :return:
        """
        columns = cls.get_columns(export_format)

        table = Table(description=description, organization=organization)

        for col in sorted(columns.values()):
            table.add_column(col)

        for index, value in enumerate(data):
            row = Row(index=index)

            assert isinstance(value, Mapping), f'Each data must be dictionary, it is {type(value)}'

            for k, col in columns.items():
                # v = value.get(k, {} if col.children else None)
                v = nested_get(value, k)
                if v is None and col.children:
                    v = {}
                if not isinstance(v, list):
                    v = [v]

                for item in v:
                    Cell(value=item, column=col, row=row)

            table.add_row(row)

        table.load_workbook()
        return table.wb


class PayrollExcelExport(ExportBase):

    @classmethod
    def skip_column(cls, heading_name, heading_map):
        """
        :param heading_name: Basic Salary
        :param heading_map:  Base -> Basic
        :param mapping: Basic Salary -> heading_amounts.basic Salary
        :return:
        """
        if not heading_map:
            return False
        if heading_name not in heading_map:
            return True
        return False

    @classmethod
    def process(cls, qs, title='Sheet1', columns=None,
                description=None, prepare_export_object=None,
                prepare_export_object_context=None, freeze_first_column=False,
                heading_map=None, footer_data=None):
        """
        :param qs : Iterable of export objects
        :param title :  Title of the worksheet
        :param columns: columns to fetch from the qs ,
                can be iterable of field names or mapping of title and field name

                eg:
                Iterable of fields: ('id','user.name')
                defaults to qs.model._meta.fields

                eg.
                Mapping of title and field name
                {
                    "Name": "user.name"
                }

                set value to #SN to include serial number

        :param description: list of lines describing the export
        :param prepare_export_object: function that will be called before accessing export data
        :param prepare_export_object_context: dict to be unpacked while calling
            prepare_export_object
        :param freeze_first_column: True if freeze first column of export
        :param heading_map: Dictionary of Display Name against Pretty Display Name
        :param footer_data: Dictionary of any but 'inline' keyword with boolean value is suggested
        :return openpyxl , Workbook object

        """

        # if passed custom header to column mapping use
        columns, mapping = cls.get_column_mapping(columns)
        colors = ['CFD8DC', '90A4AE', '78909C']
        color = itertools.cycle(colors)
        grand_total_data = prepare_export_object_context.pop('grand_total_data', None)
        prepare_export_object_context = prepare_export_object_context or dict()

        wb = Workbook()
        ws = wb.active
        ws.title = title.replace('Unnamed', str(timezone.now().date()))
        lines_used = 0

        if description and isinstance(description, (list, tuple)):
            for line in description:
                lines_used += 1
                _heading_map = getattr(settings, 'PAYROLL_HEADING_MAP', None)
                end_column = len(_heading_map) if heading_map else 7
                ws.merge_cells(start_row=lines_used, start_column=1, end_row=lines_used,
                               end_column=end_column)
                cell = ws.cell(row=lines_used, column=1, value=line)
                center_aligned = Alignment(horizontal='center')
                cell.alignment = center_aligned
                if lines_used == 1:
                    # bold the first line
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        columns = columns or [i.__str__().split('.')[-1] for i in qs.model._meta.fields]
        lines_used = lines_used + 2 if lines_used != 0 else 1
        actual_lines_used = lines_used
        if hasattr(settings, 'TAX_EXEMPT_FILTERS'):
            for index, exempt_filter in enumerate(getattr(settings, 'TAX_EXEMPT_FILTERS')):
                queryset = qs.filter(exempt_filter)
                if freeze_first_column:
                    freeze_first_column = (index == 0)

                if queryset:
                    cls.fill_tax_exempt_data(
                        color, columns, freeze_first_column, heading_map, actual_lines_used,
                        mapping, prepare_export_object, prepare_export_object_context,
                        queryset, ws
                    )
                    actual_lines_used += queryset.count() + 3
        else:
            cls.fill_tax_exempt_data(
                color, columns, freeze_first_column, heading_map,
                lines_used, mapping, prepare_export_object, prepare_export_object_context, qs, ws
            )
            actual_lines_used += qs.count() + 3

        lines_used = lines_used + 2
        if grand_total_data:
            cls.fill_grand_total(
                columns, lines_used, actual_lines_used - 1, qs, ws,
                grand_total_data, heading_map, mapping
            )
            actual_lines_used += 4

        cls.fill_footer(2, actual_lines_used, ws, footer_data)

        return wb

    @classmethod
    def fill_tax_exempt_data(cls, color, columns, freeze_first_column, heading_map,
                             lines_used, mapping, prepare_export_object,
                             prepare_export_object_context, qs, ws):
        row_offset = cls.fill_headings(columns, lines_used, ws, mapping, color, heading_map)
        if freeze_first_column:
            ws.freeze_panes = f"{'B' if freeze_first_column else 'A'}{lines_used + 1}"
        cls.fill_data(columns, row_offset + 2, qs, ws,
                      prepare_export_object, prepare_export_object_context, heading_map, mapping)

    @classmethod
    def fill_footer(cls, column_offset, lines_used, ws, footer_data):
        row_count = lines_used
        for index, footer_datum in enumerate(footer_data):
            inline = footer_datum.pop('inline', False)

            if index != 0:
                if inline:
                    column_offset += 3
                else:
                    row_count += 2

            footer_text = footer_datum.get('footer_text')
            if not isinstance(footer_text, (list, tuple)):
                footer_text = [footer_text]

            for row_offset, text_item in enumerate(footer_text):
                ws.merge_cells(
                    start_row=row_count+row_offset, end_row=row_count + row_offset,
                    start_column=column_offset, end_column=column_offset + 1
                )
                cell = ws.cell(row=row_count+row_offset, column=column_offset, value=text_item)
                center_aligned = Alignment(horizontal='center')
                bold = Font(bold=True)
                cell.alignment = center_aligned
                cell.font = bold

    @classmethod
    def fill_grand_total(
        cls, columns, data_begin, row_begin,
        queryset, worksheet, footer_data, heading_map, mapping
    ):
        if not footer_data:
            return
        footer_text = footer_data.get('footer_text')
        footer_offset = footer_data.get('footer_offset')
        footer_color = footer_data.get('footer_color')
        worksheet.merge_cells(
            start_row=row_begin + 1, end_row=row_begin + 1,
            start_column=1, end_column=footer_offset - 1
        )
        cell = worksheet.cell(row=row_begin + 1, column=1, value=footer_text)
        footer_fill = Fill(start_color=footer_color, end_color=footer_color, fill_type='solid')
        center_aligned = Alignment(horizontal='center')
        bold = Font(bold=True)
        cell.fill = footer_fill
        cell.alignment = center_aligned
        cell.font = bold
        visible_columns = list(columns)[(footer_offset - 1):]
        if heading_map:
            non_skip_headings = list()
            for column_attr in visible_columns:
                skip = cls.skip_column(mapping.get(column_attr), heading_map)
                if skip:
                    continue
                non_skip_headings.append(column_attr)
        else:
            non_skip_headings = visible_columns

        for index, column in enumerate(non_skip_headings, start=footer_offset):
            queryset_count = queryset.count()
            addition_formula = "=SUM({}:{})".format(
                f"{get_column_letter(index)}{data_begin}",
                f"{get_column_letter(index)}{data_begin + queryset_count}",
            )
            cell = worksheet.cell(
                row=row_begin + 1,
                column=index,
                value=addition_formula,
            )
            cell.fill = footer_fill
            cell.font = bold
            if mapping and heading_map and mapping.get(column) in heading_map:
                number_format = heading_map.get(mapping.get(column)).get('number_format', None)
                color_scheme = heading_map.get(mapping.get(column)).get('fill', None)
                if number_format:
                    cell.number_format = number_format
                if color_scheme:
                    custom_fill = Fill(
                        start_color=color_scheme,
                        end_color=color_scheme,
                        fill_type='solid'
                    )
                    cell.fill = custom_fill

    @classmethod
    def append_data_in_worksheet(cls, columns, ws, row, row_offset, obj, heading_map=None,
                                 mapping=None):
        if heading_map:
            non_skip_headings = list()
            for column_attr in columns:
                is_parent = heading_map.get(
                    mapping.get(column_attr),
                    {}
                ).get('child')
                skip = cls.skip_column(mapping.get(column_attr), heading_map) and not is_parent
                if skip:
                    continue
                non_skip_headings.append(column_attr)
        else:
            non_skip_headings = columns
        for col, column in enumerate(non_skip_headings, start=1):
            if column == '#SN':
                # Insert Serial Number if needed
                value = (row - row_offset) + 1
            else:
                value = cls.get_column_cell(obj, column)
            cell = ws.cell(row=row, column=col, value=value)
            if mapping and heading_map and mapping.get(column) in heading_map:
                number_format = heading_map.get(mapping.get(column)).get('number_format', None)
                if number_format:
                    cell.number_format = number_format

    @classmethod
    def multiprocessing_fill_data(cls, columns, row_offset, qs, ws,
                                  prepare_export_object=None,
                                  prepare_export_object_context=None, heading_map=None,
                                  mapping=None):

        def _on_result(result):
            r_row = result[0]
            r_row_offset = result[1]
            r_obj = result[2]
            cls.append_data_in_worksheet(columns, ws, r_row, r_row_offset, r_obj, heading_map,
                                         mapping)

        def _on_done():
            print("all done")
            pass

        def _export_process(args):
            _row, _obj, _row_offset = args
            if prepare_export_object is not None and callable(prepare_export_object):
                _obj = prepare_export_object(_obj, **prepare_export_object_context)

            return [_row, _row_offset, _obj]

        qs_list = []
        for row, obj in enumerate(qs, start=row_offset):
            qs_list.append((row, obj, row_offset))
        task = fast_map_async(_export_process, qs_list, on_result=_on_result, on_done=_on_done,
                              threads_limit=settings.Q_CLUSTER_AFFINITY,
                              procs_limit=settings.Q_CLUSTER_AFFINITY)
        task.join()

    @classmethod
    def normal_fill_data(cls, columns, row_offset, qs, ws,
                         prepare_export_object=None,
                         prepare_export_object_context=None, heading_map=None, mapping=None):
        for row, obj in enumerate(qs, start=row_offset):
            if prepare_export_object is not None and callable(prepare_export_object):
                obj = prepare_export_object(obj, **prepare_export_object_context)
            cls.append_data_in_worksheet(columns, ws, row, row_offset, obj, heading_map,
                                         mapping)

    @classmethod
    def fill_data(cls, columns, row_offset, qs, ws,
                  prepare_export_object=None,
                  prepare_export_object_context=None, heading_map=None, mapping=None):
        """
        Fill data rows in given worksheet
        :param columns: list field names to extract data from object, eg. ('id','user.name')
        :param row_offset: row offset to start filling data
        :param qs: Iterable of export objects
        :param ws: WorkSheet
        :param prepare_export_object: function that will be called before accessing export data
        :param prepare_export_object_context: dict to be unpacked while calling
            prepare_export_object
        :param heading_map: For formatting in data.
        :param mapping: To find out heading and its format.
        :return:
            It returns nothing. It writes into ws
        """

        if settings.USE_MULTIPROCESSING:
            cls.multiprocessing_fill_data(columns, row_offset, qs, ws, prepare_export_object,
                                          prepare_export_object_context, heading_map, mapping)
        else:
            cls.normal_fill_data(columns, row_offset, qs, ws, prepare_export_object,
                                          prepare_export_object_context, heading_map, mapping)

    @classmethod
    def fill_headings(cls, columns, row_offset, ws, mapping=None, color=None, heading_map=None):
        """
        Fill headings in given worksheet
        :param columns: list field names to extract data from object, eg. ('id','user.name')
        :param row_offset: row offset to leave before headers
        :param ws: worksheet instance
        :param mapping: [optional] Mapping of title and field name
            {
                "Name": "user.name"
            }
        :param color: [optional] cycle object of color codes to fill into headers
        :param heading_map: [optional] for custom heading
        :return:
            It returns nothing. It writes into ws
        """

        bold = Font(bold=True)
        center_aligned = Alignment(horizontal='center')

        if heading_map:
            return cls.generate_heading_with_map(bold, center_aligned, heading_map, row_offset, ws,
                                                 mapping)
        else:
            return cls.generate_heading_without_map(bold, center_aligned, color, columns, mapping,
                                                    row_offset, ws)

    @classmethod
    def generate_heading_without_map(cls, bold, center_aligned, color, columns, mapping,
                                     row_offset, ws):
        for x, column in enumerate(columns, start=1):
            value = cls.get_column_head(column, mapping)
            cell = ws.cell(row=row_offset if row_offset > 0 else 1, column=x, value=value)

            cell.font = bold
            cell.alignment = center_aligned

            if color:
                color_ = next(color)
                cell.fill = Fill(start_color=color_, end_color=color_, fill_type='solid')

        return row_offset

    @classmethod
    def get_response(cls, qs, title='Sheet1', columns=None, filename="export.xlsx"):
        wb = PayrollExcelExport.process(qs=qs, columns=columns, title=title)
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="' + \
                                          filename + '"'
        return response

    @classmethod
    def generate_heading_with_map(cls, bold, center_aligned, mapping, row_offset, ws, column_map):
        def add_heading_to_file(_value, col_span, row_span, color, _row_index, _column_index):
            ws.merge_cells(
                start_row=_row_index, end_row=_row_index + row_span,
                start_column=_column_index, end_column=_column_index + col_span
            )
            cell = ws[get_column_letter(_column_index) + str(_row_index)]
            cell.value = _value
            cell.font = bold
            cell.alignment = center_aligned
            cell.fill = Fill(start_color=Color(color), end_color=Color(color), fill_type='solid')

        additional_index = 0
        no_of_child = 0

        non_skip_headings = list()
        for column_attr in column_map.keys():
            skip = cls.skip_column(column_map.get(column_attr), mapping)
            if skip:
                continue
            non_skip_headings.append(mapping.get(column_map.get(column_attr)))

        for column_index, column in enumerate(non_skip_headings, start=1):
            column_index += additional_index - 1 if additional_index > 0 else 0
            value, child = cls.get_column_from_map(column)
            row_index = row_offset if row_offset > 0 else 1
            add_heading_to_file(*value, row_index, column_index)
            if child:
                no_of_child += 1
                for child_value in child.values():
                    _child_generated_value = cls.get_column_from_map(child_value)[0]
                    add_heading_to_file(*_child_generated_value, row_index + 1, column_index)
                    column_index += 1
                    additional_index += 1
        return row_offset + no_of_child


class PayrollDisbursementExcel(ExcelExport):

    @classmethod
    def process(cls, qs, title='Sheet1', columns=None,
                description=None, prepare_export_object=None,
                prepare_export_object_context=None, freeze_first_column=False,
                organization=None):

        columns, mapping = cls.get_column_mapping(columns)
        colors = ['CFD8DC', '90A4AE', '78909C']
        color = itertools.cycle(colors)
        grand_total_data = prepare_export_object_context.pop('grand_total_data', None)
        report_narration = prepare_export_object_context.pop('report_narration', None)
        prepare_export_object_context = prepare_export_object_context or dict()

        wb = Workbook()
        ws = wb.active
        ws.title = title
        lines_used = 0

        if not report_narration and organization:
            lines_used = cls.insert_org_info(ws, organization, lines_used)

        if description and isinstance(description, (list, tuple)):
            for line in description:
                lines_used += 1
                ws.merge_cells(start_row=lines_used, start_column=1, end_row=lines_used, end_column=len(columns))
                cell = ws.cell(row=lines_used, column=1, value=line)
                if not report_narration and lines_used == 1:
                    # bold the first line
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                if "subject" in line.lower():
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = Fill(start_color="90A4AE", end_color="90A4AE", fill_type='solid')

        columns = columns or [i.__str__().split('.')[-1] for i in qs.model._meta.fields]

        if report_narration:
            lines_used = lines_used + 3 if lines_used != 0 else 2
        else:
            lines_used = lines_used + 2 if lines_used != 0 else 1


        cls.fill_headings(columns, lines_used, ws, mapping, color)

        ws.freeze_panes = f"{'B' if freeze_first_column else 'A'}{lines_used+1}"

        cls.fill_data(columns, lines_used + 1, qs, ws,
                      prepare_export_object, prepare_export_object_context)

        queryset_count = qs.count()
        actual_lines_used = lines_used + queryset_count
        if grand_total_data and queryset_count > 0:
            cls.fill_grand_total(
                columns, lines_used, actual_lines_used, qs, ws,
                grand_total_data, mapping, report_narration
            )
            actual_lines_used += 4
        return wb

    @classmethod
    def fill_grand_total(
        cls, columns, data_begin, row_begin,
        queryset, worksheet, footer_data, mapping, report_narration
    ):
        if not footer_data:
            return
        queryset_count = queryset.count()
        footer_text = footer_data.get('footer_text')
        footer_offset = footer_data.get('footer_offset')
        footer_color = footer_data.get('footer_color')
        worksheet.merge_cells(
            start_row=row_begin + 1, end_row=row_begin + 1,
            start_column=1, end_column=footer_offset - 1
        )
        cell = worksheet.cell(row=row_begin + 1, column=1, value=footer_text)
        footer_fill = Fill(start_color=footer_color, end_color=footer_color, fill_type='solid')
        center_aligned = Alignment(horizontal='center')
        bold = Font(bold=True)
        cell.fill = footer_fill
        cell.alignment = center_aligned
        cell.font = bold
        visible_columns = list(columns)[(footer_offset - 1):]

        for index, column in enumerate(visible_columns, start=footer_offset):
            addition_formula = "=SUM({}:{})".format(
                f"{get_column_letter(index)}{data_begin + 1}",
                f"{get_column_letter(index)}{data_begin + queryset_count}",
            )

            cell = worksheet.cell(
                row=row_begin + 1,
                column=index,
                value=addition_formula,
            )
            cell.fill = footer_fill
            cell.font = bold

        last_column_letter = get_column_letter(len(columns))
        last_cell_name = f"{last_column_letter}{worksheet.max_row}"

        numtowords = (
            '=IF({0}<=0,"zero",TRIM(PROPER(SUBSTITUTE(CONCATENATE(CHOOSE(MID(TEXT(INT({0}),'
            'REPT(0,12)),1,1)+1,"","one hundred ","two hundred ","three hundred ","four hundred ",'
            '"five hundred ","six hundred ","seven hundred ","eight hundred ","nine hundred "),'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),2,1)+1,"",CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),3,1)+1,'
            '"ten","eleven","twelve","thirteen","fourteen","fifteen","sixteen","seventeen",'
            '"eighteen","nineteen"),"twenty","thirty","forty","fifty","sixty","seventy","eighty",'
            '"ninety"),IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),2,1))>1,CHOOSE(MID(TEXT(INT({0}),'
            'REPT(0,12)),3,1)+1,"","-one","-two","-three","-four","-five","-six","-seven","-eight",'
            '"-nine"),IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),2,1))=0,CHOOSE(MID(TEXT(INT({0}),'
            'REPT(0,12)),3,1)+1,"","one","two","three","four","five","six","seven","eight","nine"),'
            '"")),IF({0}>=10^9," billion ",""),CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),4,1)+1,"",'
            '"one hundred ","two hundred ","three hundred ","four hundred ","five hundred ",'
            '"six hundred ","seven hundred ","eight hundred ","nine hundred "),'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),5,1)+1,"",CHOOSE(MID(TEXT(INT({0}),'
            'REPT(0,12)),6,1)+1,"ten","eleven","twelve","thirteen","fourteen","fifteen","sixteen",'
            '"seventeen","eighteen","nineteen"),"twenty","thirty","forty","fifty","sixty",'
            '"seventy","eighty","ninety"),IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),5,1))>1,'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),6,1)+1,"","-one","-two","-three","-four",'
            '"-five","-six","-seven","-eight","-nine"),IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),5,1))=0,'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),6,1)+1,"","one","two","three","four","five",'
            '"six","seven","eight","nine"),"")),IF(VALUE(MID(TEXT(INT({0}),'
            'REPT(0,12)),4,3))>0," million ",""),CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),7,1)+1,"",'
            '"one hundred ","two hundred ","three hundred ","four hundred ","five hundred ",'
            '"six hundred ","seven hundred ","eight hundred ","nine hundred "),'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),8,1)+1,"",CHOOSE(MID(TEXT(INT({0}),'
            'REPT(0,12)),9,1)+1,"ten","eleven","twelve","thirteen","fourteen","fifteen","sixteen",'
            '"seventeen","eighteen","nineteen"),"twenty","thirty","forty","fifty","sixty",'
            '"seventy","eighty","ninety"),IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),8,1))>1,'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),9,1)+1,"","-one","-two","-three","-four",'
            '"-five","-six","-seven","-eight","-nine"),IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),8,1))=0,'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),9,1)+1,"","one","two","three","four","five","six",'
            '"seven","eight","nine"),"")),IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),7,3))," thousand ",""),'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),10,1)+1,"","one hundred ","two hundred ",'
            '"three hundred ","four hundred ","five hundred ","six hundred ","seven hundred ",'
            '"eight hundred ","nine hundred "),CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),11,1)+1,"",'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),12,1)+1,"ten","eleven","twelve","thirteen",'
            '"fourteen","fifteen","sixteen","seventeen","eighteen","nineteen"),"twenty","thirty",'
            '"forty","fifty","sixty","seventy","eighty","ninety"),'
            'IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),11,1))>1,CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),'
            '12,1)+1,"","-one","-two","-three","-four","-five","-six","-seven","-eight","-nine"),'
            'IF(VALUE(MID(TEXT(INT({0}),REPT(0,12)),11,1))=0,'
            'CHOOSE(MID(TEXT(INT({0}),REPT(0,12)),12,1)+1,'
            '"","one","two","three","four","five","six","seven","eight","nine"),""))),"  ",'
            '" ")&IF(FLOOR({0},1)>1," rupees"," ")&IF(AND({0}>=1,{0}<2),"rupee",'
            '""))&IFERROR(IF({0}<1,SUBSTITUTE(IF(ISERROR(FIND(".",{0},1)),""," and "),"and",""),'
            '" and ")&PROPER(IF(LEN(LEFT(TRIM(MID(SUBSTITUTE({0},".",REPT(" ",255)),255,200)),2))=1,'
            'CHOOSE(1*LEFT(TRIM(MID(SUBSTITUTE({0},".",REPT(" ",255)),255,200)),2),"ten","twenty",'
            '"thirty","forty","fifty","sixty","seventy","eighty","ninety")&" paisa",'
            '"")&CONCATENATE(CHOOSE(MID(TEXT(INT(LEFT(TRIM(MID(SUBSTITUTE({0},".",'
            'REPT(" ",255)),255,200)),2)),REPT(0,12)),11,1)+1,"",'
            'CHOOSE(MID(TEXT(INT(LEFT(TRIM(MID(SUBSTITUTE({0},".",'
            'REPT(" ",255)),255,200)),2)),REPT(0,12)),12,1)+1,"ten","eleven","twelve","thirteen",'
            '"fourteen","fifteen","sixteen","seventeen","eighteen","nineteen")&" paisa","twenty",'
            '"thirty","forty","fifty","sixty","seventy","eighty","ninety"),'
            'IF(VALUE(MID(TEXT(INT(LEFT(TRIM(MID(SUBSTITUTE({0},".",REPT(" ",255)),255,200)),2)),'
            'REPT(0,12)),11,1))>1,CHOOSE(MID(TEXT(INT(LEFT(TRIM(MID(SUBSTITUTE({0},".",'
            'REPT(" ",255)),255,200)),2)),REPT(0,12)),12,1)+1,"","-one","-two","-three","-four",'
            '"-five","-six","-seven","-eight","-nine")&" paisa",IF(LEFT(TRIM(MID(SUBSTITUTE({0},".",'
            'REPT(" ",255)),255,200)),2)="01","one paisa",IF(LEFT(TRIM(MID(SUBSTITUTE({0},".",'
            'REPT(" ",255)),255,200)),1)="0",CHOOSE(MID(TEXT(INT(LEFT(TRIM(MID(SUBSTITUTE({0},".",'
            'REPT(" ",255)),255,200)),2)),REPT(0,12)),12,1)+1,"","one","two","three","four",'
            '"five","six","seven","eight","nine")&" paisa",""))))),"")))').format(
            last_cell_name)

        # Set the new formula in the desired cell
        if report_narration:
            worksheet.merge_cells(start_row=7, end_row=7, start_column=1, end_column=2)
            worksheet['A7'] = "Kindly request you to debit the total amount of NRs."
            worksheet['C7'].value = worksheet[last_cell_name].value
            worksheet['C7'].font = Font(bold=True)
            worksheet['D7'] = "In Words"
            worksheet['E7'] = numtowords
            worksheet['E7'].alignment = Alignment(wrap_text=True)
            worksheet.merge_cells(start_row=8, end_row=8, start_column=1, end_column=5)
            worksheet['A8'] = report_narration
            worksheet['A8'].alignment = Alignment(wrap_text=True)


class SalaryDifferneceExport(ExcelExport):

    @classmethod
    def get_column_head(cls, name, mapping=None):
        return re.sub(r'\b(early_payroll |late_payroll |difference )\b', '',
                      super().get_column_head(name, mapping))

    @staticmethod
    def apply_common_style(ws, cell, value, start_col, heading_setting_len, color=None):
        thick_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.value = value
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        start_cell_column_letter = get_column_letter(start_col)
        start_cell_reference = f"{start_cell_column_letter}{cell.row}"
        end_cell_column_letter = get_column_letter(start_col + heading_setting_len - 1)
        end_cell_reference = f"{end_cell_column_letter}{cell.row}"
        ws.merge_cells(f'{start_cell_reference}:{end_cell_reference}')
        if color:
            color_ = next(color)
            cell.fill = Fill(start_color=color_, end_color=color_, fill_type='solid')

    @classmethod
    def process(cls, qs, title='Sheet1', columns=None, description=None,
                prepare_export_object=None, prepare_export_object_context=None,
                freeze_first_column=False, organization=None):
        salary_month_info = prepare_export_object_context.pop('salary_months_info')
        heading_setting_len = len(prepare_export_object_context.get('heading_setting_ids'))

        wb = super().process(qs, title, columns, description, prepare_export_object,
                             prepare_export_object_context, freeze_first_column, organization)
        ws = wb.active
        colors = ['90A4AE', 'CFD8DC', '78909C']
        color = itertools.cycle(colors)

        if organization:
            ws.insert_rows(4)
            start_row, start_col = 4, 6
        else:
            ws.insert_rows(1)
            start_row, start_col = 1, 6

        for value in salary_month_info:

            cell = ws.cell(row=start_row, column=start_col)
            cls.apply_common_style(ws, cell, value, start_col, heading_setting_len, color)
            start_col += heading_setting_len
        return wb
