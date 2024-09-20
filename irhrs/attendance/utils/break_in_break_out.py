from datetime import timedelta

from django.conf import settings

import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from irhrs.export.constants import  PROCESSING, \
    COMPLETED
from django.core.files.base import ContentFile



def get_pair(qs):
    entries = list(qs.order_by('timestamp'))
    limit = len(entries) if len(entries) % 2 == 0 else len(entries) - 1
    pairs = [(entries[i], entries[i + 1],) for i in
             range(0, limit, 2)]
    # check if the entries have absolute pair
    if len(pairs) * 2 == len(entries):
        return pairs
    return pairs + [
        (entries[-1], None),
    ]


def get_total_lost(all_entries):
    """
    return total lost from time sheet entries
    :param all_entries: time sheet entries qs
    :return: total lost hours
    """
    time_out = 0

    # get distinct dates to iterate over:
    dates = all_entries.order_by().values(
        'timesheet__timesheet_for'
    ).distinct().values_list('timesheet__timesheet_for', flat=True)
    pairs = list()
    for date in dates:
        pair = get_pair(
            all_entries.filter(timesheet__timesheet_for=date, is_deleted=False)
        )
        pairs.extend(pair)
    for out_timestamp, in_timestamp in pairs:
        time_out += (
            (
                in_timestamp.timestamp
                - out_timestamp.timestamp
            ).total_seconds() if in_timestamp else 0)
    return time_out // 60


def get_total_paid_breaks(all_entries):
    """
    return total lost from time sheet entries
    :param all_entries: time sheet entries qs
    :return: total lost hours
    """
    time_out = 0

    # get distinct dates to iterate over:
    dates = all_entries.order_by().values(
        'timesheet__timesheet_for'
    ).distinct().values_list('timesheet__timesheet_for', flat=True)
    pairs = list()
    for date in dates:
        pair = get_pair(
            all_entries.exclude(is_deleted=True).filter(timesheet__timesheet_for=date)
        )
        pairs.extend(pair)
    for out_timestamp, in_timestamp in pairs:
        if out_timestamp.remark_category in settings.UNPAID_BREAK_TYPES:
            continue
        time_out += (
            (
                in_timestamp.timestamp
                - out_timestamp.timestamp
            ).total_seconds() if in_timestamp else 0)
    timeout_in_mins = time_out //60
    return timeout_in_mins

def time_convert(minute):
    if isinstance(minute,float):
        minute = int(minute)
    minute_left = (minute % 60)
    hour = ((minute-minute_left)//60)
    return f"{minute_left} mins" if hour == 0 else f"{hour} hrs {minute_left} mins"

def get_time(minute):
    second = round(minute*60)
    left_sec = second % 60
    minute = (second-left_sec)//60
    left_min = minute % 60
    hour = (minute-left_min)//60
    return "%02d:%02d:%02d"%(hour,left_min,left_sec)

def create_report_workbook(start, end, break_details, qs, export, user, export_name):
    export.status = PROCESSING
    export.save()
    font_bold = openpyxl.styles.Font(bold=True)
    from irhrs.attendance.api.v1.reports.serializers.break_in_break_out import \
    BreakInBreakOutDetailReportSerializer
    break_out_data = BreakInBreakOutDetailReportSerializer(qs, many=True).data
    workbook = Workbook()

    max_row = 1
    max_col = 1
    ws = workbook.active
    c1 = ws.cell(row=max_row, column=1, value="Employee Name")
    c1.font = font_bold
    ws.cell(row=max_row, column=2, value=user.full_name)
    c2 = ws.cell(row=max_row, column=4, value="Username")
    c2.font = font_bold
    ws.cell(row=max_row, column=5, value=user.username)
    col = max_col
    row = max_row + 1
    heading_mapper = {
        'total_lost': 'Total Time Out',
        'total_paid_breaks': 'Total paid break',
        'break_in_out_count': 'No. of times Break Out',
        'total_unpaid_breaks': 'Total unpaid break',
    }

    for key, value in break_details.items():
        if key != 'break_in_out_count':
            value = time_convert(value)
        c = ws.cell(row=row, column=col, value=heading_mapper.get(key))
        ws.cell(row=row, column=col + 1, value=value)

        c.font = font_bold
        col += 3
        if col == 7:
            max_col = 7
            row += 1
            col = 1

    c1 = ws.cell(row=max_row + 1, column=max_col, value="From Date")
    c1.font = font_bold
    ws.cell(row=max_row + 1, column=max_col + 1, value=start)

    c2 = ws.cell(row=max_row + 2, column=max_col, value="To Date")
    c2.font = font_bold
    ws.cell(row=max_row + 2, column=max_col + 1, value=end)

    headers = [
        'Category', 'Date', 'Day', 'Break Out Time', 'Latitude, Longitude', 'Break In Time',
        'Latitude, Longitude', 'Total Time Out','Break In Remarks','Break Out Remarks'
    ]
    ws.append([])
    ws.append(headers)
    max_row = len(ws['A'])
    for cell in ws[max_row]:
        cell.font = font_bold
    for data in break_out_data:
        pairs = data['pairs']
        for pair in pairs:
            ws.append([
                pair['category'],
                pair['date'].strftime("%Y-%m-%d"),
                pair['date'].strftime("%A"),
                pair['break_out'].strftime("%H:%M:%S %p"),
                str(pair['break_out_latitude']) + ", " + str(pair['break_out_longitude']),
                pair['break_in'].strftime("%H:%M:%S %p"),
                str(pair['break_in_latitude']) + ", " + str(pair['break_in_longitude']),
                get_time(pair['lost']),
                pair['break_in_remark'],
                pair['break_out_remark'],
            ])


    export.export_file.save(
        export_name,
        ContentFile(save_virtual_workbook(workbook))
    )
    export.status = COMPLETED
    export.save()