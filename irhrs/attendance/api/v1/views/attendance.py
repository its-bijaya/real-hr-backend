from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError as DjValidationError
from django.db.models import OuterRef, Count, Subquery, Q, Prefetch, Avg, F, Sum, Exists, When, \
    Case, Value, \
    fields as dj_fields, BooleanField, FloatField, DurationField
from django.db.models.functions import Extract, Coalesce
from django.http import Http404
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.generics import get_object_or_404
from rest_framework.permissions import SAFE_METHODS
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet, GenericViewSet

from config import settings
from irhrs.attendance.api.v1.permissions import (
    IndividualAttendanceSettingPermission, OfflineAttendancePermission
)
from irhrs.attendance.utils import attendance as attendance_utils
from irhrs.attendance.api.v1.serializers.adjustment import \
    AttendanceAdjustmentCreateSerializer
from irhrs.attendance.api.v1.serializers.attendance import \
    IndividualAttendanceSettingSerializer, TimeSheetSerializer, \
    AttendanceUserMapSerializer, WebAttendanceSerializer, \
    ManualAttendanceSerializer, TimeSheetEntrySerializer, \
    UserTimeSheetSerializer, BulkIndividualAttendanceSettingSerializer, \
    TimeSheetEntryApprovalSerializer, IndividualUserShiftSerializer
from irhrs.attendance.api.v1.serializers.timesheet import TimeSheetForNoticeboardSerializer
from irhrs.attendance.constants import CONFIRMED, UNCLAIMED, DELETE, CANCELLED
from irhrs.attendance.constants import DECLINED
from irhrs.attendance.constants import LATE_IN, EARLY_OUT, WORKDAY, FULL_LEAVE, PUNCH_IN, \
    PUNCH_OUT, NO_LEAVE, FIRST_HALF, SECOND_HALF
from irhrs.attendance.models import IndividualAttendanceSetting, TimeSheet, \
    AttendanceUserMap, IndividualUserShift, TimeSheetEntry, TimeSheetEntryApproval, \
    AttendanceAdjustment, AttendanceAdjustmentHistory
from irhrs.attendance.models.overtime import OvertimeEntry
from irhrs.attendance.models.travel_attendance import TravelAttendanceDays
from irhrs.attendance.utils import allow_current_ip
from irhrs.attendance.utils.attendance import humanize_interval,get_row_data
from irhrs.core.constants.common import ATTENDANCE
from irhrs.core.constants.organization import (
    ATTENDANCE_ADJUSTMENT_IS_APPROVED_DECLINED_DELETED_BY_HR_EMAIL,
    LEAVE
)
from irhrs.core.mixins.serializers import DummySerializer
from irhrs.core.mixins.viewset_mixins import OrganizationMixin, \
    DisallowPatchMixin, ListRetrieveViewSetMixin, CreateViewSetMixin, \
    HRSOrderingFilter, UserMixin, ListCreateRetrieveUpdateViewSetMixin, ListViewSetMixin, \
    OrganizationCommonsMixin
from irhrs.core.utils import nested_getattr
from irhrs.core.utils.common import apply_filters, get_today, validate_permissions, get_yesterday
from irhrs.core.utils.email import send_email_as_per_settings
from irhrs.core.utils.filters import FilterMapBackend
from irhrs.core.utils.user_activity import create_user_activity
from irhrs.export.mixins.export import BackgroundExcelExportMixin
from irhrs.leave.constants.model_constants import APPROVED, CREDIT_HOUR as LEAVE_CREDIT_HOUR, \
    TIME_OFF as LEAVE_TIME_OFF
from irhrs.leave.models import MasterSetting
from irhrs.leave.models.account import LeaveAccount
from irhrs.organization.models import Organization, FiscalYear
from irhrs.permission.constants.permissions import HAS_OBJECT_PERMISSION, \
    ATTENDANCE_PERMISSION, HAS_PERMISSION_FROM_METHOD, ATTENDANCE_ADJUSTMENTS_REQUEST_PERMISSION, \
    ATTENDANCE_INDIVIDUAL_ATTENDANCE_SETTINGS_PERMISSION, ATTENDANCE_OFFLINE_PERMISSION, \
    ATTENDANCE_DEVICE_SETTINGS_PERMISSION
from irhrs.permission.permission_classes import permission_factory
from irhrs.users.api.v1.serializers.thin_serializers import UserThinSerializer
from irhrs.users.models import UserSupervisor

from irhrs.attendance.api.v1.serializers.attendance import ExcelCreateUpdateSerializer

from irhrs.export.models.export import Export
from irhrs.export.constants import ADMIN
from irhrs.export.constants import FAILED
from django.core.files.base import ContentFile
from irhrs.export.utils.helpers import save_virtual_workbook
from irhrs.core.utils.common import get_complete_url


from rest_framework.parsers import MultiPartParser
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from django.core.exceptions import ValidationError
from irhrs.core.utils.excel import ExcelList
from rest_framework import status
from irhrs.attendance.models import IndividualAttendanceSetting, AttendanceUserMap, AttendanceSource
from openpyxl import Workbook
from django.http import HttpResponse

USERS = get_user_model()
HOURLY_LEAVES = (LEAVE_CREDIT_HOUR, LEAVE_TIME_OFF)
INVISIBLE_BALANCE = getattr(settings, 'SHOW_INVISIBLE_LEAVE_BALANCE', False)


class IndividualAttendanceSettingViewSet(
    DisallowPatchMixin, OrganizationMixin,
    ListCreateRetrieveUpdateViewSetMixin
):
    """
    create:

    Create attendance settings for individuals

    data

        {
            "id": 1,
            "userdetail": 1,
            "work_shift": 1,
            "web_attendance": true,
            "subordinate_email": true,
            "attendance_report": true,
            "self_attendance_email": true,
            "late_in_notification": true,
            "absent_notification": true,
            "enable_overtime": true,
            "overtime_remainder": false,
            "ip_filters": [
                {
                     "allow": true,  // whether to allow ip or block
                     "cidr": "192.168.1.0/24"  // classless inter domain routing
                }
            ]
        }

    Note*: Pass ip_filters if attendance is required

    list:
    -- filters --
        branch: branch_slug
        division: division_slug
        employment_status: employment_status_slug
        work_shift: work_shift_id
        overtime_applicable: 'false'/'true'
        employment_level: employment_level_slug
    """
    queryset = IndividualAttendanceSetting.objects.all()
    serializer_class = IndividualAttendanceSettingSerializer
    filter_backends = [DjangoFilterBackend, FilterMapBackend, SearchFilter,
                       OrderingFilter]
    search_fields = (
        "user__first_name",
        "user__middle_name",
        "user__last_name",
        "user__username"
    )
    ordering_fields = (
        "user__first_name",
        "user__middle_name",
        "user__last_name",
        "web_attendance",
    )
    filter_map = {
        'id': 'user__id',
        'branch': 'user__detail__branch__slug',
        'division': 'user__detail__division__slug',
        'employment_status': 'user__detail__employment_status__slug',
        'employment_level': 'user__detail__employment_level__slug',
        'username': 'user__username',
    }
    permission_classes = [IndividualAttendanceSettingPermission]

    def check_permissions(self, request):
        super().check_permissions(request)
        if not (self.user_type == 'current' or self.request.method in SAFE_METHODS):
            self.permission_denied(request, "Updates are allowed only on current employees.")

    @property
    def user_type(self):
        user_status = self.request.query_params.get('user_status')
        if user_status in ['past', 'all']:
            return user_status
        return 'current'

    def get_queryset(self):
        user_qs = USERS.objects.filter(
            detail__organization_id=self.organization.id,
        )
        if self.user_type == 'current':
            user_qs = user_qs.current()
        elif self.user_type == 'past':
            user_qs = user_qs.past()

        return super().get_queryset().filter(
            user__in=user_qs
        ).select_related(
            'user',
            'user__detail',
            'user__detail__organization',
            'user__detail__division',
            'user__detail__division',
            'user__detail__job_title',
            'user__detail__employment_level',
            'overtime_setting',
            'credit_hour_setting',
            'penalty_setting',
        ).prefetch_related(
            'ip_filters',
            Prefetch(
                'user__supervisors',
                queryset=UserSupervisor.objects.filter(
                    authority_order=1
                ).select_related(
                    'supervisor',
                    'supervisor__detail',
                    'supervisor__detail__organization',
                    'supervisor__detail__job_title',
                    'supervisor__detail__division',
                    'supervisor__detail__employment_level'
                ),
                to_attr='user_supervisors'
            ),
            Prefetch(
                'individual_setting_shift',
                queryset=IndividualUserShift.objects.filter(
                    Q(applicable_to__gte=get_today()) | Q(
                        applicable_to__isnull=True)
                ).select_related(
                    'supervisor',
                    'supervisor__detail',
                    'supervisor__detail__organization',
                    'supervisor__detail__job_title',
                    'supervisor__detail__division',
                    'supervisor__detail__employment_level'
                ),
                to_attr='work_shift'
            ),

        )

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        work_shift = self.request.query_params.get('work_shift')
        overtime_applicable = self.request.query_params.get(
            'overtime_applicable')

        if work_shift or overtime_applicable in ['true', 'false']:
            queryset = self.annotate_shift_id(queryset)
            if work_shift:
                try:
                    queryset = queryset.filter(work_shift_id=int(work_shift))
                except (ValueError, TypeError, DjValidationError):
                    queryset = queryset.none()
            if overtime_applicable == 'false':
                # currently users who do not have shifts assigned
                # will not be applicable for overtime
                queryset = queryset.filter(work_shift_id=None)
            elif overtime_applicable == 'true':
                queryset = queryset.exclude(work_shift_id=None)

        return queryset

    @staticmethod
    def annotate_shift_id(queryset):
        _date_time_for_work_shift = timezone.now()
        return queryset.annotate(
            # annotate queryset for workshift slug
            work_shift_id=Subquery(
                IndividualUserShift.objects.filter(
                    individual_setting_id=OuterRef('pk')
                ).filter(
                    Q(
                        applicable_from__lte=_date_time_for_work_shift) & Q(
                        Q(
                            applicable_to__gte=_date_time_for_work_shift
                        ) | Q(
                            applicable_to__isnull=True
                        )
                    )
                ).order_by().values('shift_id')[:1]
            )
        )

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return BulkIndividualAttendanceSettingSerializer
        return super().get_serializer_class()

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx.update({
            'organization': self.get_organization()
        })
        return ctx


class IndividualSettingShiftView(UserMixin, OrganizationMixin, ModelViewSet):
    queryset = IndividualUserShift.objects.all()
    serializer_class = IndividualUserShiftSerializer
    permission_classes = [IndividualAttendanceSettingPermission]

    def get_queryset(self):
        return super().get_queryset().filter(
            individual_setting__user=self.user,
            individual_setting__user__detail__organization=self.organization,
        )

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['organization'] = self.organization
        ctx['user'] = self.user
        return ctx


class AttendanceUserMapViewSet(HRSOrderingFilter, OrganizationMixin,
                               ModelViewSet, BackgroundExcelExportMixin):
    """
    list:
    User Map list

    filters:
       setting: Individual Settings ID
       bio_user_id: bio id
       source: source_id
       division: division_slug

    ordering:
      full_name
    """
    serializer_class = AttendanceUserMapSerializer
    queryset = AttendanceUserMap.objects.all()
    filter_backends = (FilterMapBackend, SearchFilter)
    filter_map = {
        'setting': 'setting',
        'bio_user_id': 'bio_user_id',
        'source': 'source',
        'division': 'setting__user__detail__division__slug',
        'username': 'setting__user__username'
    }
    search_fields = (
        'setting__user__first_name',
        'setting__user__middle_name',
        'setting__user__last_name',
        'setting__user__username'
    )
    ordering_fields_map = {
        'full_name': (
            'setting__user__first_name',
            'setting__user__middle_name',
            'setting__user__last_name',
        )
    }
    permission_classes = [
        permission_factory.build_permission(
            "AttendanceSourcePermission",
            limit_write_to=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_INDIVIDUAL_ATTENDANCE_SETTINGS_PERMISSION
            ],
            limit_read_to=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_DEVICE_SETTINGS_PERMISSION,
                ATTENDANCE_INDIVIDUAL_ATTENDANCE_SETTINGS_PERMISSION
            ],

        )
    ]
    export_type = "bulk_excel_update"
    def get_queryset(self):
        return super().get_queryset().filter(
            setting__user__detail__organization=self.get_organization()
        )

    @action(methods=['POST'],
            detail=False,
            parser_classes=[MultiPartParser],
            serializer_class=ExcelCreateUpdateSerializer,
            url_path='bulk-excel-update')
    def bulk_excel_update(self, request, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        excel_update_file = serializer.validated_data['file']
        workbook = load_workbook(excel_update_file)
        ws = workbook.active
        excel_list = ExcelList(workbook)
        ft = Font(bold=True)
        excel_data = excel_list[1:]
        error_exists = False
        assigned_bio_id_of_particular_device = set()
        duplicate_detail = set()
        duplicate_data = set()

        for index, row in enumerate(excel_data, 1):
            errors = {}
            email, bio_id, device = get_row_data(row)
            setting = IndividualAttendanceSetting.objects.filter(
                user__email=email).first()

            if not setting:
                errors["setting"] = "User does not exist with provided email"

            source = AttendanceSource.objects.filter(name=device).first()
            if not source:
                errors["source"] = "Device is not available please select the available device"

            if not (bio_id and email and device):
                errors["Fields"] = "Field missing in excel sheet"

            if (bio_id, device) in assigned_bio_id_of_particular_device:
                errors["Assigned Error"] = "This bio_id is already assigned to device."

            if (email, device) in duplicate_detail:
                errors["Duplicate"] = 'This email is already assigned for this device.'

            if (email, bio_id, device) in duplicate_data:
                errors["Duplicate"] = 'Given user credential already exist in excel sheet(duplicate present).'

            if AttendanceUserMap.objects.filter(
                source=source,
                bio_user_id=bio_id
            ).exists():
                errors["Duplicate"] = "User already exists with same bio_id in same device."

            if errors:
                error_exists = True
                excel_list[index].append(",".join(errors.values()))
                continue

            AttendanceUserMap.objects.update_or_create(
                setting=setting,
                source=source,
                defaults={'bio_user_id': bio_id})
            assigned_bio_id_of_particular_device.add((bio_id, device))
            duplicate_detail.add((email, device))
            duplicate_data.add((email, bio_id, device))

        if error_exists:
            excel_list[0].append('Remark')
            error_wb = excel_list.generate_workbook()
            ws = error_wb.active
            device_list = [d.name for d in AttendanceSource.objects.all()]
            dv = DataValidation(type="list",formula1='"' + ','.join(device_list) + '"')
            col_device = ws['C']
            ws.add_data_validation(dv)
            for row in col_device[1:]:
                dv.add(row)
            for row in ws["A1:D1"]:
                for cell in row:
                    cell.font = ft
            export = Export.objects.filter(export_type=self.export_type).first()
            if not export:
                export=Export.objects.create(
                    user=self.request.user,
                    name="Bulk Excel Update",
                    exported_as=ADMIN,
                    export_type=self.export_type,
                    organization=self.get_organization(),
                    status=FAILED,
                    remarks='Bulk excel update failed.'
                )
            export.export_file.save(
                "bulk_payroll_update.xlsx",
                ContentFile(save_virtual_workbook(error_wb))
            )
            export.save()
            export_url = get_complete_url(export.export_file.url)
            return Response(
                {'file': {'error_file': export_url }},
                status=status.HTTP_400_BAD_REQUEST
            )
        return Response(
            {'msg': 'Bulk Attendance User Map completed successfully'},
            status=status.HTTP_200_OK
        )

    @action(methods=['GET'], detail=False, url_path='sample-file',
            parser_classes=[MultiPartParser],
            serializer_class=ExcelCreateUpdateSerializer)
    def download_sample_file(self, request,**kwargs):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Email'
        worksheet['B1'] = 'Bio ID'
        worksheet['C1'] = 'Device'
        worksheet.append(['example1@example.com', '1234', 'Device 1'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sample_file.xlsx'
        workbook.save(response)
        return response
    
    def get_export_type(self):
        return 'Export summary report of Device'
    
    export_fields = {
        'Device Name': 'source.name',
        'Employee Name': 'setting.user.full_name',
        'Username': 'setting.user.username',
        'Bio User Id': 'bio_user_id',
        'Employee Code': 'setting.user.detail.code',
        'Email Address': 'setting.user.email',   
    }



class UserTimeSheetViewSet(UserMixin, ListRetrieveViewSetMixin):
    """
    Attendance of a user
    """
    queryset = TimeSheet.objects.all().select_related("work_shift", "work_time")
    serializer_class = TimeSheetSerializer
    permission_classes = [permission_factory.build_permission(
        "AdjustmentPermission",
        allowed_to=[
            ATTENDANCE_PERMISSION,
            ATTENDANCE_ADJUSTMENTS_REQUEST_PERMISSION,
            HAS_PERMISSION_FROM_METHOD,
        ],
        actions={'adjust': [HAS_OBJECT_PERMISSION]},
        allowed_user_fields=['timesheet_user']
    )]
    filter_fields = ('timesheet_for', 'coefficient')
    filter_backends = (
        DjangoFilterBackend, SearchFilter, OrderingFilter, FilterMapBackend
    )
    search_fields = (
        'timesheet_for',
    )
    ordering_fields = (
        'timesheet_for',
        'created_at',
        'modified_at'
    )
    filter_map = {
        'start_date': 'timesheet_for__gte',
        'end_date': 'timesheet_for__lte'
    }

    def has_user_permission(self):
        return self.user == self.request.user

    def get_queryset(self):
        return super().get_queryset().filter(
            timesheet_user=self.user,
            timesheet_for__lte=get_today()
        ).annotate(
            adjustments_count=Count('adjustment_requests', filter=~Q(
                adjustment_requests__status=DECLINED), distinct=True)
        )

    def get_serializer_class(self):
        if self.action == 'adjust':
            return AttendanceAdjustmentCreateSerializer
        elif self.action == 'entries':
            if self.request.user.attendance_setting.enable_approval:
                return TimeSheetEntryApprovalSerializer
            return TimeSheetEntrySerializer
        return super().get_serializer_class()

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(
            self.get_queryset()
        )
        ret = super().list(request, *args, **kwargs)
        aggregates = {
            display: humanize_interval(
                value
            ) for display, value
            in qs.aggregate(
                average_in=Coalesce(
                    Avg(Extract('punch_in', 'Hour'),
                        output_field=FloatField()), 0.0
                ) * 60 * 60 + Coalesce(
                    Avg(Extract('punch_in', 'Minute'),
                        output_field=FloatField()), 0.0
                ) * 60 + Coalesce(
                    Avg(Extract('punch_in', 'Second'),
                        output_field=FloatField()), 0.0
                ),
                average_out=Coalesce(
                    Avg(Extract('punch_out', 'Hour'),
                        output_field=FloatField()), 0.0
                ) * 60 * 60 + Coalesce(
                    Avg(Extract('punch_out', 'Minute'),
                        output_field=FloatField()), 0.0
                ) * 60 + Coalesce(
                    Avg(Extract('punch_out', 'Second'),
                        output_field=FloatField()), 0.0
                ),
                average_work_time=Avg(
                    F('worked_hours'),
                    output_field=DurationField()
                )
            ).items()
        }
        aggregates['web_attendance'] = False
        user = self.request.user
        user_attendance_setting = getattr(user, 'attendance_setting', None)
        if user_attendance_setting:
            aggregates['web_attendance'] = user_attendance_setting.web_attendance
        # The multiplication is done because the Hour is extracted and
        # averaged. But, the `humanize_interval` expects seconds.
        ret.data.update(
            **aggregates
        )
        return ret

    @action(methods=['POST'], detail=True, url_name='adjust', url_path='adjust')
    def adjust(self, request, *args, **kwargs):
        context = self.get_serializer_context()
        context.update({'time_sheet': self.get_object()})
        serializer = self.get_serializer_class()(
            data=request.data,
            context=context
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        create_user_activity(
            request.user,
            f"requested an adjustment request.",
            ATTENDANCE
        )

        return Response(serializer.data, 201)

    @action(methods=['GET'], detail=True)
    def entries(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.timesheet_entries.all()
        if self.request.user.attendance_setting.enable_approval:
            if hasattr(instance, 'timesheet_approval'):
                qs = instance.timesheet_approval.timesheet_entry_approval.all()
            else:
                qs = TimeSheetEntryApproval.objects.none()
        paginated_qs = self.paginate_queryset(qs)
        data = self.get_serializer(paginated_qs, many=True).data
        return self.get_paginated_response(data)


class TimeSheetViewSet(HRSOrderingFilter, ListRetrieveViewSetMixin):
    """
    list:
    All timesheets

    ordering

        updated
        full_name
        timesheet_for

    filters
        timesheet_for=2018-01-01
        timesheet_user=user_id
        supervisor=user_id

    """
    queryset = TimeSheet.objects.all()
    serializer_class = UserTimeSheetSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    filter_fields = ('timesheet_for',)
    ordering_fields = ('timesheet_for', 'updated')
    search_fields = ('timesheet_user__first_name',
                     'timesheet_user__middle_name',
                     'timesheet_user__last_name')
    ordering = '-modified_at',

    ordering_fields_map = {
        'full_name': ('timesheet_user__first_name',
                      'timesheet_user__middle_name',
                      'timesheet_user__last_name')
    }

    def get_queryset(self):
        return super().get_queryset().select_related(
            'timesheet_user',
            'timesheet_user__detail',
            'timesheet_user__detail__organization',
            'timesheet_user__detail__division',
            'timesheet_user__detail__job_title',
            'timesheet_user__detail__employment_level',
            'work_time',
            'work_shift'
        ).prefetch_related(
            'adjustment_requests',
            Prefetch('timesheet_entries', queryset=TimeSheetEntry.objects.all(),
                     to_attr='_prefetched_timesheet_entries')
        )

    def get_organization(self):
        return get_object_or_404(
            Organization.objects.filter(
                id__in=self.request.user.switchable_organizations_pks
            ),
            slug=self.request.query_params.get('organization_slug')
        )

    def filter_queryset(self, queryset, *args, **kwargs):
        # Check Permission
        supervisor_id = self.request.query_params.get("supervisor")

        has_permission_from_code = validate_permissions(
            self.request.user.get_hrs_permissions(self.get_organization()),
            ATTENDANCE_PERMISSION,
            ATTENDANCE_OFFLINE_PERMISSION
        )

        queryset = super().filter_queryset(queryset)

        if has_permission_from_code or (
            supervisor_id == str(self.request.user.id)
        ):
            if supervisor_id:
                queryset = queryset.filter(
                    timesheet_user_id__in=self.request.user.subordinates_pks
                )
            else:
                queryset = queryset.filter(
                    timesheet_user__detail__organization_id__in=(
                        self.request.user.switchable_organizations_pks
                    )
                )
            queryset = apply_filters(
                self.request.query_params,
                {
                    'timesheet_user': 'timesheet_user'
                },
                queryset
            )
        else:
            # for normal user return only her/his timesheets
            queryset = queryset.filter(timesheet_user=self.request.user)

        return queryset


class WebAttendanceViewSet(CreateViewSetMixin):
    queryset = TimeSheet.objects.all()
    serializer_class = WebAttendanceSerializer

    def create(self, request, *args, **kwargs):
        if not hasattr(request.user, 'attendance_setting'):
            raise PermissionDenied
        if not request.user.attendance_setting.web_attendance:
            raise PermissionDenied
        if not allow_current_ip(request):
            raise PermissionDenied
        return super().create(request, *args, **kwargs)

    @action(methods=["GET", "PATCH"],
            detail=False,
            url_path="remote-working-status-for-today")
    def remote_working_status_for_today(self, request, **kwargs):
        if request.method.upper() == "PATCH":
            timesheet = TimeSheet.objects.filter(id=request.data.get('timesheet_id')).first()
            if not timesheet:
                return Response(status=400, data={'error': 'timesheet not found'})
            timesheet.working_remotely = request.data.get('working_remotely', False)
            timesheet.save()
            return Response(f"working remotely changed to {timesheet.working_remotely}.")

        timesheet = TimeSheet.objects.filter(
            timesheet_user=request.user,
            timesheet_for=get_today()
        ).first()
        return Response({
            'working_remotely': timesheet.working_remotely if timesheet else False
        })

    def get_organization(self):
        return get_object_or_404(
            Organization.objects.filter(
                id__in=self.request.user.switchable_organizations_pks
            ),
            slug=self.request.query_params.get('organization_slug')
        )


class ManualAttendanceViewSet(CreateViewSetMixin):
    """
    create:

    Manual Attendance by HR

    data

        {
            "timesheet_user": user_id,
            "date": YYYY-MM-DD, date of attendance
            "punch_in": YYYY-MM-DDTHH:MM:SS
            "punch_out": YYYY-MM-DDTHH:MM:SS
            "work_time": work_time_id // leave this empty for off day
                                        //attendance
        }
    """
    queryset = TimeSheet.objects.all()
    serializer_class = ManualAttendanceSerializer
    permission_classes = [OfflineAttendancePermission]

    def get_organization(self):
        return get_object_or_404(
            Organization.objects.filter(
                id__in=self.request.user.switchable_organizations_pks
            ),
            slug=self.request.query_params.get('organization_slug')
        )


class NoticeboardDetailOnLeaveFieldVisitAbsentMixin:
    def get_queryset(self):
        return TimeSheet.objects.filter(
            timesheet_for=get_today(),
            timesheet_user__detail__organization=self.organization
        )

    @staticmethod
    def user_data(user):
        return UserThinSerializer(
            USERS.objects.filter(
                id__in=user
            ).order_by("first_name", "middle_name", "last_name").select_related(
                'detail', 'detail__organization', 'detail__job_title',
                'detail__division', 'detail__branch', 'detail__employment_level'
            ),
            fields=(
                'id', 'full_name', 'profile_picture',
                "cover_picture", "is_online", 'is_current', 'organization', "job_title"
            ),
            many=True
        ).data

    def get_user_on_leave(self):
        on_leave = self.get_queryset().exclude(
            leave_coefficient=NO_LEAVE,
            hour_off_coefficient=''
        ).order_by(
            "timesheet_user__first_name",
            "timesheet_user__middle_name",
            "timesheet_user__last_name"
        )

        return {
            'count': on_leave.count(),
            'result': TimeSheetForNoticeboardSerializer(on_leave, many=True).data
        }

    def get_user_on_field_visit(self):
        on_field_visit = set(
            TravelAttendanceDays.objects.filter(
                day=get_today(),
                user__detail__organization=self.organization,
                is_archived=False
            ).values_list('user', flat=True)
        )
        return {
            'count': len(on_field_visit),
            'results': self.user_data(user=on_field_visit)
        }

    def get_absent_user(self):
        absent_user = set(
            self.get_queryset().exclude(
                Q(leave_coefficient=FULL_LEAVE) |
                Q(expected_punch_in__gt=get_today(with_time=True))
            ).filter(
                is_present=False,
                coefficient=WORKDAY,
            ).annotate(
                zero_working_time=Case(
                    When(
                        expected_punch_in__isnull=False,
                        expected_punch_out__isnull=False,
                        expected_punch_in=F('expected_punch_out')
                    ), then=True,
                    default=False,
                    output_field=BooleanField()
                )
            ).exclude(
                zero_working_time=True
            ).values_list('timesheet_user', flat=True)
        )
        return {
            'count': len(absent_user),
            'users': self.user_data(user=absent_user)
        }

    def get_user_working_remotely(self):
        remote_user = self.get_queryset().filter(
            Q(timesheet_approval__status=APPROVED) | Q(timesheet_approval__isnull=True),
            working_remotely=True,
        ).values_list('timesheet_user', flat=True)
        return {
            'count': len(remote_user),
            'users': self.user_data(user=remote_user)
        }


class SingleAPIForLeaveFieldVisitAndAbsentViewSet(
    OrganizationMixin, NoticeboardDetailOnLeaveFieldVisitAbsentMixin,
    ListViewSetMixin
):
    def list(self, request, *args, **kwargs):
        return Response(
            {
                'absent': self.get_absent_user(),
                'leave': self.get_user_on_leave(),
                'field_visit': self.get_user_on_field_visit(),
                'remote_work': self.get_user_working_remotely()
            }
        )


class EmployeeOnLeaveFieldVisitAndAbsentViewSet(
    OrganizationMixin, NoticeboardDetailOnLeaveFieldVisitAbsentMixin,
    ListViewSetMixin
):
    def list(self, request, *args, **kwargs):
        _action = self.kwargs.get('action')
        function_map = {
            'absent': self.get_absent_user,
            'leave': self.get_user_on_leave,
            'field-visit': self.get_user_on_field_visit
        }
        return Response(function_map.get(_action)())


class OverTimeAndLeaveNoticeboardStatsViewset(ListViewSetMixin):

    def get_queryset(self):
        return OvertimeEntry.objects.none()

    def get_overtime_detail(self):
        fiscal_year = self.get_fiscal_year()
        if fiscal_year:
            queryset = OvertimeEntry.objects.filter(
                user=self.request.user,
                timesheet__timesheet_for__range=(
                    fiscal_year.applicable_from, fiscal_year.applicable_to),
            )
            stats = queryset.filter(
                timesheet__timesheet_for__lte=get_yesterday()
            ).aggregate(
                total_generated=Sum("overtime_detail__claimed_overtime"),
                total_confirmed=Sum(
                    "overtime_detail__claimed_overtime", filter=Q(claim__status=CONFIRMED))
            )
            response = {key: humanize_interval(
                value) for key, value in stats.items()}
            response['last_generated'] = humanize_interval(
                nested_getattr(
                    queryset.order_by(
                        "-timesheet__timesheet_for"
                    ).first(),
                    'overtime_detail.claimed_overtime'
                )
            )
            return response
        return {
            'last_generated': "00:00:00",
            'total_generated': "00:00:00",
            'total_confirmed': "00:00:00"
        }

    def get_leave_detail(self):
        fiscal_year = self.get_fiscal_year()
        if fiscal_year:
            active_master_setting = MasterSetting.objects.filter(
                organization=self.request.user.detail.organization
            ).active()
            queryset = LeaveAccount.objects.filter(
                rule__leave_type__master_setting__in=active_master_setting,
                user=self.request.user,
                is_archived=False
            ).exclude(
                rule__leave_type__category__in=HOURLY_LEAVES,
            )
            consumed_balance = queryset.aggregate(
                consumed_balance=Coalesce(Sum('leave_requests__sheets__balance', filter=Q(
                    leave_requests__sheets__leave_for__gte=fiscal_year.applicable_from,
                    leave_requests__sheets__leave_for__lte=fiscal_year.applicable_to,
                    leave_requests__status=APPROVED,
                    leave_requests__is_deleted=False,
                    leave_requests__is_archived=False,
                ) & ~Q(
                    leave_requests__leave_rule__leave_type__category__in=HOURLY_LEAVES
                )), Value(0.0)),
            ).get('consumed_balance') or 0

            if not INVISIBLE_BALANCE:
                queryset = queryset.filter(
                    rule__leave_type__visible_on_default=True, rule__is_paid=True
                )
            remaining_qs = queryset.order_by('-created_at').values('usable_balance')
            remaining = remaining_qs.aggregate(
                remaining_balance=Sum('usable_balance')
            ).get('remaining_balance') or 0
            stats = dict(
                consumed_balance=consumed_balance,
                remaining_balance=remaining
            )
            stats['total_balance'] = consumed_balance + remaining
            return stats
        return {
            'consumed_balance': 0,
            'remaining_balance': 0,
            'total_balance': 0
        }

    def get_fiscal_year(self):
        if FiscalYear.objects.for_category_exists(
            organization=self.request.user.detail.organization,
            category=LEAVE
        ):
            fiscal_year = FiscalYear.objects.current(
                organization=self.request.user.detail.organization,
                category=LEAVE
            )
        else:
            fiscal_year = FiscalYear.objects.current(
                organization=self.request.user.detail.organization
            )
        return fiscal_year

    def get_lost_hour_detail(self):
        yesterday = get_yesterday()
        total_lost_data = TimeSheet.objects.filter(
            timesheet_user=self.request.user,
            timesheet_for__gte=yesterday.replace(day=1),
            timesheet_for__lte=yesterday
        ).annotate(
            absent=Case(When(
                is_present=False,
                coefficient=WORKDAY,
                leave_coefficient__in=[FIRST_HALF, SECOND_HALF, NO_LEAVE],
                then=True
            ), default=False, output_field=dj_fields.BooleanField()),

            late_in=Exists(
                TimeSheetEntry.objects.filter(timesheet_id=OuterRef('pk'),
                                              is_deleted=False).filter(
                    Q(
                        category=LATE_IN,
                        entry_type=PUNCH_IN,
                        timesheet__coefficient=WORKDAY
                    ) & ~Q(
                        timesheet__leave_coefficient=FULL_LEAVE
                    )
                )),
            early_out=Exists(
                TimeSheetEntry.objects.filter(timesheet_id=OuterRef('pk'),
                                              is_deleted=False).filter(
                    Q(
                        category=EARLY_OUT, entry_type=PUNCH_OUT, timesheet__coefficient=WORKDAY
                    ) & ~Q(
                        timesheet__leave_coefficient=FULL_LEAVE
                    ),
                )
            )

        ).exclude(absent=False, late_in=False, early_out=False).annotate(
            lost_late_in=Case(When(
                late_in=True,
                then=F('punch_in_delta')
            ), default=None, output_field=dj_fields.DurationField(null=True)),

            lost_early_out=Case(When(
                early_out=True,
                then=-F('punch_out_delta')
            ), default=None, output_field=dj_fields.DurationField(null=True)),

            lost_absent=Case(When(
                absent=True,
                expected_punch_in__isnull=False,
                expected_punch_out__isnull=False,
                then=F('expected_punch_out') - F('expected_punch_in')
            ), When(
                absent=False,
                then=None
            ), default=timezone.timedelta(0), output_field=dj_fields.DurationField(null=True))
        ).aggregate(
            total_lost=Sum(Coalesce(
                F('lost_late_in'), timezone.timedelta(0)) + Coalesce(
                F('lost_early_out'), timezone.timedelta(0)) + Coalesce(
                F('lost_absent'), timezone.timedelta(0))
                           )
        )
        return {"total_lost_hour": humanize_interval(total_lost_data.get('total_lost'))}

    def list(self, request, *args, **kwargs):
        types = ['leave', 'lost_hour', 'overtime']
        response = {}
        for type in types:
            response.update({type: getattr(self, f'get_{type}_detail')()})
        return Response(response)


class TimeSheetEntrySoftDeleteViewSet(OrganizationMixin, OrganizationCommonsMixin, GenericViewSet):
    queryset = TimeSheet.objects.all()
    organization_field = 'timesheet_user__detail__organization'
    permission_classes = [permission_factory.build_permission(
        "TimeSheetEntrySoftDeletePermission",
        allowed_to=[ATTENDANCE_PERMISSION]
    )]
    serializer_class = DummySerializer()

    def get_timesheet_entry(self):
        timesheet = self.get_object()
        entry = timesheet.timesheet_entries.filter(
            id=self.kwargs.get('timesheet_entry_id')
        ).first()
        if not entry:
            raise Http404
        return entry

    def create_cancel_adjustment(self) -> None:
        # Earlier adjustment was not created when HR delete Attendance Entries.
        # After HRIS-3582, new adjustment is created and adjustment will be in CANCELLED status.
        timesheet = self.get_object()
        remarks = self.request.data.get("description")
        timesheet_entry = self.get_timesheet_entry()
        sender = self.request.user
        receiver = attendance_utils.get_adjustment_request_receiver(
            timesheet.timesheet_user
        )
        adjustment_data = {
            "timesheet": timesheet,
            "sender": sender,
            "receiver": receiver,
            "status": CANCELLED,
            "action": DELETE,
            "timestamp": timesheet_entry.timestamp,
            "description": remarks,
            "category": timesheet_entry.entry_type
        }
        adjustment = AttendanceAdjustment.objects.create(
            **adjustment_data
        )
        AttendanceAdjustmentHistory.objects.create(
            adjustment=adjustment,
            action_performed=CANCELLED,
            action_performed_by=sender,
            remark=adjustment.description
        )

    @action(
        methods=['POST'],
        detail=True,
        url_path=r'soft-delete-entry/(?P<timesheet_entry_id>\d+)',
        url_name='delete'
    )
    def soft_delete_entry(self, *args, **kwargs):
        entry = self.get_timesheet_entry()
        if entry.entry_method not in settings.DELETE_ALLOWED_TIMESHEET_ENTRY_METHODS:
            raise serializers.ValidationError({
                'non_field_errors': [f'Entry of type {entry.entry_method} can not be deleted']
            })
        if entry.is_deleted:
            raise serializers.ValidationError({'non_field_errors': ['Entry already deleted']})

        if entry and entry.entry_method not in settings.DELETE_ALLOWED_TIMESHEET_ENTRY_METHODS:
            raise serializers.ValidationError(f"{entry.entry_method} entries can not be removed.")

        overtime_status = nested_getattr(entry.timesheet, 'overtime.claim.status')
        if overtime_status and overtime_status != UNCLAIMED:
            raise serializers.ValidationError(
                f"Cannot remove entries with overtime in {overtime_status} status"
            )
        credit_entries = getattr(entry.timesheet, 'credit_entries', None)
        if credit_entries:
            credit_entry = credit_entries.first()
            credit_hour_request = getattr(
                credit_entry, 'credit_request', None
            )
            credit_hour_status = getattr(credit_hour_request, 'status', None)

            is_deleted = getattr(credit_hour_request, 'is_deleted', False)
            if credit_hour_status == APPROVED and not is_deleted:
                raise serializers.ValidationError(
                    f"Cannot remove entries with credit hour in {credit_hour_status} status"
                )
        self.create_cancel_adjustment()
        entry.soft_delete()
        send_email_as_per_settings(
            recipients=entry.timesheet.timesheet_user,
            subject='Timesheet Entry Deleted',
            email_text=(
                f"Your timesheet entry for {entry.timesheet.timesheet_for}"
                f" has been deleted by {self.request.user}"
            ),
            email_type=ATTENDANCE_ADJUSTMENT_IS_APPROVED_DECLINED_DELETED_BY_HR_EMAIL
        )
        return Response({'message': 'Successfully deleted entry'})

    @action(
        methods=['POST'],
        detail=True,
        url_path=r'undo-soft-delete-entry/(?P<timesheet_entry_id>\d+)',
        url_name='undo-delete'
    )
    def undo_soft_delete_entry(self, *args, **kwargs):
        entry = self.get_timesheet_entry()
        if not entry.is_deleted:
            raise serializers.ValidationError({'non_field_errors': ['Entry not in deleted state']})
        entry.revert_soft_delete()
        return Response({'message': 'Undo delete complete'})
