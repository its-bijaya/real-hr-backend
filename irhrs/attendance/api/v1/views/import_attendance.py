from datetime import date, time

from django.core.cache import cache as dj_cache
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from irhrs.attendance.api.v1.serializers.import_attendance import \
    AttendanceImportSerializer
from irhrs.attendance.models import TimeSheet
from irhrs.core.mixins.viewset_mixins import OrganizationMixin
from irhrs.permission.constants.permissions import ATTENDANCE_PERMISSION, ATTENDANCE_IMPORT_PERMISSION
from irhrs.permission.permission_classes import permission_factory


class AttendanceImportView(OrganizationMixin, ModelViewSet):
    serializer_class = AttendanceImportSerializer
    queryset = TimeSheet.objects.none()
    permission_classes = [permission_factory.build_permission(
        "AttendanceCalendarPermission",
        allowed_to=[
            ATTENDANCE_PERMISSION,
            ATTENDANCE_IMPORT_PERMISSION,
        ]
    )]

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['organization'] = self.organization
        return ctx

    @action(detail=False, methods=['GET'], url_path='sample')
    def download_sample(self, request, **kwargs):
        fields = [
            'user',  # user identifier
            'timesheet_for',  # date identifier
            'punch_in',  # data 1
            'punch_out',  # data 2
            'extends'
        ]
        value = [
            'someone@example.com',
            date(2019, 1, 1),
            time(9, 0),
            time(19, 0),
            'f'
        ]

        wb = Workbook()
        ws = wb.active
        ws.append(fields)

        for i in range(1, len(fields) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30
        ws.append(value)

        response = HttpResponse(
            content=save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=sample-att.xlsx'
        return response

    def list(self, request, *args, **kwargs):
        last_exports = dj_cache.get(
            'attendance_fails_export'
        )
        return Response(
            last_exports
        )
