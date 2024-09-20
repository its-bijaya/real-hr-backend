import calendar
from datetime import date, datetime

from django.db import models
from django.utils.timezone import timedelta
import openpyxl
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db.models import (F, DurationField, Q,
                              Avg, TimeField, FloatField, Sum, Case, When, Value, Subquery,
                              OuterRef,
                              Prefetch, FilteredRelation, CharField, Exists, Count)
from django.db.models.functions import Coalesce, Extract, Cast, ExtractWeekDay
from django.forms.utils import pretty_name
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.functional import cached_property
from django_filters.rest_framework import DjangoFilterBackend
from django_q.tasks import async_task
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import serializers, status
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.filters import SearchFilter
from rest_framework.generics import get_object_or_404
from rest_framework.response import Response
from rest_framework.serializers import Serializer
from tablib import Dataset

from irhrs.attendance.api.v1.reports.serializers.individual_attendance_report import \
    AttendanceHeadingReportSettingSerializer
from irhrs.attendance.api.v1.reports.views.mixins import \
    AttendanceReportPermissionMixin
from irhrs.attendance.constants import (BREAK_IN, BREAK_OUT, WORKDAY, FULL_LEAVE, NO_LEAVE,
                                        FIRST_HALF, SECOND_HALF, PUNCH_IN,
                                        PUNCH_OUT, WEEK_DAYS_CHOICES, MISSING, REQUESTED,
                                        UNCLAIMED, CONFIRMED, APPROVED, SYNC_PENDING, SYNC_FAILED
                                        )
from irhrs.attendance.models.attendance_extraheadings import AttendanceHeadingReportSetting
from irhrs.attendance.utils.attendance import humanize_interval
from irhrs.core.constants.organization import GLOBAL
from irhrs.core.mixins.serializers import DummySerializer
from irhrs.core.mixins.viewset_mixins import (ListViewSetMixin,
                                              OrganizationMixin, DateRangeParserMixin,
                                              ModeFilterQuerysetMixin,
                                              SupervisorQuerysetViewSetMixin,
                                              PastUserTimeSheetFilterMixin)
from irhrs.core.utils import prettify_headers, nested_get, nested_getattr
from irhrs.core.utils.common import get_today, validate_permissions, format_timezone
from irhrs.core.utils.custom_mail import custom_mail
from irhrs.core.utils.filters import (NullsAlwaysLastOrderingFilter,
                                      FilterMapBackend, OrderingFilterMap)
from irhrs.export.constants import ADMIN, QUEUED, PROCESSING
from irhrs.export.mixins.export import BackgroundExcelExportMixin
from irhrs.export.models import Export
from irhrs.export.utils.export import ExcelExport
from irhrs.export.utils.helpers import has_pending_export
from irhrs.organization.models import FiscalYear, FiscalYearMonth
from irhrs.payroll.api.v1.views.reports import ReportSettingViewSetMixin
from irhrs.permission.constants.permissions import (ATTENDANCE_PERMISSION,
                                                    ATTENDANCE_REPORTS_PERMISSION,
                                                    HAS_PERMISSION_FROM_METHOD,
                                                    ATTENDANCE_IMPORT_PERMISSION)
from irhrs.permission.models import HRSPermission
from irhrs.permission.permission_classes import permission_factory
from irhrs.users.api.v1.serializers.thin_serializers import UserThinSerializer
from irhrs.users.models import UserSupervisor
from ..serializers.individual_attendance_report import \
    (IndividualAttendanceReportSerializer,
     IndividualAttendanceOverviewSerializer, OvertimeDetailReportSerializer,
     ComparativeOvertimeReportSerializer, AttendanceGeoLocationSerializer,
     IndividualAttendanceOverviewExportSerializer, DailyAttendanceReconciliationSerializer,
     EmployeeAttendanceInsightSerializer)
from .....models import TimeSheet, AttendanceUserMap, AttendanceEntryCache, TimeSheetEntry
from .....utils.reconciliation import get_attendance_entries_for_given_timesheet, \
    get_total_lost_hours_from_timesheet, get_early_out_from_timesheet, get_late_in_from_timesheet, \
    break_in_out_lost_hour, get_ktm_time, has_unsynced_attendance_entries

INFO_EMAIL = getattr(settings, 'INFO_EMAIL', 'noreply@realhrsoft.com')
USER = get_user_model()
ZERO_DURATION = timezone.timedelta(minutes=0)
WHITE = openpyxl.styles.colors.Color(indexed=1)
RED = openpyxl.styles.colors.Color(indexed=2)
blue = openpyxl.styles.colors.Color(rgb='28688C')
green = openpyxl.styles.colors.Color(rgb='147066')
blue_fill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor=blue
)
green_fill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor=green
)
red_fill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor=RED
)
white_font = openpyxl.styles.Font(
    b=True, color=WHITE
)
white_font_bold = openpyxl.styles.Font(
    b=True, color=WHITE, bold=True
)

center_align = openpyxl.styles.Alignment(
    horizontal="center", vertical="center", wrap_text=True
)


class IndividualAttendanceReport(
    PastUserTimeSheetFilterMixin,
    BackgroundExcelExportMixin,
    OrganizationMixin,
    AttendanceReportPermissionMixin,
    ListViewSetMixin
):
    """
    list:

    filters

        branch=branch_slug
        division=division_slug
        supervisor=user_id
    """
    serializer_class = IndividualAttendanceReportSerializer
    queryset = TimeSheet.objects.all()
    filter_backends = (
        DjangoFilterBackend, SearchFilter, NullsAlwaysLastOrderingFilter, FilterMapBackend)
    filter_map = {
        'username': 'timesheet_user__username',
    }
    search_fields = (
        'timesheet_user__first_name',
        'timesheet_user__middle_name',
        'timesheet_user__last_name',
        'timesheet_user__username'
    )
    ordering_fields_map = {
        'timesheet_user': (
            'timesheet_user__first_name', 'timesheet_user__middle_name',
            'timesheet_user__last_name'),
        'timesheet_for': 'timesheet_for',
        'punch_in': 'punch_in',
        'punch_out': 'punch_out',
        'duration': 'duration',
    }
    export_type = "Daily Attendance"
    export_fields = {
        "Username":"timesheet_user.username",
        "User": "timesheet_user.full_name",
        "Date": "timesheet_for",
        "Day": "day",
        "Job Title": "timesheet_user.detail.job_title",
        "Division": "timesheet_user.detail.division",
        "Employment Level": "timesheet_user.detail.employment_level",
        "Punch In Date": "punch_in.date",
        "Punch In Time": "punch_in_time",
        "Punch Out Date": "punch_out.date",
        "Punch Out Time": "punch_out_time",
        "Duration": "duration",
        "Punctuality": "punctuality",
        "Shift Remarks": "get_coefficient_display",
        "Leave Remarks": "get_leave_coefficient_display",
        "Punch In Category": "punch_in_category",
        "Punch Out Category": "punch_out_category",
    }
    permission_classes = [
        permission_factory.build_permission(
            "IndividualAttendanceReportPermission",
            limit_read_to=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_REPORTS_PERMISSION,
                ATTENDANCE_IMPORT_PERMISSION,
                HAS_PERMISSION_FROM_METHOD
            ]
        )
    ]
    notification_permissions = [ATTENDANCE_REPORTS_PERMISSION]

    def get_queryset(self):

        supervisor_id = self.request.query_params.get('supervisor')
        fil = dict(
            timesheet_user__attendance_setting__isnull=False,
            timesheet_for__lte=get_today(),
            timesheet_user__detail__organization=self.organization
        )

        if supervisor_id:
            if supervisor_id == str(self.request.user.id):
                fil.update({
                    'timesheet_user_id__in':
                        self.request.user.subordinates_pks
                })
            else:
                # if supervisor does not match return none
                return super().get_queryset().none()

        return super().get_queryset().filter(**fil).annotate(
            duration=F('worked_hours')
        ).select_related(
            'timesheet_user',
            'timesheet_user__detail',
            'timesheet_user__detail__organization',
            'timesheet_user__detail__division',
            'timesheet_user__detail__job_title',
            'timesheet_user__detail__employment_level'
        )

    def filter_queryset(self, queryset):
        branch = self.request.query_params.get("branch")
        division = self.request.query_params.get("division")
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date", get_today())
        punch_in = self.request.query_params.get("punch_in_category")
        punch_out = self.request.query_params.get("punch_out_category")

        queryset = super().filter_queryset(queryset)
        if branch:
            queryset = queryset.filter(
                timesheet_user__detail__branch__slug=branch
            )

        if division:
            queryset = queryset.filter(
                timesheet_user__detail__division__slug=division
            )

        if punch_in:
            punch_in = punch_in.replace("_", " ").title()
            if punch_in == MISSING:
                queryset = queryset.filter(punch_in__isnull=True,)
            else:
                queryset = queryset.filter(
                    timesheet_entries__entry_type=PUNCH_IN,
                    timesheet_entries__category=punch_in
                )

        if punch_out:
            punch_out = punch_out.replace("_", " ").title()
            if punch_out == MISSING:
                queryset = queryset.filter(punch_out__isnull=True,)
            else:
                queryset = queryset.filter(
                    timesheet_entries__entry_type=PUNCH_OUT,
                    timesheet_entries__category=punch_out
                )

        if start_date and end_date:
            try:
                queryset = queryset.filter(
                    timesheet_for__gte=start_date,
                    timesheet_for__lte=end_date
                )
            except (ValidationError, TypeError):
                pass
        return queryset.distinct()

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        queryset = self.filter_queryset(self.get_queryset())
        aggregates = {
            display: humanize_interval(
                value
            ) for display, value
            in queryset.aggregate(
            average_in=Avg(Extract('punch_in', 'Hour')) * 60 * 60,
            average_out=Avg(Extract('punch_out', 'Hour')) * 60 * 60,
            average_work_time=Avg(F('punch_out') - F('punch_in'))
        ).items()
        }
        # The multiplication is done because the Hour is extracted and
        # averaged. But, the `humanize_interval` expects seconds.
        response.data.update({
            'punctuality': queryset.filter(
                Q(
                    Q(coefficient=WORKDAY) &
                    ~Q(leave_coefficient=FULL_LEAVE)
                )
            ).aggregate(
                punc=Avg(Coalesce(F('punctuality'), 0.0))
            ).get('punc'),
            **aggregates
        })
        return response

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        def punch_time(time):
            return time.astimezone().strftime('%H:%M:%S %p') if time else None

        obj.punch_in_time = punch_time(obj.punch_in)
        obj.punch_out_time = punch_time(obj.punch_out)
        obj.duration = humanize_interval(obj.duration) if obj.duration else None
        return obj

    def get_frontend_redirect_url(self):
        return f'/admin/{self.organization.slug}/attendance/reports/daily-attendance'


class IndividualAttendanceOverviewView(
    BackgroundExcelExportMixin,
    PastUserTimeSheetFilterMixin,
    DateRangeParserMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    queryset = TimeSheet.objects.filter()
    serializer_class = IndividualAttendanceOverviewSerializer
    filter_backends = (
        FilterMapBackend, NullsAlwaysLastOrderingFilter
    )
    filter_map = {
        'start_date': 'timesheet_for__gte',
        'end_date': 'timesheet_for__lte',
        'user': 'timesheet_user',
        'username': 'timesheet_user__username',
    }
    ordering_fields_map = {
        'date': 'timesheet_for',
        'punch_in': 'punch_in_time',
        'punch_out': 'punch_out_time',
        'worked_hours': 'worked_hours',
        'working_hours': 'expected_work_hours',
        'overtime': 'overtime__overtime_detail__claimed_overtime',
        'punctuality': 'punctuality'
    }
    export_type = 'IndividualAttendanceOverview'
    export_fields = []

    def get_queryset(self):
        qs = super().get_queryset()
        fil = {
            'timesheet_user__detail__organization': self.organization
        }
        action = self.request.query_params.get('as')
        if action:
            if action == 'supervisor':
                fil = {
                    'timesheet_user_id__in': self.request.user.subordinates_pks
                }
            if action == 'hr':
                # if not self.request.user.is_audit_user:
                if not validate_permissions(
                    self.request.user.get_hrs_permissions(self.organization),
                    ATTENDANCE_PERMISSION,
                    ATTENDANCE_REPORTS_PERMISSION
                ):
                    raise PermissionDenied
        else:
            fil.update({
                'timesheet_user_id': self.request.user.id
            })
        return qs.filter(**fil)

    def filter_queryset(self, queryset):
        qs = super().filter_queryset(queryset)
        qs = qs.annotate(
            expected_work_hours=F('expected_punch_out') - F(
                'expected_punch_in'),
            punch_in_time=Cast(
                F('punch_in'), TimeField()
            ),
            punch_out_time=Cast(
                F('punch_out'), TimeField()
            ),
        )
        return qs.filter(
            timesheet_for__lte=timezone.now()
        ).select_related(
            'overtime'
        )

    def get_frontend_redirect_url(self):
        return '/user/attendance/reports/individual-daily-attendance'

    def collect_statistics(self):
        qs = self.filter_queryset(
            self.get_queryset()
        )
        if not qs:
            return {
                'average_in': 'N/A',
                'average_out': 'N/A',
                'average_worked_hours': 'N/A',
                'most_worked': 'N/A',
                'least_worked': 'N/A'
            }
        most_worked = getattr(qs.filter(
            worked_hours__isnull=False
        ).order_by('-worked_hours').first(), 'timesheet_for', None)
        least_worked = getattr(qs.filter(
            worked_hours__isnull=False
        ).order_by('worked_hours').first(), 'timesheet_for', None)
        # [HRIS-1974#2] For Average In Time Calculation, Off-day Is also included. [It shouldn't]
        # Modifications
        # Off Day and Full Leave shall be ignored.
        # For Avg. In, (First Half Leave will be excluded) and so on for Second half leave request.
        # For Average worked hours, select Workday with No Leave days only.
        base_aggregate_qs = qs.filter(
            coefficient=WORKDAY
        )
        average_in_filter = Q(
            leave_coefficient__in=(
                NO_LEAVE, SECOND_HALF
            )
        )
        average_out_filter = Q(
            leave_coefficient__in=(
                NO_LEAVE, FIRST_HALF
            )
        )
        aggregates = base_aggregate_qs.aggregate(
            average_in=Coalesce(
                Avg(
                    Extract('punch_in', 'Hour'),
                    filter=average_in_filter,
                    output_field=FloatField()
                ), 0.0
            ) * 60 * 60 + Coalesce(
                Avg(
                    Extract('punch_in', 'Minute'),
                    filter=average_in_filter,
                    output_field=FloatField()
                ), 0.0
            ) * 60 + Coalesce(
                Avg(
                    Extract('punch_in', 'Second'),
                    filter=average_in_filter,
                    output_field=FloatField()
                ), 0.0
            ),
            average_out=Coalesce(
                Avg(
                    Extract('punch_out', 'Hour'),
                    filter=average_out_filter,
                    output_field=FloatField()
                ), 0.0
            ) * 60 * 60 + Coalesce(
                Avg(
                    Extract('punch_out', 'Minute'),
                    filter=average_out_filter,
                    output_field=FloatField()
                ), 0.0
            ) * 60 + Coalesce(
                Avg(
                    Extract('punch_out', 'Second'),
                    filter=average_out_filter,
                    output_field=FloatField()
                ), 0.0
            ),
            average_worked_hours=Avg(
                'worked_hours',
                filter=Q(coefficient=WORKDAY, leave_coefficient=NO_LEAVE),
                output_field=DurationField()
            ),
        )
        stats = {
            'average_in': humanize_interval(aggregates.get('average_in')),
            'average_out': humanize_interval(aggregates.get('average_out')),
            'average_worked_hours': humanize_interval(
                aggregates.get('average_worked_hours')
            ),
            'most_worked': most_worked,
            'least_worked': least_worked
        }
        return stats

    def prepare_dataset(self):
        qs = self.filter_queryset(self.get_queryset())
        if not qs:
            raise ValidationError({
                'message': 'No data to export/email.'
            })
        stats = self.collect_statistics()
        data = Dataset()
        data.title = 'Individual Overview Report'
        ctx = self.get_serializer_context()
        ctx.update({
            'identify_user': True
        })
        data.dict = IndividualAttendanceOverviewSerializer(
            qs,
            exclude_fields=['expected_punch_in', 'expected_punch_out'],
            context=ctx,
            many=True
        ).data
        headers_length = len(data.headers)
        data.headers = prettify_headers(
            data.headers
        )
        # Add new lines, for visual effect
        data.append(('',) * headers_length)
        data.append(('',) * headers_length)

        stats_length = len(stats.keys())
        data.append(prettify_headers(
            [*stats.keys(), *[''] * (headers_length - stats_length)]
        ))
        data.append(
            [*stats.values(), *[''] * (headers_length - stats_length)]
        )
        return data

    def get_extra_export_data(self):
        """
        Extra export data
        type dict (like context for template)
        default --> {
            "serializer_class_params": self.get_serializer_class_params(),
        }
        :return:
        """
        extra = super().get_extra_export_data()
        extra.update({
            'statistics': self.collect_statistics(),
            'organization': self.get_organization()
        })
        return extra

    @classmethod
    def get_exported_file_content(cls, queryset, title, columns, extra_content, description=None, **kwargs):
        qs = queryset
        ctx = {
            'identify_user': True
        }
        stats = extra_content['statistics']
        data = IndividualAttendanceOverviewExportSerializer(
            qs,
            exclude_fields=['expected_punch_in', 'expected_punch_out'],
            context=ctx,
            many=True
        ).data
        headers = [
            'user',
            'username',
            'timesheet_for',
            'day',
            'punch_in_date',
            'punch_in_time',
            'punch_out_date',
            'punch_out_time',
            'worked_hours',
            'expected_work_hours',
            'overtime',
            'punctuality',
            'coefficient',
            'leave_coefficient',
        ]
        organization = extra_content['organization']
        wb = ExcelExport.process(data, columns=headers, organization=organization)
        ws = wb.active

        ws.append([])
        ws.append([])
        ws.append(prettify_headers(stats.keys()))
        ws.append(list(stats.values()))

        return ContentFile(save_virtual_workbook(wb))

    def list(self, request, *args, **kwargs):
        # export = request.query_params.get('export')
        # if export and export == 'xlsx':
        #     wb = self.prepare_export()
        #     return ExcelExport.get_response_for_workbook(wb)

        stats = self.collect_statistics()
        ret = super().list(self.request, *args, **kwargs)
        ret.data.update({
            'statistics': stats
        })
        return ret

    @action(
        methods=['POST'],
        detail=False,
        serializer_class=type(
            'UserSelectSerializer',
            (Serializer,),
            {
                'user': serializers.PrimaryKeyRelatedField(
                    queryset=get_user_model().objects.all()
                )
            }
        ),
        url_path='send-email'
    )
    def send_email(self, request, **kwargs):
        ser = self.serializer_class(
            data=request.data
        )
        ser.is_valid(raise_exception=True)
        data = self.prepare_dataset().html
        user_full_name = get_object_or_404(
            get_user_model().objects.filter(),
            pk=request.query_params.get('user')
        ).full_name
        st_dt = request.query_params.get('start_date')
        ed_dt = request.query_params.get('end_date')
        dt_range = None
        if st_dt:
            dt_range = f"{st_dt}"
        if ed_dt:
            dt_range += f' to {ed_dt}'
        html_message = render_to_string(
            'overview_report.html',
            context={
                'datatable': data,
                'date_range': dt_range,
                'report_type': 'Individual',
                'full_name': user_full_name
            }
        )
        async_task(
            custom_mail,
            subject='Individual Overview Report',
            message=data,
            from_email=INFO_EMAIL,
            recipient_list=[
                ser.validated_data.get('user').email
            ],
            html_message=html_message
        )
        return Response({
            'message': 'Email has been processed'
        })


class OvertimeDetailReport(
    ModeFilterQuerysetMixin,
    PastUserTimeSheetFilterMixin,
    BackgroundExcelExportMixin,
    DateRangeParserMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    queryset = TimeSheet.objects.all()
    serializer_class = OvertimeDetailReportSerializer
    filter_backends = (
        FilterMapBackend, NullsAlwaysLastOrderingFilter
    )
    filter_map = {
        'start_date': 'timesheet_for__gte',
        'end_date': 'timesheet_for__lte',
        'user': 'timesheet_user'
    }
    ordering_fields_map = {
        'date': 'timesheet_for',
        'punch_in': 'punch_in_time',
        'punch_out': 'punch_out_time',
        'worked_hours': 'worked_hours',
        'working_hours': 'expected_work_hours',
        'overtime': 'overtime__overtime_detail__claimed_overtime',
        'punctuality': 'punctuality',
        'overtime_worked': 'overtime_worked',
        'overtime_claimed': 'overtime_claimed'
    }

    export_type = 'Overtime detail report'
    export_fields = []
    permission_to_check = (
        ATTENDANCE_PERMISSION,
        ATTENDANCE_REPORTS_PERMISSION
    )
    user_definition = 'timesheet_user'

    def filter_queryset(self, queryset, remove_non_ot=True):
        user_id = self.request.query_params.get('user')
        if user_id and user_id.isdigit() and USER.objects.filter(
                pk=user_id
        ).exists():
                qs = super().filter_queryset(queryset)
        else:
            qs = queryset.none()
        qs = qs.annotate(
            expected_work_hours=F('expected_punch_out') - F(
                'expected_punch_in'),
            punch_in_time=Cast(
                F('punch_in'), TimeField()
            ),
            punch_out_time=Cast(
                F('punch_out'), TimeField()
            ),
            overtime_claimed=Case(
                When(
                    overtime__claim__status=UNCLAIMED,
                    then=Value(timezone.timedelta(minutes=0))
                ),
                default=F('overtime__overtime_detail__claimed_overtime')
            ),
        ).annotate(
            ot_worked=Case(
                When(
                    coefficient=WORKDAY,
                    then=F('worked_hours') - F('expected_work_hours')
                ),
                default=F('worked_hours'),
                output_field=DurationField()
            ),
        ).annotate(
            overtime_worked=Case(
                When(ot_worked__gt=ZERO_DURATION, then=F('ot_worked')),
                default=Value(ZERO_DURATION),
                output_field=DurationField()
            )
        )
        qs = qs.filter(
            timesheet_for__lte=timezone.now()
        ).select_related(
            'overtime'
        )
        if remove_non_ot:
            return qs.filter(overtime__isnull=False)
        return qs

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        worked_hours = self.filter_queryset(
            self.get_queryset(),
            remove_non_ot=False
        ).aggregate(
            sum_worked_hours=Sum('worked_hours'),
        ).get('sum_worked_hours')
        qs = self.filter_queryset(self.get_queryset())
        agg = qs.aggregate(
            sum_total_overtime=Sum(
                'overtime_worked'
            ),
            sum_confirmed_overtime=Sum(
                'overtime__overtime_detail__claimed_overtime',
                filter=Q(overtime__claim__status=CONFIRMED)
            )
        )
        response.data.update({
            'information_bar': {
                **{
                    i: humanize_interval(v) for i, v in agg.items()
                },
                'sum_worked_hours': humanize_interval(worked_hours)
            }
        })
        return response

    @staticmethod
    def generate_export(cls, qs, export, aggregate_queryset, **kwargs):
        report_title = 'Overtime Detailed Daily Report'
        workbook = openpyxl.Workbook()
        workbook.active.title = report_title
        report_sheet = workbook.active

        ExcelExport.insert_org_info(report_sheet, organization=export.organization)
        report_sheet.append([])

        queryset = qs.annotate(
            expected_work_hours=F('expected_punch_out') - F(
                'expected_punch_in'),
            punch_in_time=Cast(
                F('punch_in'), TimeField()
            ),
            punch_out_time=Cast(
                F('punch_out'), TimeField()
            ),
            overtime_claimed=Case(
                When(
                    overtime__claim__status=UNCLAIMED,
                    then=Value(timezone.timedelta(minutes=0))
                ),
                default=F('overtime__overtime_detail__claimed_overtime'),
                output_field=DurationField()
            ),
            overtime_approved=Case(
                When(
                    overtime__claim__status__in=['Approved', 'Confirmed'],
                    then=F('overtime__overtime_detail__claimed_overtime')
                ),
                default=Value(timezone.timedelta(minutes=0))
            ),
            overtime_confirmed=Case(
                When(
                    overtime__claim__status='Confirmed',
                    then=F('overtime__overtime_detail__claimed_overtime')
                ),
                default=Value(timezone.timedelta(minutes=0))
            ),
            overtime_declined=Case(
                When(
                    overtime__claim__status='Declined',
                    then=F('overtime__overtime_detail__claimed_overtime')
                ),
                default=Value(timezone.timedelta(minutes=0))
            ),
        ).annotate(
            ot_worked=Case(
                When(
                    coefficient=WORKDAY,
                    then=F('worked_hours')
                ),
                default=F('worked_hours'),
                output_field=DurationField()
            ),
        ).annotate(
            overtime_worked=Case(
                When(ot_worked__gt=ZERO_DURATION, then=F('ot_worked')),
                default=Value(ZERO_DURATION),
                output_field=DurationField()
            )
        )

        user_id = kwargs.get('user')
        if user_id and user_id.isdigit():
            employee = get_object_or_404(
                get_user_model(), pk=user_id
            )
            employee_name = employee.full_name
            selected_division = nested_getattr(employee, 'detail.division.name')
        else:
            employee = None
            employee_name = ''
            selected_division = kwargs.get('division')

        if not employee:
            queryset = queryset.none()
            aggregate_queryset = queryset
        agg = queryset.aggregate(
            sum_total_overtime=Sum(
                'overtime__overtime_detail__claimed_overtime'
            ),
            sum_confirmed_overtime=Sum(
                'overtime__overtime_detail__claimed_overtime',
                filter=Q(overtime__claim__status=CONFIRMED)
            ),
            sum_expected_hours=Sum('expected_work_hours'),
        )
        worked_overtime = queryset.aggregate(
            worked_ot=Sum('overtime_worked')
        ).get('worked_ot') or ZERO_DURATION

        aggregates = {
            i: humanize_interval(v) for i, v in agg.items()
        }

        start_date, end_date = kwargs.get('start_date'), kwargs.get(
            'end_date'
        )
        selected_range = f'{start_date}  to {end_date}'

        worked_hours = aggregate_queryset.aggregate(
            sum_worked_hours=Sum('worked_hours'),
        ).get(
            'sum_worked_hours'
        ) or ZERO_DURATION

        worked_overtime = humanize_interval(worked_overtime)
        confirmed_overtime = aggregates.get('sum_confirmed_overtime')

        report_sheet.append([report_title])
        report_sheet.append(['Employee Name', employee_name])
        report_sheet.append(['Worked Hours', humanize_interval(worked_hours)])
        report_sheet.append(['Worked Overtime', worked_overtime])
        report_sheet.append(['Confirmed Overtime', confirmed_overtime])

        report_sheet.append(['Division', selected_division])
        report_sheet.append(['Date Range', selected_range])

        headers = (
            ('Date', 'timesheet_for'),
            ('Day', 'day'),
            ('Punch In', 'punch_in'),
            ('Punch Out', 'punch_out'),
            ('Total Worked Hours', 'worked_hours'),
            ('Total Overtime Worked', 'overtime_worked'),
            ('Total Claimed', 'overtime_claimed'),
            ('Total Approved', 'overtime_approved'),
            ('Total Confirmed', 'overtime_confirmed'),
            ('Total Declined', 'overtime_declined')
        )

        report_sheet.append([])
        report_sheet.append([])

        report_sheet.append([h[0] for h in headers])
        day_display = Case(
            *[
                When(
                    wday=Value(day_value),
                    then=Value(day_display)
                ) for day_value, day_display in WEEK_DAYS_CHOICES],
            output_field=CharField()
        )
        queryset = queryset.annotate(
            wday=ExtractWeekDay('timesheet_for')
        ).annotate(day=day_display)
        for row in queryset.values_list(*[h[1] for h in headers]):
            normalized = [
                c.replace(microsecond=0).astimezone().strftime(
                    '%Y-%m-%d %I:%M:%S %p'
                ) if isinstance(c, datetime) else c for c in row
            ]
            report_sheet.append(normalized)

        cls.save_file_content(
            export_instance=export,
            file_content=ContentFile(save_virtual_workbook(workbook))
        )
        export.status = PROCESSING
        export.associated_permissions.add(
            *HRSPermission.objects.filter(
                code__in=[x.get('code') for x in [ATTENDANCE_REPORTS_PERMISSION]]
            )
        )
        export.save()
        cls.send_success_notification(
            obj=export,
            url=f'/admin/{export.organization.slug}/attendance/reports/overtime-details',
            exported_as='Admin',
            permissions=[ATTENDANCE_REPORTS_PERMISSION]
        )

    def _export_post(self):
        """
            Start task export process in background
        """
        if has_pending_export(
                export_type=self.get_export_type(),
                user=self.request.user,
                exported_as=self.get_exported_as()
        ):
            return Response({
                'message': 'Previous request for generating report is being '
                           'currently processed, Please try back later'},
                status=status.HTTP_202_ACCEPTED)

        if hasattr(self,
                   'get_organization') and self.get_exported_as() == ADMIN:
            organization = self.get_organization()
        else:
            organization = None

        export = Export.objects.create(
            user=self.request.user,
            name=self.get_export_name(),
            exported_as=self.get_exported_as(),
            export_type=self.get_export_type(),
            organization=organization,
            status=QUEUED,
            remarks=''
        )
        st, ed = self.get_parsed_dates()
        _ = async_task(
            OvertimeDetailReport.generate_export,
            self.__class__,
            self.filter_queryset(self.get_queryset()),
            export,
            self.filter_queryset(self.get_queryset(), remove_non_ot=False),
            user=self.request.query_params.get('user'),
            division=self.request.query_params.get('division'),
            start_date=st,
            end_date=ed,
        )
        return Response({
            'message': 'Your request is being processed in the background . '
                       'Please check back later'})


class OvertimeDetailExportMetaDataSerializer(UserThinSerializer):
    code = serializers.ReadOnlyField(source='detail.code')
    division = serializers.ReadOnlyField(source='detail.division.name')
    job_title = serializers.ReadOnlyField(source='detail.job_title.title')

    # first level authority, second level authority, third level authority
    fla = serializers.SerializerMethodField()
    sla = serializers.SerializerMethodField()
    tla = serializers.SerializerMethodField()

    class Meta(UserThinSerializer.Meta):
        fields_map = {
            'id': 'ID',
            'code': 'Employee Code',
            'full_name': 'Full Name',
            'division': 'Division',
            'job_title': 'Job Title',
            'fla': 'First Level Supervisor',
            'sla': 'Second Level Supervisor',
            'tla': 'Third Level Supervisor',
        }
        fields = [
            'id',
            'code',
            'full_name',
            'division',
            'job_title',
            'fla',
            'sla',
            'tla',
        ]

    @cached_property
    def supervisors(self):
        user = self.instance
        return {
            authority.authority_order: authority.supervisor.full_name
            for authority in user.supervisors.all()
        }

    def get_fla(self, _):
        return self.supervisors.get(1)

    def get_sla(self, _):
        return self.supervisors.get(2)

    def get_tla(self, _):
        return self.supervisors.get(3)


class OvertimeDetailExportSerializer(DummySerializer):
    fields_map = {
        'requested_date': "Requested Date",
        'claim_date': "Requested For",
        'shift_hours': "Shift Hours",
        'punch_in_overtime': "Punch In Overtime",
        'punch_out_overtime': "Punch Out Overtime",
        'offday_overtime': "Offday / Holiday Overtime",
        'coefficient': "Day",
        'worked_hours': "Total Worked Hours",
        'actual': "Actual Overtime",
        'claimed': "Claimed Overtime",
        'confirmed': "Confirmed Overtime",
        'description': "Reason for Overtime",
        'status': "Status",
        'first_level_supervisor_action': "Action By[1st Level]",
        'second_level_supervisor_action': "Action By[2nd Level]",
        'third_level_supervisor_action': "Action By[3rd Level]",
        'confirmed_by': "Confirmed By",
        'confirmed_date': "Confirmed Date",
    }
    fields_list = [
        'requested_date',
        'claim_date',
        'shift_hours',
        'punch_in_overtime',
        'punch_out_overtime',
        'offday_overtime',
        'coefficient',
        'worked_hours',
        'actual',
        'claimed',
        'confirmed',
        'description',
        'status',
        'first_level_supervisor_action',
        'second_level_supervisor_action',
        'third_level_supervisor_action',
        'confirmed_by',
        'confirmed_date',
    ]

    def pretty_name(self, field_name):
        return self.fields_map.get(field_name)

    def get_fields(self):
        return {
            key: serializers.SerializerMethodField() for key in self.fields_list
        }

    def get_requested_date(self, overtime_claim):
        return format_timezone(
            overtime_claim.overtime_histories.filter(
                action_performed__in=[REQUESTED, UNCLAIMED],
            ).first().created_at.astimezone()
        )

    def get_claim_date(self, overtime_claim):
        return overtime_claim.overtime_entry.timesheet.timesheet_for

    def get_shift_hours(self, overtime_claim):
        timesheet = overtime_claim.overtime_entry.timesheet
        alt = timesheet.punch_in.astimezone().time(), timesheet.punch_out.astimezone().time()
        return timesheet.work_time.display if timesheet.work_time else '%s - %s' % alt

    def get_punch_in_overtime(self, overtime_claim):
        if overtime_claim.overtime_entry.timesheet.coefficient == WORKDAY:
            return humanize_interval(
                overtime_claim.overtime_entry.overtime_detail.punch_in_overtime
            )
        return ""

    def get_punch_out_overtime(self, overtime_claim):
        if overtime_claim.overtime_entry.timesheet.coefficient == WORKDAY:
            return humanize_interval(
                overtime_claim.overtime_entry.overtime_detail.punch_out_overtime
            )
        return ""

    def get_offday_overtime(self, overtime_claim):
        if overtime_claim.overtime_entry.timesheet.coefficient == WORKDAY:
            return ""
        return humanize_interval(
            overtime_claim.overtime_entry.overtime_detail.punch_in_overtime
        )

    def get_coefficient(self, overtime_claim):
        return overtime_claim.overtime_entry.timesheet.get_coefficient_display()

    def get_actual(self, overtime_claim):
        return humanize_interval(
            overtime_claim.overtime_entry.overtime_detail.total_seconds
        )

    def get_confirmed(self, overtime_claim):
        return humanize_interval(
            overtime_claim.overtime_entry.overtime_detail.claimed_overtime
        )

    def get_description(self, overtime_claim):
        # get the requested remark
        requested_remark = overtime_claim.overtime_histories.filter(
            action_performed=REQUESTED
        ).values_list('remark', flat=True).first()
        if not action:
            return ""
        return requested_remark

    def get_action_by(self, overtime_claim, user):
        action = overtime_claim.overtime_histories.filter(
            action_performed_by=user,
        ).first()
        if not action:
            return ""
        return "%s\n%s" % (
            action.action_performed_by.full_name,
            action.action_performed_by.detail.job_title.title
        )

    def get_worked_hours(self, overtime_claim):
        timesheet = overtime_claim.overtime_entry.timesheet
        return humanize_interval(
            timesheet.punch_out - timesheet.punch_in
        )

    def get_claimed(self, overtime_claim):
        return humanize_interval(
            overtime_claim.overtime_entry.overtime_detail.claimed_overtime
        )

    def get_status(self, overtime_claim):
        return overtime_claim.get_status_display()

    def get_first_level_supervisor_action(self, overtime_claim):
        return self.get_action_by(
            overtime_claim,
            overtime_claim.overtime_entry.user.first_level_supervisor
        )

    def get_second_level_supervisor_action(self, overtime_claim):
        return self.get_action_by(
            overtime_claim,
            overtime_claim.overtime_entry.user.user_supervisors.filter(
                authority_order=2
            ).values_list('supervisor', flat=True).first()
        )

    def get_third_level_supervisor_action(self, overtime_claim):
        return self.get_action_by(
            overtime_claim,
            overtime_claim.overtime_entry.user.user_supervisors.filter(
                authority_order=3
            ).values_list('supervisor', flat=True).first()
        )

    def get_confirmed_by(self, overtime_claim):
        confirmed = overtime_claim.overtime_histories.filter(
            action_performed=CONFIRMED,
        ).first()
        if confirmed:
            return "%s\n%s" % (
                confirmed.action_performed_by.full_name,
                confirmed.action_performed_by.detail.job_title.title
            )
        return ""

    def get_confirmed_date(self, overtime_claim):
        confirmed = overtime_claim.overtime_histories.filter(
            action_performed=CONFIRMED,
        ).first()
        if confirmed:
            return format_timezone(confirmed.created_at.astimezone())
        return ""


class MonthlyAttendanceOverview(
    PastUserTimeSheetFilterMixin,
    DateRangeParserMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    queryset = TimeSheet.objects.all()
    filter_backends = (FilterMapBackend, OrderingFilterMap)
    ordering_fields_map = {
        'month': 'timesheet_for__month',
        'user': 'timesheet_user'
    }
    filter_map = {
        'user': 'timesheet_user',
        'start_date': 'timesheet_for__gte',
        'end_date': 'timesheet_for__lte',
    }

    def get_queryset(self):
        qs = super().get_queryset()
        fil = {
            'timesheet_user_id': self.request.user.id,
            'timesheet_user__detail__organization': self.organization
        }
        action = self.request.query_params.get('as')
        if action:
            if action == 'supervisor':
                fil = {
                    'timesheet_user_id__in': self.request.user.subordinates_pks
                }
            if action == 'hr':
                # if not self.request.user.is_audit_user:
                if not validate_permissions(
                    self.request.user.get_hrs_permissions(self.organization),
                    ATTENDANCE_PERMISSION,
                    ATTENDANCE_REPORTS_PERMISSION
                ):
                    raise PermissionDenied
                fil = {
                    'timesheet_user__detail__organization': self.organization
                }
        qs = qs.annotate(
            expected_work_hours=F('expected_punch_out') - F(
                'expected_punch_in'),
            punch_in_time=Cast(
                F('punch_in'), TimeField()
            ),
            punch_out_time=Cast(
                F('punch_out'), TimeField()
            ),
        ).filter(**fil)
        return qs

    def prepare_data_set(self, ret):
        results = ret.get('results')
        if not results:
            raise ValidationError({
                'message': 'No data to export/email.'
            })
        sub = ret.get('statistics')
        data = Dataset()
        data.title = 'Monthly Overview Report'
        data.dict = results
        data.headers = prettify_headers(
            data.headers
        )

        user_defined = self.request.query_params.get('user')
        if user_defined and user_defined.isdigit():
            user = get_user_model().objects.filter(
                pk=user_defined
            ).first()
            if user:
                data.title = user.full_name

        data.append(('',) * len(data.headers))
        data.append(('',) * len(data.headers))
        data.append(prettify_headers(
            [*sub.keys(), '', '']
        ))
        data.append([*sub.values(), '', ''])
        return data

    def prepare_export(self, results, stats):
        headers = {
            'Month': 'month.name',
            'Average In': 'average_in',
            'Average Out': 'average_out',
            'Punctuality': 'punctuality',
            'Total Worked Hours': 'total_worked_hours',
            'Total Working Hours': 'total_working_hours',
            'Overtime': 'overtime'
        }

        organization = self.get_organization()

        user_defined = self.request.query_params.get('user')
        title = 'Monthly Attendance'
        description = []
        if user_defined and user_defined.isdigit():
            user = get_user_model().objects.filter(
                pk=user_defined
            ).first()
            if user:
                description.append(f'User: {user.full_name}')
                description.append(f'Username: {user.username}')
        wb = ExcelExport.process(
            results,
            title=title,
            columns=headers,
            organization=organization,
            description=description
        )

        ws = wb.active
        ws.append([])
        ws.append([])

        stat_headers = []
        stat_values = []
        for key, value in stats.items():
            stat_headers.append(pretty_name(key))
            if key in ['most_worked_month', 'least_worked_month']:
                stat_values.append(value['name'])
            else:
                stat_values.append(value)

        ws.append(stat_headers)
        ws.append(stat_values)

        return wb

    def list(self, request, *args, **kwargs):
        now = timezone.now()
        queryset = super().filter_queryset(
            self.get_queryset()
        )
        show_year = self.get_parsed_dates()[1].year
        month_order = request.query_params.get('ordering')
        if month_order and month_order == '-month':
            ord = 'timesheet_for__month'
        else:
            ord = '-timesheet_for__month'
        months_available = queryset.order_by(
            ord
        ).values_list(
            'timesheet_for__year',
            'timesheet_for__month',
        ).distinct()
        ret = list()
        timesheets_filter = {
            'coefficient': WORKDAY,
            'timesheet_for__range': (
                self.get_parsed_dates()
            ),
            'leave_coefficient__in': [
                NO_LEAVE, FIRST_HALF, SECOND_HALF
            ]
        }
        for year, month in months_available:
            this_month = now.replace(
                day=1,  # omits the possibility of happening (2,1) -> will break.
                month=month
            ).strftime('%b')
            month_qs = queryset.filter(
                timesheet_for__year=year,
                timesheet_for__month=month
            )
            agg = {
                # avg of in, out, punctuality depends on workday
                # summations, overtime, total_working_hours, and expected dont.
                **month_qs.filter(
                    coefficient=WORKDAY
                ).aggregate(
                    average_in=Coalesce(
                        Avg(Extract('punch_in', 'Hour')), 0.0
                    ) * 60 * 60 + Coalesce(
                        Avg(Extract('punch_in', 'Minute')), 0.0
                    ) * 60 + Coalesce(
                        Avg(Extract('punch_in', 'Second')), 0.0
                    ),
                    average_out=Coalesce(
                        Avg(Extract('punch_out', 'Hour')), 0.0
                    ) * 60 * 60 + Coalesce(
                        Avg(Extract('punch_out', 'Minute')), 0.0
                    ) * 60 + Coalesce(
                        Avg(Extract('punch_out', 'Second')), 0.0
                    ),
                    punctuality=Avg(
                        Coalesce(F('punctuality'), 0.0),
                        filter=Q(
                            **timesheets_filter,
                            work_shift__isnull=False
                        ),
                        output_field=FloatField()
                    )
                ),
                **month_qs.aggregate(
                    total_worked_hours=Sum(
                        'worked_hours',
                        filter=Q(**timesheets_filter),
                        output_field=DurationField()
                    ),
                    total_working_hours=Sum(
                        'expected_work_hours',
                        filter=Q(**timesheets_filter),
                        output_field=DurationField()
                    ),
                    overtime=Sum(
                        F('overtime__overtime_detail__claimed_overtime'),
                        filter=Q(
                            timesheet_for__range=self.get_parsed_dates()
                        ),
                        output_field=DurationField()
                    )
                )
            }

            last_day = calendar.monthrange(year, month)[1]
            ret.append({
                'month':
                    {
                        'name': this_month,
                        'start_date': date(year, month, 1).strftime('%Y-%m-%d'),
                        'end_date': date(year, month, last_day).strftime('%Y-%m-%d')
                    },
                **{
                    'average_in': agg.get('average_in'),
                    'average_out': agg.get('average_out'),
                    'total_worked_hours': agg.get('total_worked_hours').total_seconds() if agg.get(
                        'total_worked_hours') else 0,
                    'total_working_hours': agg.get(
                        'total_working_hours').total_seconds() if agg.get(
                        'total_working_hours') else 0,
                    'overtime': agg.get('overtime'),
                    'punctuality': round(agg.get('punctuality'), 2)
                    if agg.get('punctuality') != None else 'N/A'
                    # intentional check with None, to avoid with 0
                }
            })

        def find_max(month):
            return month.get('total_worked_hours')

        def average(li):
            if not li:
                return 0
            from statistics import mean
            return mean(li)

        aggregates = {
            'average_in': humanize_interval(average([
                r.get('average_in') for r in ret
            ])),
            'average_out': humanize_interval(average([
                r.get('average_out') for r in ret
            ])),
            'average_worked_hours': humanize_interval(average([
                r.get('total_worked_hours') for r in ret
            ]))
        }

        def sort_data(arg1, reverse=False):
            return sorted(ret, key=lambda x: x[arg1], reverse=reverse)

        ordering_mapper = {
            'average_in': sort_data('average_in', reverse=True),
            '-average_in': sort_data('average_in'),
            'average_out': sort_data('average_out', reverse=True),
            '-average_out': sort_data('average_out'),
            'average_worked_hours': sort_data('total_worked_hours', reverse=True),
            '-average_worked_hours': sort_data('total_worked_hours'),
            'average_working_hours': sort_data('total_working_hours', reverse=True),
            '-average_working_hours': sort_data('total_working_hours'),
        }
        if month_order in ordering_mapper.keys():
            ret = ordering_mapper.get(month_order)

        for item in ret:
            for key, value in item.items():
                if key not in ["month", "punctuality"]:
                    item[key] = humanize_interval(value)
        page = self.paginate_queryset(ret)
        output_data = self.get_paginated_response(page)
        statistics = {
            'average_in': humanize_interval(
                aggregates.get('average_in')),
            'average_out': humanize_interval(
                aggregates.get('average_out')),
            'total_worked_hours': humanize_interval(
                aggregates.get('average_worked_hours')
            ),
            'most_worked_month': max(
                ret,
                key=find_max
            ).get('month') if ret else 'N/A',
            'least_worked_month': min(
                ret,
                key=find_max
            ).get('month') if ret else 'N/A',
        }
        output_data.data["statistics"] = statistics
        export = request.query_params.get('export')
        if export and export == 'xlsx' and not getattr(
                self, 'from_mail', False
        ):
            wb = self.prepare_export(ret, statistics)
            return ExcelExport.get_response_for_workbook(wb)
        return output_data

    @action(
        methods=['POST'],
        detail=False,
        serializer_class=type(
            'UserSelectSerializer',
            (Serializer,),
            {
                'user': serializers.PrimaryKeyRelatedField(
                    queryset=get_user_model().objects.all()
                )
            }
        ),
        url_path='send-email'
    )
    def send_email(self, request, **kwargs):
        ser = self.serializer_class(
            data=request.data
        )
        ser.is_valid(raise_exception=True)
        setattr(self, 'from_email', True)
        ret = self.list(request, **kwargs)
        if not ret.data.get('results'):
            raise ValidationError({
                'message': 'No Data was generated.'
            })
        data = self.prepare_data_set(ret.data).html
        user_full_name = get_object_or_404(
            get_user_model().objects.filter(),
            pk=self.request.query_params.get('user')
        ).full_name
        st_dt = request.query_params.get('start_date')
        ed_dt = request.query_params.get('end_date')
        dt_range = None
        if st_dt:
            dt_range = f"{st_dt}"
        if ed_dt:
            dt_range += f' to {ed_dt}'
        html_message = render_to_string(
            'overview_report.html',
            context={
                'datatable': data,
                'date_range': dt_range,
                'report_type': 'Monthly',
                'full_name': user_full_name
            }
        )
        custom_mail(
            subject='Monthly Overview Report',
            message=data,
            from_email=INFO_EMAIL,
            recipient_list=[
                ser.validated_data.get('user').email
            ],
            html_message=html_message
        )
        return Response({
            'message': 'Email has been processed'
        })


class ComparativeOvertimeReport(
    BackgroundExcelExportMixin,
    SupervisorQuerysetViewSetMixin,
    ModeFilterQuerysetMixin,
    DateRangeParserMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    """
    Allowed filters:
    as=hr displays all users in organization,
    as=supervisor displays all subordinates of logged in user.
    as= displays self results in that organization

    supervisor=int displays subordinates of selected user.
    """
    serializer_class = ComparativeOvertimeReportSerializer
    allow_non_supervisor_user = True
    immediate_only = True
    export_type = 'Overtime Comparative Report'
    export_fields = []
    queryset = USER.objects.all()
    permission_to_check = (
        ATTENDANCE_PERMISSION,
        ATTENDANCE_REPORTS_PERMISSION
    )
    notification_permissions = [ATTENDANCE_REPORTS_PERMISSION]
    filter_backends = (
        FilterMapBackend, SearchFilter, NullsAlwaysLastOrderingFilter
    )
    filter_map = {
        'user': 'id',
        'division': 'detail__division__slug',
    }
    search_fields = (
        'first_name', 'middle_name', 'last_name'
    )

    def get_ordering_fields_map(self):
        base_map = dict(
            full_name=(
                'first_name', 'middle_name', 'last_name',
            ),
            total_overtime='total_overtime'
        )
        base_map.update({
            mth.lower(): f'worked_{mth}'.lower() for mth in
            self.fiscal.values_list('display_name', flat=True)
        })
        return base_map

    @cached_property
    def fiscal_year(self):
        fis_filter = self.request.query_params.get('fiscal')
        if fis_filter and fis_filter.isdigit():
            fiscal = FiscalYear.objects.filter(
                organization=self.organization,
                pk=fis_filter,
                category=GLOBAL
            ).first()
        else:
            fiscal = FiscalYear.objects.current(
                organization=self.organization
            )
        return fiscal

    @cached_property
    def fiscal(self):
        fiscal = self.fiscal_year
        if not fiscal:
            return FiscalYearMonth.objects.none()
        return fiscal.fiscal_months.filter(
            start_at__lte=timezone.now()
        )

    @staticmethod
    def annotate_queryset(queryset, ffs, fiscal_start, fiscal_end):
        t_queries = {
            f'worked_{ff.display_name}'.lower(): Subquery(
                TimeSheet.objects.filter(
                    timesheet_for__range=(ff.start_at, ff.end_at)
                ).order_by().values('timesheet_user_id').annotate(
                    total_worked=Sum(
                        'overtime__overtime_detail__claimed_overtime',
                        filter=Q(overtime__claim__status=CONFIRMED)
                    )
                ).filter(
                    timesheet_user_id=OuterRef('pk')
                ).values('total_worked')[:1], output_field=DurationField()
            ) for ff in ffs
        }

        ffs_pair = {
            f'difference_{ffs[i].display_name}'.lower():
                Coalesce(
                    F(f'worked_{ffs[i].display_name}'.lower()), ZERO_DURATION
                ) - Coalesce(
                    F(f'worked_{ffs[i-1].display_name}'.lower()), ZERO_DURATION
            ) for i in range(1, len(ffs))
        }
        return queryset.annotate(**t_queries).annotate(
            **dict(ffs_pair)
        ).annotate(
            total_overtime=Subquery(
                TimeSheet.objects.filter(
                    timesheet_for__range=(fiscal_start, fiscal_end)
                ).order_by().values('timesheet_user_id').annotate(
                    total_worked=Sum(
                        'overtime__overtime_detail__claimed_overtime',
                        filter=Q(overtime__claim__status=CONFIRMED)
                    )
                ).filter(
                    timesheet_user_id=OuterRef('pk')
                ).values(
                    'total_worked'
                )[:1], output_field=DurationField()
            )
        )

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'export':
            return queryset
        fiscal_start, fiscal_end = getattr(
            self.fiscal_year, 'start_at', None
        ), getattr(
            self.fiscal_year, 'end_at', None
        )
        qs = self.annotate_queryset(
            queryset,
            self.fiscal,
            fiscal_start,
            fiscal_end
        )
        return qs.select_essentials().prefetch_related(
            Prefetch('supervisors',
                     queryset=UserSupervisor.objects.filter(authority_order=1)
                     .select_related('supervisor',
                                     'supervisor__detail',
                                     'supervisor__detail__organization',
                                     'supervisor__detail__job_title',
                                     'supervisor__detail__division',
                                     'supervisor__detail__employment_level'),
                     to_attr='user_supervisors')
        )

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx.update({
            'fiscal': self.fiscal
        })
        return ctx

    def list(self, request, *args, **kwargs):
        ret = super().list(request, *args, **kwargs)
        fiscal = self.fiscal_year
        ret.data.update({
            'current_fiscal': getattr(fiscal, 'id', None)
        })
        return ret

    def get_extra_export_data(self):
        ctx = super().get_extra_export_data()
        ctx.update({
            'fiscal': self.fiscal,
            'fiscal_range': self.fiscal_range
        })
        return ctx

    @classmethod
    def get_exported_file_content(cls, queryset, title, columns, extra_content, description=None, **kwargs):

        qs = cls.annotate_queryset(
            queryset, extra_content.get('fiscal'),
            *extra_content.get('fiscal_range')
        )
        json_data = ComparativeOvertimeReportSerializer(
            instance=qs,
            many=True,
            context=extra_content
        ).data
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = cls.export_type
        organization = extra_content.get('organization')
        lines_used = ExcelExport.insert_org_info(ws, organization=organization)

        si, p_row, ei = 3, lines_used + 1, 0
        headers = extra_content.get('fiscal')
        if not headers:
            return ContentFile(save_virtual_workbook(workbook))
        col = get_column_letter
        subheaders = ['worked', 'difference']
        for pre_header in headers:
            ei = si + len(subheaders) - 1
            ws.merge_cells(
                f'{col(si)}{p_row}:{col(ei)}{p_row}'
            )
            cell = ws[f'{col(si)}{p_row}']
            cell.value = str(pre_header.display_name).title()
            cell.alignment = center_align
            cell.font = white_font_bold
            si = ei + 1
            cell.fill = green_fill
        c = ws.cell(row=p_row, column=ei+1, value='Total Overtime')
        c.font = white_font
        c.fill = green_fill
        pretty_header = [
            ' '.join(hdr.split('_')).title() for hdr in subheaders
        ]
        subheaders_merged = pretty_header * len(headers)
        subheaders_merged.insert(0, 'Supervisor')
        subheaders_merged.insert(0, 'Employee Name')
        subheaders_row = 2 + lines_used
        for ind, subheader in enumerate(subheaders_merged):
            cell = ws.cell(row=subheaders_row, column=ind+1, value=subheader)
            cell.fill = blue_fill
            cell.font = white_font_bold
        parent_row = 3 + lines_used
        total_cell = None
        for index, jsn_dt in enumerate(json_data):
            data_row = parent_row + index
            row = [
                nested_get(jsn_dt, 'user.full_name'),
                nested_get(jsn_dt, 'supervisor.full_name'),
            ]
            ws.append(row)
            restructured = {
                res.get(
                    'month_name'
                ).lower(): res for res in jsn_dt.get('results')
            }
            data_column, column = 3, 0
            for ind, month in enumerate(headers):
                column = data_column + ind * 2
                curr_month = restructured.get(month.display_name.lower())
                ws.cell(
                    row=data_row,
                    column=column,
                    value=curr_month.get('worked')
                )
                difference = curr_month.get('difference')
                rcell = ws.cell(row=data_row, column=column + 1,
                                value=difference)
                if difference.startswith('-'):
                    rcell.font = white_font
                    rcell.fill = red_fill
                elif difference.startswith('+'):
                    rcell.font = white_font
                    rcell.fill = green_fill
            total_cell = ws.cell(
                row=data_row, column=column + 2,
                value=jsn_dt.get('total_overtime')
            )
        merge_vertical_rows = [1, 2, total_cell.column]
        for col_index in merge_vertical_rows:
            val = ws.cell(row=2 + lines_used, column=col_index).value
            ws.merge_cells(
                start_row=1 + lines_used, start_column=col_index,
                end_row=2 + lines_used, end_column=col_index
            )
            cell = ws.cell(1 + lines_used, col_index, value=val)
            cell.font = white_font
            cell.fill = blue_fill
            cell.alignment = center_align
        return ContentFile(save_virtual_workbook(workbook))

    def get_frontend_redirect_url(self):
        return f'/admin/{self.organization.slug}/attendance/reports/comparative-overtime'


class AttendanceGeoLocationReport(
    DateRangeParserMixin,
    ModeFilterQuerysetMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    queryset = TimeSheet.objects.all()
    permission_to_check = [ATTENDANCE_PERMISSION, ATTENDANCE_REPORTS_PERMISSION]
    serializer_class = AttendanceGeoLocationSerializer
    filter_backends = (FilterMapBackend, SearchFilter, OrderingFilterMap)
    filter_map = {
        'start_date': 'timesheet_for__gte',
        'end_date': 'timesheet_for__lte',
        'user': 'timesheet_user'
    }

    search_fields = (
        'timesheet_user__first_name',
        'timesheet_user__middle_name',
        'timesheet_user__last_name',
        'timesheet_user__username',
    )

    ordering_fields_map = {
        'timesheet_user': (
            'timesheet_user__first_name', 'timesheet_user__middle_name',
            'timesheet_user__last_name'),
        'timesheet_for': 'timesheet_for',
        'punch_in': 'punch_in',
        'punch_out': 'punch_out'
    }

    user_definition = 'timesheet_user'

    def get_queryset(self):
        queryset = super().get_queryset().filter(
            timesheet_user__detail__organization=self.organization
        ).select_related(
            'timesheet_user',
            'timesheet_user__detail',
            'timesheet_user__detail__employment_level',
            'timesheet_user__detail__job_title',
            'timesheet_user__detail__organization',
            'timesheet_user__detail__division',
        )

        queryset = queryset.annotate(
            _punch_in_entry=FilteredRelation(
                'timesheet_entries',
                condition=Q(timesheet_entries__entry_type=PUNCH_IN)
            ),
            _punch_out_entry=FilteredRelation(
                'timesheet_entries',
                condition=Q(timesheet_entries__entry_type=PUNCH_OUT)
            )
        ).annotate(
            punch_in_latitude=F('_punch_in_entry__latitude'),
            punch_in_longitude=F('_punch_in_entry__longitude'),
            punch_in_category=F('_punch_in_entry__category'),
            punch_out_latitude=F('_punch_out_entry__latitude'),
            punch_out_longitude=F('_punch_out_entry__longitude'),
            punch_out_category=F('_punch_out_entry__category')
        )
        return queryset


class DailyAttendanceReconciliationReport(
    PastUserTimeSheetFilterMixin,
    BackgroundExcelExportMixin,
    OrganizationMixin,
    AttendanceReportPermissionMixin,
    ListViewSetMixin
):
    serializer_class = DailyAttendanceReconciliationSerializer
    queryset = TimeSheet.objects.all()
    filter_backends = (
        DjangoFilterBackend, SearchFilter, NullsAlwaysLastOrderingFilter, FilterMapBackend)
    filter_map = {
        'username': 'timesheet_user__username',
        'branch': 'timesheet_user__detail__branch__slug',
        'division': 'timesheet_user__detail__division__slug',
        'start_date': 'timesheet_for__gte',
        'end_date': 'timesheet_for__lte'
    }
    search_fields = (
        'timesheet_user__first_name',
        'timesheet_user__middle_name',
        'timesheet_user__last_name'
    )
    ordering_fields_map = {
        'timesheet_user': (
            'timesheet_user__first_name', 'timesheet_user__middle_name',
            'timesheet_user__last_name'),
        'timesheet_for': 'timesheet_for',
        'punch_in': 'punch_in',
        'punch_out': 'punch_out',
        'duration': 'duration',
    }
    export_type = "Daily Attendance Reconciliation"

    export_fields = {
        "User": "timesheet_user.full_name",
        "Code": "timesheet_user.detail.code",
        "Timesheet For": "timesheet_for",
        "Day": "day",
        "Division": "timesheet_user.detail.division",
        "Punch In Date": "punch_in.date",
        "Punch In Time": "punch_in_time",
        "Punch Out Date": "punch_out.date",
        "Punch Out Time": "punch_out_time",
        "Expected Punch In": "expected_punch_in",
        "Expected Punch Out": "expected_punch_out",
        "Worked Hours": "worked_hours",
        "Expected Work Hours": "expected_work_hours",
        "Late In": "late_in",
        "Early Out": "early_out",
        "Overtime": "total_overtime",
        "Logs": "logs",
        "Timesheet Entries": "timesheet_entry",
        "Total Lost Hour": "total_lost_hours",
        "Punctuality": "punctuality",
        "Shift Remarks": "get_coefficient_display",
    }
    permission_classes = [
        permission_factory.build_permission(
            "DailyAttendanceReconciliationPermission",
            limit_read_to=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_REPORTS_PERMISSION,
                ATTENDANCE_IMPORT_PERMISSION,
                HAS_PERMISSION_FROM_METHOD
            ]
        )
    ]
    notification_permissions = [ATTENDANCE_REPORTS_PERMISSION]

    def get_queryset(self):
        supervisor_id = self.request.query_params.get('supervisor')
        fil = dict(
            timesheet_user__attendance_setting__isnull=False,
            timesheet_for__lte=get_today(),
            timesheet_user__detail__organization=self.organization
        )

        if supervisor_id:
            if supervisor_id == str(self.request.user.id):
                fil.update({
                    'timesheet_user_id__in':
                        self.request.user.subordinates_pks
                })
            else:
                # if supervisor does not match return none
                return super().get_queryset().none()

        return super().get_queryset().filter(**fil).select_related(
            'timesheet_user',
            'timesheet_user__detail',
            'timesheet_user__detail__organization',
            'timesheet_user__detail__division',
            'timesheet_user__detail__job_title',
            'timesheet_user__detail__employment_level'
        )

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)

        return queryset.annotate(
            expected_work_hours=F('expected_punch_out') - F(
                'expected_punch_in'),
            punch_in_time=Cast(
                F('punch_in'), TimeField()
            ),
            punch_out_time=Cast(
                F('punch_out'), TimeField()
            ),
            late_in=F('punch_in') - F('expected_punch_in'),
            early_out=F('expected_punch_out') - F('punch_out')
        ).distinct()

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        def format_datetime_to_string(entries):
            return ",".join([entry.strftime("%H:%M:%S") for entry in entries])

        _entries = list(
            obj.timesheet_entries.order_by('timestamp').values_list('timestamp__time', flat=True)
        )
        logs = get_attendance_entries_for_given_timesheet(obj)
        overtime = nested_getattr(
            obj,
            'overtime.overtime_detail.claimed_overtime'
        )
        expected_punch_in = obj.expected_punch_in
        if expected_punch_in:
            expected_punch_in = expected_punch_in.astimezone().strftime("%Y-%M-%d %H:%M")
        expected_punch_out = obj.expected_punch_out
        if expected_punch_out:
            expected_punch_out = expected_punch_out.astimezone().strftime("%Y-%M-%d %H:%M")

        setattr(obj, 'expected_punch_in', expected_punch_in)
        setattr(obj, 'expected_punch_out', expected_punch_out)
        setattr(obj, 'logs', format_datetime_to_string(logs))
        setattr(obj, 'timesheet_entry', format_datetime_to_string(_entries))
        setattr(obj, 'total_overtime', humanize_interval(overtime)[:-3])
        setattr(obj, 'total_lost_hours', get_total_lost_hours_from_timesheet(obj))
        setattr(obj, 'late_in', get_late_in_from_timesheet(obj, humanized=True))
        setattr(obj, 'early_out', get_early_out_from_timesheet(obj, humanized=True))
        return obj


    # def list(self, request, *args, **kwargs):
    #     mandatory_fields = ['timesheet_for']
    #     selected_fields = self.request.query_params.get("selected_fields", "")
    #     selected_fields = selected_fields.split(",") if selected_fields else []
    #     fields = mandatory_fields + selected_fields
    #     data = DailyAttendanceReconciliationSerializer(self.filter_queryset(self.get_queryset()),
    #                                             many=True).data
    #     page = self.paginate_queryset(data)
    #     return self.get_paginated_response(page)


class EmployeeAttendanceInsight(
    PastUserTimeSheetFilterMixin,
    BackgroundExcelExportMixin,
    OrganizationMixin,
    AttendanceReportPermissionMixin,
    ListViewSetMixin
):
    serializer_class = EmployeeAttendanceInsightSerializer
    queryset = TimeSheet.objects.all()
    filter_backends = (
        DjangoFilterBackend, SearchFilter, NullsAlwaysLastOrderingFilter, FilterMapBackend)
    filter_map = {
        'username': 'timesheet_user__username',
        'branch': 'timesheet_user__detail__branch__slug',
        'division': 'timesheet_user__detail__division__slug',
        'start_date': 'timesheet_for__gte',
        'end_date': 'timesheet_for__lte',
        'shift': 'work_shift'
    }
    search_fields = (
        'timesheet_user__first_name',
        'timesheet_user__middle_name',
        'timesheet_user__last_name',
        'timesheet_user__username'
    )
    ordering_fields_map = {
        'timesheet_user': (
            'timesheet_user__first_name', 'timesheet_user__middle_name',
            'timesheet_user__last_name'),
        'timesheet_for': 'timesheet_for',
        'punch_in': 'punch_in',
        'punch_out': 'punch_out',
        'duration': 'duration',
    }
    permission_classes = [
        permission_factory.build_permission(
            "EmployeeAttendanceInsight",
            limit_read_to=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_REPORTS_PERMISSION,
                ATTENDANCE_IMPORT_PERMISSION,
                HAS_PERMISSION_FROM_METHOD
            ]
        )
    ]
    export_type = "Employee Attendance insight"

    @property
    def field_mapping(self):
        return {
            "day": "day",
            "leave_coefficient": "get_leave_coefficient_display",
            "logs": "logs",
            "expected_punch_in": "expected_punch_in",
            "expected_punch_out": "expected_punch_out",
            "punch_in_category": "punch_in_category",
            "punch_out_category": "punch_out_category",
            "total_lost_hours": "total_lost_hours",
            "break_in_out_lost_hours" : "break_in_out_lost_hours",
            "overtime": "total_overtime",
            "approved_overtime": "approved_overtime",
            "confirmed_overtime": "confirmed_overtime",
            "worked_hours": "worked_hours",
            "expected_work_hours": "expected_work_hours",
            "coefficient": "get_coefficient_display",
        }

    def get_export_fields(self):
        export_fields = {
            "User": "timesheet_user.full_name",
            "Code": "timesheet_user.detail.code",
            "Job Title": "timesheet_user.detail.job_title",
            "Division": "timesheet_user.detail.division",
            "Branch": "timesheet_user.detail.branch",
            "Date": "timesheet_for",
            "Punch In": "punch_in_time",
            "Punch Out": "punch_out_time",
            "Work Shift": "work_shift.name",
            "Attendance Status": "attendance_status",
            "Break In Count": "break_in_count",
            "Timesheet Entries": "timesheet_entry",
            "Attendance Sync": "attendance_sync"

        }
        setting_headings = AttendanceHeadingReportSetting.objects.filter(organization=self.organization).first()
        choice_headings = getattr(setting_headings,'headings',None)
        if choice_headings:
            export_fields.update({key.replace('_', ' ').title(): self.field_mapping.get(key) for key in choice_headings})
        return export_fields

    def get_serializer(self, *args, **kwargs):
        mandatory_fields = ['timesheet_user', 'branch_name', 'division_name', 'work_shift',
                            'timesheet_for', 'punch_in_category', 'punch_out_category', \
                            'break_in_count', 'break_out_count', 'punch_in', 'punch_out',
                            'leave_coefficient', 'timesheet_entries', 'attendance_sync']
        setting_headings = AttendanceHeadingReportSetting.objects.filter(organization=self.organization).first()
        choice_headings = getattr(setting_headings,'headings',None)
        if choice_headings:
            mandatory_fields += [heading for heading in choice_headings]
            kwargs.setdefault('fields',mandatory_fields)
        return super().get_serializer(*args, **kwargs)

    notification_permissions = [ATTENDANCE_REPORTS_PERMISSION]

    def get_queryset(self):
        qs = super().get_queryset()
        supervisor_id = self.request.query_params.get('supervisor')
        selected_employees = self.request.query_params.get('selected_employees', '').split(',')
        selected_employees = set(map(int,filter(str.isdigit, selected_employees)))
        fil = dict(
            timesheet_user__attendance_setting__isnull=False,
            timesheet_for__lte=get_today(),
            timesheet_user__detail__organization=self.organization
        )

        if selected_employees:
            fil.update({
                'timesheet_user_id__in': selected_employees
            })
        if supervisor_id:
            timesheet_user_id = self.request.user.subordinates_pks
            if selected_employees:
                timesheet_user_id = timesheet_user_id.intersection(selected_employees)
            if supervisor_id == str(self.request.user.id):
                fil.update({
                    'timesheet_user_id__in':
                        timesheet_user_id
                })
            else:
                # if supervisor does not match return none
                return qs.none()

        return qs.filter(**fil).select_related(
            'timesheet_user',
            'timesheet_user__detail',
            'timesheet_user__detail__organization',
            'timesheet_user__detail__division',
            'timesheet_user__detail__job_title',
            'timesheet_user__detail__employment_level'
        )

    def filter_queryset(self, queryset):
        punch_in = self.request.query_params.get("punch_in_category")
        punch_out = self.request.query_params.get("punch_out_category")
        attendance_sync = self.request.query_params.get("attendance_sync")
        sub_query = AttendanceUserMap.objects.filter(
            setting=OuterRef(OuterRef('timesheet_user__attendance_setting')))
        queryset = super().filter_queryset(queryset)

        if punch_in:
            punch_in = punch_in.replace("_", " ").title()
            if punch_in == MISSING:
                queryset = queryset.filter(punch_in__isnull=True,)
            else:
                queryset = queryset.filter(
                    timesheet_entries__entry_type=PUNCH_IN,
                    timesheet_entries__category=punch_in
                )

        if punch_out:
            punch_out = punch_out.replace("_", " ").title()
            if punch_out == MISSING:
                queryset = queryset.filter(punch_out__isnull=True,)
            else:
                queryset = queryset.filter(
                    timesheet_entries__entry_type=PUNCH_OUT,
                    timesheet_entries__category=punch_out
                )

        if attendance_sync in ['true', 'false']:
            filter_mapper = {
                "true": True,
                "false": False
            }
            queryset = queryset.annotate(
                sync_attendance=Case(
                    When(Exists(
                        AttendanceEntryCache.objects.filter(
                            timestamp__date=OuterRef('timesheet_for'),
                            bio_id__in=Subquery(sub_query.values('bio_user_id')),
                            source__id__in=Subquery(sub_query.values('source')),
                            reason__in=[SYNC_PENDING, SYNC_FAILED]
                        )[:1]

                    ) | ~Exists(
                        TimeSheetEntry.objects.filter(timesheet_id=OuterRef('id'),
                                                      is_deleted=False)[:1]),
                         then=False),
                    default=True,
                    output_field=models.BooleanField()
                )
            ).filter(sync_attendance=filter_mapper.get(attendance_sync))

        return queryset.annotate(
            expected_work_hours=F('expected_punch_out') - F(
                'expected_punch_in'),
            punch_in_time=Cast(
                F('punch_in'), TimeField()
            ),
            punch_out_time=Cast(
                F('punch_out'), TimeField()
            ),
            late_in=F('punch_in') - F('expected_punch_in'),
            early_out=F('expected_punch_out') - F('punch_out')
        ).distinct()

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        def format_datetime_to_string(entries):
            return ",".join([entry.strftime("%H:%M:%S") for entry in entries])

        logs = get_attendance_entries_for_given_timesheet(obj)
        _entries = list(
            obj.timesheet_entries.order_by('timestamp').values_list('timestamp__time', flat=True)
        )
        _sync = not has_unsynced_attendance_entries(obj)
        unclaimed_overtime = nested_getattr(
            obj,
            'overtime.overtime_detail.claimed_overtime'
        )

        approved_overtime = None
        if hasattr(obj, 'overtime'):
            overtime = obj.overtime
            if hasattr(overtime, 'claim') and overtime.claim.status == APPROVED:
                approved_overtime = nested_getattr(
                    obj,
                    'overtime.overtime_detail.claimed_overtime'
                )

        confirmed_overtime = None
        if hasattr(obj, 'overtime'):
            overtime = obj.overtime
            if hasattr(overtime, 'claim') and overtime.claim.status == CONFIRMED:
                confirmed_overtime = nested_getattr(
                    obj,
                    'overtime.overtime_detail.claimed_overtime'
                )

        expected_punch_in = obj.expected_punch_in
        if expected_punch_in:
            expected_punch_in = expected_punch_in.astimezone().strftime("%Y-%M-%d %H:%M")
        expected_punch_out = obj.expected_punch_out
        if expected_punch_out:
            expected_punch_out = expected_punch_out.astimezone().strftime("%Y-%M-%d %H:%M")
        if obj.leave_coefficient == FULL_LEAVE:
            expected_punch_in = expected_punch_out = ' '
        attendance_status = "Present" if obj.is_present else "Absent"
        if obj.leave_coefficient == FULL_LEAVE:
            expected_punch_in = expected_punch_out = ' '
            attendance_status = "On Leave"

        setattr(obj, 'attendance_status', attendance_status)
        setattr(obj, 'expected_punch_in', expected_punch_in)
        setattr(obj, 'expected_punch_out', expected_punch_out)
        setattr(obj, 'logs', format_datetime_to_string(logs))
        setattr(obj, 'total_overtime', humanize_interval(unclaimed_overtime)[:-3])
        setattr(obj, 'approved_overtime', humanize_interval(approved_overtime)[:-3])
        setattr(obj, 'confirmed_overtime', humanize_interval(confirmed_overtime)[:-3])
        setattr(obj, 'total_lost_hours', get_total_lost_hours_from_timesheet(obj))
        setattr(obj, 'late_in', get_late_in_from_timesheet(obj, humanized=True))
        setattr(obj, 'early_out', get_early_out_from_timesheet(obj, humanized=True))
        setattr(obj, 'punch_in_time', get_ktm_time(obj.punch_in_time))
        setattr(obj, 'punch_out_time', get_ktm_time(obj.punch_out_time))
        setattr(obj, 'break_in_out_lost_hours', humanize_interval(
            timezone.timedelta(minutes=break_in_out_lost_hour(obj))
        )[:-3])
        setattr(obj, 'break_out_count', obj.timesheet_entries.filter(entry_type = BREAK_OUT).count())
        setattr(obj, 'break_in_count', obj.timesheet_entries.filter(entry_type = BREAK_IN).count())
        setattr(obj, 'timesheet_entry', format_datetime_to_string(_entries))
        setattr(obj, 'attendance_sync', _sync)

        return obj

    def _export_post(self):
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        if end_date > (start_date + timedelta(days=6)):
            raise ValidationError("Can only be downloaded for duration less than 7 Days.")
        return super()._export_post()


class AttendanceHeadingReportSettingViewSet(ReportSettingViewSetMixin):
    queryset = AttendanceHeadingReportSetting.objects.all()
    serializer_class = AttendanceHeadingReportSettingSerializer

    permission_classes = [
        permission_factory.build_permission(
            "AttendanceHeadingReportSettingViewSet",
            limit_read_to=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_REPORTS_PERMISSION,
                ATTENDANCE_IMPORT_PERMISSION,
                HAS_PERMISSION_FROM_METHOD
            ]
        )
    ]

