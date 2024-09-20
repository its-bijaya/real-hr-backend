import uuid
import logging
from zipfile import BadZipFile

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.cache import cache as dj_cache
from django.core.files.storage import default_storage
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill

from irhrs.attendance.api.v1.serializers.attendance import \
    TimeSheetImportSerializer
from irhrs.attendance.constants import METHOD_IMPORT
from irhrs.attendance.models import TimeSheet
from irhrs.export.utils.helpers import save_workbook, load_workbook
from irhrs.notification.utils import notify_organization
from irhrs.permission.constants.permissions import (ATTENDANCE_IMPORT_PERMISSION,
                                                    ATTENDANCE_PERMISSION)
logger = logging.getLogger(__name__)


def parse(workbook):
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]

    for row in sheet:
        values = []
        for col in row:
            values.append(col.value)
        yield values


def _save_file(in_memory_file):
    """
    Save in memory file to file system

    :param in_memory_file:
    :return: path to file
    """
    extension = in_memory_file.name.split('.')[-1]
    return default_storage.save("{}.{}".format(str(uuid.uuid4()), extension), in_memory_file)


def validate_export_file(file):
    """
    Parse excel file and return rows

    :param file: File opened in binary read mode
    :type file: File
    :return: generator returning values of rows of excel files
    """
    try:
        workbook = load_workbook(file, read_only=True)
    except (OSError, BadZipFile):
        logger.error("The file seems to be corrupted.", exc_info=True)
        return False, "The file seems to be corrupted."
    if not len(workbook.sheetnames) > 0:
        return False, "The File is empty"
    return True, workbook


def save_records(rows, inputs, user, org):
    """
    Takes rows of the excel file and process it and save it.

    :param rows: rows of the file
    :type rows: generator
    :return: errors
    """
    headers = rows[0]  # remove the first row
    if not inputs.issubset(set(headers)):
        notify_organization(
            text="Could not import attendance. The file formatting was not correct.",
            organization=org,
            action=user,
            permissions=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_IMPORT_PERMISSION
            ]
        )
        return {"detail": "The file formatting was not correct.", "status": "bad-formatting"}
    validation_errors = {}
    # zip the headers and values
    create_data = list()

    # Fetch all users, so wont have to hit db every time.
    users = get_user_model().objects.filter(
        attendance_setting__isnull=False,
        attendance_setting__is_disabled=False,
        detail__organization_id=org.id
    ).current()
    users_map = {}
    for user in users:
        users_map[user.email] = user
        users_map[user.username] = user
    for row in rows[1:]:
        create_data.append(
            dict(
                zip(
                    headers,
                    row
                )
            )
        )
    # validate rows
    for row_id, data in enumerate(create_data):
        user_error = {}
        # we do not validate user here, as we require object for this.
        serializer = TimeSheetImportSerializer(
            data=data
        )
        valid_record = serializer.is_valid(raise_exception=False)
        email_or_username = data.get('user')
        valid_user = users_map.get(
            email_or_username
        )
        if not valid_user:
            valid_record = False
            user_error.update({
                email_or_username: [
                    f'The user with this email/username does not exist in {org.name} or '
                    f'is not valid for import.'
                ]
            })
        timesheet_for = data.get('timesheet_for')
        if valid_record:
            timesheet_params = {
                'user': valid_user,
                'entry_method': METHOD_IMPORT,
                'remarks': 'Imported',
                'manual_user': user
            }
            for entry_data in [
                serializer.validated_data.get('punch_in'),
                serializer.validated_data.get('punch_out'),
            ]:
                timesheet = valid_user.timesheets.filter(timesheet_for=timesheet_for).first()
                if timesheet:
                    timesheet_params['timesheet'] = timesheet
                TimeSheet.objects.clock(
                    **entry_data,
                    **timesheet_params
                )
        else:
            validation_errors.update({
                row_id: {
                    **serializer.errors,
                    **user_error
                }
            })
    return validation_errors


def export_failed(file_p, recipient, org):
    status, workbook = validate_export_file(file_p)

    if not status:
        message = workbook
        notify_organization(
            text=message,
            action=recipient,
            organization=org,
            permissions=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_IMPORT_PERMISSION
            ]
        )
        return

    rows = list(parse(workbook))
    valid_input_fields = {
        'user',
        'timesheet_for',
        'punch_in',
        'punch_out'
    }
    errors = save_records(
        rows,
        valid_input_fields,
        recipient,
        org
    )
    file_name = uuid.uuid4().hex + uuid.uuid4().hex + '.xlsx'
    file_path = file_name
    if not errors:
        notify_organization(
            text="Attendance import has completed successfully.",
            action=recipient,
            organization=org,
            url=f'/admin/{org.slug}/attendance/reports/daily-attendance',
            permissions=[
                ATTENDANCE_PERMISSION,
                ATTENDANCE_IMPORT_PERMISSION
            ]
        )
        return

    if errors.get('status') == "bad-formatting":
        # Already notified in save_records
        return
    failed_row_ids = errors.keys()

    parsed = parse(workbook)
    header = next(parsed)
    header.append('Cause')

    wb = Workbook()
    ws = wb.active
    ws.append(header)
    fill = PatternFill("solid", fgColor="FF0040")
    for index, row in enumerate(parsed):
        insert = []
        if index not in failed_row_ids:
            continue

        def stringify_err(error_dict):
            error_str = ''
            for key, val in error_dict.items():
                if isinstance(val, dict):
                    for k, v in val.items():
                        error_str += f"{k}: {','.join(v) if isinstance(v, list) else v}\n"
                else:
                    error_str += f"{key}: {','.join(val) if isinstance(val, list) else val}\n"
            return error_str
        if errors.get(index):
            row.append(
                stringify_err(errors.get(index))
            )
        highlight_cells = errors.get(index).keys()
        for cell in row:
            if cell in highlight_cells:
                cell = WriteOnlyCell(ws, value=cell)
                cell.fill = fill
                insert.append(cell)
            else:
                insert.append(WriteOnlyCell(ws, value=cell))
        ws.append(insert)
    file_path = save_workbook(wb, file_path)

    _prepare_cache = {
        'url': settings.BACKEND_URL + '/media/' + file_path,
        'created_on': timezone.now()
    }
    dj_cache.set('attendance_fails_export', _prepare_cache)
    notify_organization(
        text="Attendance import has completed with errors."
             f"{len(errors)} failed out of {len(rows)-1} records.",
        organization=org,
        action=recipient,
        url=f'/admin/{org.slug}/attendance/import/attendance?status=failed',
        permissions=[
            ATTENDANCE_PERMISSION,
            ATTENDANCE_IMPORT_PERMISSION
        ]
    )
    if default_storage.exists(file_p):
        default_storage.delete(file_p)
