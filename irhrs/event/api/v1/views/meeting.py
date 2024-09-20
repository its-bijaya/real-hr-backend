import os

from django.db.models import Exists, OuterRef
from django.http import HttpResponse
from django.utils.functional import cached_property
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.exceptions import MethodNotAllowed
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import SAFE_METHODS
from rest_framework.response import Response

from irhrs.core.mixins.viewset_mixins import (
    ListCreateDestroyViewSetMixin,
    MeetingMixin, ListCreateUpdateDestroyViewSetMixin,
    ListCreateUpdateViewSetMixin,
    UpdateViewSetMixin,
    ListRetrieveUpdateDestroyViewSetMixin, CreateUpdateViewSetMixin)
from irhrs.core.utils import nested_getattr
from irhrs.core.utils.common import get_today
from irhrs.event.api.v1.permissions import (
    MeetingPermission, MeetingPermissionForTimeKeeperMixin,
    MeetingPermissionForMinuterMixin
)
from irhrs.event.api.v1.serializers.event import EventDetailExportSerializer
from irhrs.event.api.v1.serializers.meeting import (
    MeetingDocumentSerializer,
    AgendaSerializer, AgendaCommentSerializer,
    MeetingAttendanceSerializer,
    MeetingSerializer, MeetingAgendaTaskSerializer)
from irhrs.event.constants import MEETING_DOCUMENT_MAX_UPLOAD_SIZE
from irhrs.event.models import (
    MeetingDocument, MeetingAgenda, AgendaComment,
    MeetingAttendance, EventDetail, MeetingNotification, AgendaTask,
    MeetingAcknowledgeRecord)
from irhrs.notification.api.v1.serializers.notification import \
    NotificationSerializer


class AgendaCommentMixin:
    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['agenda'] = self.agenda
        return ctx

    @cached_property
    def agenda(self):
        return MeetingAgenda.objects.filter(
            id=self.kwargs.get('agenda_id')
        ).select_related(
            'meeting'
        ).first()


class MeetingNotificationViewSet(MeetingMixin,
                                 ListCreateUpdateDestroyViewSetMixin):
    serializer_class = NotificationSerializer
    queryset = MeetingNotification.objects.all()
    permission_classes = [MeetingPermission]

    def get_queryset(self):
        meeting = self.meeting
        if meeting and meeting.meeting_attendances.filter(
            member=self.request.user).exists():
            return super().get_queryset()
        else:
            return self.queryset.none()

    def has_user_permission(self):
        meeting = self.meeting
        if meeting and meeting.created_by == self.request.user:
            return True
        return False

    def create(self, request, *args, **kwargs):
        raise MethodNotAllowed(method='post')


class MeetingAgendaTaskViewSet(
    MeetingPermissionForMinuterMixin,
    AgendaCommentMixin,
    MeetingMixin,
    ListCreateUpdateDestroyViewSetMixin
):
    serializer_class = MeetingAgendaTaskSerializer
    queryset = AgendaTask.objects.all()
    permission_classes = [MeetingPermission]

    def get_queryset(self):
        agenda = self.agenda
        if agenda and self.meeting.meeting_attendances.filter(member=self.request.user).exists():
            return self.queryset.filter(agenda=self.agenda)
        else:
            return self.queryset.none()


class MeetingViewSet(CreateUpdateViewSetMixin):
    queryset = EventDetail.objects.all()
    serializer_class = MeetingSerializer

    def update(self, request, *args, **kwargs):
        meeting = self.get_object()
        if self.request.user == meeting.created_by or (
           meeting.minuter and self.request.user == meeting.minuter.user):
            return super().update(request, *args, **kwargs)
        return Response(
            {
                'detail': 'You don\'t have permission to perform this action.'},
            status=status.HTTP_403_FORBIDDEN
        )

    @action(methods=['POST'], detail=True, url_path="export-event-details",
            serializer_class=EventDetailExportSerializer)
    def export_event_details(self, request, pk=None):
        event_details = self.get_object()
        event_title = event_details.event.title
        event_start_date = event_details.event.start_at.astimezone()
        event_end_date = event_details.event.end_at.astimezone()
        event_location = event_details.event.location
        meeting_attendances = event_details.meeting_attendances.order_by('member__first_name')
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="event-detail.xlsx"'

        wb = Workbook()
        sheet = wb['Sheet']
        bold = Font(bold=True)
        align_center = Alignment(horizontal='center')
        strp_time_format = "%Y-%m-%d %H:%M:%S"

        organization = event_details.created_by.detail.organization

        # for organization logo in downloaded excel report:
        # logo = nested_getattr(organization, 'appearance.logo')
        # if logo:
        #     sheet.merge_cells(
        #         start_row=1,
        #         start_column=1, end_row=1,
        #         end_column=7
        #     )
        #     image_obj = Image(logo)
        #     sheet.add_image(image_obj, anchor="A1")
        #     dimension = sheet.row_dimensions[1]
        #     dimension.height = image_obj.height * 0.75

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        sheet.cell(row=2, column=1, value=organization.name)
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
        sheet.cell(row=3, column=1,
                value=f"Downloaded at: ({get_today(with_time=True).strftime(strp_time_format)})")

        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
        cell = sheet.cell(column=1, row=4, value=f'Event Title: {event_title}')
        cell.font = bold
        cell.alignment = align_center

        sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
        cell = sheet.cell(column=1, row=5, value=f'Start Date: '
                                                 f'{event_start_date.strftime(strp_time_format)}')
        cell.alignment = align_center

        sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=7)
        cell = sheet.cell(column=1, row=6, value=f'End Date:  '
                                                 f'{event_end_date.strftime(strp_time_format)}')
        cell.alignment = align_center

        sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=7)
        cell = sheet.cell(column=1, row=7, value=f'Event Location: {event_location}')
        cell.alignment = align_center

        cell = sheet.cell(column=3, row=9, value="Full Name")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=4, row=9, value="Arrival Time")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=5, row=9, value="Remarks")
        cell.font = bold
        cell.alignment = align_center

        for column_index, meeting_attendance in enumerate(meeting_attendances):
            full_name = meeting_attendance.member.full_name
            arrival_time = getattr(meeting_attendance, 'arrival_time', None)
            if arrival_time:
                arrival_time = arrival_time.astimezone().strftime(strp_time_format)
            remarks = meeting_attendance.remarks or 'N/A'
            sheet.cell(column=3, row=column_index + 10, value=full_name)
            sheet.cell(column=4, row=column_index + 10, value=arrival_time or 'N/A')
            sheet.cell(column=5, row=column_index + 10, value=remarks)

        wb.save(response)
        return response


class MeetingAgendaViewSet(MeetingMixin, MeetingPermissionForMinuterMixin,
                           ListRetrieveUpdateDestroyViewSetMixin):
    serializer_class = AgendaSerializer
    queryset = MeetingAgenda.objects.all()
    permission_classes = [MeetingPermission]

    def get_queryset(self):
        meeting = self.meeting
        if meeting and meeting.meeting_attendances.filter(
           member=self.request.user).exists():
            return super().get_queryset()
        else:
            return self.queryset.none()


class AgendaCommentViewSet(AgendaCommentMixin,
                           ListCreateUpdateDestroyViewSetMixin):
    serializer_class = AgendaCommentSerializer
    queryset = AgendaComment.objects.all()
    permission_classes = [MeetingPermission]

    def check_permissions(self, request):
        super().check_permissions(request)
        if self.request.method.upper() not in SAFE_METHODS and not self.agenda.meeting.prepared:
            self.permission_denied(
                request,
                message="Making any comment until meeting is prepared"
            )

    def has_user_permission(self):
        if self.request.method.lower() == 'delete':
            comment = self._get_comment()
            if comment and comment.commented_by_id == self.request.user.id:
                return True

            if self.agenda and self.request.user in [self.agenda.meeting.created_by,
                                                     self.agenda.meeting.minuter.user]:
                return True
        else:
            user = self.request.user
            meeting = self.agenda.meeting if self.agenda else None
            meeting_member = MeetingAttendance.objects.filter(member=user).exists()
            if meeting and (meeting_member or meeting.created_by == user):
                return True
        return False

    def get_queryset(self):
        agenda = self.agenda
        if agenda and agenda.meeting.meeting_attendances.filter(
           member=self.request.user).exists():
            return super().get_queryset().filter(
                agenda=self.agenda,
                agenda__meeting__prepared=True
            )
        else:
            return self.queryset.none()

    def _get_comment(self):
        comment_id = self.kwargs.get('pk')
        try:
            return AgendaComment.objects.get(id=comment_id)
        except (AgendaComment.DoesNotExist, TypeError, ValueError):
            return None


class MeetingAttendanceViewSet(MeetingMixin, MeetingPermissionForTimeKeeperMixin,
                               ListCreateUpdateViewSetMixin):
    serializer_class = MeetingAttendanceSerializer
    queryset = MeetingAttendance.objects.all()
    permission_classes = [MeetingPermission]

    def get_queryset(self):
        meeting = self.meeting
        if meeting and meeting.meeting_attendances.filter(
           member=self.request.user).exists():
            return super().get_queryset().annotate(
                acknowledged=Exists(
                    MeetingAcknowledgeRecord.objects.filter(
                        member_id=OuterRef('member_id'),
                        meeting_id=OuterRef('meeting_id')
                    )
                )
            ).select_related(
                'member', 'member__detail', 'member__detail__job_title',
                'member__detail__division', 'meeting',
                'meeting__time_keeper__user', 'meeting__minuter__user'
            ).order_by(
                'member__first_name',
                'member__middle_name',
                'member__last_name'
            )
        else:
            return self.queryset.none()

    def create(self, request, *args, **kwargs):
        raise MethodNotAllowed(method='post')

    def update(self, request, *args, **kwargs):
        remarks = request.data.get('remarks')
        arrival_time = request.data.get('arrival_time')
        if remarks and not arrival_time:
            return Response({
                'remarks': 'Can\'t add remarks without arrival time.'
            }, status=status.HTTP_400_BAD_REQUEST)
        return super().update(request, *args, **kwargs)


class MeetingDocumentViewSet(MeetingMixin, MeetingPermissionForMinuterMixin,
                             ListCreateDestroyViewSetMixin):
    serializer_class = MeetingDocumentSerializer
    queryset = MeetingDocument.objects.all()
    parser_classes = (MultiPartParser, FormParser,)
    permission_classes = [MeetingPermission]

    def get_queryset(self):
        meeting = self.meeting
        if meeting and meeting.meeting_attendances.filter(
           member=self.request.user).exists():
            return super().get_queryset()
        else:
            return self.queryset.none()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['meeting'] = self.meeting
        return context


MeetingDocumentViewSet.__doc__ = """
            create:

                    Create Documents for Meeting
                    Max Size : {} MB

            """.format(MEETING_DOCUMENT_MAX_UPLOAD_SIZE / (1024 * 1024))
