from django.utils.html import strip_tags
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportMixin:
    ws = None
    FONTS = {
        '12B': Font(size=12, bold=True),
        '11B': Font(size=11, bold=True),
        '10B': Font(size=10, bold=True),
        '9B': Font(size=10, bold=True),
        '9N': Font(size=10, bold=False),
    }

    ALIGNMENT = {
        'center': Alignment(horizontal='center', vertical='center')
    }

    def extend_cell(self, cell, right=0, down=0):
        assert self.ws is not None
        self.ws.merge_cells(
            f"{cell.column_letter}{cell.row}:{get_column_letter(cell.column+right)}{cell.row+down}"
        )

    @classmethod
    def add_font(cls, cells, font_key):
        if not isinstance(cells, list):
            cells = [cells]
        for cell in cells:
            setattr(cell, 'font', cls.FONTS.get(font_key, Font()))

    @classmethod
    def align(cls, cells, align):
        if not isinstance(cells, list):
            cells = [cells]
        for cell in cells:
            setattr(cell, 'alignment', cls.ALIGNMENT.get(align, Alignment()))

    @classmethod
    def add_border(cls, cells, borders, border_style='dotted'):
        valid_borders = {'right', 'left', 'top', 'bottom'}
        applicable_borders = set(borders).intersection(valid_borders)

        side = Side(border_style=border_style)
        borders_parm = dict()
        for i in applicable_borders:
            borders_parm[i] = side

        border = Border(**borders_parm)

        if not isinstance(cells, list):
            cells = [cells]
        for cell in cells:
            setattr(cell, 'border', border)

    def set_height(self, row, height):
        self.ws.row_dimensions[row].height = height

    def set_width(self, col, width):
        self.ws.column_dimensions[get_column_letter(col)].width = width


class Question(ReportMixin):

    def __init__(self, ws, row, col, num, question):
        self.ws = ws
        self.row = row
        self.col = col
        self.num = num
        self.title = question.get('title')
        self.weightage = question.get('weightage')
        self.create_cell()

    def create_cell(self):
        cell = self.ws.cell(self.row, self.col)
        cell.value = f'Q#{self.num} \nMax {self.weightage}'
        self.align(cell, 'center')
        self.add_font(cell, '9B')


class Candidate(ReportMixin):

    def __init__(self, ws, candidate_id, name, number, row, has_weightage=False):
        self.ws = ws
        self.id = candidate_id
        self.num = number
        self.row = row
        self.end_row = row
        self.name = name

        self.remarks_cell = None
        self.has_weightage = has_weightage

        self.create_cell()
        self.add_remarks()

    def create_cell(self):

        # SN Number
        sn = self.ws.cell(column=1, row=self.row)
        sn.value = self.num
        self.extend_cell(sn, down=1)
        self.set_width(sn.column, 5)
        self.align(sn, 'center')

        # Candidate Name
        name = self.ws.cell(column=2, row=self.row)
        name.value = self.name
        self.add_font(name, '10B')

    def add_remarks(self):
        self.end_row += 1
        remarks = self.ws.cell(column=2, row=self.end_row)
        remarks.value = 'Remarks'
        self.add_font(remarks, '9N')
        self.set_height(self.row, 25)
        self.remarks_cell = remarks

    def __str__(self):
        return self.name


class CandidateQuestionAnswer(ReportMixin):

    def __init__(self, ws, candidate, question, question_answer, interviewer_weightage=None):
        self.ws = ws
        self.candidate = candidate
        self.question = question
        self.question_answer = question_answer
        self.score = self.question_answer.get('score')
        self.interviewer_weightage = interviewer_weightage
        self.create_cell()

    def create_cell(self):
        cell = self.ws.cell(self.candidate.row, self.question.col)
        cell.value = self.score
        self.add_font(cell, '9B')

    def __str__(self):
        return f"{str(self.candidate)} {str(self.question)}"


class CandidateRemarks(ReportMixin):

    def __init__(self, ws, candidate, table, remarks):
        self.ws = ws
        self.candidate = candidate
        self.table = table
        self.remarks = remarks
        self.create_cell()

    def create_cell(self):
        cell = self.ws.cell(self.candidate.row + 1, self.table.initial_col)
        self.extend_cell(cell, right=(self.table.col - self.table.initial_col))
        cell.value = strip_tags(self.remarks).strip()
        self.align(cell, 'center')


class SummaryTable(ReportMixin):

    def __init__(self, ws, row, col, question_answer):
        self.ws = ws
        self.row = row
        self.col = col
        self.initial_col = col

        self.name = 'Summary Evaluation'
        self.code = 'summary_evaluation'
        self.header_title = 'Summary Evaluation'

        self.question_answers = question_answer.get('question_answers')
        self.total = question_answer.get('total_score')

        self.questions = list()
        self.candidate_question_answers = list()
        self.candidate = None

    def add_candidate(self, candidate):
        self.candidate = candidate

    def populate_question(self):
        for index, question_answer in enumerate(self.question_answers, 1):
            self.questions.append(
                Question(self.ws, self.row + 1, self.col, index, question_answer)
            )
            self.col += 1

    def create_header(self):
        evaluation_panel = self.ws.cell(column=self.col, row=self.row)
        evaluation_panel.value = self.header_title
        self.align(evaluation_panel, 'center')
        self.add_font(evaluation_panel, '9B')
        self.extend_cell(evaluation_panel, right=len(self.question_answers) - 1)

    def create_total_marks(self):
        total_marks = self.ws.cell(column=self.col, row=self.row)
        total_marks.value = f'Total\nMax\n{self.total}'
        self.align(total_marks, 'center')
        self.add_font(total_marks, '9B')
        self.extend_cell(total_marks, down=1)

    def populate_score(self, report=None):
        assert report is not None
        assert self.candidate is not None
        for question in self.questions:
            aggregate_value = report.aggregate_report(question=question, candidate=self.candidate)
            question_answer = {
                'score': aggregate_value
            }
            candidate_question_answer = CandidateQuestionAnswer(
                self.ws,
                self.candidate,
                question,
                question_answer
            )
            self.candidate_question_answers.append(candidate_question_answer)

        total_cell = self.ws.cell(row=self.candidate.row, column=self.col)
        total_cell.value = self._aggregate_total()

    def _aggregate_total(self):
        return sum([
            qa.score for qa in
            self.get_candidate_question_answers(candidate=self.candidate)
        ])

    def get_candidate_question_answers(self, question=None, candidate=None):

        try:
            candidate_question_answers = list(
                filter(
                    lambda x: (x.candidate == (candidate or x.candidate)) and (
                        x.question.title == question.title if question else x.question.title
                    ),
                    self.candidate_question_answers
                )
            )
            if candidate and candidate.has_weightage:
                for data in candidate_question_answers:
                    weightage = getattr(data, 'interviewer_weightage', None)
                    if weightage:
                        setattr(data, 'score', data.score * weightage * 0.01)
                        data.question_answer['score'] = data.score * weightage * 0.01
                return candidate_question_answers
            return candidate_question_answers
        except IndexError:
            return None

    def populate_headers(self):
        self.create_header()
        self.populate_question()
        self.create_total_marks()

    def get_question(self, title):
        try:
            return list(filter(lambda question: question.title == title, self.questions))[0]
        except IndexError:
            return None

    def __str__(self):
        return self.name


class EvaluationTable(SummaryTable):

    def __init__(self, ws, row, col, given_index, question_answer):
        super().__init__(ws, row, col, question_answer)
        self.given_index = given_index
        self.header_title = f'Evaluation Panel {self.given_index}'
        self.name = question_answer.get('assigned_person')
        self.code = question_answer.get('assigned_person_code')
        self.given = question_answer.get('given_score')

        self.candidate_remarks = list()
        self.interviewer_weightage = question_answer.get('interviewer_weightage', None)

    def populate_score(self, report=None):
        assert self.candidate is not None

        for question_answer in self.question_answers:
            question = self.get_question(question_answer.get('title'))
            candidate_question_answer = CandidateQuestionAnswer(
                self.ws,
                self.candidate,
                question,
                question_answer,
                self.interviewer_weightage
            )
            self.candidate_question_answers.append(candidate_question_answer)

        total_cell = self.ws.cell(row=self.candidate.row, column=self.col)
        total_cell.value = self.given

    def add_remarks(self, remarks=''):
        assert self.candidate is not None
        self.candidate_remarks.append(
            CandidateRemarks(self.ws, candidate=self.candidate, table=self, remarks=remarks)
        )

    def has_candidate_remarks(self, candidate):
        return bool(list(filter(
            lambda candidate_remark: candidate_remark.candidate == candidate,
            self.candidate_remarks
        )))


class QuestionAnswerReport(ReportMixin):

    def __init__(self, question_answers, job_position='', scheduled_at=""):
        self.question_answers = question_answers
        self.job_position = job_position
        self.scheduled_at = scheduled_at

        self.wb = Workbook()
        self.ws = self.wb.active

        self.row = 1
        self.table_row = 2
        self.table_column = 1

        self.tables = list()
        self.candidates = list()
        self.summary_table = None

    def add_title(self):
        title = self.ws.cell(column=2, row=self.row)
        title.value = 'Interview Conducted'
        self.set_width(title.column, 25)
        scheduled_at = self.ws.cell(column=3, row=self.row)
        scheduled_at.value = self.scheduled_at
        self.set_width(title.column, 25)
        self.add_font(scheduled_at, '12B')
        scheduled_at.alignment = Alignment(horizontal='left')
        return title

    def add_position(self):
        position = self.ws.cell(column=4, row=self.row)
        position.value = 'Position:'

        self.extend_cell(position, right=1)
        return position

    def add_job_position(self):
        job_position = self.ws.cell(column=6, row=self.row)
        job_position.value = self.job_position

        self.extend_cell(job_position, right=3)
        return job_position

    def create_header(self):
        self.set_height(1, 25)

        title, position, job_position = (self.add_title(), self.add_position(),
                                         self.add_job_position())
        self.add_font([title, position, job_position], '12B')

    def create_candidate_detail_header(self):
        self.row = 3
        self.table_row = 3
        self.table_column = 3

        # SN number
        sn = self.ws.cell(column=1, row=self.row)
        sn.value = 'SN'
        self.extend_cell(sn, down=1)

        candidate = self.ws.cell(column=2, row=self.row)
        candidate.value = 'Name of the Candidate'

        self.extend_cell(candidate, down=1)

        self.row = 4

        self.set_height(4, 25)
        self.add_font([sn, candidate], '11B')
        self.align([sn, candidate], 'center')
        self.add_border([sn, candidate], ['right'])

    def get_candidate(self, candidate_id):
        try:
            return list(
                filter(lambda x: x.id == candidate_id, self.candidates)
            )[0]
        except IndexError:
            return None

    def add_candidate(self, candidate_info):
        candidate = self.get_candidate(candidate_info.get('candidate_id'))

        if not candidate:
            self.row += 1
            candidate = Candidate(
                self.ws,
                candidate_info.get('candidate_id'),
                candidate_info.get('candidate_name'),
                len(self.candidates) + 1,
                self.row,
                candidate_info.get('has_weightage', False)
            )
            self.candidates.append(candidate)

            self.row = candidate.end_row
        return candidate

    def fill_candidate_data(self):
        for candidate in self.question_answers:
            candidate_obj = self.add_candidate(candidate)

            # for evaluation table
            for question_answer in candidate.get('question_answers'):
                table = self.get_or_create_evaluation_table(question_answer)
                table.add_candidate(candidate_obj)
                table.populate_score()
                table.add_remarks(question_answer.get('remarks'))

        summary_table = self.get_or_create_summary_table()
        if summary_table:
            for candidate in self.candidates:
                summary_table.add_candidate(candidate)
                summary_table.populate_score(self)

        self.fill_empty_remarks()

    def fill_empty_remarks(self):
        for candidate in self.candidates:
            for table in self.tables:
                if not table.has_candidate_remarks(candidate):
                    table.add_candidate(candidate)
                    table.add_remarks(None)

    def get_or_create_summary_table(self):
        if not self.summary_table:
            gen = (x for x in self.question_answers if bool(x.get('question_answers')))

            try:
                candidate_info = next(gen)
            except StopIteration:
                candidate_info = None

            if candidate_info and candidate_info.get('question_answers'):
                candidate_question_answers = candidate_info.get('question_answers')
                table = SummaryTable(
                    self.ws, self.table_row,
                    self.table_column, candidate_question_answers[0]
                )
                table.populate_headers()
                self.summary_table = table
                return self.summary_table
        return None

    def get_or_create_evaluation_table(self, question_answer):
        table = self.get_table(question_answer.get('assigned_person_code'))
        if not table:
            table = EvaluationTable(
                self.ws, self.table_row,
                self.table_column, len(self.tables) + 1,
                question_answer
            )
            table.populate_headers()

            self.tables.append(table)
            self.table_column = table.col + 2

        else:
            setattr(table, 'question_answers', question_answer.get('question_answers'))
            setattr(table, 'given', question_answer.get('given_score'))
            setattr(table, 'interviewer_weightage', question_answer.get('interviewer_weightage', None))

        return table

    def get_table(self, assigned_person_code):
        try:
            return list(filter(lambda table: table.code == assigned_person_code, self.tables))[0]
        except IndexError:
            return None

    def aggregate_report(self, question=None, candidate=None):
        assert (question or candidate) is not None
        sums = [
            [qa.score for qa in table.get_candidate_question_answers(question, candidate)]
            for table in self.tables
        ]
        import itertools
        flatten_sums = list(itertools.chain(*sums))
        total_score = sum(flatten_sums)
        has_weightage = getattr(candidate, 'has_weightage', False)
        return round(total_score, 2) if has_weightage else round(total_score/len(flatten_sums), 2)
