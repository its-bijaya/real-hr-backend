"""
Organization Related Views. This view contains Organization Related Views.
i. Organization retrieve-update view.
ii. Holiday Create, List, Update, Retrieve, Delete View.
"""
from datetime import datetime as dt

import dateutil.parser as dateparser
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from irhrs.common.models import ReligionAndEthnicity, HolidayCategory
from irhrs.core.constants.common import RELIGION, ETHNICITY
from irhrs.core.mixins.file_import_mixin import BackgroundFileImportMixin
from irhrs.core.mixins.viewset_mixins import \
    (ListRetrieveCreateDestroyViewSetMixin, ListRetrieveUpdateViewSetMixin, OrganizationMixin, HRSOrderingFilter, OrganizationCommonsMixin)
from irhrs.core.utils.common import apply_filters
from irhrs.organization.models import OrganizationDivision, OrganizationBranch
from ..permissions import (OrganizationWritePermission,
                           HolidayPermission, DisallowUpdateOfPastHoliday)
from ....api.v1.serializers.organization import (OrganizationSerializer,
                                                 HolidaySerializer)
from ....models import Organization, Holiday
from ....utils.holiday import refresh_timesheets


class OrganizationViewSet(ListRetrieveUpdateViewSetMixin, OrganizationMixin):
    """
    list:

    List of organization with short details

    retrieve:

    Retrieve an organization detail given the slug.
    The organization that was earlier created with initial-setup contains the
    minimal detail about an organization. This method displays only the basic
    info - on first install, later shows updated details.

    ```javascript
    {
        "name": "ABC Company",
        "abbreviation": "ABC",
        "industry": {
            "slug": "logistic-courier-air-express-companies",
            "name": "Logistic / Courier / Air Express Companies"
        },
        "appearance": {
            "header_logo": "http://localhost:8000/media/organization/
                header-logo/download.png",
            "logo": "http://localhost:8000/media/organization/logo/
                download_6Suyvsg.png"
        },
        "ownership": "Government",
        "size": "50 - 100 employees",
        "established_on": null,
        "registration_number": '4446545465',
        "vat_pan_number": '54544654654',
        "address": {
            "address": "asdfasdf",
            "mailing_address": "asdfasdfasdfasdf"
        },
        "contacts": "{\"a\":\"123456789\"}",
        "email": "nachos@nachos.com",
        "about": "This is ABC Don",
        "website": "http://champions.com",
        "parent": null,
        "organization_head": null,
        "administrators": [
            {
                "full_name": "abc cba",
                "profile_picture": "N/A",
                "code": "abc1"
            },
            {
                "full_name": "xyz zyx",
                "profile_picture": "N/A",
                "code": "xyz1"
            }
        ],
        "slug": "abc-company"
    }
    ```
    update:

    Update Organization Details
    The organization details initially has minimal details -
    as complete details are required PUT method is used as
    a method for full update.
    # Field Info:
    # parent
    * Slug of parent.

    # industry
    * Slug of industry: /api/v1/commons/industry

    # organization_head
    * id of the user

    # administrators
    * list of user ids

    # contacts
    * JSON Field
    * Example
    ```
    {
        "Phone": "98123456789",
        "Fax": "019887654321"
    }
    ```

    # registration_number
    * Must contain at least one digit.Special characters except '-' disallowed.

    # vat_pan_number
    * Must contain at least one digit. Special characters except '-' disallowed.

    # social_link
    * List of JSON.
    * if no slug provided, creates. If provided updates.
    * Example
    ```
    [
        {
            "title": "facebook",
            "link": "https://facebook.com"
        },
        {
            "title": "You Tube",
            "slug": "youtubes",
            "link": "https://youtube.com"
        }
    ]
    ```

    Data
    ```javascript
        {
            "name": "ABC Company",
            "abbreviation": "ABC",
            "industry": 28,
            "appearance": {
                "header_logo": "http://localhost:8000/media/organization/
                    header-logo/download.png",
                "logo": "http://localhost:8000/media/organization/logo/
                    download_6Suyvsg.png"
            },
            "ownership": "Government",
            "size": "50 - 100 employees",
            "established_on": null,
            "registration_number": '4446545465',
            "vat_pan_number": '54544654654',
            "address": {
                "address": "asdfasdf",
                "mailing_address": "asdfasdfasdfasdf"
            },
            "contacts": "{\"a\":\"123456789\"}",
            "email": "nachos@nachos.com",
            "about": "This is ABC Don",
            "website": "http://champions.com",
            "parent": null,
            "organization_head": null,
            "administrators": [
                1,
                2
            ],
        }
    ```

    partial_update:

    Updates Organization Details partially.

    Accepts the same parameters as ```.update()``` but not all fields required.

    """
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    lookup_field = 'slug'
    lookup_url_kwarg = 'organization_slug'
    permission_classes = [OrganizationWritePermission]

    def get_serializer(self, *args, **kwargs):
        if self.action == 'list':
            kwargs.update({'fields': ["name", "abbreviation", "slug",
                                      "appearance"]})
        return super().get_serializer(*args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()

        allowed_organizations = self.request.user.switchable_organizations_pks

        if self.request.method == 'GET' and self.request.user.detail.organization:
            allowed_organizations.add(self.request.user.detail.organization.id)

        return queryset.filter(id__in=allowed_organizations)

    def get_organization(self):
        if self.action in ['list', 'create']:
            return None
        return super().get_organization()


class HolidayDateRangeFilter(filters.BaseFilterBackend):
    """
    Filter the time to be lt current time.
    """

    def filter_queryset(self, request, queryset, view):
        # From Query Parameter
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date:
            try:
                start_date_parsed = dateparser.parse(start_date)
                queryset = queryset.filter(date__gte=start_date_parsed)
            except (ValueError, OverflowError):
                pass
        if end_date:
            try:
                end_date_parsed = dateparser.parse(end_date)
                queryset = queryset.filter(date__lte=end_date_parsed)
            except (ValueError, OverflowError):
                pass
        return queryset


class HolidayViewSet(
    OrganizationCommonsMixin, BackgroundFileImportMixin,
    HRSOrderingFilter, OrganizationMixin, ListRetrieveCreateDestroyViewSetMixin
):
    """
    list:

    Lists all the available holidays for the organization.

    filters->

        start_date, end_date  (Holiday date range filter)

    data

        {
            "slug": "asdfsdf",
            "name": "asdfsdf",
            "date": "2019-01-01",
            "category": "public",
            "image": "http://localhost:8000/media/organization/holiday/abc.jpg",
            "rule": {
                "name": "XYZ",
                "division": null,
                "religion": "cao-dai",
                "ethnicity": "badi",
                "gender": "Male",
                "age": 22,
                "slug": "xyz"
            }
        }


    create:
    Creates a now holiday category by selecting holiday category.
    Optionally Religion and Ethnicity.

        {
            "name": "asdfasdffsdf",
            "date": "2019-01-01",
            "category": "public",
            "image": null,
            "rule": {
                "name": "asldkfj",
                "division": "alskd",
                "religion": "asldjf",
                "ethnicity": "LSJDF",
                "gender": "Male",
                "age": 21
            }
        }


    retrieve:
    Provides detailed Holiday View provided the slug.

         {
            "slug": "asdfsdf",
            "name": "asdfsdf",
            "date": "2019-01-01",
            "category": "public",
            "image": "http://localhost:8000/media/organization/holiday/abc.jpg",
            "rule": {
                "name": "XYZ",
                "division": null,
                "religion": "cao-dai",
                "ethnicity": "badi",
                "gender": "Male",
                "age": 22,
                "slug": "xyz"
            }
        }

    delete:
    Deletes a holiday instance provided a slug.

    """
    queryset = Holiday.objects.all()
    lookup_field = 'slug'
    serializer_class = HolidaySerializer
    filter_backends = (HolidayDateRangeFilter, filters.SearchFilter,
                       DjangoFilterBackend)
    search_fields = ('name',)
    all_fields = ('name', 'date', 'category', 'image', 'slug')
    ordering_fields_map = {
        'name': 'name',
        'date': 'date',
        'category': 'category__name',
        'applicable_to': 'rule__gender',
        'lower_age': 'lower_age',
        'upper_age': 'upper_age'
    }
    permission_classes = [HolidayPermission, DisallowUpdateOfPastHoliday]

    # field for used for file import
    import_fields = [
        'NAME',
        'CATEGORY',
        'START DATE',
        'END DATE',
        'GENDER',
        'DIVISION',
        'BRANCH',
        'RELIGION',
        'ETHNICITY',
        'LOWER_AGE',
        'UPPER_AGE',
        'DESCRIPTION'
    ]
    values = [
        'Dashain',
        '',
        '2019-01-01',
        '2019-01-01',
        'All',
        '',
        '',
        '',
        '',
        16,
        99,
        'Holiday_description'
    ]
    background_task_name = 'holiday'
    sample_file_name = 'holiday'
    non_mandatory_field_value = {
        'description': '',
        'gender': 'All',
        'lower_age': 16,
        'upper_age': 99,
        'division': [],
        'branch': [],
        'religion': [],
        'ethnicity': []
    }
    many_to_many_fields = ['division', 'branch', 'ethnicity', 'religion']
    nested_serializer_fields = {
        'rule': [
            'gender', 'division', 'branch', 'religion', 'ethnicity', 'lower_age', 'upper_age'
        ]
    }

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({
            'organization': self.get_organization(),
            'action': self.action
        })
        return context

    def filter_queryset(self, queryset, *args, **kwargs):
        filter_map = {
            'name': 'name',
            'category': 'category__slug',
            'start_date': 'date__gte',
            'end_date': 'date__lte'
        }
        queryset = apply_filters(
            self.request.query_params,
            filter_map,
            queryset
        )
        return super().filter_queryset(queryset)

    def get_queryset_fields_map(self):
        return {
            'gender': ['All', 'Male', 'Female', 'Other'],
            'division': OrganizationDivision.objects.filter(organization=self.organization),
            'branch': OrganizationBranch.objects.filter(organization=self.organization),
            'religion': ReligionAndEthnicity.objects.filter(category=RELIGION),
            'ethnicity': ReligionAndEthnicity.objects.filter(category=ETHNICITY),
            'category': HolidayCategory.objects.all()
        }

    def get_failed_url(self):
        return f'/admin/{self.organization.slug}/organization/settings/holiday/?status=failed'

    def get_success_url(self):
        return f'/admin/{self.organization.slug}/organization/settings/holiday/?status=success'

    @classmethod
    def get_data_for_serializer(cls, data, serializer_class):
        start_date = data.get('start_date')
        if isinstance(start_date, dt):
            start_date = start_date.date()
        data['start_date'] = start_date
        data['end_date'] = data.get('end_date', None) or start_date
        return super().get_data_for_serializer(data, serializer_class)

    def destroy(self, request, *args, **kwargs):
        # previously deleting past holiday was not supported.
        # From hris-3618, deleting past holiday is supported.
        holiday = self.get_object()
        holiday.delete()
        # refreshes timesheet after deleting the holiday.
        refresh_timesheets(holiday.date, self.get_organization())
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['GET'], url_path='export')
    def download_list(self, request, **kwargs):
        fields = [
            'NAME',
            'CATEGORY',
            'DATE',
            'GENDER',
            'DIVISION',
            'BRANCH',
            'RELIGION',
            'ETHNICITY',
            'LOWER_AGE',
            'UPPER_AGE',
            'DESCRIPTION'
        ]
        queryset = self.get_queryset()
        holidays = self.filter_queryset(queryset)

        wb = Workbook()
        ws = wb.active
        ws.append(fields)

        for i in range(1, len(fields) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30

        def get_detail(_queryset, field='name'):
            return ', '.join(_queryset.values_list(field, flat=True)) if _queryset else '-'

        for holiday in holidays:
            name = holiday.name
            category = holiday.category.name if holiday.category else ''
            date = holiday.date
            applicable_to = holiday.rule.gender
            division = get_detail(holiday.rule.division.all())
            branch = get_detail(holiday.rule.branch.all())
            religion = get_detail(holiday.rule.religion.all())
            ethnicity = get_detail(holiday.rule.ethnicity.all())
            lower_age = holiday.rule.lower_age
            upper_age = holiday.rule.upper_age
            description = holiday.description

            ws.append([name, category, date, applicable_to, division, branch,
                       religion, ethnicity, lower_age, upper_age, description])

        response = HttpResponse(
            content=save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=holiday-list.xlsx'
        return response
