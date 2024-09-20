from django.http import HttpResponse
from django.contrib.auth import get_user_model
import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet

from irhrs.core.constants.user import CONTACT_OF, DEPENDENT_DOCUMENT_TYPES, CONTACT_CHOICES
from irhrs.core.mixins.viewset_mixins import (
    OrganizationCommonsMixin,
    OrganizationMixin,
)
from irhrs.users.api.v1.permissions import UserContactDetailPermission
from irhrs.users.api.v1.serializers.import_contact_detail import UserContactDetailImportSerializer
from irhrs.users.api.v1.utils import UserImportBase

USER = get_user_model()


class UserContactDetailImportViewSet(OrganizationMixin, OrganizationCommonsMixin, ViewSet):
    permission_classes = [UserContactDetailPermission]
    fields = [
        "email/username",
        "contact_of",
        "name",
        "address",
        "emergency",
        "email",
        "is_dependent",
        "date_of_birth",
        "occupation",
        "dependent_id_type",
        "dependent_id_number",
        "number_type",
        "number"
    ]
    values = [
        "example@email.com",
        "Father",
        "random persom",
        "Kathmandu",
        "TRUE",
        "example@email.com",
        "FALSE",
        "YYYY-MM-DD",
        "Accountant",
        "citizenship certificate",
        "2078-900-33044",
        "Mobile",
        "9844280399"
    ]

    def create(self, request, *args, **kwargs):
        context = {"request": request, "organization": self.get_organization()}
        serializer = UserContactDetailImportSerializer(data=request.data, context=context)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("Successfully uploaded user contact details.")

    @action(methods=["get"], detail=False, url_path="sample")
    def get_sample(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(self.fields)
        worksheet.append(self.values)
        self.add_validators(worksheet, workbook=workbook)
        self.add_column_type(worksheet)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=sample.xlsx"

        workbook.save(response)
        return response

    def add_validators(self, ws, workbook):
        key_qs = {
            'contact_of': [c for c, _ in CONTACT_OF],
            'emergency': [True, False],
            'is_dependent': [True, False],
            'dependent_id_type': [d for _, d in DEPENDENT_DOCUMENT_TYPES]+["-"],
            'number_type': [c for c, _ in CONTACT_CHOICES],
        }
        for key in key_qs:
            ws.add_data_validation(self.get_dv(key, key_qs, workbook))

        return ws

    def add_column_type(self, ws):
        selected_fields = [{'name': 'number', 'type': '@'}]
        for field in selected_fields:
            index = UserImportBase.get_index(field.get('name'), self.fields)
            column_letter = get_column_letter(index)
            ws.column_dimensions[column_letter].number_format = field.get('type')
        return ws

    def get_dv(self, fieldname, key_qs, workbook):
        qs = key_qs[fieldname]
        index = UserImportBase.get_index(fieldname, self.fields)
        return UserImportBase.get_dv_for_qs(qs, index, fieldname, workbook)

