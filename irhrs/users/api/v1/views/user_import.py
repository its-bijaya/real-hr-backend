import json

from django.core.cache import cache
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from irhrs.core.mixins.viewset_mixins import CreateViewSetMixin, ListCreateViewSetMixin
from irhrs.organization.models import Organization
from irhrs.users.api.v1.permissions import UserDetailPermission, UserImportPermission
from irhrs.users.api.v1.serializers.user_import import UserImportSerializer, FailedUserImportSerializer
from irhrs.users.models import UserDetail
from irhrs.users.api.v1.utils import UserImportBase


class UserImportView(UserImportBase, ListCreateViewSetMixin):
    """
    Import users

    Upload a file to import users

    required fields are

    `file`: file containing user details.

    `organization`: organization_slug

    response or output fields:

    `status`: "imported with errors" or "successfully imported"

    `errors`: "list of unsuccessful values"

    valid fields to import users are:

        'code',
        'first_name',
        'middle_name',
        'last_name',
        'email',
        'gender'
    """
    queryset = UserDetail.objects.none()
    serializer_class = UserImportSerializer
    permission_classes = [UserImportPermission]

    def create(self, request, *args, **kwargs):
        organization = self.get_organization(raise_exception=True)
        if cache.get(f'block_user_import_{organization.id}', False):
            status = cache.get(f'user_import_status_{organization.id}', 'processing')
            raise ValidationError({"non_field_errors": f"Could not start import. Previous import is {status}"})
        return super().create(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        organization = self.get_organization(raise_exception=True)
        lock = cache.get(f'block_user_import_{organization.id}', False)
        status = cache.get(f'user_import_status_{organization.id}', 'processing' if lock else 'successful')
        failed_import_file = cache.get(f'failed_import_file_{organization.id}', None)

        return Response({
            'lock': lock,
            'status': status,
            'failed_import_file': failed_import_file
        })

    # @cache_page(60*60)
    @action(methods=['GET', 'POST'], detail=False,
            url_path='sample', url_name='download-sample')
    def download_sample(self, request):
        wb = Workbook()
        ws = wb.active
        ws.append(self.export_fields)

        for i in range(1, len(self.export_fields) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30

        ws.append(self.value)
        self.add_validators(ws, self.get_organization(raise_exception=True), workbook=wb)

        response = HttpResponse(
            content=save_virtual_workbook(wb),
            content_type=
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response['Content-Disposition'] = \
            'attachment; filename=sample-import.xlsx'
        return response

    def get_organization(self, raise_exception=False):
        if self.request.method == 'GET':
            organization_slug = self.request.query_params.get('organization', None)
        else:
            organization_slug = self.request.data.get('organization', None)

        organization = Organization.objects.filter(slug=organization_slug).first()

        if organization is None and raise_exception:
            if self.request.method == 'GET':
                if not organization_slug:
                    raise ValidationError({"non_field_errors": ["`organization`  required in query params."]})
                else:
                    raise ValidationError({'organization': [f"Organization with slug {organization_slug} does"
                                                            " not exist."]})

        return organization
