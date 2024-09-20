from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from irhrs.core.constants.user import GENDER_CHOICES, MARITAL_STATUS_CHOICES
from irhrs.hris.models import ChangeType
from irhrs.organization.models import EmploymentJobTitle, OrganizationDivision, EmploymentStatus, EmploymentLevel


class UserImportBase:
    export_fields = [
        'code',
        'username',
        'first_name',
        'middle_name',
        'last_name',
        'email',
        'gender',

        'date_of_birth',

        'joined_date',
        'experience_start_date',
        'experience_end_date',
        'marital_status',

        'job_title',
        'division',
        'employment_type',
        'employment_level',
        'employment_step',
        'change_type',
    ]
    value = ['EMP1',
             'john',
             'John',
             'Kumar',
             'Doe',
             'john.doe@example.com',
             'Male',
             '1990-01-01',
             '2000-01-01']

    @staticmethod
    def get_index(field, l):
        return l.index(field) + 1

    @staticmethod
    def get_dv_for_qs(qs, index, fieldname, workbook, field='slug'):
        if isinstance(qs, list):
            autocomplete = qs
        else:
            autocomplete = list(qs.values_list(field, flat=True))
        # v = ",".join(autocomplete)

        worksheet = workbook.create_sheet(fieldname)
        max_row = len(autocomplete)
        for value in autocomplete:
            worksheet.append([value])

        cl = get_column_letter(index)

        # dv = DataValidation(type="list", formula1=f'"{v}"', allow_blank=False)
        dv = DataValidation(type="list", formula1=f'{fieldname}!$A$1:$A${max_row}')
        dv.add(f'{cl}2:{cl}1048576')

        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        worksheet.sheet_state = 'hidden'
        worksheet.protection.sheet = True
        worksheet.protection.password = f"{fieldname}{max_row}"
        return dv

    @classmethod
    def get_dv(cls, fieldname, key_qs, workbook):
        qs = key_qs[fieldname]
        index = cls.get_index(fieldname, cls.export_fields)
        return cls.get_dv_for_qs(qs, index, fieldname, workbook)

    @classmethod
    def add_validators(cls, ws, organization, workbook):
        key_qs = {
            'job_title': EmploymentJobTitle.objects.filter(organization=organization),
            'division': OrganizationDivision.objects.filter(is_archived=False, organization=organization),
            'employment_type': EmploymentStatus.objects.filter(is_archived=False, organization=organization),
            'employment_level': EmploymentLevel.objects.filter(is_archived=False, organization=organization),
            'gender': [g for g, _ in GENDER_CHOICES],
            'marital_status': [m for m, _ in MARITAL_STATUS_CHOICES],
            'change_type': ChangeType.objects.filter(organization=organization),
        }

        for key in key_qs:
            ws.add_data_validation(cls.get_dv(key, key_qs, workbook))

        return ws

