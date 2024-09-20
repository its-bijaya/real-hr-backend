from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.validators import FileExtensionValidator
from rest_framework import serializers
from rest_framework.exceptions import ValidationError

from irhrs.core.constants.user import DEPENDENT_DOCUMENT_TYPES
from irhrs.core.utils.common import DummyObject
from irhrs.core.utils.excel import ExcelList, ExcelDict
from irhrs.core.validators import ExcelFileValidator
from irhrs.export.constants import ADMIN, FAILED
from irhrs.export.models.export import Export
from irhrs.export.utils.helpers import save_virtual_workbook
from irhrs.notification.utils import notify_organization
from irhrs.permission.constants.permissions import USER_CONTACT_DETAIL_IMPORT_PERMISSION

from irhrs.users.api.v1.serializers.contact_detail import UserContactDetailSerializer
from irhrs.users.models import UserContactDetail

USER = get_user_model()


class UserContactDetailImportSerializer(serializers.Serializer):
    excel_file = serializers.FileField(
        max_length=100,
        validators=[
            FileExtensionValidator(allowed_extensions=["xlsx", "xlsm", "xltx", "xltm"]),
            ExcelFileValidator()
        ],
        write_only=True,
    )

    @property
    def organization(self):
        return self.context["organization"]

    @property
    def request(self):
        return self.context["request"]

    @transaction.atomic
    def create(self, validated_data):
        excel_file = validated_data["excel_file"]
        workbook = load_workbook(excel_file)
        excel_list = ExcelList(workbook)
        user_contact_details = excel_list[1:]
        headers = excel_list[0]
        for contact_record in user_contact_details:
            email_or_username = contact_record[0]
            user = USER.objects.filter(
                Q(email=email_or_username) | Q(username=email_or_username)).distinct().first()
            contact_detail = self.get_contact_detail_with_heading(
                headers[1:],
                contact_record[1:]
            )
            serializer = UserContactDetailSerializer(
                data=contact_detail,
                context={'user': user, 'send_notification': False, **self.context}
            )
            serializer.is_valid()
            serializer.save()
        notify_organization(
            text="Successfully imported contact detail.",
            organization=self.organization,
            action=self.organization,
            actor=self.request.user,
            permissions=[USER_CONTACT_DETAIL_IMPORT_PERMISSION],
            url=f"/admin/{self.organization.slug}/hris/employees/import/contact-details"
        )
        return DummyObject()

    def validate(self, attrs):
        excel_file = attrs["excel_file"]
        workbook = load_workbook(excel_file)
        excel_list = ExcelList(workbook)
        error_exists = False
        user_contact_details = excel_list[1:]
        headers = excel_list[0]

        allowed_fields = [
            "email",
            "contact_of",
            "name",
            "address",
            "emergency",
            "email",
            "is_dependent",
            "date_of_birth",
            "occupation",
            "dependent_id_type",
            "dependent_id_number",
            "number",
            "number_type"
        ]

        if len(headers) > len(allowed_fields):
            raise ValidationError(f"Duplicate headings.")
        heading_error = {}
        for header in headers[1:]:
            if header is None:
                heading_error['error'] = "Can not send empty heading."
            elif header not in allowed_fields[1:]:
                heading_error[header] = "Incorrect heading."
        if heading_error:
            raise ValidationError(heading_error)
        for index, contact_record in enumerate(user_contact_details, 1):
            errors = {}
            email_or_username = contact_record[0]

            user = USER.objects.filter(
                Q(email=email_or_username) | Q(username=email_or_username)).distinct().first()
            if not user:
                errors[email_or_username] = "User not found."

            # turning our excel record into dictionary and only taking those
            # fields whose value exists, ("-" is also considered null)
            contact_detail_serializer_data = self.get_contact_detail_with_heading(
                headers[1:],
                contact_record[1:]
            )

            serializer = UserContactDetailSerializer(
                data=contact_detail_serializer_data,
                context={'user': user, 'send_notification': False, **self.context}
            )

            # turning serializer error into more readable format for user
            if not serializer.is_valid():
                serializer_errors = {
                    field: str(detail[0]) if isinstance(detail, list) else str(detail)
                    for field, detail in serializer.errors.items()
                }
                errors = {**errors, **serializer_errors}

            if errors:
                error_exists = True
                excel_list[index].append(str(errors))

        if error_exists:
            excel_list[0].append("Errors")
            error_wb = excel_list.generate_workbook()
            export_type = "ContactDetailImport"
            export = Export.objects.filter(export_type=export_type).first()
            # only create a new
            if not export:
                export = Export.objects.create(
                    user=self.request.user,
                    name="User Contact Detail Import",
                    exported_as=ADMIN,
                    export_type=export_type,
                    organization=self.organization,
                    status=FAILED,
                    remarks="User Contact Detail Import failed.",
                )

            export.export_file.save(
                "contact_detail_import.xlsx",
                ContentFile(save_virtual_workbook(error_wb)),
            )
            export.save()
            url = settings.BACKEND_URL + export.export_file.url
            notify_organization(
                text="Contact detail import failed.",
                organization=self.organization,
                action=self.organization,
                actor=self.request.user,
                permissions=[USER_CONTACT_DETAIL_IMPORT_PERMISSION],
                url=f"/admin/{self.organization.slug}/hris/employees/import/contact-details"
            )
            raise ValidationError({"error_file": url})
        return attrs

    @staticmethod
    def get_contact_detail_with_heading(headers, contact_record):
        field_value_mapper = {
            **{item[1]: item[0] for item in DEPENDENT_DOCUMENT_TYPES}
        }
        contact_data = dict(zip(headers, contact_record))
        result = {}
        for key, val in contact_data.items():
            if val in [None, "-"]:
                continue
            if key == 'date_of_birth':
                value = val.strftime("%Y-%m-%d")
            else:
                value = field_value_mapper.get(val, val)
            result[key] = value
        return result
