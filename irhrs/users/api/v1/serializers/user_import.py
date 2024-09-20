import uuid
from datetime import datetime as dtclass
from zipfile import BadZipFile

from django.conf import settings
from django.contrib.auth.models import AnonymousUser
from django.core.files.storage import default_storage
from django.core.validators import FileExtensionValidator
from django.core.cache import cache
from django_q.tasks import async_task
from openpyxl import Workbook
from rest_framework.fields import FileField, CharField, JSONField
from rest_framework.relations import SlugRelatedField
from rest_framework import serializers

from irhrs.core.mixins.serializers import DummySerializer
from irhrs.core.utils.common import DummyObject
from irhrs.core.utils.filters import inverse_mapping
from irhrs.export.utils.helpers import save_workbook, load_workbook
from irhrs.notification.utils import notify_organization
from irhrs.organization.models import Organization
from irhrs.permission.constants.permissions import HRIS_IMPORT_EMPLOYEE_PERMISSION, HRIS_PERMISSION
from irhrs.users.api.v1.utils import UserImportBase
from ..serializers.user_detail import UserCreateSerializer


class UserImportSerializer(UserImportBase, serializers.Serializer):
    """
    Serializer to import users from excel file.

    required fields are

    `file`: file containing user details.

    `organization`: organization_id

    response or output fields:

    `status`: "imported with errors" or "successfully imported"

    `errors`: "list of unsuccessful values"

    valid fields to import users are:

        'code',
        'first_name',
        'middle_name',
        'last_name',
        'email',
        'gender'
    """
    file = FileField(write_only=True, validators=[
        FileExtensionValidator(
            allowed_extensions=['xlsx'],
            message='Not a valid file format. Options are [.xlsx]',
            code='Invalid Format'
        )])
    organization = SlugRelatedField(queryset=Organization.objects.all(),
                                    write_only=True,
                                    slug_field='slug')
    message = CharField(read_only=True)
    imported = CharField(read_only=True)

    name_field_map = {
        'experience_start_date': 'start_date',
        'experience_end_date': 'end_date',
        'employment_type': 'employment_status',
        'employment_level': 'employee_level',
        'employment_step': 'current_step'
    }

    valid_input_fields = {
        'code',
        'username',
        'first_name',
        'middle_name',
        'last_name',
        'email',
        'gender',
        'date_of_birth',
        'joined_date'
    }

    def create(self, validated_data):

        organization = validated_data.get('organization')
        # save uploaded file to memory and process it
        path = self._save_file(validated_data.get('file'))

        context = dict(self.context)
        context["request"] = DummyObject(method="POST", user=getattr(
            self.context.get('request', None), 'user', AnonymousUser))
        context["view"] = DummyObject()

        cache.set(f'user_import_status_{organization.id}', 'queued', None)
        cache.set(f'block_user_import_{organization.id}', True, None)
        async_task(UserImportSerializer.start_import,
                   self.__class__, path, organization, context)

        return DummyObject(imported="success", message="File sent for background processing.")

    def update(self, instance, validated_data):
        return instance

    @staticmethod
    def start_import(cls, path, organization, context):
        cache.set(f'user_import_status_{organization.id}', 'processing', None)
        UserImportSerializer.save_records(cls, path, organization, context)
        cache.set(f'block_user_import_{organization.id}', False, None)

    @staticmethod
    def save_records(cls, path, organization, context):
        """
        Takes rows of the excel file and process it and save it.

        :param cls: class name
        :type cls: type
        :param path: path of file
        :type path: str
        :param organization: organization
        :return: errors
        :param context: Serializer Context
        :type context: dict
        """
        headers = None
        validation_errors = []

        sheet = cls.get_sheet(path, organization)

        if sheet is None:
            cache.set(f'user_import_status_{organization.id}', 'failed', None)
            cache.set(f'failed_import_file_{organization.id}', None, None)
            return

        rows = cls.parse(sheet)

        total_users = 0
        imported_users = 0
        for row in rows:
            if not headers:
                # get first rows as headers from excel
                headers = row

                # check if all required headers are provided
                if not cls.valid_input_fields.issubset(set(headers)):
                    notify_organization(
                        text=('User Import Failed. File formatting not matched.'
                              ' Please pass all required fields'),
                        action=organization,
                        organization=organization,
                        url=f'/admin/{organization.slug}/hris/employees/import',
                        permissions=[
                            HRIS_PERMISSION,
                            HRIS_IMPORT_EMPLOYEE_PERMISSION
                        ]
                    )
                    cache.set(
                        f'user_import_status_{organization.id}', 'failed', None)
                    cache.set(
                        f'failed_import_file_{organization.id}', None, None)
                    return

                elif len(set(headers)) != len(headers):
                    notify_organization(
                        text='User Import Failed. Duplicate headers found.',
                        action=organization,
                        organization=organization,
                        url=f'/admin/{organization.slug}/hris/employees/import',
                        permissions=[
                            HRIS_PERMISSION,
                            HRIS_IMPORT_EMPLOYEE_PERMISSION
                        ]
                    )
                    cache.set(
                        f'user_import_status_{organization.id}', 'failed', None)
                    cache.set(
                        f'failed_import_file_{organization.id}', None, None)
                    return

            else:
                # if row is empty or contains all values are None
                if len(row) == 0 or not any(row):
                    continue

                values = dict()

                # generate dictionary of field:value from list of values
                for i in range(0, len(headers)):
                    value = row[i]
                    if isinstance(value, dtclass):
                        value = value.date()
                    values.update({headers[i]: value})
                values.update({'organization': organization.slug})
                data = cls._get_data_for_serializer(values)
                data['employment']['organization'] = organization.slug

                total_users += 1
                serializer = UserCreateSerializer(data=data, context=context)
                if serializer.is_valid():
                    try:
                        ud = serializer.save()
                        ud.organization = organization
                        ud.save()
                        imported_users += 1
                    except serializers.ValidationError as e:
                        # some checks are only performed on save, such as
                        # max_users_count
                        values.update({'errors': str(e.detail)})
                        validation_errors.append(values)
                else:
                    values.update({'errors': cls.stringify_errors(serializer)})
                    validation_errors.append(values)

        cls._delete_path(path)

        filename = cls.generate_failed_exports(
            validation_errors, organization=organization)

        if len(validation_errors) > 0:
            if imported_users > 0:
                data = {
                    "message": f"User Import Failed. Imported {imported_users}"
                               f" out of {total_users} users.",
                    "imported": "partial",
                    "filename": settings.BACKEND_URL + settings.MEDIA_URL + filename,
                    "url": f'/admin/{organization.slug}/hris/employees/import/?status=failed'
                }
            else:
                data = {
                    "message": f"Could not import any of {total_users}"
                               " users.",
                    "imported": "partial",
                    "filename": settings.BACKEND_URL + settings.MEDIA_URL + filename,
                    "url": f'/admin/{organization.slug}/hris/employees/import/?status=failed'
                }
        else:
            if imported_users == 0 and total_users == 0:
                data = {
                    "message": f"User Import failed. No users were found in the file.",
                    "imported": "failed",
                    "filename": None,
                    "url": f'/admin/{organization.slug}/hris/employees/import/?status=failed'
                }
            else:
                data = {
                    "message": f"User Import complete. Successfully imported {total_users} users.",
                    "imported": "success",
                    "filename": None,
                    "url": f"/admin/{organization.slug}/hris/employees/current-employee-list"
                }

        cache.set(
            f'user_import_status_{organization.id}', data['imported'], None)
        cache.set(
            f'failed_import_file_{organization.id}', data['filename'], None)
        notify_organization(
            text=data['message'],
            action=organization,
            organization=organization,
            url=data["url"],
            permissions=[
                HRIS_PERMISSION,
                HRIS_IMPORT_EMPLOYEE_PERMISSION
            ]
        )

    @classmethod
    def generate_failed_exports(cls, failed, organization):

        fields = cls.export_fields + ['errors']

        wb = Workbook()
        ws = wb.active
        ws.append(fields)

        cls.add_validators(ws, organization=organization, workbook=wb)

        for fail in failed:

            row = []
            for field in fields:
                row.append(fail.get(field))
            ws.append(row)

        filename = f"{uuid.uuid4().hex}.xlsx"
        save_workbook(wb, filename)
        return filename

    @classmethod
    def stringify_errors(cls, serializer):
        """Build string from dictionary of errors"""
        error_dict = serializer.errors
        field_name_map = inverse_mapping(cls.name_field_map)
        error_str = ''
        for key, val in error_dict.items():
            if isinstance(val, dict):
                for k, v in val.items():
                    error_str += f"{field_name_map.get(k, k)}:"\
                        f"{','.join(v) if isinstance(v, list) else v}\n"
            else:
                error_str += f"{field_name_map.get(key, key)}:"\
                    f" {','.join(val) if isinstance(val, list) else val}\n"
        return error_str.strip()

    @classmethod
    def get_sheet(cls, file, organization):
        """
        :param file: File opened in binary read mode
        :type file: File
        :param organization: Organization Instance
        :type organization: Organization
        """
        try:
            workbook = load_workbook(filename=file, read_only=True)
        except (OSError, BadZipFile):
            notify_organization(
                text='User Import Failed. Corrupt file was sent.',
                action=organization,
                organization=organization,
                url=f'/admin/{organization.slug}/hris/employees/import',
                permissions=[
                    HRIS_PERMISSION,
                    HRIS_IMPORT_EMPLOYEE_PERMISSION
                ]
            )
            return None

        if len(workbook.sheetnames) > 0:
            sheet_name = workbook.sheetnames[0]
        else:
            notify_organization(
                text='User Import Failed. File had no data in it.',
                action=organization,
                organization=organization,
                url=f'/admin/{organization.slug}/hris/employees/import',
                permissions=[
                    HRIS_PERMISSION,
                    HRIS_IMPORT_EMPLOYEE_PERMISSION
                ]
            )
            return None

        return workbook[sheet_name]

    @staticmethod
    def parse(sheet):
        """
        Parse excel file and return rows

        :param sheet: worksheet
        :type sheet: WorkSheet
        :return: generator returning values of rows of excel files
        """

        for row in sheet:
            values = []
            for col in row:
                values.append(col.value)
            yield values

    @classmethod
    def _get_data_for_serializer(cls, data):
        """
        Get fields from excel row and generate structured data for
        UserDetailSerializer

        :param data: excel fields and values
        :type data: dict
        :return: data
        """
        user_fields = [
            "username",
            "email",
            "first_name",
            "middle_name",
            "last_name",
            "password",
            "organization"
        ]
        experience_fields = [
            'experience_start_date',
            'experience_end_date',
            'job_title',
            'division',
            'branch',
            'employment_level',
            'employment_type',
            'employment_step',
            'change_type'
        ]
        return_data = {
            "user": {},
            "employment": {}
        }

        for field in data:
            if field in user_fields:
                if field == 'middle_name' and not data[field]:
                    data[field] = ''
                return_data['user'].update({field: data[field]})
            elif field in experience_fields:
                return_data['employment'].update(
                    {cls.name_field_map.get(field, field): data[field]})
            else:
                return_data.update({field: data[field]})
        return return_data

    @staticmethod
    def _save_file(in_memory_file):
        """
        Save in memory file to file system

        :param in_memory_file:
        :return: path to file
        """
        extension = in_memory_file.name.split('.')[-1]
        return default_storage.save("{}.{}".format(str(uuid.uuid4()), extension), in_memory_file)

    @staticmethod
    def _delete_path(path):
        """Delete given file path from the system"""
        if default_storage.exists(path):
            default_storage.delete(path)


class FailedUserImportSerializer(DummySerializer):
    """
    Serializer used for exporting user fields
    """
    organization = SlugRelatedField(
        queryset=Organization.objects.all(),
        write_only=True, slug_field='slug'
    )
    failed = JSONField()
