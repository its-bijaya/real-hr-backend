import datetime

import openpyxl
from dateutil.parser import parse
import pytz
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db.models import (
    F,
    Count,
    Sum,
    Q,
    Subquery,
    OuterRef,
    FloatField,
    Exists,
    Value,
)
from django.db.models.functions import Coalesce, Concat
from django.http import Http404
from django.utils import timezone
from django.utils.functional import cached_property
from django_filters.rest_framework import DjangoFilterBackend

# Rest_framework imports
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import filters, serializers

# Project other app imports
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.fields import ReadOnlyField
from rest_framework.response import Response
from rest_framework.serializers import Serializer

from irhrs.attendance.api.v1.reports.views.individual_attendance_report import (
    center_align,
    white_font,
    blue_fill,
)
from irhrs.core.constants.user import MALE, FEMALE, OTHER
from irhrs.core.mixins.serializers import (
    DummySerializer,
    create_read_only_dummy_serializer,
)
from irhrs.core.mixins.viewset_mixins import (
    ListViewSetMixin,
    OrganizationMixin,
    DateRangeParserMixin,
    HRSOrderingFilter,
    UserMixin,
    ModeFilterQuerysetMixin,
    PastUserFilterMixin,
    PastUserParamMixin,
)

# Project current app imports
from irhrs.core.pagination import NoCountLimitZeroNoResultsPagination
from irhrs.core.utils import this_month_range, get_system_admin, nested_get
from irhrs.core.utils.common import get_today
from irhrs.core.utils.filters import (
    get_applicable_filters,
    FilterMapBackend,
    NullsAlwaysLastOrderingFilter,
    OrderingFilterMap,
)
from irhrs.export.mixins.export import BackgroundExcelExportMixin
from irhrs.export.utils.export import ExcelExport
from irhrs.leave.api.v1.permissions import AdminOnlyLeaveReportPermission
from irhrs.leave.api.v1.reports.views.mixins import LeaveReportPermissionMixin
from irhrs.leave.api.v1.serializers.account import LeaveAccountHistorySerializer
from irhrs.leave.api.v1.serializers.on_leave import (
    UserOnLeaveSerializer,
    UserOnLeaveThinSerializer,
    FutureOnLeaveSerializer,
)
from irhrs.leave.constants.model_constants import APPROVED, TIME_OFF, CREDIT_HOUR
from irhrs.leave.constants.model_constants import DEDUCTED, COMPENSATORY
from irhrs.leave.models import (
    LeaveType,
    LeaveAccount,
    LeaveAccountHistory,
    MasterSetting,
)
from irhrs.leave.models.request import LeaveSheet
from irhrs.leave.tasks import get_active_master_setting
from irhrs.leave.utils.mixins import LeaveRequestPastUserFilterMixin
from irhrs.organization.models import FiscalYear
from irhrs.permission.constants.permissions import (
    LEAVE_PERMISSION,
    LEAVE_REPORT_PERMISSION,
)
from irhrs.users.api.v1.serializers.thin_serializers import (
    UserThinSerializer,
    UserFieldThinSerializer,
)
from .....models import LeaveRequest

# Django imports
from .....utils.balance import get_fiscal_year_for_leave

User = get_user_model()
col = get_column_letter
HOURLY_CATEGORY = (TIME_OFF, CREDIT_HOUR)


class UserOnLeaveViewSet(
    PastUserFilterMixin, OrganizationMixin, LeaveReportPermissionMixin, ListViewSetMixin
):
    queryset = User.objects.all()
    serializer_class = UserOnLeaveSerializer
    filter_backends = (filters.SearchFilter, OrderingFilterMap, FilterMapBackend)
    ordering_fields_map = {
        "full_name": ("first_name", "middle_name", "last_name"),
        "num_leaves": "num_leaves",
    }
    filter_map = {
        "branch": "detail__branch__slug",
        "marital_status": "detail__marital_status",
        "gender": "detail__gender",
        "employment_level": "detail__employment_level__slug",
        "employment_status": "detail__employment_status__slug",
    }
    search_fields = ("first_name", "middle_name", "last_name")

    def _get_request_filters(self):
        filter_map = {
            "leave_type": "leave_requests__leave_rule__leave_type",
            "division": "detail__division__slug",
            "branch": "detail__branch__slug",
            "gender": "detail__gender__iexact",
            "marital_status": "detail__marital_status__iexact",
            "start_date": "leave_requests__start__date__gte",
            "end_date": "leave_requests__end__date__lte",
        }

        applicable_filters = get_applicable_filters(
            self.request.query_params, filter_map
        )
        return applicable_filters

    def get_serializer_class(self):
        if self.action == "summary":
            return UserOnLeaveThinSerializer
        return super().get_serializer_class()

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx.update({"applicable_filters": self._get_request_filters()})
        return ctx

    def age_range_filter_queryset(self, queryset):
        age_range = self.request.query_params.get("age_group")
        try:
            today = timezone.now().date()
            lower_bound, upper_bound = map(int, age_range.split("-"))
            lower_dob, upper_dob = (
                today - timezone.timedelta(days=lower_bound * 365),
                today - timezone.timedelta(days=upper_bound * 365),
            )
        except (ValueError, AttributeError):
            return queryset
        queryset = queryset.filter(
            detail__date_of_birth__lte=lower_dob,
            detail__date_of_birth__gt=upper_dob,
        )
        return queryset

    def filter_queryset(self, queryset):
        applicable_filters = self._get_request_filters()
        qs = super().filter_queryset(queryset)
        qs = qs.filter(
            Q(Q(leave_requests__status=APPROVED) & Q(leave_requests__is_deleted=False)),
            **applicable_filters,
        )
        leave_for = self.request.query_params.get("leave_for")

        if self.action == "summary":
            leave_for = "today"

        if leave_for in ["today", "future"]:
            if leave_for == "today":
                leave_sheet = LeaveSheet.objects.filter(
                    leave_for=get_today(),
                    request__is_deleted=False,
                    request__status=APPROVED,
                ).filter(request__user=OuterRef("id"))
                qs = qs.annotate(leave_sheet_exists=Exists(leave_sheet)).filter(
                    leave_sheet_exists=True
                )
            elif leave_for == "future":
                leave_sheet = LeaveSheet.objects.filter(
                    leave_for__gt=get_today(),
                    request__is_deleted=False,
                    request__status=APPROVED,
                ).filter(request__user=OuterRef("id"))
                qs = (
                    qs.annotate(leave_sheet_exists=Exists(leave_sheet))
                    .filter(leave_sheet_exists=True)
                    .distinct()
                )

        if self.action != "summary":
            qs = self.age_range_filter_queryset(qs)
            qs = qs.annotate(
                num_leaves=Sum(
                    "leave_requests__balance",
                    filter=Q(
                        Q(leave_requests__status=APPROVED)
                        & Q(leave_requests__is_deleted=False)
                    ),
                )
            )

        month = self.request.query_params.get("month")
        if month:
            month = int(month)
            if not (0 <= month <= 11):
                raise Http404

            try:
                year = int(self.request.query_params.get("year", get_today().year))
                start = parse(f"{year}-{month + 1}-01")
                end = parse(
                    f"{year}-{month + 2}-01" if month != 11 else f"{year + 1}-01-01"
                )
                qs = qs.filter(
                    leave_requests__end__date__lte=end,
                    leave_requests__start__date__gte=start,
                )
            except TypeError:
                return ValidationError({"year": ["Bad Year"]})
        return qs.distinct()

    def get_queryset(self):
        supervisor_id = self.request.query_params.get("supervisor")
        fil = dict(
            leave_requests__status=APPROVED, detail__organization=self.organization
        )

        if supervisor_id:
            if supervisor_id == str(self.request.user.id):
                fil.update({"id__in": self.request.user.subordinates_pks})
            else:
                # if supervisor does not match return none
                return super().get_queryset().none()
        return (
            super()
            .get_queryset()
            .current()
            .filter(**fil)
            .select_related(
                "detail",
                "detail__employment_level",
                "detail__job_title",
                "detail__organization",
                "detail__division",
            )
        )

    @action(
        methods=["GET"],
        detail=False,
        url_path="today-summary",
        url_name="summary",
        pagination_class=NoCountLimitZeroNoResultsPagination,
    )
    def summary(self, request, *args, **kwargs):
        """On Leave Users today"""
        today = get_today()

        leave_sheet = (
            LeaveSheet.objects.exclude(
                # Except deleted ones.
                request__is_deleted=True
            )
            .filter(
                # Leave for today.
                leave_for=today,
            )
            .filter(request__user_id=OuterRef("pk"))
        )
        queryset = self.filter_queryset(self.get_queryset()).annotate(
            leave_type=Subquery(
                leave_sheet.values("request__leave_rule__leave_type__name")[:1]
            ),
            part_of_day=Subquery(leave_sheet.values("request__part_of_day")[:1]),
        )

        page = self.paginate_queryset(queryset)
        return self.get_paginated_response(self.get_serializer(page, many=True).data)

    def leave_for_fixed_dates(self, start_date, end_date=None):
        fil = {"start__date__gte": start_date}
        if end_date:
            fil["end__date__lte"] = end_date
        leave_sheets = LeaveSheet.objects.filter(**fil)
        qs = (
            LeaveRequest.objects.filter(
                status=APPROVED,
                is_deleted=False,
                user__detail__organization=self.organization,
                sheets__in=leave_sheets,
            )
            .distinct()
            .order_by("start")
        )
        page = self.paginate_queryset(qs)
        return self.get_paginated_response(
            FutureOnLeaveSerializer(
                page, many=True, context=self.get_serializer_context()
            ).data
        )

    @action(
        methods=["GET"],
        detail=False,
    )
    def future(self, request, **kwargs):
        start = timezone.now().date()
        return self.leave_for_fixed_dates(start_date=start)

    @action(methods=["GET"], detail=False, url_path="this-month")
    def this_month(self, request, **kwargs):
        start, end = this_month_range()
        return self.leave_for_fixed_dates(start_date=start, end_date=end)


class OnLeaveViewSet(
    LeaveRequestPastUserFilterMixin,
    OrganizationMixin,
    LeaveReportPermissionMixin,
    HRSOrderingFilter,
    ListViewSetMixin,
):
    """
    list:

    ## Lists all the users on leave today.
    ### Filters:
    * leave_type: id
    * branch:
    * division: Division Slug
    * gender: Case Insensitive Gender
    * marital_status: Care insensitive marital status
    * employment_level : employment_level_slug
    * employment_status : slug
    * supervisor: supervisor_id

    ### Ordering
    * full_name

    """

    queryset = LeaveRequest.objects.filter(is_deleted=False).all()
    serializer_class = type(
        "OnLeaveSerializer",
        (Serializer,),
        {
            "user": UserThinSerializer(),
            "supervisor": UserThinSerializer(
                source="user.first_level_supervisor", read_only=True
            ),
            "num_leaves": serializers.ReadOnlyField(),
        },
    )
    filter_backends = (
        filters.SearchFilter,
        filters.OrderingFilter,
        DjangoFilterBackend,
    )
    ordering_fields_map = {
        "full_name": ("user__first_name", "user__middle_name", "user__last_name")
    }
    ordering = "-num_leaves"
    search_fields = (
        "user__first_name",
        "user__middle_name",
        "user__last_name",
    )

    def age_range_filter_queryset(self, queryset):
        age_range = self.request.query_params.get("age_group")
        try:
            today = timezone.now().date()
            lower_bound, upper_bound = map(int, age_range.split("-"))
            lower_dob, upper_dob = (
                today - timezone.timedelta(days=lower_bound * 365),
                today - timezone.timedelta(days=upper_bound * 365),
            )
        except (ValueError, AttributeError):
            return queryset
        queryset = queryset.filter(
            user__detail__date_of_birth__lte=lower_dob,
            user__detail__date_of_birth__gt=upper_dob,
        )
        return queryset

    def filter_queryset(self, queryset):
        filter_map = {
            "leave_type": "leave_rule__leave_type",
            "division": "user__detail__division__slug",
            "branch": "user__detail__branch__slug",
            "gender": "user__detail__gender__iexact",
            "marital_status": "user__detail__marital_status",
            "employment_level": "user__detail__employment_level__slug",
            "employment_status": "user__detail__employment_status__slug",
            "start_date": "start__gte",
            "end_date": "end__lte",
        }

        applicable_filters = get_applicable_filters(
            self.request.query_params, filter_map
        )
        # filter queryset for dob_group
        qs = self.age_range_filter_queryset(
            queryset
        )
        qs = qs.filter(
            **applicable_filters).annotate(
            num_leaves=Sum('balance')
        )
        return super().filter_queryset(qs)

    def get_queryset(self):
        supervisor_id = self.request.query_params.get("supervisor")
        fil = dict(status=APPROVED)

        if supervisor_id:
            if supervisor_id == str(self.request.user.id):
                fil.update(
                    {"leave_account__user_id__in": self.request.user.subordinates_pks}
                )
            else:
                # if supervisor does not match return none
                return super().get_queryset().none()
        else:
            # only use organization filter if supervisor is not passed in
            # query params, else filter by subordinates
            fil.update({"user__detail__organization": self.get_organization()})

        return (
            super()
            .get_queryset()
            .exclude(leave_rule__leave_type__category__in=HOURLY_CATEGORY)
            .filter(**fil)
            .select_related(
                "user__detail",
                "user__detail__organization",
                "user__detail__division",
                "user__detail__job_title",
                "user__detail__employment_level",
            )
        )

    @action(methods=["get"], detail=False, url_path="today", url_name="on-leave-today")
    def get_on_leave_today(self, *args, **kwargs):
        serializer = type(
            "TodayOnLeaveSerializer",
            (Serializer,),
            {
                "type": serializers.ReadOnlyField(),
                "num_leaves": serializers.ReadOnlyField(),
                "id": serializers.ReadOnlyField(),
            },
        )
        qs = (
            self.get_queryset()
            .filter(
                start__date__lte=timezone.now().date(),
                end__date__gte=timezone.now().date(),
            )
            .order_by()
            .values("leave_rule__leave_type__name")
            .annotate(
                num_leaves=Count("id"),
                type=F("leave_rule__leave_type__name"),
                id=F("leave_rule__leave_type"),
            )
        )
        data = serializer(qs, many=True).data
        return Response(data)

    @action(
        methods=["get"], detail=False, url_path="future", url_name="on-leave-future"
    )
    def get_on_leave_future(self, request, *args, **kwargs):

        start_date = request.query_params.get("start_date", "")
        end_date = request.query_params.get("end_date") or start_date
        try:
            start = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            if pytz.utc.localize(start) < timezone.now():
                raise ValueError
        except ValueError:
            start = timezone.now().date() + datetime.timedelta(days=1)
        try:
            end = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            if pytz.utc.localize(end) < timezone.now():
                raise ValueError
        except ValueError:
            end = start
        serializer = type(
            "TomorrowOnLeaveSerializer",
            (Serializer,),
            {
                "division": serializers.ReadOnlyField(),
                "slug": serializers.ReadOnlyField(),
                "num_leaves": serializers.ReadOnlyField(),
            },
        )
        qs = (
            self.get_queryset()
            .filter(**{"start__date__gte": start, "end__date__lte": end})
            .order_by()
            .values("user__detail__division__name")
            .annotate(
                num_leaves=Count("id"),
                division=F("user__detail__division__name"),
                slug=F("user__detail__division__slug"),
            )
        )
        data = serializer(qs, many=True).data
        return Response(data)

    @action(
        methods=["get"],
        detail=False,
        url_path="gender-report",
        url_name="gender-report",
    )
    def get_gender_leave_report(self, *args, **kwargs):
        serializer = type(
            "GenderLeaveReportSerializer",
            (Serializer,),
            {
                "gender": serializers.ReadOnlyField(source="user__detail__gender"),
                "num_leaves": serializers.ReadOnlyField(),
            },
        )
        qs = (
            self.get_queryset()
            .order_by()
            .values("user__detail__gender")
            .annotate(
                num_leaves=Sum("balance"),
            )
            .order_by("-num_leaves")
        )
        return Response(serializer(qs, many=True).data)

    @action(
        methods=["get"],
        detail=False,
        url_path="marital-status-report",
        url_name="marital-status-report",
    )
    def marital_status_leave_report(self, *args, **kwargs):
        serializer = type(
            "MaritalStatusLeaveReportSerializer",
            (Serializer,),
            {
                "marital_status": serializers.ReadOnlyField(
                    source="user__detail__marital_status"
                ),
                "num_leaves": serializers.ReadOnlyField(),
            },
        )
        qs = (
            self.get_queryset()
            .order_by()
            .values("user__detail__marital_status")
            .annotate(
                num_leaves=Sum("balance"),
            )
            .order_by("-num_leaves")
        )
        return Response(serializer(qs, many=True).data)

    @action(methods=["get"], detail=False, url_path="status-report")
    def status_leave_report(self, *args, **kwargs):
        filter_map = {
            "branch": "user__detail__branch__slug",
            "marital_status": "user__detail__marital_status",
            "gender": "user__detail__gender",
            "employment_level": "user__detail__employment_level__slug",
            "employment_status": "user__detail__employment_status__slug",
        }
        display = self.request.query_params.get("category")
        group_by = filter_map.get(display)
        if not display or not group_by:
            return Response("No Filters selected.")
        serializer = type(
            "MaritalStatusLeaveReportSerializer",
            (Serializer,),
            {
                display: serializers.ReadOnlyField(source=group_by),
                "num_leaves": serializers.ReadOnlyField(),
            },
        )
        qs = (
            self.get_queryset()
            .order_by()
            .exclude(
                **{f"{group_by}__isnull": True}  # exclude null values from exclude
            )
            .values(group_by)
            .annotate(
                num_leaves=Sum("balance"),
            )
            .order_by("-num_leaves")
        )
        return Response(serializer(qs, many=True).data)

    @action(methods=["get"], detail=False, url_path="by-category")
    def by_category(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        annotated_queryset = (
            queryset.order_by()
            .values(
                "leave_rule__leave_type__name",
                "leave_rule__leave_type",
            )
            .annotate(requests=Count("id"))
        )

        serializer = type(
            "SerializerHere",
            (DummySerializer,),
            {
                "leave_type": ReadOnlyField(source="leave_rule__leave_type__name"),
                "type_id": ReadOnlyField(source="leave_rule__leave_type"),
                "requests": ReadOnlyField(),
            },
        )

        data = serializer(self.paginate_queryset(annotated_queryset), many=True).data
        response = self.get_paginated_response(data)

        gender = request.query_params.get("gender")
        if not gender:
            response.data.update(
                {
                    "male": queryset.filter(user__detail__gender=MALE).count(),
                    "female": queryset.filter(user__detail__gender=FEMALE).count(),
                    "other": queryset.filter(user__detail__gender=OTHER).count(),
                }
            )
        return response

    @action(methods=["get"], url_path="age-report", detail=False)
    def get_age_report(self, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        result = {
            f"{lower_limit}-{upper_limit}": qs.filter(
                user__detail__date_of_birth__lte=start,
                user__detail__date_of_birth__gt=end,
            ).count()
            for (lower_limit, upper_limit), (
                start,
                end,
            ) in self.dob_ranges_from_age_groups.items()
        }
        return Response(result)

    @property
    def dob_ranges_from_age_groups(self):
        age_groups = [
            (0, 18),
            (18, 25),
            (25, 30),
            (30, 35),
            (35, 40),
            (40, 45),
            (45, 50),
            (50, 60),
            (60, 70),
        ]
        today = timezone.now().date()
        dob_ranges_from_age_groups = {
            (start, end): (
                today - timezone.timedelta(days=start * 365),
                today - timezone.timedelta(days=end * 365),
            )
            for start, end in age_groups
        }
        return dob_ranges_from_age_groups


class OnLeaveByLeaveTypeViewSet(
    PastUserParamMixin, DateRangeParserMixin, OrganizationMixin, ListViewSetMixin
):
    serializer_class = type(
        "OnLeaveByTypeReportSerializer",
        (Serializer,),
        {
            "leave_type": serializers.ReadOnlyField(source="name"),
            "num_leaves": serializers.ReadOnlyField(),
        },
    )
    permission_classes = [AdminOnlyLeaveReportPermission]
    queryset = LeaveType.objects.all()

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .filter(master_setting__organization=self.get_organization())
        )

    def filter_queryset(self, queryset):
        filter_map = {
            "gender": "leave_rules__leave_requests__user__detail__gender__iexact",
            "marital_status": "leave_rules__leave_requests__user__detail__marital_status__iexact",
            "start_date": "leave_rules__leave_requests__start__date__gte",
            "end_date": "leave_rules__leave_requests__end__date__lte",
        }

        filters = get_applicable_filters(self.request.query_params, filter_map)
        excludes = dict()
        if self.user_type == "past":
            excludes.update(
                {
                    "leave_rules__leave_requests__user__user_experiences__is_current": True
                }
            )
        else:
            filters.update(
                {
                    "leave_rules__leave_requests__user__user_experiences__is_current": True
                }
            )
        qs = (
            super()
            .filter_queryset(queryset)
            .order_by()
            .values("name")
            .annotate(
                num_leaves=Sum(
                    Coalesce("leave_rules__leave_requests__balance", 0),
                    filter=Q(**filters) & ~Q(**excludes),
                )
            )
            .order_by("-num_leaves")
            .filter(num_leaves__gt=0)
        )
        return qs


class IndividualMonthlyLeaveDetail(
    UserMixin,
    PastUserParamMixin,
    BackgroundExcelExportMixin,
    OrganizationMixin,
    ListViewSetMixin,
):
    """
    Required filters:

        fiscal_year=fiscal_year_id
        leave_types=1,2,3

        send_leave_type=true for leave types

    """

    queryset = LeaveAccount.objects.none()
    permission_classes = [AdminOnlyLeaveReportPermission]
    export_type = "Individual Monthly Leave"
    export_fields = []
    notification_permissions = [LEAVE_REPORT_PERMISSION]

    def get_selected_leave_types_ids(self):
        """
        :return: Selected leave types from query params
        """
        selected_leave_str = self.request.query_params.get("leave_types")
        if not selected_leave_str:
            return []
        try:
            selected_leaves = {int(lt_id) for lt_id in selected_leave_str.split(",")}
        except ValueError:
            raise ValidationError("Invalid leave types sent")
        return selected_leaves

    @cached_property
    def fiscal_year(self):
        fiscal_year_id = self.request.query_params.get("fiscal_year")

        if fiscal_year_id:
            fiscal = FiscalYear.objects.filter(
                organization=self.organization,
                pk=fiscal_year_id,
            ).first()
        else:
            fiscal = get_fiscal_year_for_leave(organization=self.organization)
        if not fiscal:
            raise ValidationError({"non_field_errors": ["No fiscal year selected."]})

        return fiscal

    @cached_property
    def fiscal_months(self):
        return self.fiscal_year.fiscal_months.filter(start_at__lte=timezone.now())

    @cached_property
    def master_setting(self):
        """Active master setting for given date range"""
        end_date = min(self.fiscal_year.end_at, get_today())
        ms = (
            MasterSetting.objects.filter(organization=self.get_organization())
            .active_for_date(end_date)
            .first()
        )
        if not ms:
            raise ValidationError(
                {
                    "non_field_errors": [
                        "No active master setting for given fiscal year."
                    ]
                }
            )
        return ms

    def get_available_leave_types(self):
        """Available leave types for selecting"""
        hourly = self.request.query_params.get("hourly")
        start_date = getattr(self.fiscal_year, "start_at", get_today())
        end_date = getattr(self.fiscal_year, "end_at", get_today())
        base = LeaveType.objects.filter(
            master_setting__in=MasterSetting.objects.filter(
                organization=self.get_organization(),
            ).active_between(start_date, end_date)
        )
        if hourly:
            if hourly == "true":
                base = base.filter(category__in=HOURLY_CATEGORY)
            elif hourly == "false":
                base = base.exclude(category__in=HOURLY_CATEGORY)
        return base.annotate(
            exists=Exists(
                LeaveAccount.objects.filter(
                    rule__leave_type_id=OuterRef("pk"), user=self.user
                )
            )
        ).filter(exists=True)

    def list(self, request, *args, **kwargs):
        user = self.user
        selected_leave_type_ids = self.get_selected_leave_types_ids()
        active_master_setting = self.master_setting
        months = self.fiscal_months
        result = self.get_result(
            user, selected_leave_type_ids, active_master_setting, months
        )

        fiscal = FiscalYear.objects.current(self.organization)
        if fiscal:
            fiscal = fiscal.id
        data = {"results": result, "current_fiscal": fiscal}

        if self.request.query_params.get("send_leave_types", None) == "true":
            data.update(
                {
                    "leave_type": [
                        {
                            "name": t.get("cname"),
                            "id": t.get("id"),
                            "category": t.get("category"),
                        }
                        for t in self.get_available_leave_types()
                        .annotate(
                            cname=Concat(
                                F("name"), Value("--"), F("master_setting__name")
                            )
                        )
                        .values("cname", "id", "category")
                    ]
                }
            )
            # data.update({'leave_type': list(
            #     self.get_available_leave_types().values('name', 'id', 'category'))})

        return Response(data)

    def get_export_type(self):
        export_type = super().get_export_type()
        return export_type + str(self.user.id)

    def get_export_name(self):
        return self.export_type

    def get_extra_export_data(self):
        return dict(
            user=self.user,
            selected_leave_type_ids=self.get_selected_leave_types_ids(),
            active_master_setting=self.master_setting,
            months=self.fiscal_months,
            organization=self.organization,
            redirect_url=self.get_frontend_redirect_url(),
            exported_as=self.get_exported_as(),
            notification_permissions=self.notification_permissions,
        )

    @classmethod
    def get_exported_file_content(
        cls, data, title, columns, extra_content, description=None, **kwargs
    ):
        json_data = cls.get_result(
            **{
                k: v
                for k, v in extra_content.items()
                if k
                in [
                    "user",
                    "selected_leave_type_ids",
                    "active_master_setting",
                    "months",
                ]
            }
        )
        wb = openpyxl.Workbook()
        wb.active.title = extra_content.get("user").full_name

        ws = wb.active
        organization = extra_content.get("organization")
        lines_used = ExcelExport.insert_org_info(ws, organization)

        start_index, start_row = 2, lines_used + 2
        subheaders = [
            "initial_balance",
            "remaining_balance",
            "leave_taken",
        ]

        headers = (
            extra_content.get("active_master_setting")
            .leave_types.exclude(category=TIME_OFF)
            .annotate(
                exists=Exists(
                    LeaveAccount.objects.filter(
                        rule__leave_type_id=OuterRef("pk"),
                        user=extra_content.get("user"),
                    )
                )
            )
            .filter(exists=True, id__in=extra_content.get("selected_leave_type_ids"))
        )
        for header in headers:
            end_index = start_index + len(subheaders) - 1
            ws.merge_cells(f"{col(start_index)}{start_row}:{col(end_index)}{start_row}")
            cell = ws[f"{col(start_index)}{start_row}"]
            cell.value = header.name.title()
            cell.alignment = center_align
            cell.font = white_font
            cell.fill = blue_fill
            start_index = end_index + 1

        pretty_header = [" ".join(hdr.split("_")).title() for hdr in subheaders]
        subheaders_merged = pretty_header * len(headers)
        subheaders_merged.insert(0, "Month Name")
        subheaders_merged.append("Total Remaining Leave")
        ws.append(subheaders_merged)
        for each_row in json_data:
            json_data_reformatted = {
                li.get("leave_type_id"): li for li in each_row.get("results")
            }
            li = [nested_get(each_row, "month")]
            for header in headers:
                data = json_data_reformatted.get(header.id)
                for hd in subheaders:
                    li.append(nested_get(data, hd))
            li.append(each_row.get("total_leave"))
            ws.append(li)
        return ContentFile(save_virtual_workbook(wb))

    @staticmethod
    def get_result(user, selected_leave_type_ids, active_master_setting, months):
        valid_leave_accounts = user.leave_accounts.exclude(
            rule__leave_type__category=TIME_OFF
        ).filter(
            rule__leave_type__master_setting=active_master_setting,
            rule__leave_type_id__in=selected_leave_type_ids,
            is_archived=False
        )

        result = list()
        for month in months:
            start_date, end_date = month.start_at, month.end_at

            st_ed = (start_date, end_date)
            data = list()

            for la in user.leave_accounts.filter(
                id__in=valid_leave_accounts
            ).select_related("rule", "rule__leave_type"):
                lt = la.rule.leave_type
                used = (
                    LeaveSheet.objects.filter(
                        request__user_id=user.id,
                        request__leave_rule__leave_type_id=lt.id,
                    )
                    .filter(
                        request__is_deleted=False,
                        request__status=APPROVED,
                        leave_for__range=st_ed,
                    )
                    .order_by()
                    .values("request__user_id")
                    .aggregate(total_balance=Sum("balance"))["total_balance"]
                )

                his = (
                    LeaveAccountHistory.objects.filter(
                        user_id=user.id, account__rule__leave_type_id=lt.id
                    )
                    .filter(modified_at__date__lte=end_date)
                    .order_by("-modified_at")
                    .only("new_usable_balance")
                    .first()
                )

                his_ = (
                    LeaveAccountHistory.objects.filter(
                        user_id=user.id, account__rule__leave_type_id=lt.id
                    )
                    .filter(modified_at__date__lte=start_date)
                    .order_by("-modified_at")
                    .only("new_usable_balance")
                    .first()
                )

                start_balance = his_.new_usable_balance if his_ else "-"
                end_balance = his.new_usable_balance if his else "-"
                used = "-" if end_balance == "-" else used or 0.0
                data.append(
                    {
                        "leave_type_id": lt.id,
                        "leave_type_name": lt.name,
                        "category": lt.category,
                        "initial_balance": start_balance,
                        "remaining_balance": end_balance,
                        "leave_taken": used,
                    }
                )
            total_leave = sum(
                [d["remaining_balance"] for d in data if d["remaining_balance"] != "-"]
            )
            result.append(
                {
                    "month": month.display_name,
                    "results": data,
                    "total_leave": total_leave,
                }
            )
        return result

    def get_frontend_redirect_url(self):
        return f"/admin/{self.organization.slug}/leave/reports/basic/individual-monthly-leave"


class CompensatoryLeaveReport(
    PastUserFilterMixin,
    ModeFilterQuerysetMixin,
    OrganizationMixin,
    ListViewSetMixin,
    DateRangeParserMixin,
    BackgroundExcelExportMixin,
):
    queryset = User.objects.all()
    filter_backends = (
        filters.SearchFilter,
        FilterMapBackend,
        NullsAlwaysLastOrderingFilter,
    )
    export_type = "Compensatory Leave Report"
    export_fields = {
        "Employee Name": "full_name",
        "Remaining": "remaining_balance",
        "Collapsed": "collapsed_balance",
    }
    serializer_class = type(
        "CompensatoryLeaveReportSerializer",
        (
            UserFieldThinSerializer,
            create_read_only_dummy_serializer(
                ["remaining_balance", "collapsed_balance"]
            ),
        ),
        {},
    )
    filter_map = dict(
        division="detail__division__slug",
        user="id",
    )
    search_fields = ("first_name", "middle_name", "last_name")
    permission_to_check = [LEAVE_PERMISSION, LEAVE_REPORT_PERMISSION]
    notification_permissions = [LEAVE_REPORT_PERMISSION]

    def get_ordering_fields_map(self):
        if self.action == "export":
            return {
                "full_name": ("first_name", "middle_name", "last_name"),
            }
        return {
            "remaining_balance": "remaining_balance",
            "collapsed_balance": "collapsed_balance",
            "full_name": ("first_name", "middle_name", "last_name"),
        }

    @staticmethod
    def annotate_queryset(queryset, start_date, end_date, organization):
        remaining_balance_for_range = Subquery(
            LeaveAccountHistory.objects.filter(
                user_id=OuterRef("pk"), account__rule__leave_type__category=COMPENSATORY
            )
            .filter(modified_at__date__lte=end_date)
            .order_by("-modified_at")
            .values("new_usable_balance")[:1],
            output_field=FloatField(),
        )
        compensatory_leave_accounts = LeaveAccount.objects.filter(
            rule__leave_type__master_setting=get_active_master_setting(organization),
            rule__leave_type__category=COMPENSATORY,
        )
        collapsed_balance_for_range = Subquery(
            LeaveAccountHistory.objects.filter(
                modified_at__date__range=(start_date, end_date),
                account__user=OuterRef("pk"),
                account__in=compensatory_leave_accounts,
            )
            .filter(
                action=DEDUCTED,
                actor=get_system_admin(),
                remarks__regex="^Collapsed \d{1,}\.\d{1,} balance for "
                "\d{4}-\d{2}-\d{2}$",
            )
            .annotate(
                collapsed_balance=F("previous_usable_balance") - F("new_usable_balance")
            )
            .order_by()
            .values("account__user")
            .annotate(sum_collapsed_balance=Sum("collapsed_balance"))
            .values("sum_collapsed_balance")[:1],
            output_field=FloatField(),
        )
        queryset = queryset.annotate(
            remaining_balance=remaining_balance_for_range,
            collapsed_balance=collapsed_balance_for_range,
        )
        return queryset

    def get_queryset(self):
        return super().get_queryset().filter(detail__organization=self.organization)

    def filter_queryset(self, queryset):
        self.filter_map.pop("start_date", None)
        self.filter_map.pop("end_date", None)
        if self.action == "export":
            return super().filter_queryset(queryset)
        return self.annotate_queryset(
            super().filter_queryset(queryset),
            *self.get_parsed_dates(),
            self.organization,
        ).select_essentials()

    def get_extra_export_data(self):
        start, end = self.get_parsed_dates()
        return {
            "start_date": start,
            "end_date": end,
            "organization": self.organization,
            "redirect_url": self.get_frontend_redirect_url(),
            "exported_as": self.get_exported_as(),
            "notification_permissions": self.notification_permissions,
        }

    @classmethod
    def get_exported_file_content(
        cls, queryset, title, columns, extra_content, description=None, **kwargs
    ):
        queryset = cls.annotate_queryset(
            queryset,
            extra_content.get("start_date"),
            extra_content.get("end_date"),
            extra_content.get("organization"),
        )
        return super().get_exported_file_content(
            queryset, title, columns, extra_content, description=description, **kwargs
        )

    @action(methods=["GET"], detail=False, url_path="collapsed/(?P<user_id>\d+)")
    def collapsed_balance_for_user(self, request, **kwargs):
        start_date, end_date = self.get_parsed_dates()
        qs = (
            LeaveAccountHistory.objects.filter(
                account__user=kwargs.get("user_id"),
                modified_at__date__range=(start_date, end_date),
            )
            .filter(
                action=DEDUCTED,
                actor=get_system_admin(),
                remarks__regex="^Collapsed \d{1,}\.\d{1,} balance for "
                "\d{4}-\d{2}-\d{2}$",
            )
            .annotate(
                collapsed_balance=F("previous_usable_balance") - F("new_usable_balance")
            )
        )
        page = self.paginate_queryset(qs)
        response = self.get_paginated_response(
            LeaveAccountHistorySerializer(
                page,
                many=True,
                fields=[
                    "created_at",
                    "previous_usable_balance",
                    "new_usable_balance",
                    "remarks",
                ],
            ).data
        )
        return response

    def get_frontend_redirect_url(self):
        return f"/admin/{self.organization.slug}/leave/reports/basic/compensatory"
