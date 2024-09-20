import re
from datetime import date, timedelta

import openpyxl
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db.models import OuterRef, Exists, Prefetch, Case, When, \
    IntegerField, Subquery, F
from django.utils.functional import cached_property
from openpyxl.styles.colors import BLUE, WHITE
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import filters
from rest_framework.exceptions import ValidationError
from rest_framework.filters import SearchFilter, OrderingFilter

from irhrs.attendance.utils.attendance import humanize_interval
from irhrs.core.mixins.viewset_mixins import OrganizationMixin, \
    ListViewSetMixin, ModeFilterQuerysetMixin, PastUserFilterMixin, \
    DateRangeParserMixin
from irhrs.core.utils import nested_get
from irhrs.core.utils.common import get_today
from irhrs.core.utils.filters import FilterMapBackend, \
    NullsAlwaysLastOrderingFilter
from irhrs.export.mixins.export import BackgroundExcelExportMixin
from irhrs.export.utils.export import ExcelExport
from irhrs.hris.api.v1.permissions import HRISReportPermission
from irhrs.leave.api.v1.reports.serializers.balance_report import \
    IndividualLeaveBalanceReportSerializer, CarryForwardLeaveDetailsSerializer
from irhrs.leave.api.v1.reports.views.mixins import LeaveReportPermissionMixin
from irhrs.leave.constants.model_constants import CREDIT_HOUR
from irhrs.leave.models import LeaveAccount, LeaveType, MasterSetting
from irhrs.leave.tasks import get_active_master_setting
from irhrs.leave.utils.balance import \
    get_applicable_leave_types_for_organization
from irhrs.permission.constants.permissions import LEAVE_PERMISSION, LEAVE_REPORT_PERMISSION

USER = get_user_model()
col = get_column_letter


class IndividualLeaveBalanceReport(
    PastUserFilterMixin,
    BackgroundExcelExportMixin,
    OrganizationMixin,
    LeaveReportPermissionMixin,
    ListViewSetMixin,
):
    serializer_class = IndividualLeaveBalanceReportSerializer
    filter_backends = (
        SearchFilter, OrderingFilter, FilterMapBackend
    )
    ordering = 'id'
    ordering_fields = (
        'first_name',
        'middle_name',
        'last_name'
    )
    search_fields = (
        'first_name',
        'middle_name',
        'last_name',
        'username',
    )
    filter_map = {
        'id': 'id',
        'division': 'detail__division__slug',
        'branch': 'detail__branch__slug',
        'username': 'username',
    }
    export_type = 'Individual Leave Balance Report'
    queryset = USER.objects.all()
    permission_classes = [HRISReportPermission]
    notification_permissions = [LEAVE_REPORT_PERMISSION]

    def check_permissions(self, request):
        supervisor = self.request.query_params.get("supervisor")
        if self.action == 'export':
            return supervisor and supervisor == str(
                self.request.user.id)
        return super().check_permissions(request)


    def get_selected_leave_type_ids(self):
        """
        :return: Selected leave types from query params
        """
        selected_leave_str = self.request.query_params.get('leave_types')
        if not selected_leave_str:
            return {}
        try:
            selected_leaves = {
                int(lt_id) for lt_id in selected_leave_str.split(',')
            }
        except ValueError:
            raise ValidationError('Invalid leave types sent')
        return selected_leaves

    def get_selected_leave_types(self):
        """Selected leave types queryset"""
        return list(LeaveType.objects.filter(id__in=self.get_selected_leave_type_ids()))

    def get_queryset(self):
        supervisor_id = self.request.query_params.get('supervisor')
        fil = dict(
            detail__organization=self.organization
        )

        if supervisor_id:
            if supervisor_id == str(self.request.user.id):
                fil.update({
                    'id__in':
                        self.request.user.subordinates_pks
                })
            else:
                # if supervisor does not match return none
                return USER.objects.none()
        return super().get_queryset().filter(
            **fil
        )

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        leave_archived_filter = {'is_archived': False} if self.user_type != 'past' else {}

        if self.action == 'export':
            return queryset

        selected_leave_type = self.get_selected_leave_type_ids()
        has_leave_account = LeaveAccount.objects.filter(
            user=OuterRef('id'),
            **leave_archived_filter
        ).order_by('rule__leave_type', '-modified_at').distinct('rule__leave_type')
        leave_account = LeaveAccount.objects.filter(
            rule__leave_type__master_setting=get_active_master_setting(
                self.organization
            ),
            **leave_archived_filter
        ).select_related(
            'rule',
            'rule__leave_type'
        ).order_by('rule__leave_type', '-modified_at')

        if selected_leave_type:
            has_leave_account = has_leave_account.filter(
                rule__leave_type__id__in=selected_leave_type,
            )
            leave_account = leave_account.filter(
                rule__leave_type__id__in=selected_leave_type,
            )

        queryset = queryset.annotate(
            has_leave_account=Exists(has_leave_account)
        ).filter(
            has_leave_account=True
        ).select_essentials(
        ).prefetch_related(
            Prefetch('leave_accounts', leave_account)
        )
        return self.order_by_leave_type(queryset)

    def order_by_leave_type(self, queryset):
        self.applicable_leaves = get_applicable_leave_types_for_organization(
            self.organization
        )
        ordering = self.request.query_params.get('ordering')
        desc = False
        if ordering and ordering.startswith('-'):
            ordering = ordering[1:]
            desc = True
        available_sorts = [
            leave_type.get('name') for leave_type in self.applicable_leaves
        ]

        if ordering in available_sorts:
            queryset = queryset.annotate(ordering_field=Subquery(
                LeaveAccount.objects.filter(user_id=OuterRef(
                    'pk'), rule__leave_type__name=ordering).values('usable_balance')[:1]
            ))

            order_by = F('ordering_field').desc(nulls_last=True) if desc else F(
                'ordering_field').asc(nulls_last=True)
            queryset = queryset.order_by(
                order_by
            )
        return queryset

    def list(self, request, *args, **kwargs):
        ret = super().list(request, *args, **kwargs)
        applicable_leaves = getattr(
            self,
            'applicable_leaves',
            None
        ) or get_applicable_leave_types_for_organization(
            self.organization
        )
        ret.data.update({
            'applicable_leaves': applicable_leaves
        })
        return ret

    def get_export_fields(self):
        export_fields = {
                "Username":"username",
                "Full Name": "full_name",
                "Job Title": "detail.job_title.title",
                "Employment Level": "detail.employment_level.title",
            }
        if self.get_selected_leave_type_ids():
            for lt in self.get_selected_leave_types():
                export_fields.update(
                    {
                        str(lt.name): str(lt.id),
                    }
                )
        else:
            for lt in get_applicable_leave_types_for_organization(self.organization):
                export_fields.update(
                    {
                        str(lt['name']): str(lt['id']),
                    }
                )
        return export_fields

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        for lt in get_applicable_leave_types_for_organization(obj.detail.organization):
            leave_account = obj.leave_accounts.filter(rule__leave_type__id=lt['id'])
            if leave_account.count() == 0:
                usable_balance = ''
            else:
                usable_balance = leave_account[0].usable_balance
                if leave_account[0].rule.leave_type.category == CREDIT_HOUR:
                    usable_balance = humanize_interval(usable_balance * 60)
            setattr(obj, str(lt['id']), usable_balance)
        return obj

    def get_extra_export_data(self):
        extra_data = super().get_extra_export_data()
        extra_data.update({
            'selected_leave_type': self.get_selected_leave_type_ids(),
            'organization': self.organization,
            'user_type': self.user_type
        })
        return extra_data

    @classmethod
    def get_exported_file_content(cls, queryset, title, columns, extra_content, description=None, **kwargs):
        selected_leave_type = extra_content.get('selected_leave_type')
        organization = extra_content.get('organization')
        user_type = extra_content.get('user_type')

        leave_archived_filter = {'is_archived': False} if user_type != 'past' else {}

        has_leave_account = LeaveAccount.objects.filter(
            user=OuterRef('id'), **leave_archived_filter
        ).order_by('rule__leave_type', '-modified_at').distinct('rule__leave_type')

        leave_account = LeaveAccount.objects.filter(
            rule__leave_type__master_setting=get_active_master_setting(organization),
            **leave_archived_filter
        ).select_related(
            'rule',
            'rule__leave_type'
        ).order_by('rule__leave_type', '-modified_at').distinct('rule__leave_type')
        if selected_leave_type:
            has_leave_account = has_leave_account.filter(
                rule__leave_type__id__in=selected_leave_type,
            )
            leave_account = leave_account.filter(
                rule__leave_type__id__in=selected_leave_type,
            )
        queryset = queryset.annotate(
            has_leave_account=Exists(has_leave_account)
        ).filter(
            has_leave_account=True
        ).prefetch_related(Prefetch('leave_accounts', leave_account))

        return super().get_exported_file_content(
            queryset, title, columns, extra_content, description=description, **kwargs
        )

    def get_frontend_redirect_url(self):
        supervisor = self.request.query_params.get("supervisor")
        if supervisor:
            return f'/user/supervisor/leave/reports/individual'
        return f'/admin/{self.organization.slug}/leave/reports/basic/individual-leave-balance'


class CarryForwardLeaveDetails(
    PastUserFilterMixin,
    ModeFilterQuerysetMixin,
    BackgroundExcelExportMixin,
    DateRangeParserMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    """
    Details are

    Available filter details
    -----------------------

    Employee name search - User should be able to enter text and search employee
    `?search=first_name`

    Department - User should be able to filter employee by department.
    `?division=slug`

    Leave Type - By default all leave types are displayed.
    User can select multiple number of leave.
    ```md
    By default all leave types are not displayed, they are selected as:
    ?leave_type=1,2,3,4
    ALSO: no data is displayed if no leave type is selected.
    ```

    Localize Date Picker - By default last two year should be selected.
    Changing localize date should change english date picker accordingly.
    User should be able to select only year from this date picker.
    ```md
    UPDATE: User can select a fiscal year.
    ?fiscal_type=fiscal&fiscal=26.
    A previous fiscal year, if available is auto displayed.
    ```

    English Date Picker - By default last two year should be selected.
    Changing english date should change localize date picker accordingly.
    User should be able to select only year from this date picker.
    ```md
    UPDATE: User should select a year.
    ?fiscal_type=gregorian&year=2019
    ALSO: previous year (i.e. 2018) is also selected.
    ```

    Table details
    Employee Name - Should display list of employee name. Required sorting.
    `?ordering=full_name`

    Year - Should display Year as shown in mock-up.
    `field_name:year`

    Leave Type - Should display leave type as shown in mock-up.
    `field_name: leave`

    Carry Forward Balance (CFB) -
    Should display CFB in heading and when hovering should display full form.
    It should display carry forward balance for selected year of respective
    employee. When click on number should redirect to leave balance history
    of the respective employee.
    `field_name: carry_forward`

    Leave Encashment Balance (LEB) - Should display LEB in heading and when
    hovering should display full form. It should display leave encashment
    balance for selected year of respective employee. When click on number
    should redirect to leave balance history of the respective employee.
    `field_name: encashment_balance`

    Collapse (C) - Should display C in heading and when hovering should
    display full form. It should display leave collapse balance for selected
    year of respective employee. When click on number should redirect to leave
    balance history of the respective employee.
    `field_name: collapsed_balance`

    Pagination required. [LIMIT/OFFSET available]

    Download - User should be able to download report for selected filter.
    `use /export` to export.
    """
    queryset = USER.objects.all()
    serializer_class = CarryForwardLeaveDetailsSerializer
    permission_to_check = [LEAVE_PERMISSION, LEAVE_REPORT_PERMISSION]
    filter_backends = (
        filters.SearchFilter, NullsAlwaysLastOrderingFilter, FilterMapBackend
    )
    search_fields = (
        'first_name', 'middle_name', 'last_name', 'username'
    )
    filter_map = {
        "division":"detail__division__slug",
        "user":"id",
        "username": "username",
    }
    ordering_fields_map = dict(
        full_name=('first_name', 'middle_name', 'last_name'),
    )
    export_fields = []
    allowed_extra = {
        'balance_added', 'used', 'initial_balance', 'accrued_balance'
    }
    notification_permissions = [LEAVE_REPORT_PERMISSION]

    @cached_property
    def leave_types(self):
        leave_from_query = self.request.query_params.get(
            'leave_type', ''
        )
        if not leave_from_query:
            return LeaveType.objects.none()
        leave_type_ids = [
            num for num in leave_from_query.split(',') if num.isdigit()
        ]
        _, end = self.fiscal_range
        return LeaveType.objects.filter(
            id__in=leave_type_ids,
            master_setting=MasterSetting.objects.filter(
                organization=self.organization,
            ).active_for_date(
                min((get_today(), end))
            ).first()
        )

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset).filter(
            leave_accounts__rule__leave_type__in=self.leave_types.values_list(
                'id', flat=True
            )
        ).distinct().select_essentials()
        if self.action == 'export':
            self.get_exported_file_content(queryset, '', [],
                                           self.get_serializer_context())
            return queryset
        return queryset.prefetch_related(
            Prefetch(
                'leave_accounts',
                queryset=LeaveAccount.objects.filter(
                    rule__leave_type__in=self.leave_types.values_list(
                        'id', flat=True
                    )
                ).select_related(
                    'rule',
                    'rule__leave_type',
                ),
                to_attr='prefetch_leave_accounts'
            )
        )

    def get_serializer_context(self):
        ctx = {}
        # Create two fiscal from current fiscal year
        fiscal_type = self.request.query_params.get('fiscal_type')
        selected_fiscals = [(None,) * 3]
        if fiscal_type == 'fiscal':
            fiscal = self.fiscal_year
            fiscals = list()
            if fiscal:
                for fis in [fiscal.previous, fiscal]:
                    if not fis:
                        continue
                    fiscals.append((
                        fis.name, fis.applicable_from, fis.applicable_to,
                        fis.next.applicable_from,
                        fis.next.applicable_to,
                    ))
                selected_fiscals = fiscals
        elif fiscal_type == 'gregorian':
            ret = []
            year = self.request.query_params.get('year')
            if year and re.fullmatch('\d{4}', year):
                this_year = int(year)
                years = [this_year - 1, this_year]
                for current_year in years:
                    st = date(this_year, 1, 1)
                    ed = date(this_year + 1, 1, 1) - timedelta(days=1)
                    st_n = date(this_year + 1, 1, 1)
                    ed_n = date(this_year + 2, 1, 1) - timedelta(days=1)
                    ret.append(
                        (str(current_year), st, ed, st_n, ed_n)
                    )
                selected_fiscals = ret
        extra_requested_fields = self.request.query_params.get('extra_fields')
        if extra_requested_fields:
            extra_fields = set(
                extra_requested_fields.split(',')
            ).intersection(self.allowed_extra)
        else:
            extra_fields = set()
        ctx.update({
            'selected_leave_types': self.leave_types.values_list(
                'id', flat=True
            ),
            'date_range': self.fiscal_range,
            'organization': self.organization,
            'selected_fiscals': selected_fiscals,
            'extra_fields': extra_fields,
        })
        return ctx

    def get_extra_export_data(self):
        extra_data = super().get_extra_export_data()
        extra_data.update(self.get_serializer_context())
        return extra_data

    def get_export_type(self):
        export_type = 'Carry Forward Leave Details'
        if 'extra_fields' in self.request.query_params:
            export_type = 'Yearly Leave Report'
        return export_type

    @classmethod
    def get_exported_file_content(cls, queryset, title, columns, extra_content, description=None, **kwargs):
        blue = openpyxl.styles.colors.Color(rgb='28688C')
        green = openpyxl.styles.colors.Color(rgb='147066')
        blue_fill = openpyxl.styles.fills.PatternFill(
            patternType='solid', fgColor=blue
        )
        green_fill = openpyxl.styles.fills.PatternFill(
            patternType='solid', fgColor=green
        )
        white_font = openpyxl.styles.Font(
            b=True, color=WHITE
        )
        bold_font = openpyxl.styles.Font(
            bold=True
        )
        center_align = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )
        json_data = cls.serializer_class(
            queryset,
            many=True,
            context=extra_content
        ).data
        wb = openpyxl.Workbook()
        ws = wb.active

        organization = extra_content.get('organization')
        lines_used = ExcelExport.insert_org_info(ws, organization)

        start_index, si, p_row, h_row = 3, 3, 2+lines_used, 3+lines_used
        subheaders = [
            'carry_forward',
            'collapsed_balance',
            'encashment_balance',
        ] + list(extra_content.get('extra_fields'))

        pre_headers = extra_content.get('selected_fiscals')
        headers = LeaveType.objects.filter(
            master_setting__organization=extra_content.get('organization'),
            id__in=extra_content.get('selected_leave_types'),
        )
        if not headers:
            return ContentFile(save_virtual_workbook(wb))
        for pre_header in pre_headers:
            ei = si + len(headers) * len(subheaders) - 1
            ws.merge_cells(
                f'{col(si)}{p_row}:{col(ei)}{p_row}'
            )
            cell = ws[f'{col(si)}{p_row}']
            cell.value = str(pre_header[0]).title()
            cell.alignment = center_align
            cell.font = white_font
            si = ei + 1
            cell.fill = green_fill
            for header in headers:
                end_index = start_index + len(subheaders) - 1
                ws.merge_cells(
                    f'{col(start_index)}{h_row}:{col(end_index)}{h_row}'
                )
                cell = ws[f'{col(start_index)}{h_row}']
                cell.value = header.name.title()
                cell.alignment = center_align
                start_index = end_index + 1
                cell.fill = blue_fill
                cell.font = white_font
        pretty_header = [
            ' '.join(hdr.split('_')).title() for hdr in subheaders
        ]
        subheaders_merged = pretty_header * len(headers) * len(pre_headers)
        subheaders_merged.insert(0, 'Username')
        subheaders_merged.insert(1, 'Employee Name')
        ws.append(subheaders_merged)

        subheaders_merged_index_in_excel=h_row +1
        for cell in ws[f"{subheaders_merged_index_in_excel}"]:
            cell.font = bold_font
        for row_data in json_data:
            restructure_one = {
                yr.get('year'): yr for yr in row_data.get('results')
            }
            li = list()
            for yr_identifier in pre_headers:
                row_year = restructure_one.get(yr_identifier[0])
                json_data_reformatted = {
                    li.get('id'): li for li in row_year.get('balance_details')
                }
                for header in headers:
                    data = json_data_reformatted.get(header.id)
                    for hd in subheaders:
                        li.append(nested_get(data, hd))
            li.insert(0, nested_get(row_data, 'user.username'))
            li.insert(1, nested_get(row_data, 'user.full_name'))
            ws.append(li)
        return ContentFile(save_virtual_workbook(wb))

    def get_frontend_redirect_url(self):
        if 'extra_fields' in self.request.query_params:
            return f'/admin/{self.organization.slug}/leave/reports/basic/yearly-leave'
        return f'/admin/{self.organization.slug}/leave/reports/basic/carry-forward'


