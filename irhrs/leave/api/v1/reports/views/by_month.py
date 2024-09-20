import openpyxl
from dateutil.parser import parse
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db.models import Q, Sum, Value, Prefetch
from django.db.models.functions import Coalesce
from django.http import Http404
from django.utils.functional import cached_property
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import filters, status
from rest_framework.exceptions import ValidationError
from rest_framework.fields import ReadOnlyField
from rest_framework.response import Response

from irhrs.export.utils.export import ExcelExport
from irhrs.leave.constants.model_constants import CREDIT_HOUR, TIME_OFF
from irhrs.core.constants.user import MALE, FEMALE, OTHER
from irhrs.core.mixins.serializers import DummySerializer
from irhrs.core.mixins.viewset_mixins import DateRangeParserMixin, \
    OrganizationMixin, ListRetrieveViewSetMixin, ListViewSetMixin, \
    ModeFilterQuerysetMixin, PastUserFilterMixin
from irhrs.core.utils import nested_get
from irhrs.core.utils.common import get_today
from irhrs.core.utils.filters import FilterMapBackend, \
    NullsAlwaysLastOrderingFilter
from irhrs.export.mixins.export import BackgroundExcelExportMixin
from irhrs.leave.api.v1.permissions import AdminOnlyLeaveReportPermission
from irhrs.leave.api.v1.reports.serializers.balance_report import \
    SummarizedYearlyLeaveReportSerializer
from irhrs.leave.models import LeaveRequest, LeaveAccount, LeaveType, \
    MasterSetting
from irhrs.leave.utils.mixins import LeaveRequestPastUserFilterMixin
from irhrs.permission.constants.permissions import LEAVE_PERMISSION, LEAVE_REPORT_PERMISSION

USER = get_user_model()


class LeaveByMonth(
    LeaveRequestPastUserFilterMixin,
    DateRangeParserMixin,
    OrganizationMixin,
    ListRetrieveViewSetMixin
):
    lookup_url_kwarg = 'month_key'
    lookup_field = 'pk'
    serializer_class = type(
        "SerializerHere",
        (DummySerializer,),
        {
            "leave_type": ReadOnlyField(
                source='leave_rule__leave_type__name'),
            "type_id": ReadOnlyField(
                source='leave_rule__leave_type'
            ),
            'leave_days': ReadOnlyField()
        }
    )
    permission_classes = [AdminOnlyLeaveReportPermission]
    filter_backends = (
        FilterMapBackend,
    )
    filter_map = {
        'branch': 'user__detail__branch__slug',
        'division': 'user__detail__division__slug',
        'employee_level': 'user__detail__employment_level__slug'
    }
    queryset = LeaveRequest.objects.all()

    def get_queryset(self):
        return super().get_queryset().filter(
            user__detail__organization=self.organization
        ).exclude(
            leave_rule__leave_type__category__in=(CREDIT_HOUR, TIME_OFF)
        )

    def list(self, request, *args, **kwargs):
        agg = {}
        month = 0
        for start, end in self.get_months():
            agg.update({
                str(month): Coalesce(Sum(
                    'balance',
                    filter=Q(
                        start__date__lt=end,
                        end__date__gte=start
                    )
                ), Value(0.0))
            })
            month += 1
        qs = self.filter_queryset(self.get_queryset())
        res = qs.aggregate(**agg)

        agg2 = {
            "male": Coalesce(
                Sum('balance', filter=Q(user__detail__gender=MALE)),
                Value(0.0)),
            "female": Coalesce(
                Sum('balance', filter=Q(user__detail__gender=FEMALE)),
                Value(0.0)),
            "other": Coalesce(
                Sum('balance', filter=Q(user__detail__gender=OTHER)),
                Value(0.0))
        }

        try:
            year = int(self.request.query_params.get('year', get_today().year))
            start = parse(f"{year}-01-01")
            end = parse(
                f"{year}-12-01"
                if month != 11 else
                f"{year + 1}-01-01"
            )
        except TypeError:
            return ValidationError({'year': ['Bad Year']})
        others = qs.filter(start__date__lt=end, end__date__gte=start).aggregate(**agg2)
        return Response({
            "results": res,
            **others
        })

    def get_months(self):
        try:
            year = int(self.request.query_params.get('year', get_today().year))
            # validate year
            parse(f"{year}-01-01")
        except (ValueError, TypeError):
            return ValidationError({'year': ['Bad Year']})

        for month in range(0, 12):
            start = f"{year}-{month + 1}-01"
            if month != 11:
                end = f"{year}-{month + 2}-01"
            else:
                end = f"{year+1}-01-01"
            yield start, end

    def get_month_queryset(self):
        month = int(self.kwargs.get('month_key'))
        if not (0 <= month <= 11):
            raise Http404

        try:
            year = int(self.request.query_params.get('year', get_today().year))
            start = parse(f"{year}-{month+1}-01")
            end = parse(
                f"{year}-{month+2}-01"
                if month != 11 else
                f"{year+1}-01-01"
            )
        except TypeError:
            return ValidationError({'year': ['Bad Year']})

        return self.filter_queryset(self.get_queryset()).filter(
            start__date__lt=end,
            end__date__gte=start
        )

    def retrieve(self, request, *args, **kwargs):
        queryset = self.get_month_queryset()
        annotated_queryset = queryset.order_by().values(
            'leave_rule__leave_type__name',
            'leave_rule__leave_type',
        ).annotate(
            leave_days=Coalesce(Sum('balance'), Value(0.0))
        )

        data = self.get_serializer(
            self.paginate_queryset(
                annotated_queryset
            ),
            many=True).data
        response = self.get_paginated_response(data)

        agg = {
            "male": Coalesce(
                Sum('balance', filter=Q(user__detail__gender=MALE)),
                Value(0.0)),
            "female": Coalesce(
                Sum('balance', filter=Q(user__detail__gender=FEMALE)),
                Value(0.0)),
            "other": Coalesce(
                Sum('balance', filter=Q(user__detail__gender=OTHER)),
                Value(0.0))
        }

        response.data.update(
            queryset.aggregate(**agg)
        )

        return response


class SummarizedYearlyLeaveReport(
    PastUserFilterMixin,
    BackgroundExcelExportMixin,
    DateRangeParserMixin,
    OrganizationMixin,
    ModeFilterQuerysetMixin,
    ListViewSetMixin
):
    """
    This page displays the Yearly Summarized Leave Report.

    Filter by division is available for user filtering.

    For data filtering, filter of year and leave_type is available.
    ### Select Leave types with ?leave_type=1,2,3,4,5,6,7
    #### for selecting a fiscal year, ?fiscal_type=fiscal&fiscal=<id>
    #### for selecting a year in gregorian, ?fiscal_type=gregorian&year=2019
    """
    queryset = USER.objects.all()
    permission_to_check = [LEAVE_PERMISSION, LEAVE_REPORT_PERMISSION]
    filter_backends = (
        filters.SearchFilter, NullsAlwaysLastOrderingFilter, FilterMapBackend
    )
    search_fields = ['username']
    filter_map = dict(
        division='detail__division__slug',
        user='id'
    )
    ordering_fields_map = dict(
        full_name=('first_name', 'middle_name', 'last_name'),
    )
    serializer_class = SummarizedYearlyLeaveReportSerializer
    export_fields = []
    export_type = 'Summarized Yearly Leave Report'
    notification_permissions = [LEAVE_REPORT_PERMISSION]

    @cached_property
    def leave_types(self):
        leave_from_query = self.request.query_params.get(
            'leave_type', ''
        )
        if not leave_from_query:
            return LeaveType.objects.none()
        leave_type_ids = [
            num for num in leave_from_query.split(',') if num.isdigit()
        ]
        start, end = self.fiscal_range
        return LeaveType.objects.filter(
            id__in=leave_type_ids,
            master_setting__in=MasterSetting.objects.filter(
                organization=self.organization,
            ).get_between(start, end)
        )

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset).filter(
            leave_accounts__rule__leave_type__in=self.leave_types.values_list(
                'id', flat=True
            )
        ).distinct().select_essentials()
        if self.action == 'export':
            return queryset
        return queryset.prefetch_related(
            Prefetch(
                'leave_accounts',
                queryset=LeaveAccount.objects.filter(
                    rule__leave_type__in=self.leave_types.values_list(
                        'id', flat=True
                    )
                ).select_related(
                    'rule',
                    'rule__renewal_rule',
                    'rule__leave_type',
                ),
                to_attr='prefetch_leave_accounts'
            )
        )

    def get_serializer_context(self):
        ctx = {}
        next_fiscal = self.fiscal_year.next if self.fiscal_year else None
        ctx.update({
            'selected_leave_types': self.leave_types.values_list(
                'id', flat=True
            ),
            'date_range': self.fiscal_range,
            'next_fiscal': next_fiscal,
            'organization': self.organization,
        })
        return ctx

    def get_extra_export_data(self):
        extra_data = super().get_extra_export_data()
        extra_data.update(self.get_serializer_context())
        return extra_data

    @classmethod
    def get_exported_file_content(cls, queryset, title, columns, extra_content, description=None, **kwargs):
        json_data = cls.serializer_class(
            queryset,
            many=True,
            context=extra_content
        ).data
        wb = openpyxl.Workbook()
        wb.active.title = cls.export_type
        ws = wb.active

        organization = extra_content.get('organization')
        lines_used = ExcelExport.insert_org_info(ws, organization)

        start_index, start_row = 3, lines_used + 2
        subheaders = [
            'carry_forward',
            'initial_balance',
            'total',
            'collapsed_balance',
            'used',
            'remaining_balance',
            'carry_forward_max_limit',
            'carry_forward_to_next_year',
            'encashment_balance',
        ]
        subheaders_mapper = {'initial_balance': 'balance_added', 'total': 'total_leave_balance'}

        headers = LeaveType.objects.filter(
            master_setting__organization=extra_content.get('organization'),
            id__in=extra_content.get('selected_leave_types'),
        )
        for header in headers:
            end_index = start_index + len(subheaders) - 1
            ws.merge_cells(
                f'{get_column_letter(start_index)}{start_row}:{get_column_letter(end_index)}{start_row}'
            )
            cell = ws[f'{get_column_letter(start_index)}{start_row}']
            cell.value = header.name.title()
            center_align = openpyxl.styles.Alignment(
                horizontal="center", vertical="center"
            )
            cell.alignment = center_align
            start_index = end_index + 1

        pretty_header = [' '.join(hdr.split('_')).title() for hdr in subheaders]
        subheaders_merged = pretty_header * len(headers)
        subheaders_merged.insert(0, 'Employee Name')
        subheaders_merged.insert(1, 'Username')
        ws.append(subheaders_merged)
        for each_row in json_data:
            json_data_reformatted = {
                li.get('id'): li for li in each_row.get('results')
            }
            li = list()
            li.append(nested_get(each_row, 'user.full_name'))
            li.append(nested_get(each_row, 'user.username'))
            for header in headers:
                data = json_data_reformatted.get(header.id)
                for hd in subheaders:
                    li.append(nested_get(data, f'balance_details.{subheaders_mapper.get(hd, hd)}'))
            ws.append(li)
        return ContentFile(save_virtual_workbook(wb))

    def get_frontend_redirect_url(self):
        return f'/admin/{self.organization.slug}/leave/reports/basic/summarized-yearly'
