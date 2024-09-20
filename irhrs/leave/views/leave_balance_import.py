from django.db import transaction
from django.db.models import Q
from django.forms import ValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.views import View

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.reader.excel import load_workbook

from irhrs.core.mixins.admin import AdminFormMixin, AdminMixin
from irhrs.leave.api.v1.serializers.account import LeaveAccountSerializer
from irhrs.leave.constants.model_constants import COMPENSATORY
from irhrs.leave.forms import LeaveBalanceImportForm
from irhrs.leave.models.account import LeaveAccount
from irhrs.organization.models.organization import Organization
from irhrs.leave.models import MasterSetting


def get_leave_types_for_organization(organization: Organization):
    master_setting = MasterSetting.objects.filter(
        organization=organization
    ).active().first()

    if not master_setting:
        raise ValidationError(
            "This organization does not have active master setting")

    return master_setting.leave_types.exclude(
        category=COMPENSATORY,
    ).exclude(
        is_archived=True
    )


class LeaveBalanceImportView(AdminFormMixin):
    form_class = LeaveBalanceImportForm
    template_name = 'leave_admin/leave_balance_import.html'

    success_url = '/dj-admin'

    headers = None
    organization = None

    def get_leave_data(self, ws: Worksheet):
        """convert worksheet rows  to input dictionary"""
        rows = ws.rows

        self.headers = [
            col.value for col in next(rows)
        ]

        if 'User' not in self.headers:
            raise ValidationError(
                'Invalid headers. `User` should be on first header.')

        return (
            dict(
                zip(
                    self.headers, [col.value for col in row]
                )
            )
            for row in rows
        )

    def prepare_data_for_serializer(self, data):
        """convert {'User': 'email or username', 'lt1': 1, 'lt2': 2}
        to [{'instance': LeaveAccount, data={'balance': 1, 'remark': 'Updated'}}, ...]
        """

        raw_data = list(data.values())
        user = data.pop('User')
        serializer_data = []
        valid = True

        for leave_type, balance in data.items():
            if balance is None:
                # skip if balance is None
                continue

            errors = []

            instance = LeaveAccount.objects.filter(
                Q(user__email=user) | Q(user__username=user,),
                rule__leave_type__name=leave_type,
                is_archived=False,
                user__detail__organization=self.organization,
            ).exclude(rule__leave_type__category=COMPENSATORY).first()

            if not instance:
                errors.append(
                    f'Leave {leave_type} not found for {user}. Possible'
                    ' problems can be "leave is not assigned", "email is wrong", '
                    '"leave type is compensatory", "user is inactive", '
                    '"leave type does not exist"'
                )

            if errors:
                raw_data.append("\n".join(errors))
                valid = False
            else:
                serializer_data.append({
                    'instance': instance,
                    'data': {
                        'balance': balance,
                        'remark': 'Leave Balance Import'
                    }
                })
        return serializer_data, raw_data, valid

    def form_valid(self, form):

        import_file = form.cleaned_data['import_file']
        self.organization = form.cleaned_data['organization']
        wb = load_workbook(import_file)
        ws = wb.active

        serializers = []

        return_rows = []
        valid = True

        try:
            for serializer_data, raw_data, is_valid in map(
                self.prepare_data_for_serializer,
                self.get_leave_data(ws)
            ):

                if is_valid:
                    for data in serializer_data:
                        serializer = LeaveAccountSerializer(**data)
                        if not serializer.is_valid():
                            raw_data.append(str(serializer.errors))
                        serializers.append(serializer)
                else:
                    valid = False

                return_rows.append(raw_data)

        except ValidationError as e:
            form.add_error('import_file', e.error_list)
            return self.form_invalid(form)

        if not valid:
            return self.get_invalid_response(return_rows)

        self.perform_create(serializers)

        return super().form_valid(form)

    def get_invalid_response(self, return_rows):
        """build invalid response from errors while creating"""
        workbook = Workbook()
        worksheet = workbook.active

        headers = self.headers + ['errors (remove this column)']
        worksheet.append(headers)

        for row in return_rows:
            worksheet.append(row)

        response = HttpResponse(
            content=save_virtual_workbook(workbook),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment;  filename=Leave Balance Import Error.xlsx'
        )
        return response

    @transaction.atomic()
    def perform_create(self, serializers):
        for serializer in serializers:
            serializer.save()


class LeaveBalanceSampleDownloadView(AdminMixin, View):

    def get_leave_types_for_organization(self, organization: Organization):

        return list(
            get_leave_types_for_organization(
                organization
            ).values_list('name', flat=True)
        )

    def get_organization(self):
        return get_object_or_404(
            Organization,
            id=self.kwargs['organization_id']
        )

    def get(self, request, *args, **kwargs):
        organization = self.get_organization()

        import_fields = ["User", ] + \
            self.get_leave_types_for_organization(organization)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(import_fields)

        response = HttpResponse(
            content=save_virtual_workbook(workbook),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment;  filename={organization.name}'
            '_leave_balance_import_sample.xlsx'
        )
        return response
