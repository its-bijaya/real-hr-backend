import re
from django.core.cache import cache
from io import BytesIO
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Prefetch, F, Value, Count, Q, Case, When
from django.db.models.functions import Concat

from django.http import FileResponse
from django.utils.functional import cached_property
from django_q.tasks import async_task
from openpyxl import Workbook, load_workbook
from rest_framework import viewsets, status, mixins, serializers
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.generics import get_object_or_404
from rest_framework.response import Response
from rest_framework.viewsets import ReadOnlyModelViewSet
from rest_framework.parsers import MultiPartParser

from irhrs.attendance.utils.payroll import get_salary_holdings_count,\
    get_employee_left_and_joined_summary
from irhrs.core.constants.organization import ACTION_ON_PAYROLL_APPROVAL_BY_APPROVAL_LEVELS, PAYROLL_CONFIRMATION_BY_HR
from irhrs.core.constants.payroll import PENDING
from irhrs.core.utils.common_utils import get_users_list_from_permissions
from irhrs.core.utils.email import send_email_as_per_settings
from irhrs.core.utils.excel import ExcelList
from irhrs.export.constants import COMPLETED
from irhrs.core.mixins.serializers import create_dummy_serializer, DummySerializer
from irhrs.core.mixins.viewset_mixins import (
    GetStatisticsMixin,
    ListCreateRetrieveDestroyViewSetMixin,
    OrganizationMixin,
    OrganizationCommonsMixin,
    ListViewSetMixin
)
from irhrs.core.utils import nested_getattr, nested_get
from irhrs.core.utils.common import validate_permissions
from irhrs.core.utils.filters import OrderingFilterMap, FilterMapBackend
from irhrs.export.constants import QUEUED, NORMAL_USER, ADMIN
from irhrs.export.mixins.export import BackgroundExcelExportMixin
from irhrs.export.models.export import Export
from irhrs.export.utils.export import PayrollExcelExport
from irhrs.export.utils.helpers import save_virtual_workbook
from irhrs.notification.utils import add_notification, notify_organization
from irhrs.organization.models import Organization, FiscalYear
from irhrs.payroll.api.v1.serializers import PayrollSerializer, \
    EmployeeThinSerializer as PayrollEmployeeSerializer, \
    EmployeePayrollSerializer, EmployeePayrollListSerializer, EmployeePayrollHistorySerializer
from irhrs.payroll.api.v1.serializers.payroll import (
    EmployeePayrollExportSerializer,
    ExcelPayrollPackageSerializer,
    ExcelUpdateSerializer,
    PayrollUpdateSerializer,
    PayrollGenerationHistorySerializer,
    PayrollCreateSerializer
)
from irhrs.payroll.api.v1.serializers.payroll_approval import PayrollApprovalHistorySerializer
from irhrs.payroll.api.v1.serializers.payroll_serializer import PayrollNoteSerializer, \
    SignedPayrollHistorySerializer, EmployeePayrollCommentSerializer
from irhrs.payroll.api.v1.serializers.report.overview import \
    OverViewReportDataSerializer
from irhrs.payroll.models import (
    GENERATED,
    APPROVED,
    CONFIRMED,
    REJECTED,
    APPROVAL_PENDING,
    GENERATED_PAYROLL_STATUS,
    PAYSLIP_ACKNOWLEDGEMENT_PENDING,
    PAYSLIP_ACKNOWLEDGED,
    PROCESSING,
    DONE,
    FAILED,
    Payroll,
    EmployeePayroll,
    Heading,
    ReportRowRecord,
    PayrollEditHistoryAmount,
    PayrollGenerationHistory,
    PayrollApproval,
    PayrollApprovalHistory,
    OrganizationPayrollConfig,
    PayrollApprovalSetting,
    SignedPayrollHistory,
    EmployeePayrollComment,
    PayrollExcelUpdateHistory, PayrollCollectionDetailReportSetting, ExtraHeadingReportSetting
)
from irhrs.payroll.models.payroll import PAYSLIP_GENERATED, ExcelPayrollPackage, \
    PayrollEmployeeAddedHistory, PACKAGE_DELETED
from irhrs.payroll.models.user_voluntary_rebate_requests import DELETED, UserVoluntaryRebate
from irhrs.payroll.tasks import create_package_activity
from irhrs.payroll.utils.calculator import EmployeeSalaryCalculator
from irhrs.payroll.utils.employee_payroll import get_employee_payrolls_via_settings, \
    get_extra_addition_and_deduction, get_filtered_employee_payrolls_from_query_params
from irhrs.payroll.utils.exceptions import CustomValidationError
from irhrs.payroll.utils.generate import PayrollGenerator, check_in_progress_payroll
from irhrs.payroll.utils.generate_payroll import (
    generate_payroll,
    payroll_excel_update
)
from irhrs.payroll.utils.helpers import ExtendedPageNumberPagination, create_payroll_edit_remarks
from irhrs.payroll.utils.user_detail_for_payroll import get_user_detail_for_payroll
from irhrs.payroll.utils.user_voluntary_rebate import update_rebate_settings_from_payroll_edit, \
    revert_fiscal_months_amount_to_zero_when_rebate_is_archived
from irhrs.permission.constants.permissions import GENERATE_PAYROLL_PERMISSION, \
    PAYROLL_REPORT_PERMISSION, HAS_PERMISSION_FROM_METHOD, \
    ASSIGN_PAYROLL_PACKAGES_PERMISSION
from irhrs.permission.constants.permissions.hrs_permissions import DELETE_CONFIRMED_EMPLOYEE_PAYROLL_PERMISSION
from irhrs.permission.permission_classes import permission_factory
from irhrs.users.api.v1.serializers.thin_serializers import UserThinSerializer

Employee = get_user_model()
ADD_VALUES_TO_BASIC = getattr(settings, 'ADD_VALUES_TO_BASIC', False)
BASIC_SALARY_NAME = getattr(settings, 'BASIC_SALARY_NAME', 'Basic Salary')
RemarksRequiredSerializer = create_dummy_serializer({
    'remarks': serializers.CharField(max_length=600, allow_blank=False)
})


def payroll_excel_update_callback(task):
    payroll = task.args[0]
    edit_log = task.kwargs['edit_log']
    if task.result == payroll:
        edit_log.status = DONE
        edit_log.save()
        notify_organization(
            f"Payroll of {payroll.organization} from {payroll.from_date} "
            f"{payroll.to_date} has been updated via excel file input.",
            organization=payroll.organization,
            action=edit_log,
            permissions=[GENERATE_PAYROLL_PERMISSION],
            url=f'/admin/{payroll.organization.slug}/payroll/collection/detail/{payroll.id}'
        )
    else:
        edit_log.status = FAILED
        edit_log.save()

        rgx = re.search(r'.+?(?=: Traceback)', task.result)

        error_message = rgx.group()

        notify_organization(
            f"Payroll update of {payroll.organization} from {payroll.from_date} "
            f"{payroll.to_date} via excel file input has been failed."
            f"Error: {error_message}",
            organization=payroll.organization,
            action=edit_log,
            permissions=[GENERATE_PAYROLL_PERMISSION],
            url=f'/admin/{payroll.organization.slug}/payroll/collection/detail/{payroll.id}'
        )


class PayrollAPIViewSet(viewsets.ModelViewSet, GetStatisticsMixin):
    pagination_class = ExtendedPageNumberPagination
    queryset = Payroll.objects.all().select_related(
        'approval_pending',
        'approval_pending__detail',
        'approval_pending__detail__organization',
        'approval_pending__detail__job_title',
    )
    serializer_class = PayrollSerializer
    permission_classes = [permission_factory.build_permission(
        "PayrollViewPermission",
        actions={
            'list': [GENERATE_PAYROLL_PERMISSION, PAYROLL_REPORT_PERMISSION],
            'create': [GENERATE_PAYROLL_PERMISSION],
            'retrieve': [GENERATE_PAYROLL_PERMISSION, PAYROLL_REPORT_PERMISSION,
                         HAS_PERMISSION_FROM_METHOD],
            'update': [GENERATE_PAYROLL_PERMISSION],
            'excel_update': [GENERATE_PAYROLL_PERMISSION],
            'get_excel_update_sample': [GENERATE_PAYROLL_PERMISSION],
            'partial_update': [GENERATE_PAYROLL_PERMISSION],
            'destroy': [GENERATE_PAYROLL_PERMISSION],
            'chart_report': [PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION],
            'dashboard_data': [PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION],
            'overview_report_data': [PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION],
            'cit_report': [PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION],
            'disbursement_report': [PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION],
            'export': [GENERATE_PAYROLL_PERMISSION, PAYROLL_REPORT_PERMISSION],
            'approval_history': [GENERATE_PAYROLL_PERMISSION, PAYROLL_REPORT_PERMISSION,
                                 HAS_PERMISSION_FROM_METHOD],
            'stat': [GENERATE_PAYROLL_PERMISSION],
        }
    )]
    filter_backends = (FilterMapBackend, OrderingFilter, SearchFilter)
    search_fields = ('title',)
    filter_map = {
        'organization__slug': 'organization__slug',
        'status': 'status',
        "from_date": "from_date__gte",
        "to_date": "to_date__lte",
    }
    ordering_fields = (
        'from_date',
        'to_date',
        'timestamp',
        'status'
    )
    export_type = "excel_update"
    statistics_field = "status"

    def __get_payroll(self):
        pk = self.kwargs.get('pk')
        return get_object_or_404(Payroll.objects.all(), pk=pk)

    def has_user_permission(self):
        if self.action in ['retrieve', 'approval_history']:
            # if user has pending payroll approval, then allow to view
            payroll = self.__get_payroll()
            return payroll.approval_pending == self.request.user
        return False

    def get_organization(self):
        return get_object_or_404(
            Organization.objects.all(),
            slug=self.request.query_params.get(
                'organization__slug'
            )
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.status != 'Approved':
            fiscal_year = FiscalYear.objects.active_for_date(
                instance.organization, instance.to_date
            )
            employee_payrolls = list(instance.employee_payrolls.all())
            self.perform_destroy(instance)
            for employee_payroll in employee_payrolls:
                user_rebate = UserVoluntaryRebate.objects.filter(
                    user=employee_payroll.employee,
                    fiscal_year=fiscal_year,
                    statuses__action=DELETED
                ).first()
                if user_rebate:
                    revert_fiscal_months_amount_to_zero_when_rebate_is_archived(user_rebate)

            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return Response(
                {'error_message': 'Cannot delete approved payroll'},
                status=status.HTTP_403_FORBIDDEN
            )

    def perform_destroy(self, instance):
        instance.delete()

    def get_paginated_response(self, data, extra_data):
        """
        Return a paginated style `Response` object for the given output data.
        """
        assert self.paginator is not None
        return self.paginator.get_paginated_response(data, extra_data)


    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        stats = self.statistics
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data, stats)

        serializer = self.get_serializer(queryset, many=True)
        serializer.data['extra_data'] = stats
        return Response(serializer.data)

    def get_serializer_class(self):
        if self.action == 'overview_report_data':
            return OverViewReportDataSerializer
        if self.action == "create":
            return PayrollCreateSerializer
        return super().get_serializer_class()

    def create(self, request, *args, **kwargs):

        try:
            serializer, data = PayrollGenerator.sync_validate(request.data)
        except CustomValidationError as e:
            return Response(
                e.error_dict,
                status.HTTP_400_BAD_REQUEST
            )

        if check_in_progress_payroll(serializer.validated_data['organization_slug']):
            raise ValidationError('Previous payroll generation is in progress please wait'
                                  ' until that is completed.')

        payroll_log = PayrollGenerationHistory(
            # store initial data to regenerate payroll using same params
            data=serializer.initial_data,
            # organization instance
            organization=serializer.validated_data['organization_slug'],
            status=QUEUED,
            errors={}
        )
        payroll_log.save()

        # quick hack, send datework=None and reinitialize in async task start
        # as datework can not be pickled
        data = list(data)
        data[1] = None
        async_task(generate_payroll, payroll_log, *data, timeout=10000)

        return Response({
            "message": "Payroll has been sent for background processing"
        })

    @action(detail=True, methods=['POST'], url_path="add-employee")
    def add_employees(self, request, *args, **kwargs):
        payroll = self.get_object()
        payroll_data = {**request.data, **PayrollSerializer(instance=payroll).data}
        payroll_data['cutoff_date'] = str(payroll.simulated_from) if payroll.simulated_from else None
        if payroll.status in [COMPLETED, REJECTED]:
            raise ValidationError({'error': f"Can't add employee in {payroll.status} Payroll."})
        try:
            serializer, data = PayrollGenerator.sync_validate(payroll_data)
        except CustomValidationError as e:
            return Response(
                e.error_dict,
                status.HTTP_400_BAD_REQUEST
            )

        if check_in_progress_payroll(serializer.validated_data['organization_slug']):
            raise ValidationError('Previous payroll generation is in progress please wait'
                                  ' until that is completed.')

        payroll_log = getattr(payroll, 'generation', None)
        if not payroll_log:
            payroll_log = PayrollGenerationHistory(
                # store initial data to regenerate payroll using same params
                data=serializer.initial_data,
                # organization instance
                organization=serializer.validated_data['organization_slug'],
                status=QUEUED,
                errors={}
            )
        payroll_log.status = PROCESSING
        payroll_log.payroll = payroll
        payroll_log.save()
        # quick hack, send datework=None and reinitialize in async task start
        # as datework can not be pickled
        data = list(data)
        data[1] = None
        async_task(generate_payroll, payroll_log, *data, payroll)
        PayrollEmployeeAddedHistory.objects.create(payroll=payroll, data=serializer.initial_data, errors={})
        return Response({"message": "Employee adding process has been sent for background processing."})

    def retrieve(self, request, pk=None, **kwargs):
        # TODO: update this
        payroll = self.get_object()
        res = self.calculate_employee_payrolls(payroll)
        response = self.get_paginated_response(
            res,
            PayrollSerializer(
                instance=payroll,
                many=False,
                context=self.get_serializer_context()
            ).data
        )
        response.data.update({
            'payslip_template': OrganizationPayrollConfig.get_payslip_template(
                self.get_organization()
            )
        })
        return response

    @action(methods=['GET'], detail=True, url_path='approval-history',
            serializer_class=PayrollApprovalHistorySerializer)
    def approval_history(self, request, pk=None):
        payroll = self.get_object()
        qs = payroll.payroll_approval_histories.all().order_by('created_at')
        page = self.paginate_queryset(qs)
        data = PayrollApprovalHistorySerializer(page, many=True).data

        response = self.get_paginated_response(data, {})
        export = payroll.excel_updates.filter(status=FAILED).first()
        if export:
            response.data["excel_error"] = {
                "error_file": settings.BACKEND_URL + export.excel_file.url,
                "created_on": export.created_at,
                "actor": UserThinSerializer(export.created_by).data,
                "remarks": 'Payroll excel update failed.'
            }
        return response

    @action(methods=['GET'], detail=True)
    def cit_report(self, request, pk=None):
        res = []
        payroll = self.get_object()
        payroll_employees = payroll.employee_payrolls.all()
        # employees = employees.exclude(**payroll.employee_exclude)

        for payroll_employee in payroll_employees:
            employee_ser = PayrollEmployeeSerializer(
                instance=payroll_employee.employee, many=False)
            heading_amounts = dict()

            package = payroll_employee.package
            package_pf_headings = package.package_headings.filter(
                heading__payroll_setting_type__in=[
                    'Self CIT Office Addition', 'Self CIT'
                ]
            ).exclude(heading__type__in=['Type1Cnst', 'Type2Cnst'])

            for package_pf_heading in package_pf_headings:
                heading_amounts[
                    package_pf_heading.heading.id] = payroll.get_heading_amount(
                    payroll_employee.employee,
                    package_pf_heading
                )
            employee_data = employee_ser.data
            employee_data['designation_name'] = str(
                payroll_employee.user_experience_package_slot.user_experience.job_title
            )
            res.append({
                'employee': employee_data,
                'heading_amounts': heading_amounts
            })
        return Response({
            'results': res,
            'payroll': PayrollSerializer(instance=payroll, many=False).data
        }, status=status.HTTP_200_OK)

    @action(methods=['GET'], detail=True)
    def disbursement_report(self, request, pk=None):
        res = []
        payroll = self.get_object()
        payroll_employees = payroll.employee_payrolls.all()
        # employees = employees.exclude(**payroll.employee_exclude)

        for payroll_employee in payroll_employees:
            employee_ser = PayrollEmployeeSerializer(
                instance=payroll_employee.employee, many=False)
            heading_amounts = dict()

            package = payroll_employee.package
            package_pf_headings = package.package_headings.filter(
                heading__payroll_setting_type__in=[
                    'Salary TDS'
                ]
            )

            for package_pf_heading in package_pf_headings:
                heading_amounts[
                    package_pf_heading.heading.id] = payroll.get_heading_amount(
                    payroll_employee.employee,
                    package_pf_heading
                )

            cash_in_hand = payroll.get_cash_in_hand(payroll_employee.employee)
            employee_data = employee_ser.data
            employee_data['designation_name'] = str(
                payroll_employee.user_experience_package_slot.user_experience.job_title
            )
            res.append({
                'employee': employee_data,
                'cash_in_hand': cash_in_hand,
                'heading_amounts': heading_amounts
            })
        return Response({
            'results': res,
            'payroll': PayrollSerializer(
                instance=payroll,
                many=False
            ).data
        }, status=status.HTTP_200_OK)

    @action(methods=['POST'], detail=False, url_path='excel-update-sample')
    def get_excel_update_sample(self, request):
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Payroll update data sample.'

        required_headings = request.data.get('headings', [])

        heading_names = Heading.objects.filter(
            organization=self.get_organization(),
            id__in=required_headings
        ).values_list(
            'name',
            flat=True
        )

        write_data = [
            ['User'] + list(heading_names),
            [
                'employee_example_email@gmail.com'
            ] + ['N/A'] * len(heading_names)

        ]

        for data in write_data:
            ws.append(data)

        wb.save(buffer)

        buffer.seek(0)

        return FileResponse(buffer, as_attachment=True, filename='payroll_update_excel_sample.xlsx')

    @action(
        methods=['PUT'],
        detail=True,
        parser_classes=[MultiPartParser],
        serializer_class=ExcelUpdateSerializer,
        url_path='excel-update'
    )
    def excel_update(self, request, pk=None):

        sync = True if self.request.query_params.get(
            'sync') in ['true', '1'] else False

        heading_name_to_id_mapping = dict(
            Heading.objects.filter(
                organization=self.get_organization()
            ).values_list('name', 'id')
        )

        payroll = self.get_object()

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        excel_update_file = serializer.validated_data['file']
        remarks = serializer.validated_data['remarks']

        wb = load_workbook(excel_update_file)
        ws = wb.active

        sheet_values_rows = ws.values

        first_row = next(sheet_values_rows)
        if not first_row[0]:
            raise ValidationError({"excel": "Heading cannot be empty."})

        all_headings = Heading.objects.filter(organization=payroll.organization).values_list('name', flat=True)
        headings_length = len(all_headings)
        if headings_length < len(first_row[1:]):
            raise ValidationError({'Heading length': f'Heading limit exceeded max count {headings_length}.'})

        for heading in first_row[1:]:
            if not heading:
                raise ValidationError({"Error":"Heading field cannot be null."})
            if heading not in all_headings:
                raise ValidationError({f"{heading}":"Is incorrect heading."})

        employees_edited_headings = dict()
        employees_email = list()
        row_records = dict()
        for row in ReportRowRecord.objects.filter(
            employee_payroll__payroll=payroll,
            heading__name__in=first_row[1:len(first_row)]
        ).values(
            'employee_payroll__employee__email', 'heading', 'amount'
        ):
            emp_email = row.get('employee_payroll__employee__email')
            heading = str(row.get('heading'))

            amount = row.get('amount')
            heading_type = Heading.objects.get(id=heading).type
            heading_values = {
                heading: {
                    "type": heading_type,
                    "amount": amount
                }
            }
            row_records[emp_email] = heading_values

        # iterating to values rows
        excel_list = ExcelList(wb)
        header = excel_list[0]
        excel_data = excel_list[1:]
        error_exists = False
        for index, row in enumerate(excel_data, 1):
            email = str(row[0]).strip()

            errors = {}
            if email in employees_email:
                errors["email"] = "Duplicate email"

            user = payroll.employees.filter(
                Q(email=email) | Q(username=email)
            ).first()

            if not user:
                errors["email"] = "User with this email/username doesn't exist for this payroll"

            else:
                employees_email.append(user.email)
                employees_edited_headings[user.email] = dict()

            for heading_name, value in zip(header[1:], row[1:]):
                if type(value) in (int, float):
                    heading_id = str(heading_name_to_id_mapping[heading_name])
                    if not user:
                        continue
                    employees_edited_headings[
                        user.email
                    ][
                        heading_id
                    ] = dict(
                        initialValue=nested_get(
                            row_records, f'{email},{heading_id},amount', separator=','
                        ),
                        currentValue=value,
                        type=nested_get(
                            row_records, f'{email},{heading_id},type', separator=','
                        )
                    )
                else:
                    if value is None:
                        continue
                    errors[heading_name] = f"Invalid number, {value}"

            if errors:
                error_exists = True
                excel_list[index].append(str(errors))

        if error_exists:
            excel_list[0].append("Errors")
            error_wb = excel_list.generate_workbook()
            excel_history = PayrollExcelUpdateHistory()
            excel_history.payroll = payroll
            excel_history.status = FAILED
            excel_history.excel_file.save(
                "excel_payroll_update.xlsx",
                ContentFile(save_virtual_workbook(error_wb))
            )
            excel_history.save()
            raise ValidationError(
                "Error in excel file, click on 'Show Remarks' button"
            )

        payroll = self.get_object()

        if payroll.status not in [GENERATED, REJECTED]:
            return Response(
                dict(
                    non_field_errors='Only generated and rejected payroll can be edited'
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        edit_log = PayrollExcelUpdateHistory.objects.create(
            payroll=payroll,
            excel_file=excel_update_file,
            status=PROCESSING
        )

        if sync:
            ''' This block is only for testing purpose
            as there is something wrong with async_task sync=True
            option.
            https://github.com/Koed00/django-q/issues/421

            In this workaround hook or callback is not tested
            '''

            payroll_excel_update(
                payroll,
                edited_general_headings=employees_edited_headings,
                edit_log=edit_log,
                remarks=remarks
            )
        else:
            async_task(
                payroll_excel_update,
                payroll,
                hook=payroll_excel_update_callback,
                edited_general_headings=employees_edited_headings,
                edit_log=edit_log,
                remarks=remarks,
                employees_email=employees_email,
            )
            payroll.excel_updates.filter(status=FAILED).delete()
        notify_organization(
            f"Payroll of {payroll.organization} from {payroll.from_date} "
            f"{payroll.to_date} is being updated via excel file input.",
            organization=payroll.organization,
            action=edit_log,
            permissions=[GENERATE_PAYROLL_PERMISSION],
            url=f'/admin/{payroll.organization.slug}/payroll/collection/detail/{payroll.id}'
        )

        return Response(status=status.HTTP_200_OK)

    # @action(methods=['POST'], detail=False)
    # def chart_report(self, request):
    #     post_data = request.data
    #
    #     error_messages = dict()
    #     payroll_filter = post_data.get('payroll_filter', {})
    #     payroll_filter['status'] = 'Approved'
    #
    #     from_date = payroll_filter.get('from_date')
    #     to_date = payroll_filter.get('to_date')
    #
    #     if from_date:
    #         from_date = datetime.datetime.strptime(
    #             from_date, "%Y-%m-%d").date()
    #     else:
    #         error_messages['from_date'] = 'This field is required'
    #     if to_date:
    #         to_date = datetime.datetime.strptime(to_date, "%Y-%m-%d").date()
    #     else:
    #         error_messages['to_date'] = 'This field is required'
    #
    #     if (from_date and to_date) and (from_date > to_date):
    #         error_messages['to_date'] = 'Must be less than from date'
    #
    #     if error_messages.keys():
    #         return Response(error_messages, status=status.HTTP_400_BAD_REQUEST)
    #
    #     payrolls = Payroll.objects.filter(**payroll_filter)
    #
    #     res = []
    #     for payroll in payrolls:
    #         payroll_employees = payroll.employee_payrolls.all()
    #         heading_amounts = dict()
    #         # ctc = 0
    #         for payroll_employee in payroll_employees:
    #             # employee_ser = PayrollEmployeeSerializer(instance=employee, many=False)
    #             package = payroll_employee.package
    #             package_headings = package.package_headings.all()
    #             for package_heading in package_headings:
    #                 if heading_amounts.get(package_heading.heading.id):
    #                     heading_amounts[
    #                         package_heading.heading.id] += payroll.get_heading_amount(
    #                         payroll_employee.employee,
    #                         package_heading
    #                     )
    #                 else:
    #                     heading_amounts[
    #                         package_heading.heading.id] = payroll.get_heading_amount(
    #                         payroll_employee.employee,
    #                         package_heading
    #                     )
    #
    #             # ctc += payroll.get_ctc_amount(
    #             #     payroll_employee.employee
    #             # )
    #         res.append({
    #             'payroll': PayrollSerializer(
    #                 instance=payroll,
    #                 many=False
    #             ).data,
    #             'heading_amounts': heading_amounts,
    #             # 'ctc': ctc
    #         })
    #     return Response(res, status=status.HTTP_200_OK)

    # @action(methods=['POST'], detail=False)
    # def dashboard_data(self, request):
    #     post_data = request.data
    #
    #     error_messages = dict()
    #     payroll_filter = post_data.get('payroll_filter', {})
    #     payroll_filter['status'] = 'Approved'
    #     employee_filter = post_data.get('employee_filter', {})
    #
    #     from_date = payroll_filter.get('from_date__gte')
    #     to_date = payroll_filter.get('to_date__lte')
    #
    #     if from_date:
    #         from_date = datetime.datetime.strptime(
    #             from_date, "%Y-%m-%d").date()
    #     else:
    #         error_messages['from_date__gte'] = 'This field is required'
    #     if to_date:
    #         to_date = datetime.datetime.strptime(to_date, "%Y-%m-%d").date()
    #     else:
    #         error_messages['to_date__lte'] = 'This field is required'
    #
    #     if (from_date and to_date) and (from_date > to_date):
    #         error_messages['to_date__lte'] = 'Must be less than from date'
    #
    #     if error_messages.keys():
    #         return Response(error_messages, status=status.HTTP_400_BAD_REQUEST)
    #
    #     payrolls = Payroll.objects.filter(**payroll_filter)
    #
    #     heading_amounts = dict()
    #     ctc = 0
    #     cash_in_hand = 0
    #     salary_range_breakdown_data = {
    #         '10000 - 20000': 10,
    #         '20001 - 30000': 30,
    #         '40001 - 50000': 40,
    #         '50001 - 60000': 50,
    #         '60001 - 70000': 70,
    #         '70001 - 80000': 25,
    #     }
    #     for payroll in payrolls:
    #         payroll_employees = payroll.employee_payrolls.all()
    #         heading_amounts = dict()
    #         total_payroll = 0
    #         for payroll_employee in payroll_employees:
    #             # employee_ser = PayrollEmployeeSerializer(instance=employee, many=False)
    #             package = payroll_employee.package
    #             package_headings = package.package_headings.all()
    #
    #             total_basic_salary = 0
    #             for package_heading in package_headings:
    #                 if package_heading.heading.payroll_setting_type == 'Salary Structure':
    #                     total_basic_salary += payroll.get_heading_amount(
    #                         payroll_employee.employee,
    #                         package_heading
    #                     )
    #                 if heading_amounts.get(package_heading.heading.id):
    #                     heading_amounts[
    #                         package_heading.heading.id] += payroll.get_heading_amount(
    #                         payroll_employee.employee,
    #                         package_heading
    #                     )
    #                 else:
    #                     heading_amounts[
    #                         package_heading.heading.id] = payroll.get_heading_amount(
    #                         payroll_employee.employee,
    #                         package_heading
    #                     )
    #
    #             if 10000 <= total_basic_salary <= 20000:
    #                 salary_range_breakdown_data['10000 - 20000'] += 1
    #             elif 20001 <= total_basic_salary <= 30000:
    #                 salary_range_breakdown_data['20001 - 30000'] += 1
    #             elif 30001 <= total_basic_salary <= 40000:
    #                 salary_range_breakdown_data['30001 - 40000'] += 1
    #             elif 40001 <= total_basic_salary <= 50000:
    #                 salary_range_breakdown_data['40001 - 50000'] += 1
    #             elif 50001 <= total_basic_salary <= 60000:
    #                 salary_range_breakdown_data['50001 - 60000'] += 1
    #             elif 60001 <= total_basic_salary <= 70000:
    #                 salary_range_breakdown_data['60001 - 70000'] += 1
    #             elif 70001 <= total_basic_salary <= 80000:
    #                 salary_range_breakdown_data['70001 - 80000'] += 1
    #
    #             ctc += payroll.get_ctc_amount(payroll_employee.employee)
    #             cash_in_hand += payroll.get_cash_in_hand(
    #                 payroll_employee.employee)
    #     res = {
    #         'heading_amounts': heading_amounts,
    #         'cost_to_company': ctc,
    #         'cash_in_hand': cash_in_hand,
    #         'employees_detail': {
    #             'total_employee': Employee.objects.filter(
    #                 **employee_filter).count(),
    #             'male_count': Employee.objects.filter(
    #                 **employee_filter).filter(
    #                 detail__gender='Male'
    #             ).count(),
    #             'female_count': Employee.objects.filter(
    #                 **employee_filter).filter(
    #                 detail__gender='Female'
    #             ).count(),
    #         },
    #         'salary_range_breakdown_data': salary_range_breakdown_data
    #
    #     }
    #     return Response(res, status=status.HTTP_200_OK)

    @ action(detail=False, methods=['post'])
    def overview_report_data(self, request):
        serializer = OverViewReportDataSerializer(data=request.data)
        if serializer.is_valid():
            return Response(
                serializer.get_data(),
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(methods=['GET'], detail=True, url_path='edit',
            serializer_class=PayrollUpdateSerializer)
    def edit_payroll(self, request, *args, **kwargs):
        serializer_class = self.get_serializer_class()
        payroll_batch = self.get_object()

        if payroll_batch.status not in [GENERATED, REJECTED]:
            raise self.permission_denied(
                request,
                "Can only update payrolls in Generated and Rejected"
            )

        serializer = serializer_class(
            instance=payroll_batch,
            context=self.get_serializer_context()
        )
        return Response({
            'data': serializer.data,
            'employee_payroll': self.calculate_employee_payrolls(payroll_batch)
        })

    @transaction.atomic()
    @edit_payroll.mapping.post
    def edit_payroll_post(self, request, *args, **kwargs):
        payroll_batch = self.get_object()
        if payroll_batch.status not in  [GENERATED, REJECTED]:
            raise ValidationError({
                'message': 'Only Generated and Rejected payrolls can be edited.'
            })

        serializer = PayrollUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        _PAYROLL_MONTH_DAYS_SETTING = 'ORGANIZATION_CALENDAR'

        # New Extra Data
        extra_headings = serializer.validated_data.get('extra_headings') or {}
        edited_general_headings = serializer.validated_data.get(
            'edited_general_headings') or {}

        # Old Extra Data
        old_extra_headings = payroll_batch.extra_data.get('extra_headings', {})
        old_edited_general_headings = payroll_batch.extra_data.get(
            'edited_general_headings', {})

        remarks = serializer.validated_data.get('remarks') or {}
        employee_payrolls = payroll_batch.employee_payrolls.filter(
            employee_id__in=remarks.keys()
        )
        for payroll_employee in employee_payrolls:
            employee = payroll_employee.employee
            from_date = payroll_batch.from_date
            to_date = payroll_batch.to_date
            date_work, payroll_config = PayrollGenerator.get_FY(
                payroll_batch.organization.slug,
                from_date, to_date
            )

            employee_extra_headings = old_extra_headings.get(
                str(employee.id), {})
            if not employee_extra_headings:
                employee_extra_headings = get_extra_addition_and_deduction(payroll_employee)
            employee_extra_headings.update(
                extra_headings.get(str(employee.id), {})
            )
            employee_edited_general_headings = old_edited_general_headings.get(
                str(employee.id), {}
            )
            previous_edited_general_headings = payroll_employee.payroll.extra_data.get(
                'edited_general_headings'
            ) or dict()
            employee_edited_general_headings.update(
                previous_edited_general_headings.get(employee.id, {})
            )
            employee_edited_general_headings.update(
                edited_general_headings.get(str(employee.id), {})
            )
            new_income_difference = 0
            for k, v in employee_edited_general_headings.items():
                if not k.isdigit():
                    continue
                try:
                    new_income_difference += float(
                        v.get('currentValue')
                    ) - float(
                        v.get('initialValue')
                    )
                except TypeError:
                    continue

            employee_edited_general_headings['incomeDifference'] = new_income_difference

            appoint_date, dismiss_date, salary_packages = \
                PayrollGenerator.get_user_experience_calculation_data(
                    employee,
                    from_date,
                    to_date,
                    payroll_config
                )

            update_rebate_settings_from_payroll_edit(
                employee, employee_edited_general_headings, from_date, to_date,
                payroll_batch.organization, payroll_employee.package
            )

            calculation = EmployeeSalaryCalculator(
                employee,
                date_work,
                from_date,
                to_date,
                salary_packages,
                appoint_date,
                dismiss_date,
                month_days_setting=_PAYROLL_MONTH_DAYS_SETTING,
                extra_headings=employee_extra_headings,
                edited_general_headings=employee_edited_general_headings,
                edited_general_headings_difference_except_tax_heading=new_income_difference,
                simulated_from=payroll_batch.simulated_from,
                package_assigned_date=salary_packages[0]["applicable_from"],
                employee_payroll=payroll_employee
            )

            # Update Employee Payroll Instance
            payroll_employee.annual_gross_salary = calculation.payroll.annual_gross_salary
            payroll_employee.rebate_amount = calculation.payroll.rebate_amount
            payroll_employee.annual_gross_salary_after_rebate = calculation.payroll. \
                annual_gross_salary_after_rebate
            payroll_employee.annual_tax = calculation.payroll.annual_tax
            payroll_employee.paid_tax = calculation.payroll.paid_tax
            payroll_employee.tax_to_be_paid = calculation.payroll.tax_to_be_paid
            payroll_employee.tax_rule = calculation.payroll.tax_rule
            payroll_employee.tax_condition = calculation.payroll.tax_condition
            payroll_employee.tds_type = calculation.payroll.tds_type

            payroll_employee.save()

            edits = dict()
            for old_record in ReportRowRecord.objects.filter(
                employee_payroll=payroll_employee,
            ):
                # if package heading was not included during generation,
                # exclude during edit too, eg. Advance Salary Deduction
                if old_record.heading:
                    amount = calculation.payroll.get_heading_amount_from_heading(
                        old_record.heading
                    )
                    old_amount = getattr(old_record, 'amount', None)
                    edits[old_record.heading.id] = (old_amount, amount)

                    old_record.amount = amount
                    old_record.save(update_fields=['amount'])

            if calculation.payroll.backdated_calculations:
                calculation.payroll.backdated_calculations.update(
                    adjusted_payroll=payroll_employee
                )

            create_payroll_edit_remarks(
                employee_payroll=payroll_employee,
                edited_packages=edits,
                remarks=remarks.get(
                    str(employee.id)
                )
            )
            old_extra_headings[str(employee.id)] = employee_extra_headings
            old_edited_general_headings[str(
                employee.id)] = employee_edited_general_headings
            payroll_batch.extra_data = {
                'extra_headings': old_extra_headings,
                'edited_general_headings': old_edited_general_headings,
            }
            payroll_batch.save()
        return Response({
            'message': 'success'
        })

    def calculate_employee_payrolls(self, payroll):
        payroll_employees = get_filtered_employee_payrolls_from_query_params(
            payroll, self.request.query_params)
        payroll_employees = self.paginate_queryset(payroll_employees)
        res = list()
        for payroll_employee in payroll_employees:
            employee_ser = PayrollEmployeeSerializer(
                instance=payroll_employee.employee, many=False)
            heading_amounts = dict()

            fil = {}
            heading_setting = PayrollCollectionDetailReportSetting.objects.filter(
                organization=self.get_organization()).first()
            if heading_setting:
                fil['heading_id__in'] = heading_setting.headings.values_list('id', flat=True)

            for row in ReportRowRecord.objects.filter(
                    employee_payroll=payroll_employee, **fil):
                if row.heading_id:
                    heading_amounts[row.heading_id] = row.amount

            employee_data = employee_ser.data

            user_detail_response, slots = get_user_detail_for_payroll(
                payroll_employee, self.get_organization())

            # Designation_name is 'job_title' but for the experience, for
            #  which payroll was generated & not current.
            employee_data['designation_name'] = slots[-1]["job_title"]

            employee_data['package'] = nested_getattr(
                slots[-1].get('package'), 'id')
            employee_data['employee_payroll_id'] = payroll_employee.id
            employee_data['employment_status'] = slots[-1]["employment_status"]
            employee_data['step'] = slots[-1]["current_step"]

            res.append({
                'employee': employee_data,
                'heading_amounts': heading_amounts,
                'user_note': payroll_employee.user_note,
                'step': slots[-1]["current_step"],
                **user_detail_response
            })

        return res

    @action(
        methods=['PATCH'],
        detail=True,
        url_path=r'user/(?P<user_id>\d+)/note',
        serializer_class=PayrollNoteSerializer
    )
    def user_note(self, *args, **kwargs):
        payroll = self.get_object()
        instance = payroll.employee_payrolls.filter(
            employee_id=self.kwargs.get('user_id')
        ).first()
        if instance:
            serializer = self.get_serializer(
                instance=instance, data=self.request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)
        return Response({'detail': 'Employee Payroll not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    @action(
        methods=['POST'],
        detail=True,
        url_path='global/note',
        serializer_class=PayrollNoteSerializer
    )
    def global_note(self, *args, **kwargs):
        serializer = self.get_serializer(data=self.request.data)
        serializer.is_valid(raise_exception=True)
        note = serializer.data.get('user_note')

        payroll = self.get_object()
        employee_payrolls = payroll.employee_payrolls.all()
        if self.request.query_params.get('for_all', 'false') != 'true':
            employee_payrolls = employee_payrolls.filter(
                user_note__isnull=True)
        employee_payrolls.update(user_note=note)
        return Response(serializer.data)

    @action(
        methods=['GET'],
        detail=False,
        serializer_class=DummySerializer,
        url_path='stat'
    )
    def stat(self, *args, **kwargs):
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')
        return Response({
            'holding_detail': get_salary_holdings_count(
                self.get_organization().slug,
                from_date,
                to_date,
            ),
            'employee_left_and_joined': get_employee_left_and_joined_summary(
                self.get_organization().slug,
                from_date,
                to_date,
            ),
        })


class PayrollApprovalViewSet(ListViewSetMixin):
    """
    Approval List Payroll

    to approve,
    /<pk>/approve

    to reject
    /<pk>/reject
    """
    queryset = Payroll.objects.all()
    serializer_class = PayrollSerializer

    def get_queryset(self):
        return super().get_queryset().filter(
            approval_pending=self.request.user,
            status=APPROVAL_PENDING
        )

    def get_approval(self, payroll):
        return get_object_or_404(
            PayrollApproval.objects.all(),
            user=self.request.user,
            status=PENDING,
            payroll=payroll
        )

    @staticmethod
    def get_next_approval(payroll, current_approval_level):
        return PayrollApproval.objects.filter(
            status=PENDING,
            payroll=payroll,
            approval_level__gt=current_approval_level
        ).order_by('approval_level').first()

    @transaction.atomic()
    @action(detail=True, methods=['POST'], serializer_class=RemarksRequiredSerializer)
    def approve(self, request, *args, **kwargs):
        payroll = self.get_object()
        approval = self.get_approval(payroll)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        approval.status = APPROVED
        approval.save()

        PayrollApprovalHistory.objects.create(
            actor=request.user,
            payroll=payroll,
            action='approved payroll',
            remarks=serializer.validated_data.get('remarks')
        )
        next_approval = self.get_next_approval(
            payroll, approval.approval_level)
        if next_approval:
            payroll.approval_pending = next_approval.user
            payroll.save()
            add_notification(
                text=f"{self.request.user} forwarded a payroll to approve.",
                actor=self.request.user,
                action=payroll,
                url='/user/payroll/approval/request',
                recipient=payroll.approval_pending
            )

            subject = "Request for Payroll Approval."
            email_text = f"{self.request.user} forwarded a payroll from " \
                        f" {payroll.from_date} to {payroll.to_date} for approval."
            send_email_as_per_settings(
                recipients=next_approval.user,
                subject=subject,
                email_text=email_text,
                email_type=ACTION_ON_PAYROLL_APPROVAL_BY_APPROVAL_LEVELS
            )
        else:
            payroll.status = APPROVED
            payroll.approval_pending = None
            payroll.save()
            notify_organization(
                text="Payroll has been approved and awaits confirmation.",
                organization=payroll.organization,
                action=payroll,
                url=f'/admin/{payroll.organization.slug}/payroll/collection/detail/{payroll.pk}',
                permissions=[GENERATE_PAYROLL_PERMISSION, ]
            )
            hrs = get_users_list_from_permissions(
            permission_list=[GENERATE_PAYROLL_PERMISSION],
            organization=payroll.organization
            )
            email_text = f"{self.request.user} has approved payroll from " \
                         f"{payroll.from_date} to {payroll.to_date} and awaits confirmation."
            send_email_as_per_settings(
                recipients=hrs,
                subject="Payroll approved and awaits confirmation.",
                email_text=email_text,
                email_type=PAYROLL_CONFIRMATION_BY_HR
            )

        return Response({'message': 'Successfully Approved'})

    @transaction.atomic()
    @action(detail=True, methods=['POST'], serializer_class=RemarksRequiredSerializer)
    def reject(self, request, *args, **kwargs):
        payroll = self.get_object()
        approval = self.get_approval(payroll)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        approval.status = REJECTED
        approval.save()

        PayrollApprovalHistory.objects.create(
            actor=request.user,
            payroll=payroll,
            action='rejected payroll',
            remarks=serializer.validated_data.get('remarks')
        )

        payroll.status = REJECTED
        payroll.save()

        notify_organization(
            text=f"{self.request.user.full_name} has rejected payroll.",
            organization=payroll.organization,
            action=payroll,
            url=f'/admin/{payroll.organization.slug}/payroll/collection/detail/{payroll.pk}',
            permissions=[GENERATE_PAYROLL_PERMISSION, ]
        )

        hrs = get_users_list_from_permissions(
            permission_list=[GENERATE_PAYROLL_PERMISSION],
            organization=payroll.organization
        )
        send_email_as_per_settings(
            recipients=hrs,
            subject="Payroll Rejected",
            email_text=f"{self.request.user} has rejected Payroll from" \
                    f" {payroll.from_date} to {payroll.to_date}.",
            email_type=ACTION_ON_PAYROLL_APPROVAL_BY_APPROVAL_LEVELS,
        )

        return Response({'message': 'Rejected payroll'})


class SignedPayrollHistoryAPIViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    viewsets.GenericViewSet
):
    queryset = SignedPayrollHistory.objects.all()
    serializer_class = SignedPayrollHistorySerializer
    permission_classes = [permission_factory.build_permission(
        "SignedPayrollHistoryPermission",
        allowed_to=[GENERATE_PAYROLL_PERMISSION, PAYROLL_REPORT_PERMISSION],
        limit_write_to=[GENERATE_PAYROLL_PERMISSION]
    )]

    @cached_property
    def payroll(self):
        return get_object_or_404(
            Payroll,
            id=self.kwargs.get('payroll_id'),
            organization=self.get_organization()
        )

    def get_queryset(self):
        return super().get_queryset().filter(payroll=self.payroll)

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['payroll'] = self.payroll
        return ctx

    def get_organization(self):
        return get_object_or_404(
            Organization.objects.all(),
            slug=self.request.query_params.get(
                'organization__slug'
            )
        )


class PayrollDetailExportViewSet(BackgroundExcelExportMixin, viewsets.ViewSet):
    permission_classes = [permission_factory.build_permission(
        "PayrollViewPermission",
        allowed_to=[HAS_PERMISSION_FROM_METHOD],
    )]
    export_type = "Saved Payroll"
    serializer_class = EmployeePayrollExportSerializer
    export_freeze_first_column = True
    heading_map = getattr(settings, 'PAYROLL_HEADING_MAP', None)
    notification_permissions = [
        GENERATE_PAYROLL_PERMISSION, PAYROLL_REPORT_PERMISSION]

    def get_frontend_redirect_url(self):
        return (f'/admin/{self.get_organization().slug}/'
                f'payroll/collection/detail/{self.get_object().id}')

    @property
    def mode(self):
        if self.request.query_params.get('as') == 'approver':
            return 'approver'
        return 'hr'

    def get_export_type(self):
        return f"{self.export_type}-{self.kwargs.get('payroll_id')}"

    def has_user_permission(self):
        if self.mode == 'approver':
            payroll = get_object_or_404(
                self.get_queryset(), id=self.kwargs.get('payroll_id'))
            return PayrollApproval.objects.all().filter(
                user=self.request.user,
                status=PENDING,
                payroll=payroll
            ).exists()
        return validate_permissions(
            self.request.user.get_hrs_permissions(self.get_organization()),
            GENERATE_PAYROLL_PERMISSION,
            PAYROLL_REPORT_PERMISSION
        )

    def get_organization(self):
        return get_object_or_404(
            Organization.objects.all(),
            slug=self.request.query_params.get(
                'organization__slug'
            )
        )

    def get_queryset(self):
        return Payroll.objects.filter(organization=self.get_organization())

    def get_object(self):
        obj = get_object_or_404(self.get_queryset(),
                                id=self.kwargs.get('payroll_id'))
        self.check_object_permissions(self.request, obj)
        return obj

    def get_export_data(self):
        return get_filtered_employee_payrolls_from_query_params(
            self.get_object(), self.request.query_params)

    def get_export_fields(self):
        personal_heading_setting = ExtraHeadingReportSetting.objects.filter(
            organization=self.get_organization()).first()
        export_fields = {
            "Employee Code": "employee_code",
            "Name": "full_name",
            "Username": "username",
        }

        personal_heading_mapper = {
            "Username": "username",
            "Job Title": "job_title",
            "Division": "division",
            "Employment Level": "employment_level",
            "Employee Level Hierarchy": "employee_level_hierarchy",
            "Employment Type": "employment_status",
            "Branch": "branch",
            "Step": "user_detail.step",
            "Package Name": "user_detail.package_name",
            "Working Days": "user_detail.working_days",
            "Worked Days": "user_detail.worked_days",
            "Absent Days": "user_detail.absent_days",
            "Worked Hours": "user_detail.worked_hours",
            "Overtime Hours": "user_detail.overtime_hours",
            "SSF Number": "ssf_number",
            "Bank Name": "bank_name",
            "Bank Branch": "bank_branch",
            "Bank Account Number": "bank_account_number",
            "PF Number": "pf_number",
            "PAN Number": "pan_number",
            "CIT Number": "cit_number"
        }

        if personal_heading_setting:
            for heading in personal_heading_setting.headings:
                export_fields.update({
                    heading: personal_heading_mapper[heading]
                })
        else:
            export_fields.update(**personal_heading_mapper)

        payroll_heading_fil = {}
        heading_setting = PayrollCollectionDetailReportSetting.objects.filter(
            organization=self.get_organization()).first()
        if heading_setting:
            payroll_heading_fil['id__in'] = heading_setting.headings.values_list('id', flat=True)
        for heading in Heading.objects.filter(**payroll_heading_fil).only("name"):
            export_fields.update(
                {heading.name: f"heading_amounts.{heading.name}"})
            if heading.name == BASIC_SALARY_NAME and ADD_VALUES_TO_BASIC:
                export_fields.update({
                    'Percentage of Month Worked': "heading_amounts.percentage_of_month_worked",
                })
        return export_fields

    def get_extra_export_data(self):
        ctx = super().get_extra_export_data()
        ctx.update({
            'grand_total_data': {
                'footer_offset': 5,
                'footer_color': '00FF00',
                'footer_text': 'Grand Total Salary to be paid in NPR',
            }
        })
        return ctx

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        data = EmployeePayrollExportSerializer(
            obj, context={'organization': kwargs.get('organization')}).data
        return data

    def get_export_description(self):
        payroll = self.get_object()
        from_date = payroll.from_date
        to_date = payroll.to_date
        title = payroll.title
        return getattr(settings, 'PAYROLL_EXPORT_TITLE', []) + [f"Title: {title}"] + [
            f"Saved Payroll from {from_date} to {to_date} of {self.get_organization().name}."
        ]

    def get_footer_data(self):
        footer_data = PayrollApprovalSetting.objects.filter(
            organization=self.get_organization()
        ).annotate(
            full_name=Case(
                When(
                    user__middle_name='',
                    then=Concat(
                        F('user__first_name'), Value(' '),
                        F('user__last_name')
                    )
                ),
                default=Concat(
                    F('user__first_name'), Value(' '),
                    F('user__middle_name'), Value(' '),
                    F('user__last_name')
                )
            ),
            job_title=F('user__detail__job_title__title'),
        ).values('full_name', 'job_title')

        return [
            {
                'footer_text': [
                    '_' * 30,
                    item['full_name'],
                    item['job_title']
                ],
                'inline': True
            }
            for item in footer_data

        ]

    @classmethod
    def get_exported_file_content(cls, data, title, columns, extra_content,
                                  description=None, heading_map=None, footer_data=None):
        wb = PayrollExcelExport.process(
            data,
            title=title,
            columns=columns,
            description=description,
            prepare_export_object=cls.prepare_export_object,
            prepare_export_object_context=extra_content,
            freeze_first_column=cls.export_freeze_first_column,
            heading_map=heading_map,
            footer_data=footer_data
        )
        return ContentFile(save_virtual_workbook(wb))

    def get_exported_as(self):
        if self.mode == "approver":
            return NORMAL_USER
        return ADMIN


class EmployeePayrollAPIViewSet(
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet
):
    queryset = EmployeePayroll.objects.all().annotate(
        total_comment=Count('employee_payroll_comments')
    ).order_by('-total_comment')
    update_serializer_class = EmployeePayrollSerializer
    list_serializer_class = EmployeePayrollListSerializer
    # choice_fields = ('id', 'label')
    filter_map = {
        "employee__id": "employee__id",
        "payroll__status": "payroll__status",
        "acknowledgement_status": "acknowledgement_status"
    }

    filter_backends = (
        FilterMapBackend, OrderingFilterMap
    )
    ordering_fields_map = {
        'from_date': 'payroll__from_date',
        'to_date': 'payroll__to_date',
        'status': 'acknowledgement_status',
        'id': 'id',
        'acknowledged_at': 'acknowledged_at',
        'total_comment': 'total_comment'
    }

    @cached_property
    def status_counts(self):
        # Do not aggregate if filter by acknowledgement_status
        self.filter_map = {
            "employee__id": "employee__id",
            "payroll__status": "payroll__status",
        }
        return self.get_queryset().aggregate(
            total=Count('id', distinct=True),
            generated=Count(
                'id',
                filter=Q(acknowledgement_status=PAYSLIP_GENERATED)
            ),
            pending=Count(
                'id',
                filter=Q(acknowledgement_status=PAYSLIP_ACKNOWLEDGEMENT_PENDING),
                distinct=True
            ),
            acknowledged=Count(
                'id',
                filter=Q(acknowledgement_status=PAYSLIP_ACKNOWLEDGED),
                distinct=True
            )
        )

    def get_queryset(self):
        organization = self.get_organization()
        qs = super().get_queryset().filter(
            employee__detail__organization=organization
        )
        is_hr = self.request.query_params.get("as") == "hr"
        if is_hr and validate_permissions(
            self.request.user.get_hrs_permissions(self.get_organization()),
            GENERATE_PAYROLL_PERMISSION,
            PAYROLL_REPORT_PERMISSION
        ):
            return qs

        if self.action == 'payroll_edits':
            # if not hr, approval pending can only be viewed by recipient of approval request
            return qs.filter(payroll__approval_pending=self.request.user)

        fiscal_year = self.request.query_params.get('fiscal_year', None)
        if fiscal_year:
            get_fiscal_year = FiscalYear.objects.filter(
                slug=fiscal_year).first()
            qs = qs.filter(payroll__from_date__range=(
                get_fiscal_year.start_at, get_fiscal_year.end_at))

        qs = qs.filter(employee=self.request.user)
        return get_employee_payrolls_via_settings(qs, organization)

    def get_serializer_class(self):
        if self.action == 'list':
            return self.list_serializer_class
        return self.update_serializer_class

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        response.data.update(
            {'payslip_template': OrganizationPayrollConfig.get_payslip_template(
                self.get_organization())
             }
        )
        response.data.update({"counts": self.status_counts})
        return response

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if not instance.status:
            return Response({
                'success': False
            })
        return super().update(request, *args, **kwargs)

    def get_organization(self):
        is_approver = self.request.query_params.get('as') == 'approver'
        approver_id = self.request.user.id
        approver_list = PayrollApprovalSetting.objects.all().values_list('user', flat=True)
        if is_approver and approver_id in approver_list:
            return get_object_or_404(
                Organization,
                slug=self.request.query_params.get(
                    'employee__detail__organization__slug'
                )
            )
        org_list = self.request.user.switchable_organizations_pks.union({
            self.request.user.detail.organization.id
        })
        return get_object_or_404(
            Organization.objects.filter(
                id__in=org_list
            ),
            slug=self.request.query_params.get(
                'employee__detail__organization__slug'
            )
        )

    def check_object_permissions(self, request, obj):
        super().check_object_permissions(request, obj)
        if request.method.upper() == "DELETE" and obj.payroll.status in [APPROVED, CONFIRMED] \
        and self.request.query_params.get('as') == 'hr' and  not \
            validate_permissions(
                self.request.user.get_hrs_permissions(self.get_organization()),
                DELETE_CONFIRMED_EMPLOYEE_PAYROLL_PERMISSION,
                ):
            self.permission_denied(
                request, message="You do not have permission to perform this action"
                )

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        allowed_in = [GENERATED, APPROVED, CONFIRMED]
        payroll = obj.payroll
        payroll_status = payroll.status
        if payroll_status not in allowed_in:
            raise ValidationError(
                f'Can not delete individual payroll for {payroll_status} state.'
            )
        if payroll_status != GENERATED:
            PayrollApprovalHistory.objects.create(
                payroll=obj.payroll,
                actor=request.user,
                action="deleted",
                remarks=f"Payroll of {obj.employee}"
            )
        extra_data = payroll.extra_data
        # remove edited_general_heading for user
        edited_general_heading = extra_data.get('edited_general_headings', {}).pop(str(obj.employee.id), None)
        if edited_general_heading:
            payroll.save()
        return super().destroy(request, *args, **kwargs)

    @action(
        methods=['GET'],
        detail=True,
        url_path='edits'
    )
    def payroll_edits(self, *args, **kwargs):
        employee_payroll = self.get_object()
        queryset = employee_payroll.history.select_related(
            'created_by',
            'created_by__detail',
            'created_by__detail__job_title',
        ).prefetch_related(
            Prefetch(
                'amount_history',
                queryset=PayrollEditHistoryAmount.objects.all().order_by('heading__order'),
                to_attr='_amounts'
            )
        )

        page = self.paginate_queryset(queryset)
        return self.get_paginated_response(
            EmployeePayrollHistorySerializer(
                page,
                many=True,
                context=self.get_serializer_context()
            ).data
        )

    @action(
        methods=['GET'],
        detail=False,
        url_path='(?P<organization_slug>[\w\-]+)/show-generated-payslip'
    )
    def show_generated_payslip(self, request, **kwargs):
        payroll_config = OrganizationPayrollConfig.objects.filter(
            organization__slug=kwargs["organization_slug"]
        ).first()
        generated_status = False
        if payroll_config:
            generated_status = payroll_config.show_generated_payslip
        return Response({
            "show_generated_payslip": generated_status
        })


class EmployeePayrollCommentAPIViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    viewsets.GenericViewSet
):
    queryset = EmployeePayrollComment.objects.select_related(
        'commented_by__detail',
        'commented_by__detail__employment_level',
        'commented_by__detail__job_title',
        'commented_by__detail__organization',
        'commented_by__detail__division',
    )
    serializer_class = EmployeePayrollCommentSerializer
    permission_classes = [
        permission_factory.build_permission(
            "EmployeePayrollCommentPermission",
            allowed_to=[GENERATE_PAYROLL_PERMISSION,
                        HAS_PERMISSION_FROM_METHOD]
        )
    ]

    @property
    def is_hr(self):
        return validate_permissions(
            self.request.user.get_hrs_permissions(self.organization),
            GENERATE_PAYROLL_PERMISSION
        )

    def get_queryset(self):
        qs = super().get_queryset().filter(employee_payroll=self.employee_payroll)
        if not self.is_hr:
            qs = qs.filter(employee_payroll__employee=self.request.user)
        return qs

    @cached_property
    def employee_payroll(self):
        pk = self.kwargs.get('employee_payroll_id')
        return get_object_or_404(EmployeePayroll.objects.all(), pk=pk)

    def has_user_permission(self):
        return self.is_hr or (self.employee_payroll.employee == self.request.user)

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['employee_payroll'] = self.employee_payroll
        ctx['organization'] = self.organization
        return ctx

    @cached_property
    def organization(self):
        return get_object_or_404(
            Organization.objects.all(),
            slug=self.request.query_params.get(
                'organization__slug'
            )
        )


class PayrollGenerationHistoryViewSet(
    OrganizationMixin,
    OrganizationCommonsMixin,
    ReadOnlyModelViewSet
):
    queryset = PayrollGenerationHistory.objects.all()
    serializer_class = PayrollGenerationHistorySerializer
    permission_classes = [
        permission_factory.build_permission(
            "PayrollHistoryPermission",
            allowed_to=[GENERATE_PAYROLL_PERMISSION]
        )
    ]

    def get_serializer(self, *args, **kwargs):
        if not self.detail:
            kwargs['exclude_fields'] = ['data', 'errors']
        return super().get_serializer(*args, **kwargs)


class ExcelPayrollPackageViewSet(
    OrganizationMixin,
    OrganizationCommonsMixin,
    ListCreateRetrieveDestroyViewSetMixin
):
    serializer_class = ExcelPayrollPackageSerializer
    queryset = ExcelPayrollPackage.objects.all()
    permission_classes = [permission_factory.build_permission(
        "ExcelPayrollPackageImportPermission",
        allowed_to=[
            ASSIGN_PAYROLL_PACKAGES_PERMISSION,
        ],
    )]
    filter_backends = (FilterMapBackend, OrderingFilterMap, SearchFilter)
    filter_map = {
        'organization__slug': 'organization__slug',
        'cloned_from': 'cloned_from__name',
        'status': 'status',
        'name': 'name'
    }
    ordering_fields_map = {
        'name': 'name',
        'modified_at': 'modified_at',
        'status': 'status',
        'cloned_from': 'cloned_from__name'
    }

    search_fields = (
        'name', 'cloned_from__name'
    )

    def list(self, *args, **kwargs):
        response = super().list(*args, **kwargs)
        response.data["error_file"] = cache.get("failed_package_errors", None)
        response.data["created_on"] = cache.get("failed_package_errors_timestamp", None)
        return response

    def create(self, *args, **kwargs):
        if cache.get(f'block_excel_package', False):
            raise ValidationError({"non_field_errors": f"Previous bulk assignment is getting processed."})
        return super().create(*args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """
        UserExperiencePackageSlot is protected for Package deletion
        so User ExperiencePackageSlot and Package cannot be cascade after
        ExcelPayrollPackage deletion
        """
        instance = self.get_object()
        for package in instance.packages.all():
            if package.is_used_package:
                raise ValidationError("Used package cannot be deleted.")

        for package in instance.packages.all():
            title = f'{self.request.user.full_name} has {PACKAGE_DELETED} a package named "{package.name}" by bulk package assign feature.'
            create_package_activity(title=title, package=package,
                                    action=PACKAGE_DELETED)
        instance.package_slots.all().delete()
        instance.packages.all().delete()
        instance.delete()
        cache.delete("block_excel_package")
        cache.delete("payroll_generated_employee_count")
        cache.delete("payroll_generated_employee_name")
        return Response(status=status.HTTP_204_NO_CONTENT)

    def check_permissions(self, request):
        super().check_permissions(request)
        if self.request.query_params.get('as') == 'hr' and not validate_permissions(
                self.request.user.get_hrs_permissions(self.get_organization()),
                ASSIGN_PAYROLL_PACKAGES_PERMISSION
        ):
            self.permission_denied(
                request, message="You do not have permission to perform this action."
            )
