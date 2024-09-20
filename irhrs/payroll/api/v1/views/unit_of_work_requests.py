from copy import deepcopy

from django.core.files.base import ContentFile
from django.db.models import F, Sum, Func
from django.db.models.functions import Round
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.filters import SearchFilter
from rest_framework.response import Response

from irhrs.core.constants.payroll import CANCELED, DENIED, FORWARDED, CONFIRMED, REQUESTED
from irhrs.core.mixins.serializers import create_dummy_serializer
from irhrs.core.mixins.viewset_mixins import OrganizationMixin, ListCreateRetrieveViewSetMixin, \
    OrganizationCommonsMixin, GetStatisticsMixin
from irhrs.core.utils import subordinates, nested_getattr
from irhrs.core.utils.common import validate_permissions, get_today
from irhrs.core.utils.filters import FilterMapBackend, OrderingFilterMap
from irhrs.export.constants import COMPLETED
from irhrs.export.mixins.export import BackgroundExportMixin
from irhrs.payroll.api.v1.serializers.unit_of_work_requests import UnitOfWorkRequestSerializer
from irhrs.payroll.models import UnitOfWorkRequest, APPROVED
from irhrs.permission.constants.permissions import HAS_PERMISSION_FROM_METHOD, \
    ALL_PAYROLL_PERMISSIONS, UNIT_OF_WORK_REQUEST_PERMISSION
from irhrs.permission.permission_classes import permission_factory

RemarksRequiredSerializer = create_dummy_serializer({
    'remarks': serializers.CharField(max_length=600, allow_blank=False)
})


class UnitOfWorkRequestViewSet(
    BackgroundExportMixin,
    GetStatisticsMixin,
    OrganizationMixin,
    OrganizationCommonsMixin,
    ListCreateRetrieveViewSetMixin
):
    """
    create:

    Request unit of work

    By Normal User
    -------------

    data sample

        {
            "rate": rate_id,
            "quantity": 2,
            "attachment": File,
            "remarks": "Remarks Here"

        }

    By supervisor
    ------------
    *send `?as=supervisor` in url*


        {
            "user": user_id,
            "rate": rate_id,
            "quantity": 2,
            "attachment": File,
            "remarks": "Remarks",
            "status": "Approved" or "Forwarded"
        }

    By HR
    -----
    *send `?as=hr` in url*

        {
            "user": user_id,
            "rate": rate_id,
            "quantity": 2,
            "attachment": File,
            "remarks": "Remarks",
            "status": "Confirmed"
        }

    """
    serializer_class = UnitOfWorkRequestSerializer
    queryset = UnitOfWorkRequest.objects.all().select_related(
        'user',
        'user__detail',
        'user__detail__job_title',
        'created_by',
        'created_by__detail',
        'created_by__detail__job_title',
        'recipient',
        'recipient__detail',
        'recipient__detail__job_title',
        'rate',
        'rate__operation',
        'rate__operation_code'
    )
    organization_field = 'user__detail__organization'
    permission_classes = (permission_factory.build_permission(
        'UnitOfWorkRequestPermission',
        allowed_to=[HAS_PERMISSION_FROM_METHOD]
    ),)

    filter_backends = (FilterMapBackend, SearchFilter, OrderingFilterMap)
    filter_map = {
        'operation': 'rate__operation',
        'code': 'rate__operation_code',
        'start_date': 'created_at__date__gte',
        'end_date': 'created_at__date__lte',
        'status': 'status',
    }

    ordering = '-created_at'
    ordering_fields_map = {
        'operation': 'rate__operation__title',
        'code': 'rate__operation_code__title',
        'quantity': 'quantity',
        'rate': 'rate__rate',
        'created_at': 'created_at',
        'modified_at': 'modified_at',
        'full_name': (
            'user__first_name',
            'user__middle_name',
            'user__last_name'
        )
    }

    search_fields = (
        'user__first_name',
        'user__middle_name',
        'user__last_name'
    )
    statistics_field = 'status'
    export_fields = []

    @property
    def mode(self):
        mode = self.request.query_params.get('as')
        if mode in ['hr', 'supervisor']:
            return mode
        return 'user'

    def get_queryset(self):
        queryset = super().get_queryset()

        if self.mode == 'supervisor':
            queryset = queryset.filter(recipient=self.request.user)
        elif self.mode == 'user':
            queryset = queryset.filter(user=self.request.user)

        queryset = queryset.annotate(total=F('quantity')*F('rate__rate'))
        return queryset

    def has_user_permission(self):
        if self.mode == "hr":
            is_hr = validate_permissions(
                self.request.user.get_hrs_permissions(self.get_organization()),
                ALL_PAYROLL_PERMISSIONS,
                UNIT_OF_WORK_REQUEST_PERMISSION
            )
            # If not is_hr and mode is hr raise permission denied directly
            if not is_hr:
                return False

        # post actions limited to hr and supervisor now allow others as
        # get_queryset will filter
        if self.action == 'perform_action':
            action_performed = self.kwargs.get('action_performed')
            if self.mode == 'user' and action_performed != 'cancel':
                return False
            elif self.mode == 'supervisor' and action_performed not in ['approve', 'deny',
                                                                        'forward']:
                return False
            elif self.mode == 'hr' and action_performed in ['forward', 'cancel']:
                return False

        return True

    def get_serializer(self, *args, **kwargs):
        exclude = []
        if self.mode == 'user' and self.action in ['list', 'retrieve', 'create']:
            exclude.append('user')
        if self.mode == 'user' and self.action == 'create':
            exclude.append('status')

        if self.action == 'list':
            exclude.append('histories')
        if exclude:
            kwargs.update({'exclude_fields': exclude})
        return super().get_serializer(*args, **kwargs)

    def create(self, request, *args, **kwargs):
        data = deepcopy(request.data)
        if self.mode == 'user':
            data.update({'user': request.user.id})

        serializer = self.get_serializer_class()(data=data, context=self.get_serializer_context())
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=201, headers=headers)

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        response.data['stats'] = self.statistics

        response.data.get('stats').update(
            self.filter_queryset(self.get_queryset()).aggregate(
                total_sum=Round(Sum('total'), precision=2)
            )
        )
        return response

    @action(detail=True, methods=['POST'],
            url_path=r'(?P<action_performed>(approve|deny|cancel|forward|confirm))',
            serializer_class=RemarksRequiredSerializer, url_name='perform-action')
    def perform_action(self, request, action_performed, *args, **kwargs):
        instance = self.get_object()

        self.validate_permissions_for_action(action_performed, instance)

        serializer = RemarksRequiredSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action_status_map = {
            'approve': APPROVED,
            'cancel': CANCELED,
            'deny': DENIED,
            'forward': FORWARDED,
            'confirm': CONFIRMED
        }
        instance.status = action_status_map[action_performed]

        if action_performed == 'forward':
            next_level_supervisor = subordinates.get_next_level_supervisor(
                instance.user,
                instance.recipient
            )
            if not next_level_supervisor:
                raise serializers.ValidationError({
                    'non_field_errors': _(
                        "Could not forward request. Next level supervisor not set to forward.")
                })
            instance.recipient = next_level_supervisor

        elif action_performed == 'confirm':
            instance.confirmed_on = get_today()
        instance.save()

        instance.histories.create(
            action_performed_by=self.request.user,
            action_performed_to=instance.recipient,
            action_performed=instance.status,
            remark=serializer.validated_data.get('remarks', '')
        )

        return Response({'message': _(f"Successfully {instance.status.lower()} request.")})

    def validate_permissions_for_action(self, action_performed, instance):
        request = self.request

        if self.mode == 'supervisor' and not instance.recipient == request.user:
            raise self.permission_denied(request, "You are not the right recipient.")

        if (
            action_performed in ['approve', 'cancel', 'forward'] and
            instance.status not in [REQUESTED, FORWARDED]
        ):
            raise self.permission_denied(
                request,
                message=f"Can not perform {action_performed}. Request must be either in "
                        "requested or forwarded state."
            )
        elif action_performed == 'deny':
            if self.mode == 'supervisor' and instance.status not in [REQUESTED, FORWARDED]:
                raise self.permission_denied(
                    request,
                    message=f"Can not perform {action_performed}. Request must be either in "
                            "requested or forwarded state."
                )
            elif self.mode == 'hr' and instance.status in [CONFIRMED, DENIED]:
                raise self.permission_denied(
                    request,
                    message=f"Can not {action_performed} {instance.status.lower()} requests."
                )
        elif action_performed == 'confirm' and instance.status in [DENIED, CONFIRMED]:
            raise self.permission_denied(
                request,
                message=f"Can not {action_performed} {instance.status.lower()} requests."
            )

        if (
            self.mode == 'supervisor' and
            action_performed in ['approve', 'deny', 'forward'] and
            not subordinates.authority_exists(instance.user, self.request.user, action_performed)
        ):
            raise self.permission_denied(
                request, f"You do not have authority to {action_performed} this request."
            )

    def get_export_type(self):
        return f"unit_of_work_request_{self.organization.id}"

    def get_export_name(self):
        return "Unit of Work Request"

    def get_extra_export_data(self):
        extra_data = super().get_extra_export_data()
        extra_data.update({
            "start_date": self.request.query_params.get('start_date') or 'N/A',
            "end_date": self.request.query_params.get('end_date') or 'N/A'
        })
        return extra_data

    @classmethod
    def get_exported_file_content(cls, queryset, title, columns, extra_content,
                                  description=None, heading_map=None, footer_data=None
                                  ):
        # response = HttpResponse(content_type='application/vnd.ms-excel')
        # response['Content-Disposition'] = 'attachment; filename="unit-of-work.xlsx"'

        wb = Workbook()
        sheet = wb['Sheet']
        bold = Font(bold=True)
        align_center = Alignment(horizontal='center')
        strp_time_format = "%Y-%m-%d %H:%M:%S"
        organization = extra_content.get('organization')
        start_date = extra_content.get('start_date')
        end_date = extra_content.get('end_date')
        # for organization logo in downloaded excel report:
        # logo = nested_getattr(organization, 'appearance.logo')
        # if logo:
        #     sheet.merge_cells(
        #         start_row=1,
        #         start_column=1, end_row=1,
        #         end_column=7
        #     )
        #     image_obj = Image(logo)
        #     sheet.add_image(image_obj, anchor="A1")
        #     dimension = sheet.row_dimensions[1]
        #     dimension.height = image_obj.height * 0.75

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        sheet.cell(row=1, column=1, value=organization.name)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        sheet.cell(row=2, column=1,
                   value=f"Downloaded at: ({get_today(with_time=True).strftime(strp_time_format)})")

        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
        cell = sheet.cell(row=4, column=1, value="Unit Of Work Request")
        cell.font = bold
        cell.alignment = align_center

        sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
        cell = sheet.cell(row=5, column=1, value=f"Start Date: {start_date} - End Date: {end_date}")
        cell.alignment = align_center

        cell = sheet.cell(column=1, row=7, value="Full Name")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=2, row=7, value="Operation/Project")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=3, row=7, value="Code/Task")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=4, row=7, value="Rate")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=5, row=7, value="Quantity/Hour")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=6, row=7, value="Total")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=7, row=7, value="Date Requested")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=8, row=7, value="Status")
        cell.font = bold
        cell.alignment = align_center

        for column_index, work in enumerate(queryset):
            column = column_index + 8
            full_name = work.user.full_name or 'N/A'
            code = work.rate.operation_code.title or 'N/A'
            operation = work.rate.operation.title or 'N/A'
            rate = work.rate.rate or 0.0
            quantity = work.quantity or 0.0
            total = rate * quantity or 0.0
            date_requested = work.created_at.astimezone().strftime(strp_time_format)
            status = work.status
            sheet.cell(column=1, row=column, value=full_name)
            sheet.cell(column=2, row=column, value=operation)
            sheet.cell(column=3, row=column, value=code)
            sheet.cell(column=4, row=column, value=rate)
            sheet.cell(column=5, row=column, value=quantity)
            sheet.cell(column=6, row=column, value=round(total, 2))
            sheet.cell(column=7, row=column, value=date_requested)
            sheet.cell(column=8, row=column, value=status)

        last_row_count = sheet.max_row
        sheet.merge_cells(
            start_row=last_row_count+1, start_column=1, end_row=last_row_count+1, end_column=3
        )
        cell = sheet.cell(row=last_row_count + 1, column=1, value="Total Amount")
        cell.font = bold
        cell.alignment = Alignment(horizontal='center')

        def formula(col):
            return "=SUM({}:{})".format(
                f"{get_column_letter(col)}{6}",
                f"{get_column_letter(col)}{last_row_count}",
            )

        dynamic_column_count = 3
        # loop range is 3 because we need sum of three column only
        for count in range(3):
            count += 1
            cell = sheet.cell(
                row=last_row_count + 1, column=dynamic_column_count + count,
                value=formula(dynamic_column_count + count)
            )
            cell.font = bold
            cell.alignment = Alignment(horizontal='center')

        return ContentFile(save_virtual_workbook(wb))

    @classmethod
    def save_file_content(cls, export_instance, file_content):
        export_instance.export_file.save('unit-of-work-request.xlsx', file_content)
        export_instance.status = COMPLETED
        export_instance.save()

    def get_frontend_redirect_url(self):
        if self.mode == "hr":
            return f'/admin/{self.organization.slug}/payroll/unit-of-work/'
        elif self.mode == "supervisor":
            return "/user/supervisor/payroll/unit-of-work"
        return "/user/payroll/unit-of-work"
