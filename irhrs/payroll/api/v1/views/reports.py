import itertools

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile

from django.db.models import Sum, OuterRef, Q, F, Subquery, ExpressionWrapper, Value, FloatField
from django.db.models.functions import Coalesce, Extract
from django.utils import timezone
from django.utils.functional import cached_property
from django_filters.rest_framework import DjangoFilterBackend

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import serializers
from rest_framework.exceptions import ValidationError
from rest_framework.filters import SearchFilter
from rest_framework.generics import get_object_or_404
from rest_framework.permissions import BasePermission
from rest_framework.response import Response
from rest_framework.viewsets import ReadOnlyModelViewSet

from irhrs.core.constants.organization import GLOBAL
from irhrs.core.mixins.serializers import create_dummy_serializer
from irhrs.core.mixins.viewset_mixins import OrganizationCommonsMixin, OrganizationMixin, \
    ListViewSetMixin, DateRangeParserMixin, IStartsWithIContainsSearchFilter, RetrieveViewSetMixin, \
    UserMixin, ListCreateViewSetMixin
from irhrs.core.utils import nested_getattr
from irhrs.core.utils.common import get_today
from irhrs.core.utils.filters import OrderingFilterMap, FilterMapBackend, \
    NullsAlwaysLastOrderingFilter
from irhrs.core.validators import validate_comma_separated_integers
from irhrs.export.mixins.export import BackgroundExcelExportMixin, BackgroundTableExportMixin
from irhrs.export.utils.export import ExcelExport, PayrollDisbursementExcel
from irhrs.organization.models import FiscalYear, FY, FiscalYearMonth, OrganizationBank
from irhrs.payroll.api.v1.serializers import EmployeeThinSerializer
from irhrs.payroll.api.v1.serializers.payroll import SSFReportSettingSerializer, \
    DisbursementReportSettingSerializer, TaxReportSettingSerializer, \
    PayrollCollectionDetailReportSettingSerializer, PayrollExtraHeadingSettingSerializer, \
    PayrollDifferenceDetailReportSettingSerializer
from irhrs.payroll.api.v1.serializers.report.payroll_reports import TaxReportSerializer, \
    PFReportSerializer, PayrollGeneralReportSerializer, PackageWiseSalarySerializer, \
    BackdatedCalculationSerializer, SSFReportSerializer, DisbursementReportSerializer
from irhrs.payroll.models import EmployeePayroll, Payroll, ReportRowRecord, \
    Heading, ReportRowUserExperiencePackage, UserExperiencePackageSlot, CONFIRMED, \
    OrganizationPayrollConfig, YearlyHeadingDetail, BackdatedCalculation, \
    SSFReportSetting, DisbursementReportSetting, TaxReportSetting, \
    PayrollCollectionDetailReportSetting, ExtraHeadingReportSetting, \
    PayrollApproval, GENERATED, PayrollDifferenceDetailReportSetting
from irhrs.core.utils.nepdate import ad2bs
from irhrs.payroll.models.payslip_report_setting import PayslipReportSetting
from irhrs.payroll.utils.calculator import PackageSalaryCalculator
from irhrs.payroll.utils.headings import get_heading_details
from irhrs.payroll.utils.helpers import get_dismiss_date, get_appoint_date
from irhrs.hris.api.v1.permissions import ViewPayrollReportPermission
from irhrs.payroll.utils.reports import YearlyPayslipReportForRemainingMonths, has_heading_rule
from irhrs.permission.constants.permissions import PAYROLL_REPORT_PERMISSION, \
    GENERATE_PAYROLL_PERMISSION, HAS_PERMISSION_FROM_METHOD, ALL_PAYROLL_PERMISSIONS, \
    PAYROLL_SETTINGS_PERMISSION, PAYROLL_READ_ONLY_PERMISSIONS
from irhrs.permission.permission_classes import permission_factory
from irhrs.users.models import UserExperience

Employee = get_user_model()


class EmployeePayrollMixin(
    OrganizationCommonsMixin,
    OrganizationMixin
):
    queryset = EmployeePayroll.objects.all().select_related(
        'employee',
        'employee__detail',
        'employee__detail__job_title',
        'employee__legal_info'
    )

    def get_queryset(self):
        # super not called because super will have organization filter
        return self.queryset.all().filter(payroll=self.payroll)

    @cached_property
    def payroll(self):
        payroll_id = self.kwargs.get('payroll_id')
        return get_object_or_404(Payroll, id=payroll_id, organization=self.organization)


class PayrollTaxReportViewSet(
    BackgroundExcelExportMixin,
    EmployeePayrollMixin,
    ReadOnlyModelViewSet
):
    permission_classes = [
        permission_factory.build_permission(
            "PFReportPermission",
            allowed_to=[
                PAYROLL_REPORT_PERMISSION,
                GENERATE_PAYROLL_PERMISSION,
                ALL_PAYROLL_PERMISSIONS,
                PAYROLL_SETTINGS_PERMISSION,
                PAYROLL_READ_ONLY_PERMISSIONS
            ]
        )
    ]
    serializer_class = TaxReportSerializer
    filter_backends = (OrderingFilterMap,)

    ordering_fields_map = {
        "full_name": (
            "employee__first_name",
            "employee__middle_name",
            "employee__last_name",
        )
    }

    @cached_property
    def headings(self):
        setting = TaxReportSetting.objects.filter(organization=self.organization).first()
        if not setting:
            return []
        return setting.headings.values_list('name', 'id')

    def get_export_fields(self):
        fields = {
            "PAN": "pan_number",
            "Employee Name": "employee.full_name",
            "Username": "employee.username",
            "T Date": "t_date",
            "Date Type": "date_type",
            **{
                heading_name:
                    f'amount_of_{heading_id}' for heading_name, heading_id in self.headings
            }
        }
        return fields

    def get_extra_export_data(self):
        ctx = super().get_extra_export_data()
        ctx['prepare_export_object_context'] = {'tax_headings_id': self.headings}
        return ctx

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        """
        this is a method that will prepare export object. It will get object instance as its
        first parameter and should return object or dict.
        :param obj: Instance (A row) that will be exported
        :return: prepared object
        """

        date = TaxReportSerializer.get_t_date(obj)
        obj.t_date = "-".join(map(str, ad2bs(date))) if date else None
        obj.date_type = "BS"
        obj.pan_number = TaxReportSerializer.get_pan_number(obj)

        row_report_records = ReportRowRecord.objects.filter(
            employee_payroll=obj,
            heading__in=[heading_id for heading_name, heading_id in kwargs.get('tax_headings_id')],
            employee_payroll__payroll__status__in=[GENERATED, CONFIRMED]
        )

        obj.tax_headings = getattr(obj, 'tax_headings', [])
        for row_report_record in row_report_records:
            heading = row_report_record.heading
            heading_id = heading.id
            heading_name = heading.name
            setattr(obj,
                    f'amount_of_{heading_id}',
                    "-" if row_report_record.amount is None else row_report_record.amount
                    )
        return obj

    def get_export_type(self):
        return f"tax_report_payroll_{self.payroll.id}"


class PayrollPFReportViewSet(
    BackgroundExcelExportMixin,
    EmployeePayrollMixin,
    ReadOnlyModelViewSet
):
    serializer_class = PFReportSerializer
    permission_classes = [
        permission_factory.build_permission(
            "PFReportPermission",
            allowed_to=[PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION]
        )
    ]
    notification_permissions = [PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION]

    export_fields = {
        "Sno": "#SN",
        "Designation": "user_experience_package_slot.user_experience.job_title.title",
        "Name": "employee.full_name",
        "Username": "employee.username",
        "PF Number": "pf_number",
        # hack to insert blank cell
        "": "null",
        "Total Fund Deducted": "total_fund_deducted",
        "PF Deducted from Employee": "deducted_from_employee",
        "PF Contribution by Employer": "contribution_from_company",
    }

    def get_frontend_redirect_url(self):
        return f'/admin/{self.organization.slug}/payroll/collection/details/' \
               f'{self.payroll.id}/reports/pf-report'

    def get_export_type(self):
        return f"pf_report_payroll_{self.payroll.id}"

    def get_export_name(self):
        export_name = super().get_export_name()
        if export_name == self.get_export_type():
            return "PF Report"
        return export_name

    def get_extra_export_data(self):
        ctx = super().get_extra_export_data()

        ctx.update({
            "payroll": self.payroll
        })
        return ctx

    @staticmethod
    def get_two_digit_string(number):
        return f"0{number}" if number < 10 else str(number)[-2:]

    @staticmethod
    def get_total_pf_amount(payroll):
        return ReportRowRecord.objects.filter(
            employee_payroll__payroll=payroll,
            heading__payroll_setting_type='Provident Fund',
            heading__type__in=['Addition', 'Deduction']
        ).aggregate(total=Sum('amount'))['total'] or 0.0

    @classmethod
    def get_details(cls, organization, payroll):
        from_bs_date = ad2bs(payroll.from_date)
        to_bs_date = ad2bs(payroll.to_date)

        return (
            ("Phone No.", organization.contacts.get("Phone")),
            ("Email Id", organization.email),
            ("Office Code", ""),
            ("Total Amount", cls.get_total_pf_amount(payroll)),
            ("From Month", f"{cls.get_two_digit_string(from_bs_date[0])}"
                           f"{cls.get_two_digit_string(from_bs_date[1])}"),
            ("To Month", f"{cls.get_two_digit_string(to_bs_date[0])}"
                         f"{cls.get_two_digit_string(to_bs_date[1])}"),
            ("Deposit Date", None)
        )

    @staticmethod
    def prepare_export_object(obj, **kwargs):

        obj.pf_number = PFReportSerializer.get_pf_number(obj)
        obj.null = None

        pf_amounts = PFReportSerializer.get_pf_amounts(obj)

        obj.deducted_from_employee = pf_amounts['deduction']
        obj.contribution_from_company = pf_amounts['addition']
        obj.total_fund_deducted = (
            (pf_amounts['addition'] or 0.0) + (pf_amounts['deduction'] or 0.0)
        )
        return obj

    @classmethod
    def get_exported_file_content(cls, data, title, columns, extra_content, description=None, **kwargs):

        organization = extra_content["organization"]
        payroll = extra_content["payroll"]

        bold_font = Font(bold=True)
        wb = Workbook()

        ws = wb.active
        ws.title = title

        # row_1
        row_no = 1
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=8)
        cell = ws.cell(row=row_no, column=1, value="EMPLOYEES PROVIDENT FUND")
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

        # row 3
        row_no = 3
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=8)
        cell = ws.cell(row=row_no, column=1,
                       value=f"Organization: {organization.name}")
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

        details = cls.get_details(organization, payroll)

        row_no = 4
        for detail in details:
            ws.cell(row=row_no, column=1, value=detail[0])
            ws.cell(row=row_no, column=2, value=detail[1])
            row_no += 1

        row_no = 14
        ExcelExport.fill_headings(
            ["A", "B", "C16", "D16", "E", "F16", "G", "H"],
            row_no, ws
        )

        row_no = 15
        columns, mapping = ExcelExport.get_column_mapping(columns)
        ExcelExport.fill_headings(
            columns, row_no, ws, mapping)

        row_no = 16
        ExcelExport.fill_data(
            columns, row_no,
            data,
            ws,
            prepare_export_object=cls.prepare_export_object,
            prepare_export_object_context={}
        )

        return ContentFile(save_virtual_workbook(wb))

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)

        response.data["details"] = dict(self.get_details(self.organization, self.payroll))

        return response


class PayrollApprovePermission(BasePermission):
    def has_permission(self, request, view):
        return PayrollApproval.objects.filter(user=request.user).exists()


class ReportSettingViewSetMixin(OrganizationMixin, ListCreateViewSetMixin):
    permission_classes = [
        permission_factory.build_permission(
            "ReportPermission",
            methods={
                'get': [
                    PAYROLL_SETTINGS_PERMISSION,
                    GENERATE_PAYROLL_PERMISSION,
                    PAYROLL_REPORT_PERMISSION,
                    PAYROLL_READ_ONLY_PERMISSIONS
                ],
                'post': [
                    ALL_PAYROLL_PERMISSIONS,
                    PAYROLL_SETTINGS_PERMISSION,
                    GENERATE_PAYROLL_PERMISSION
                ]
            }
        ) | PayrollApprovePermission
    ]

    def get_queryset(self):
        return super().get_queryset().filter(organization=self.get_organization())

    def list(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_queryset().first())
        return Response(serializer.data)

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['organization'] = self.get_organization()
        return ctx


class PayrollCollectionDetailSettingViewSet(ReportSettingViewSetMixin):
    serializer_class = PayrollCollectionDetailReportSettingSerializer
    queryset = PayrollCollectionDetailReportSetting.objects.all()


class PayrollDifferenceDetailSettingViewSet(ReportSettingViewSetMixin):
    serializer_class = PayrollDifferenceDetailReportSettingSerializer
    queryset = PayrollDifferenceDetailReportSetting.objects.all()


class PayrollSSFReportSettingViewSet(ReportSettingViewSetMixin):
    serializer_class = SSFReportSettingSerializer
    queryset = SSFReportSetting.objects.all()


class PayrollDisbursementReportSettingViewSet(ReportSettingViewSetMixin):
    serializer_class = DisbursementReportSettingSerializer
    queryset = DisbursementReportSetting.objects.all()


class PayrollTaxReportSettingViewSet(ReportSettingViewSetMixin):
    serializer_class = TaxReportSettingSerializer
    queryset = TaxReportSetting.objects.all()


class PayrollExtraHeadingReportSettingViewSet(ReportSettingViewSetMixin):
    serializer_class = PayrollExtraHeadingSettingSerializer
    queryset = ExtraHeadingReportSetting.objects.all()

    def list(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_queryset(), many=True)
        return Response(serializer.data)


class PayrollSSFReportViewSet(
    BackgroundExcelExportMixin,
    EmployeePayrollMixin,
    ReadOnlyModelViewSet
):
    serializer_class = SSFReportSerializer
    permission_classes = [
        permission_factory.build_permission(
            "PFReportPermission",
            allowed_to=[
                PAYROLL_REPORT_PERMISSION,
                GENERATE_PAYROLL_PERMISSION,
                ALL_PAYROLL_PERMISSIONS,
                PAYROLL_SETTINGS_PERMISSION,
                PAYROLL_READ_ONLY_PERMISSIONS
            ]
        )
    ]
    filter_backends = (OrderingFilterMap,)
    ordering_fields_map = {
        "full_name": (
            "employee__first_name",
            "employee__middle_name",
            "employee__last_name",
        )
    }

    def get_export_type(self):
        return 'Export SSF of payroll id ' + str(self.payroll.id)

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        return response

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['organization'] = self.organization
        return ctx

    @cached_property
    def headings(self):
        setting = SSFReportSetting.objects.filter(organization=self.organization).first()
        if not setting:
            return []
        return setting.headings.values_list('name', 'id')

    def get_export_fields(self):
        fields = {
            'S.N.': '#SN',
            'Employee Name': 'employee.full_name',
            'Username': 'employee.username',
            'Designation': 'user_experience_package_slot.user_experience.job_title.title',
            'SSF ID No.': 'employee.legal_info.ssfid',
            **{
                heading_name: f'amount_of_{heading_id}' for heading_name, heading_id in self.headings
            }
        }
        return fields

    def get_extra_export_data(self):
        ctx = super().get_extra_export_data()
        ctx['prepare_export_object_context'] = {'ssf_headings_id': self.headings}
        return ctx

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        """
        this is a method that will prepare export object. It will get object instance as its
        first parameter and should return object or dict.
        :param obj: Instance (A row) that will be exported
        :return: prepared object
        """
        row_report_records = ReportRowRecord.objects.filter(
            employee_payroll=obj,
            heading__in=[heading_id for heading_name, heading_id in kwargs.get('ssf_headings_id')],
            employee_payroll__payroll__status__in=[GENERATED, CONFIRMED]
        )

        obj.ssf_headings = getattr(obj, 'ssf_headings', [])
        for row_report_record in row_report_records:
            heading = row_report_record.heading
            heading_id = heading.id
            heading_name = heading.name
            setattr(obj,
                    f'amount_of_{heading_id}',
                    row_report_record.amount
                    )
        return obj


class PayrollDisbursementReportViewSet(
    BackgroundExcelExportMixin,
    EmployeePayrollMixin,
    ReadOnlyModelViewSet
):
    serializer_class = DisbursementReportSerializer
    permission_classes = [
        permission_factory.build_permission(
            "PFReportPermission",
            allowed_to=[
                PAYROLL_REPORT_PERMISSION,
                GENERATE_PAYROLL_PERMISSION,
                ALL_PAYROLL_PERMISSIONS,
                PAYROLL_SETTINGS_PERMISSION,
                PAYROLL_READ_ONLY_PERMISSIONS
            ]
        )
    ]
    filter_backends = (OrderingFilterMap, DjangoFilterBackend, FilterMapBackend)
    ordering_fields_map = {
        "full_name": (
            "employee__first_name",
            "employee__middle_name",
            "employee__last_name",
        )
    }
    filter_map = {
        'bank': 'employee__userbank__bank__slug'
    }

    def get_export_type(self):
        return 'Export Disbursement of payroll id ' + str(self.payroll.id)

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['organization'] = self.organization
        ctx['headings'] = self.headings
        return ctx

    @cached_property
    def headings(self):
        setting = DisbursementReportSetting.objects.filter(organization=self.organization).first()
        return [] if not setting else setting.headings.all()

    def get_export_fields(self):
        export_fields = {
            'S.N.': '#SN',
            'Employee Name': 'employee.full_name',
            'Username': 'employee.username',
            'Designation': 'user_experience_package_slot.user_experience.job_title.title',
            'Bank Name': 'employee.userbank.bank.name',
            'Bank Account Number': 'employee.userbank.account_number',
            'Bank Branch': 'employee.userbank.branch',
            **{
                heading.name: f'amount_of_{heading.id}' for heading in self.headings
            }
        }
        if self.request.query_params.get('bank'):
            for field in ('Username', 'Designation', 'Bank Branch'):
                export_fields.pop(field)
        return export_fields

    def get_payroll_info(self):
        payroll = get_object_or_404(Payroll, id=self.kwargs.get('payroll_id'))
        from_date = payroll.from_date
        to_date = payroll.to_date
        month = FiscalYearMonth.objects.filter(start_at=from_date, end_at=to_date).first()
        if month:
            subject = f"Salary/Wages for the Month of {month.display_name} {month.fiscal_year.name}"
        else:
            subject = f"Salary/Wages from {from_date} to {to_date}"
        return subject

    def get_disbursement_narration(self):
        bank = self.request.query_params.get('bank')
        if not bank:
            return
        org_bank = OrganizationBank.objects.filter(bank__slug=bank).first()
        if org_bank:
            org_bank_account = org_bank.account_number
            return f'From A/C No. {org_bank_account} to the following  account name and numbers with the respective amount along with the Credit Narration as "{self.get_payroll_info()}".'
    def get_extra_export_data(self):
        ctx = super().get_extra_export_data()
        ctx['prepare_export_object_context'] = {'disbursement_headings': self.headings,
                                                'report_narration': self.get_disbursement_narration()
                                                }
        if self.request.query_params.get('bank'):
            ctx['prepare_export_object_context']['grand_total_data'] = {
                    'footer_offset': 5,
                    'footer_color': 'CFD8DC',
                    'footer_text': 'Total Salary in NRs',
                }
        return ctx

    def get_export_description(self, *args, **kwargs):
        bank = self.request.query_params.get('bank')
        if not bank:
            return
        subject = self.get_payroll_info()
        org_bank = OrganizationBank.objects.filter(bank__slug=bank).first()
        bank_name = org_bank.bank.name if org_bank else " "
        bank_branch = org_bank.branch if org_bank else " "
        return ["To,",
                "The Branch Manager,",
                f"{bank_name},",
                f"{bank_branch},",
                f"Subject: {subject}",
                "Dear Sir/Madam,"
        ]

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        get_heading_details(obj, kwargs.get('disbursement_headings'), set_attribute=True)
        return obj

    @classmethod
    def get_exported_file_content(cls, data, title, columns, extra_content,
                                  description=None, **kwargs):
        organization = extra_content.get('organization')
        wb = PayrollDisbursementExcel.process(
            data,
            title=title,
            columns=columns,
            description=description,
            prepare_export_object=cls.prepare_export_object,
            prepare_export_object_context=extra_content.get('prepare_export_object_context'),
            freeze_first_column=cls.export_freeze_first_column,
            organization=organization
        )
        return ContentFile(save_virtual_workbook(wb))


class PayrollGeneralReportViewSet(
    BackgroundTableExportMixin,
    DateRangeParserMixin,
    OrganizationCommonsMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    """
    API for general report

    filters
        start_date, end_date, user_ids, heading_ids

    ordering
        ordering=full_name
        ordering=-full_name
        ordering=heading_id
        ordering=-heading_id
    """
    permission_classes = [
        permission_factory.build_permission(
            "PayrollGeneralReportPermission",
            allowed_to=[PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION]
        )
    ]
    serializer_class = PayrollGeneralReportSerializer
    raise_on_invalid_dates = True
    queryset = Employee.objects.all()
    notification_permissions = [PAYROLL_REPORT_PERMISSION]

    filter_backends = (FilterMapBackend, OrderingFilterMap,)
    ordering_fields_map = {'full_name': ('first_name', 'middle_name', 'last_name')}
    filter_map = {
        'job_title': 'detail__job_title__slug',
        'branch': 'detail__branch__slug',
        'employee_level': 'detail__employment_level__slug',
        'division': 'detail__division__slug',
    }

    export_type = "Payroll General Export"
    export_freeze_first_column = True

    def get_export_fields(self):
        # return [
        #     {
        #         "name": "user",
        #         "title": "Employee",
        #         "fields": (
        #             {
        #                 "name": "full_name",
        #                 "title": "Name"
        #             },
        #             {
        #                 "name": "job_title",
        #                 "title": "Job Title"
        #             },
        #         )
        #     },
        #     {
        #         "name": "records",
        #         "title": "Headings",
        #         "fields": [
        #             {
        #                 "name": heading.id,
        #                 "title": heading.name
        #             } for heading in self.get_headings()
        #         ]
        #     },
        # ]
        results = [
                   {"name": "username", "title": "Username"},
                   {"name": "user.full_name", "title": "Name"},
                   {"name": "user.job_title", "title": "Designation"},
                   {"name": "user.division.name", "title": "Division"},
                   {"name": "branch.name", "title": "Branch"},
                   {"name": "joined_date", "title": "Joined Date"},
                   {"name": "resigned_date", "title": "Resigned Date"},
                   {"name": "marital_status", "title": "Marital Status"}]
        results += [
            {
                "name": f"records.{heading.id}",
                "title": heading.name
            } for heading in self.get_headings()
        ]
        return results

    def get_extra_export_data(self):
        data = super().get_extra_export_data()
        data["serializer_context"] = {'heading_ids': self.heading_ids}
        return data

    @classmethod
    def get_workbook_to_export_file_content(cls, data, title, columns, extra_content,
                                            description=None, **kwargs):

        bold_font = Font(bold=True)

        wb = super().get_workbook_to_export_file_content(
            data, title, columns, extra_content, **kwargs
        )

        ws = wb.active

        # Adding sum to last row
        # ws.max_row provides total number of row in the worksheet
        last_row_count = ws.max_row
        start_column = 1
        end_column = 8
        ws.merge_cells(
            start_row=last_row_count + 1, start_column=start_column,
            end_row=last_row_count + 1, end_column=end_column
        )
        cell = ws.cell(row=last_row_count + 1, column=1, value="Total Sum")
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
        data_row = [cell.row for cell in ws['A'] if cell.value == 'Username']

        data_row_begin = data_row[0] + 1
        dynamic_column_count = len(columns) - end_column

        def formula(column):
            return "=SUM({}:{})".format(
                f"{get_column_letter(column)}{data_row_begin}",
                f"{get_column_letter(column)}{last_row_count}",
            )

        for count in range(dynamic_column_count):
            count += 1
            cell = ws.cell(
                row=last_row_count + 1, column=end_column + count,
                value=formula(end_column + count)
            )
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center')

        return wb

    def get_employees(self, queryset):
        user_id_from_query_param = self.request.query_params.get('user_ids')
        if user_id_from_query_param:
            user_id_from_query_param = validate_comma_separated_integers(
                user_id_from_query_param,
                message={"detail": "Invalid user_ids. Expected list of integers Eg. 1,2,3"}
            )
            list_of_user_ids = list(map(int, user_id_from_query_param.split(",")))
            queryset = queryset.filter(id__in=list_of_user_ids)

        return queryset.filter(
            detail__organization=self.organization
        ).select_essentials()

    def get_headings(self):
        heading_ids_from_query_param = self.request.query_params.get('heading_ids')
        if not heading_ids_from_query_param:
            raise ValidationError({"detail": "Please send list of heading_ids."})
        heading_ids_from_query_param = validate_comma_separated_integers(
            heading_ids_from_query_param,
            message={"detail": "Invalid value of heading_ids. Expected list of integers Eg. 1,2,3"}
        )

        list_of_headings = list(map(int, heading_ids_from_query_param.split(",")))
        return Heading.objects.filter(id__in=list_of_headings, organization=self.organization)

    @cached_property
    def heading_ids(self):
        return list(self.get_headings().values_list('id', flat=True))

    def get_heading_annotate(self):
        return {
            str(heading_id): Coalesce(Sum(
                'employee_payrolls__report_rows__amount',
                filter=Q(
                    employee_payrolls__report_rows__heading=heading_id,
                    employee_payrolls__payroll__to_date__range=self.get_parsed_dates()
                )
            ), 0.0)
            for heading_id in self.heading_ids
        }

    def get_queryset(self):
        return self.get_employees(self.queryset).annotate(**self.get_heading_annotate())

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        ordering = self.request.query_params.get('ordering')
        allowed_heading_ordering = list(map(str, self.heading_ids)) + \
            [f"-{str(heading_id)}" for heading_id in self.heading_ids]

        if ordering in allowed_heading_ordering:
            queryset = queryset.order_by(ordering)

        return queryset

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        if self.organization:
            ctx["heading_ids"] = self.heading_ids
        return ctx

    def get_frontend_redirect_url(self):
        return f'/admin/{self.organization.slug}/payroll/reports/general-info-report'


class PackageWiseSalaryViewSet(BackgroundExcelExportMixin, ListViewSetMixin, OrganizationMixin):
    """
    ## Api for the package wise salary for all organization-related current employees.

    ## The amount per heading ID is available in this API.

    ## The heading IDs can be found at

    * `/api/v1/payroll/headings/?organization__slug=slug`.

    Filters are listed below:

    * employee Name `search`
    * Division filter as `division`.
    * Job Title filter as `job_title`.
    * Employment Level filter as `employment_level`.
    * Heading list can be sent as `heading_ids`. Defaults to all headings.

    Available Ordering filters:

    * `full_name`
    * `yos` for Years of Service Ordering
    * `current_step`
    * Individual heading ordering through heading Id from this API:
    * `/api/v1/payroll/headings/?organization__slug=slug`.

    """
    queryset = get_user_model().objects.all().select_essentials()
    permission_classes = [
        permission_factory.build_permission(
            "PayrollPackageWiseSalaryReportPermission",
            allowed_to=[PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION]
        )
    ]
    filter_backends = (
        FilterMapBackend,
        IStartsWithIContainsSearchFilter,
        NullsAlwaysLastOrderingFilter
    )
    search_fields = (
        'first_name', 'middle_name', 'last_name'
    )
    filter_map = {
        'job_title': 'detail__job_title__slug',
        'branch': 'detail__branch__slug',
        'division': 'detail__division__slug',
        'employment_level': 'detail__employment_level__slug',
        'employment_status': 'detail__employment_status__slug',
        'user_id': 'id'
    }
    serializer_class = PackageWiseSalarySerializer
    notification_permissions = [PAYROLL_REPORT_PERMISSION]

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        if self.organization:
            ctx["heading_ids"] = self.heading_ids
        return ctx

    @cached_property
    def heading_ids(self):
        heading_ids_from_query_param = self.request.query_params.get('heading_ids')
        if not heading_ids_from_query_param:
            return list(
                map(
                    str,
                    Heading.objects.filter(
                        organization=self.organization
                    ).values_list('id', flat=True)
                )
            )
        heading_ids_from_query_param = validate_comma_separated_integers(
            heading_ids_from_query_param,
            message={"detail": "Invalid value of heading_ids. Expected list of integers Eg. 1,2,3"}
        )
        list_of_headings = list(map(str, heading_ids_from_query_param.split(",")))
        return list_of_headings

    def get_queryset(self):
        base_qs = super().get_queryset().filter(detail__organization=self.organization)
        selected_employees = self.request.query_params.get('selected_employees', '').split(',')
        selected_employees = set(map(int,filter(str.isdigit,selected_employees)))
        if selected_employees:
            base_qs = base_qs.filter(id__in=selected_employees)
        if self.action == 'export':
            return base_qs
        return self.annotate_headings(base_qs, self.heading_ids)

    def get_ordering_fields_map(self):
        all_headings_map = dict(
            (heading, heading) for heading in self.heading_ids
        )
        return {
            'full_name': (
                'first_name', 'middle_name', 'last_name'
            ),
            'joined_date': 'detail__joined_date',
            'yos': 'yos',
            'division': 'detail__division__name',
            'current_step': 'current_step',
            **all_headings_map
        }

    # Begin Export mechanics

    @staticmethod
    def annotate_headings(base_qs, heading_ids):
        all_headings_amount = {
            heading: Subquery(
                ReportRowUserExperiencePackage.objects.filter(
                    package_heading__heading_id=heading,
                    package_slot=OuterRef('active_slot')
                ).values('package_amount')[:1]
            ) for heading in heading_ids
        }
        ret_qs = base_qs.current().annotate(
            active_slot=Subquery(
                UserExperiencePackageSlot.objects.filter(
                    active_from_date__lt=get_today(),
                    user_experience__user=OuterRef('pk'),
                    user_experience__is_current=True,
                ).order_by('-active_from_date').values('pk')[:1]
            )
        ).annotate(
            **all_headings_amount
        )
        return ret_qs.annotate(
            yos=ExpressionWrapper(
                Extract(Value(timezone.now().date()) - F('detail__joined_date'), 'Days') / 365.25,
                output_field=FloatField()
            )
        ).annotate(
            current_step=Subquery(
                UserExperience.objects.exclude(
                    upcoming=True
                ).filter(
                    is_current=True,
                    user_id=OuterRef('id'),
                ).order_by(
                    '-end_date'
                ).values('current_step')[:1]
            )
        )

    def get_export_fields(self):
        return {
            'Employee Name': 'full_name',
            'Username': 'username',
            'Date of Join': 'detail.joined_date',
            'Years of Service': 'yos',
            'Division': 'detail.division.name',
            'Current Step': 'current_step',
            **{
                heading_.name: str(heading_.id) for heading_ in Heading.objects.filter(
                    organization=self.organization
                ).filter(
                    id__in=self.heading_ids
                )
            }
        }

    def get_export_type(self):
        return f"package_wise_salary_{self.organization.id}"

    def get_export_name(self):
        return "Package wise salary"

    @classmethod
    def get_exported_file_content(cls, queryset, title, columns, extra_content, description=None, **kwargs):
        heading_ids = extra_content.get('heading_ids')
        queryset = cls.annotate_headings(queryset, heading_ids)
        return super().get_exported_file_content(
            queryset, title, columns, extra_content, description=None, **kwargs
        )

    def get_extra_export_data(self):
        ctx = super().get_extra_export_data()
        ctx.update({
            'heading_ids': self.heading_ids
        })

        return ctx

    @staticmethod
    def prepare_export_object(obj, **kwargs):
        obj.yos = PackageWiseSalarySerializer.get_years_of_service(obj).get('years')
        return obj

    def get_frontend_redirect_url(self):
        return f'/admin/{self.organization.slug}/payroll/reports/package-wise-salary'
    # /End Export mechanics


class HeadingWiseExpenditureViewSet(
    OrganizationCommonsMixin,
    OrganizationMixin,
    ListViewSetMixin
):
    """
    This report will give monthly amount spent on each heading.

    send `?exclude_zeros=true` to exclude headings containing 0.0 value
    """
    permission_classes = [
        permission_factory.build_permission(
            "HeadingWiseSalaryExpenditureReportPermission",
            allowed_to=[PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION]
        )
    ]
    queryset = Heading.objects.all().filter(is_hidden=False)
    serializer_class = create_dummy_serializer({
        'id': serializers.ReadOnlyField(),
        'heading': serializers.ReadOnlyField(source='name'),
        'amount': serializers.ReadOnlyField()
    })
    filter_backends = (SearchFilter, OrderingFilterMap)
    search_fields = ('name',)
    ordering_fields_map = {
        'heading': 'name',
        'order': 'order'
    }
    ordering = 'name'

    def get_queryset(self):
        active_package_slots = set(UserExperience.objects.filter(
            is_current=True,
            organization=self.get_organization()
        ).annotate(
            active_package_slot_id=Subquery(
                UserExperiencePackageSlot.objects.filter(
                    user_experience=OuterRef('pk'),
                    active_from_date__lte=get_today()
                ).order_by('-active_from_date').values('id')[:1]
            )
        ).values_list('active_package_slot_id', flat=True))

        return super().get_queryset().annotate(
            amount=Coalesce(
                Sum(
                    'packageheading__package_rows__package_amount',
                    filter=Q(
                        packageheading__package_rows__package_slot_id__in=active_package_slots
                    )
                ), 0.0
            )
        )

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)

        if self.request.query_params.get('exclude_zeros') == 'true':
            queryset = queryset.exclude(amount=0.0)

        return queryset


class YearlyPayslipReport(UserMixin, RetrieveViewSetMixin, OrganizationMixin):
    """
    Yearly Payslip Details for particular user

    `/api/v1/payroll/<organization_slug>/reports/yearly-payslip/<user_id>?fiscal_year=1`

    if fiscal_year is not sent in query params, by default current fiscal year of organization
    will be taken


    """

    queryset = Employee.objects.all()
    serializer_class = EmployeeThinSerializer
    lookup_url_kwarg = 'user_id'

    permission_classes = [
        permission_factory.build_permission(
            "HeadingWiseSalaryExpenditureReportPermission",
            allowed_to=[PAYROLL_REPORT_PERMISSION, GENERATE_PAYROLL_PERMISSION,
                        HAS_PERMISSION_FROM_METHOD]
        )
    ]

    def has_user_permission(self):
        return self.user == self.request.user

    def get_object(self):
        return self.user

    def get_fiscal_year(self):
        fiscal_id = self.request.query_params.get('fiscal_year', None)
        if fiscal_id:
            try:
                fiscal_id = int(fiscal_id)
            except (TypeError, ValueError):
                raise serializers.ValidationError({'fiscal_year': ['This value must be an integer']})

            fiscal_year = FiscalYear.objects.filter(
                organization=self.organization,
                category=GLOBAL,
                applicable_from__lte=get_today()
            ).filter(id=fiscal_id).first()

            if not fiscal_year:
                raise serializers.ValidationError(
                    {'fiscal_year': ['Selected fiscal year was not found. Perhaps you selected'
                                     ' future fiscal year or it does not exist.']})

        else:
            fiscal_year = FiscalYear.objects.current(self.organization, category=GLOBAL)
            if not fiscal_year:
                raise serializers.ValidationError(
                    {'fiscal_year': ['Active Fiscal Year not found.']})
        return fiscal_year

    @staticmethod
    def get_dismiss_date(employee):
        experience = employee.current_experience
        if not experience:
            # if not current then last
            experience = employee.user_experiences.order_by('start_date').last()
        return get_dismiss_date(employee, experience)

    @staticmethod
    def get_display_name(record, months):
        from_date = record.get('from_date') or record.get('employee_payroll__payroll__from_date')
        to_date = record.get('to_date') or record.get('employee_payroll__payroll__to_date')

        return ",".join(
            month["display_name"]
            for month in
            filter(
                lambda m: m['start_at'] < to_date and m['end_at'] > from_date,
                months
            )
        )

    def retrieve(self, request, *args, **kwargs):
        employee = self.get_object()
        fiscal_year = self.get_fiscal_year()

        dismiss_date = self.get_dismiss_date(employee)
        end_date = min(fiscal_year.applicable_to, dismiss_date) \
            if dismiss_date else fiscal_year.applicable_to

        months = list(
            fiscal_year.fiscal_months.filter(
                start_at__lte=end_date  # to exclude months starting after dismiss date
            ).all().values('display_name', 'start_at', 'end_at')
        )
        employee_payrolls = EmployeePayroll.objects.filter(
            employee=employee,
            payroll__from_date__gte=fiscal_year.applicable_from,
            payroll__from_date__lte=end_date,
            payroll__status=CONFIRMED
        )
        last_paid = nested_getattr(
            employee_payrolls.order_by('payroll__to_date').last(),
            'payroll.to_date',
            default=None
        )
        if (
            not last_paid and
            # not paid in past year
            not fiscal_year.applicable_from <= get_today() <= fiscal_year.applicable_to
        ):
            raise serializers.ValidationError({'fiscal_year': ['No payroll records found']})

        # ---------------------------------------------------------------------------------------#
        #                  PAST CALCULATION

        payroll_start_fiscal_year = OrganizationPayrollConfig.objects.filter(
            organization=self.organization
        ).first()

        show_zero_amount_row_records = False
        if payroll_start_fiscal_year:
            show_zero_amount_row_records = payroll_start_fiscal_year.display_heading_with_zero_value

        fil = Q(employee_payroll__in=employee_payrolls)
        if not show_zero_amount_row_records:
            fil = fil & ~Q(amount=0)

        report_rows = list(ReportRowRecord.objects.filter(
            fil
        ).order_by('heading__name', 'employee_payroll__payroll__from_date').values(
            'heading__id',
            'heading__name',
            'employee_payroll__payroll__from_date',
            'employee_payroll__payroll__to_date',
            'amount'
        ))

        grouped_records = itertools.groupby(
            report_rows, key=lambda x: x['heading__name']
        )

        paid_slots = list(ReportRowRecord.objects.filter(
            employee_payroll__in=employee_payrolls,
        ).order_by(
            'employee_payroll__payroll__from_date', 'employee_payroll__payroll__to_date'
        ).distinct(
            'employee_payroll__payroll__from_date', 'employee_payroll__payroll__to_date'
        ).values(
            'employee_payroll__payroll__from_date', 'employee_payroll__payroll__to_date')
        )

        month_slots = [
            {
                'from_date': record['employee_payroll__payroll__from_date'],
                'to_date': record['employee_payroll__payroll__to_date'],
                'display_name': self.get_display_name(record, months)
            }
            for record in paid_slots
        ]

        result = {
            group_name: [
                {
                    'from_date': record['employee_payroll__payroll__from_date'],
                    'to_date': record['employee_payroll__payroll__to_date'],
                    'amount': record['amount'],
                    'display_name': self.get_display_name(record, months)
                }
                for record in group
            ] for group_name, group in grouped_records
        }

        # ----------------------------------------------------------------------------------------#
        #                    FUTURE CALCULATION

        if last_paid and last_paid < end_date:
            simulate_from = last_paid + timezone.timedelta(days=1)
            simulate_to = end_date

            remaining_months = [
                {
                    'display_name': month['display_name'],
                    'start_at': max(month['start_at'], simulate_from),  # max(start_at, joined_date)
                    'end_at': min(month['end_at'], end_date)  # if dismiss_date, min(end_at, dismiss_date)
                }
                for month in filter(lambda m: m['end_at'] > last_paid, months)
            ]
            month_slots += [
                {
                    'from_date': month['start_at'],
                    'to_date': month['end_at'],
                    'display_name': month['display_name']
                }
                for month in remaining_months
            ]

            appoint_date = get_appoint_date(employee, payroll_start_fiscal_year)

            package_slot = UserExperiencePackageSlot.objects.filter(
                user_experience__user=employee,
                active_from_date__lte=simulate_from
            ).order_by('active_from_date').select_related('user_experience').last()
            if package_slot:
                calculation = PackageSalaryCalculator(
                    user_experience=package_slot.user_experience,
                    employee=employee,
                    datework=FY(organization=self.organization),
                    from_date=simulate_from,
                    to_date=simulate_to,
                    salary_package=package_slot.package,
                    appoint_date=appoint_date,
                    month_days_setting='ORGANIZATION_CALENDAR',
                    package_assigned_date=package_slot.active_from_date,
                    update_remaining_rebate_row=True
                )
                remaining_report = YearlyPayslipReportForRemainingMonths(
                    employee=employee,
                    organization=self.organization,
                    remaining_months=remaining_months,
                    include_holiday_off_day=calculation.payroll_config.include_holiday_offday_in_calculation
                )
                for row in calculation.payroll.rows:
                    heading = row.heading
                    heading_name = heading.name
                    amount = row.amount
                    duration_unit = row.package_heading.duration_unit
                    if not (show_zero_amount_row_records or amount):
                        continue
                    if duration_unit == 'Yearly':
                        payment_date = None
                        yearly_heading_detail = YearlyHeadingDetail.objects.filter(
                                heading=row.package_heading.heading,
                                fiscal_year=fiscal_year
                        ).first()
                        if yearly_heading_detail:
                            payment_date = getattr(yearly_heading_detail, 'date')

                        if payment_date:
                            payment_months = list(filter(
                                lambda m: m['start_at'] <= payment_date <= m['end_at'],
                                remaining_months
                            ))
                            if payment_months:
                                payment_month = payment_months[0]
                                result[heading_name] = [{
                                    'from_date': payment_month['start_at'],
                                    'to_date': payment_month['end_at'],
                                    'amount': amount,
                                    'display_name': payment_month['display_name']
                                }]
                    else:
                        if heading.duration_unit == "Monthly":
                            if has_heading_rule(heading, "__USER_VOLUNTARY_REBATE__"):
                                result[heading_name] = result.get(heading_name, []) + \
                                                remaining_report.get_monthly_rebate_amounts(row)
                                continue
                            if has_heading_rule(heading, "__TOTAL_LOST_HOURS__"):
                                result[heading_name] = result.get(heading_name, []) + \
                                                  remaining_report.get_monthly_lost_hours()
                                continue
                            if has_heading_rule(heading, "__TOTAL_HOLIDAY_COUNT__"):
                                result[heading_name] = result.get(heading_name, []) + \
                                                  remaining_report.get_monthly_holiday_count()
                                continue
                            if has_heading_rule(heading, "__TOTAL_WORKING_DAYS__"):
                                result[heading_name] = result.get(heading_name, []) + \
                                                  remaining_report.get_monthly_working_days()
                                continue
                        monthly_amount = amount / len(remaining_months)
                        result[heading_name] = result.get(heading_name, []) + [
                            {
                                'from_date': month['start_at'],
                                'to_date': month['end_at'],
                                'amount': monthly_amount,
                                'display_name': month['display_name']
                            }
                            for month in remaining_months
                        ]

        parsed_report_rows, earning_rows, deduction_rows = self.get_report_rows_earning_and_deduction_rows(result)

        return Response({
            'employee': EmployeeThinSerializer(employee).data,
            'last_paid': last_paid,
            'dismiss_date': dismiss_date,
            'month_slots': month_slots,
            'report_rows': parsed_report_rows,
            'earning_rows': earning_rows,
            'deduction_rows': deduction_rows
        })

    def parse_results_for_response(self, results):
        heading_type_map = {
            heading['name']: (heading['type'],heading['id'], heading['order'])
            for heading in
            Heading.objects.filter(
                organization=self.organization
            ).values('id', 'name', 'type', 'order')
        }
        for heading, detail in results.items():
            heading_type, heading_id, order = heading_type_map[heading]
            yield {
                'heading_id': heading_id,
                'heading': heading,
                'heading_type': heading_type,
                'order': order,
                'payment_months': detail
            }

    def get_report_rows_earning_and_deduction_rows(self, result):
        earning_headings = list(PayslipReportSetting.objects.filter(
            organization=self.organization,
            category="Earning"
        ).values_list("headings", flat=True))

        deduction_headings = list(PayslipReportSetting.objects.filter(
            organization=self.organization,
            category="Deduction"
        ).values_list("headings", flat=True))

        parsed_report_rows = list(self.parse_results_for_response(result))
        earning_rows = [row for row in parsed_report_rows if row['heading_id'] in earning_headings]
        deduction_rows = [row for row in parsed_report_rows if row['heading_id'] in deduction_headings]

        return parsed_report_rows, earning_rows, deduction_rows

class BackdatedCalculationReportViewSet(
        OrganizationMixin,
        ListViewSetMixin
):
    permission_classes = [ViewPayrollReportPermission]
    serializer_class = BackdatedCalculationSerializer

    def get_queryset(self):
        package_slot_id = self.kwargs.get('package_slot_id')
        return BackdatedCalculation.objects.filter(
            package_slot__id = package_slot_id
        ).order_by('heading__order')
