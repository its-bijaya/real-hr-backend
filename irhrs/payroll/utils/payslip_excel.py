from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from irhrs.payroll.models.payroll import TEMPLATE_1, TEMPLATE_2

all_side_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GRAY = PatternFill("solid", fgColor="808080")
CELADON = PatternFill("solid", fgColor="AFE1AF")
LIGHT_BLUE = PatternFill("solid", fgColor="E3F2FD")
BLUE_GRAY = PatternFill("solid", fgColor="334d6d")
ROW_COLOR = PatternFill("solid", fgColor="e0e0e0")

field_map = {
    "code": "Emp code",
    "approved_date": "Payment Date",
    "full_name": "Name",
    "job_title": "Job Title",
    "branch": "Branch",
    "pan_number": "Pan number",
    "ssfid": "SSF ID",
    "cit_number": "Cit number",
    "pf_number": "Pf number",
    "employee_level": "Designation",
    "division": "Division",
    "bank_name": "Bank Name",
    "branch": "Branch",
    "account_number": "Account Number",
    "marital_status": "Marital status",
    "joined_date": "DOJ",
    "username": "Username",
    "working_days": "Working Days",
    "worked_days": "Worked Days",
    "absent_days": "Absent Days",
    "leave_days": "Leave Days",
    "leave_days_on_workdays": "Leave Days On Workdays",
    "unpaid_leave_days": "Leave without Pay",
    "paid_days": "Paid Days",
    "days_deduction_from_penalty": "Days Deduction from Penalty",
}


def calculate_sum(income):
    return round(sum(map(lambda x: x["amount"], income)), 2)


class ExcelMixin:

    def add_table(self, name, ref):
        tab = Table(displayName=name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        self.ws.add_table(tab)

    def merge_cells(
        self, row_offset=0, column_offset=0, start_row=None, start_column=None
    ):
        start_row = start_row or self.row
        start_column = start_column or self.column
        start_coordinate = get_column_letter(start_column) + str(start_row)
        end_coordinate = get_column_letter(start_column + column_offset) + str(
            start_row + row_offset
        )
        ref = f"{start_coordinate}:{end_coordinate}"
        self.ws.merge_cells(ref)

class ExcelPayslip(ExcelMixin):

    def __init__(self, data, template=None):
        self.template = template or TEMPLATE_2

        self.data = data
        self.workbook = Workbook()
        self.ws = self.workbook.active

        self.attendance_start_row = 1
        self.start_earning_deduction_row = 1

        self.row = 1
        self.column = 1

        self.total_earnings = calculate_sum(self.data["earnings"])
        self.total_deductions = calculate_sum(self.data["deductions"])
        self.net_pay = round(self.total_earnings - self.total_deductions, 2)

    def append_payroll_date(self):
        from_date = self.data["from_date"]
        to_date = self.data["to_date"]
        payroll_date = f"{from_date:%b %d, %Y} - {to_date:%b %d, %Y}"
        self.ws.append(("Payroll Date:", payroll_date))

        cell1 = self.ws.cell(self.row, self.column, value="Payroll Date:")
        cell1.fill = GRAY
        cell1.font = Font(bold=True)

        self.ws.cell(self.row, self.column + 1, value=payroll_date)
        self.row += 2

    @property
    def payment_information(self):
        employee = self.data["employee"]
        legal_info = employee.get("legal_info")
        bank = self.data["employee"].get("userbank")
        approved_date = self.data.get("approved_date")
        return {
            "approved_date": approved_date and f"{approved_date:%Y-%m-%d}",
            "bank_name": bank and bank["bank_name"],
            "branch": bank and bank.get("branch"),
            "account_number": bank and bank["account_number"],
            "pan_number": legal_info and legal_info.get("pan_number"),
            "pf_number": legal_info and legal_info.get("pf_number"),
            "cit_number": legal_info and legal_info.get("cit_number"),
            "ssfid": legal_info and legal_info.get("ssfid"),
        }

    def append_employee_information(self):
        employee = self.data["employee"]
        division = employee.get("division")
        employee = {
            **employee,
            **self.payment_information,
            "joined_date": employee.get("joined_date"),
            "marital_status": employee.get("marital_status"),
            "username": employee.get("username"),
            "division": division and division["name"],
        }
        employee_fields = [
            ("code", "pan_number"),
            ("full_name", "pf_number"),
            ("job_title", "cit_number"),
            ("branch", "employee_level"),
            ("division", "bank_name"),
            ("branch", "account_number"),
        ]
        if self.template == TEMPLATE_2:
            employee_fields = [
                ("full_name", "code"),
                ("job_title", "username"),
                ("division", "joined_date"),
                ("marital_status", "branch"),
            ]

        employee_rows = [
            (
                field_map.get(key[0]),
                employee[key[0]],
                field_map.get(key[1]),
                employee[key[1]],
            )
            for key in employee_fields
        ]

        for row in employee_rows:
            self.column = 1
            for field in row:
                cell = self.ws.cell(self.row, self.column, value=field)
                cell.alignment = Alignment(horizontal="left")
                cell.border = all_side_border
                if self.column % 2 == 1:
                    cell.fill = GRAY
                self.column += 1
            self.row += 1
        self.column = 1

        self.row += 1

    def append_report_rows(self):
        def get_selected_headings(row):
            return (
                row["heading"],
                f'{row["package_amount"]:,.2f}',
                f'{row["amount"]:,.2f}',
                f'{abs(row["amount"] - row["package_amount"]):,.2f}',
                row.get("year_to_date") and f'{row["year_to_date"]:,.2f}',
            )

        start_coordinate = get_column_letter(self.column) + str(self.row)
        report_headers = [
            "Particulars",
            "Package Amount",
            "This Month",
            "Difference",
            "YTD",
        ]

        for field in report_headers:
            cell = self.ws.cell(self.row, self.column, value=field)
            cell.font = Font(bold=True)
            cell.fill = GRAY
            self.column += 1

        self.row += 1
        self.column = 1

        report_data = self.data["report_rows"]
        report_rows = [get_selected_headings(row) for row in report_data]
        for row in report_rows:
            self.column = 1
            for fields in row:
                cell = self.ws.cell(self.row, self.column, fields)
                cell.alignment = Alignment(horizontal="left")
                cell.border = all_side_border
                self.column += 1
            self.row += 1

        end_coordinate = get_column_letter(self.column - 1) + str(self.row - 1)
        ref = f"{start_coordinate}:{end_coordinate}"
        self.add_table(name="ReportRows", ref=ref)
        self.column = 1
        self.row += 1

    def append_earnings(self):
        self.merge_cells(column_offset=1)
        cell = self.ws.cell(self.row, self.column, value="Earnings")
        cell.fill = GRAY
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.border = all_side_border
        self.row += 1

        earnings = self.data["earnings"]
        for earning in earnings:
            self.column = 1
            for field in earning.values():
                cell = self.ws.cell(self.row, self.column, value=field)
                cell.border = all_side_border
                self.column += 1
            self.row += 1
        self.column = 1

    def append_deductions(self):
        self.row = self.start_earning_deduction_row
        self.column = 3
        self.merge_cells(column_offset=1)
        cell = self.ws.cell(self.row, self.column, value="Deductions")
        cell.fill = GRAY
        cell.alignment = Alignment(horizontal="center")
        cell.border = all_side_border
        cell.font = Font(bold=True)
        self.row += 1

        deductions = self.data["deductions"]
        for deduction in deductions:
            self.column = 3
            for field in deduction.values():
                cell = self.ws.cell(self.row, self.column, value=field)
                cell.border = all_side_border
                self.column += 1
            self.row += 1
        self.column = 1

    def append_payment_information(self):
        payment_info_start_row = self.row
        payment_info = self.payment_information
        payment_info.pop("pf_number")
        payment_info.pop("branch")
        for item in payment_info.items():
            cell1 = self.ws.cell(self.row, 1, value=field_map.get(item[0]))
            cell1.fill = GRAY
            cell1.border = all_side_border
            cell2 = self.ws.cell(self.row, 2, value=item[1])
            cell2.border = all_side_border
            self.row += 1

        payment_info_end_row = self.row
        self.column = 1

        self.column = 3
        paid_days = self.data["attendance_details"].get("paid_days")
        payment_info = ["No. of Pay Day", paid_days, "NET PAY", self.net_pay]

        self.row = payment_info_start_row
        for index, info in enumerate(payment_info, 1):
            self.merge_cells(row_offset=1 if index % 2 == 0 else 0, column_offset=1)
            cell = self.ws.cell(self.row, self.column)
            cell.value = info
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = all_side_border

            if index % 2 != 0:
                cell.fill = GRAY

            self.row += 2 if index % 2 == 0 else 1

        self.row = max(payment_info_end_row + 1, 6 + 1)
        self.column = 1

    def append_leave_details(self):
        leave_details = self.data["leave_details"]
        header = {
            "leave_type": "Leaves",
            "opening": "Initial",
            "used": "Used",
            "closing": "Closing",
        }
        leave_rows = [
            (leave["leave_type"], leave["opening"], leave["used"], leave["closing"])
            for leave in leave_details
        ]
        start_coordinate = get_column_letter(self.column) + str(self.row)
        for field in header.values():
            cell = self.ws.cell(self.row, self.column, value=field)
            cell.font = Font(bold=True)
            cell.fill = GRAY
            self.column += 1

        self.row += 1

        for row in leave_rows:
            self.column = 1
            for field in row:
                cell = self.ws.cell(self.row, self.column, value=field)
                cell.alignment = Alignment(horizontal="left")
                cell.border = all_side_border
                self.column += 1
            self.row += 1

        end_coordinate = get_column_letter(self.column - 1) + str(self.row - 1)
        ref = f"{start_coordinate}:{end_coordinate}"
        self.add_table(name="leaveDetail", ref=ref)

        self.row += 1
        self.column = 1

    def append_attendance_details(self):
        self.attendance_start_row = self.row
        attendance_details = self.data["attendance_details"]
        attendance_details.pop("simulated_from")
        attendance_details.pop("previous_payroll_adjusted_from")

        start_coordinate = get_column_letter(self.column) + str(self.row)
        for field in ["Attendance", "Days"]:
            cell = self.ws.cell(self.row, self.column, value=field)
            cell.font = Font(bold=True)
            cell.fill = GRAY
            self.column += 1
        self.row += 1

        for row in attendance_details.items():
            self.column = 1
            for index, field in enumerate(row, start=1):
                value = field if index % 2 == 0 else field_map.get(field)
                cell = self.ws.cell(self.row, self.column, value=value)
                cell.alignment = Alignment(horizontal="left")
                cell.border = all_side_border
                self.column += 1
            self.row += 1

        end_coordinate = get_column_letter(self.column - 1) + str(self.row - 1)
        ref = f"{start_coordinate}:{end_coordinate}"
        self.add_table(name="attendanceDetail", ref=ref)

        self.row += 1

    def append_hourly_attendance(self):
        hourly_attendance = {
            **self.data["hourly_attendance"],
            "expected_working_hours": self.data.get("expected_working_hours"),
        }
        self.row = self.attendance_start_row
        self.column = 3
        start_coordinate = get_column_letter(self.column) + str(self.row)

        for field in ["Particulars", "Details"]:
            cell = self.ws.cell(self.row, self.column, value=field)
            cell.font = Font(bold=True)
            cell.fill = GRAY
            self.column += 1

        self.row += 1

        title_map = {
            "total_worked_hours": "Total Worked Hours",
            "actual_overtime_hours": "Actual Overtime Hours",
            "normalized_overtime_hours": "Normalized Overtime Hours",
            "expected_working_hours": "Expected Working Hours",
            "lost_hours": "Total Lost Hours",
        }
        self.column = 3
        for field, value in hourly_attendance.items():
            cell1 = self.ws.cell(self.row, self.column, value=title_map[field])
            cell1.border = all_side_border
            cell2 = self.ws.cell(self.row, self.column + 1, value=value)
            cell2.border = all_side_border
            cell2.alignment = Alignment(horizontal="left")
            self.row += 1

        end_coordinate = get_column_letter(self.column + 1) + str(self.row - 1)
        ref = f"{start_coordinate}:{end_coordinate}"
        self.add_table(name="hourlyAttendance", ref=ref)
        self.row += 1

    def append_total_earnings_deductions_row(self):
        total_earning_deduction_row = [
            "Total Earnings",
            self.total_earnings,
            "Total Deductions",
            self.total_deductions,
        ]
        for index, field in enumerate(total_earning_deduction_row, start=1):
            cell = self.ws.cell(self.row, self.column, value=field)
            cell.border = all_side_border
            if index % 2 != 0:
                cell.fill = GRAY
            self.column += 1

        self.row += 2
        self.column = 1

    def append_earnings_and_deductions(self):
        self.start_earning_deduction_row = self.row
        self.append_earnings()
        self.append_deductions()

        self.row = (
            self.start_earning_deduction_row
            + max(len(self.data["earnings"]), len(self.data["deductions"]))
            + 2
        )
        self.append_total_earnings_deductions_row()
        self.append_payment_information()

    def create_workbook(self):
        self.append_payroll_date()
        self.append_employee_information()
        (
            self.append_report_rows()
            if self.template == TEMPLATE_1
            else self.append_earnings_and_deductions()
        )
        self.append_leave_details()
        self.append_attendance_details()
        self.append_hourly_attendance()
        return self.workbook


class ExcelTaxReport(ExcelMixin):

    def __init__(self, data):
        self.data = data
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.row = 1
        self.column = 1

    def append_employee_information(self):
        employee = self.data["employee"]
        division = employee.get("division")
        employee = {
            **employee,
            "joined_date": employee.get("joined_date"),
            "marital_status": employee.get("marital_status"),
            "username": employee.get("username"),
            "division": division and division["name"],
        }

        employee_fields = [
            ("full_name", "job_title"),
            ("branch", "username"),
            ("division", "code"),
            ("joined_date", "marital_status"),
        ]

        employee_rows = [
            (
                field_map.get(key[0]),
                employee[key[0]],
                field_map.get(key[1]),
                employee[key[1]],
            )
            for key in employee_fields
        ]

        for row in employee_rows:
            self.column = 1
            for field in row:
                cell = self.ws.cell(self.row, self.column, value=field)
                cell.alignment = Alignment(horizontal="left")
                cell.border = all_side_border
                cell.fill = LIGHT_BLUE
                self.column += 1
            self.row += 1
        self.column = 1

        self.row += 1

    def append_particulars(self):
        self.column = 1
        header = ["Particulars", "Paid", "Projected"]

        for field in header:
            if field == "Particulars":
                self.merge_cells(column_offset=1)

            cell = self.ws.cell(self.row, self.column, value=field)
            cell.border = all_side_border
            cell.font = Font(color="FFFFFF")
            cell.fill = BLUE_GRAY
            cell.alignment = Alignment(horizontal="center")
            self.column += 2 if field == "Particulars" else 1

        self.row += 1
        results = self.data["results"]
        for result in results:
            self.column = 1
            self.merge_cells(column_offset=3)
            cell = self.ws.cell(self.row, self.column, value=result["category"])
            cell.font = Font(bold=True)
            cell.fill = LIGHT_BLUE
            headings = result["headings"]
            self.row += 1
            for heading in headings:
                self.column = 1
                self.merge_cells(column_offset=1)
                value = heading['heading_name']
                cell = self.ws.cell(self.row, self.column, value=value)
                cell.alignment = Alignment(horizontal="left")
                cell.fill = ROW_COLOR
                cell.border = all_side_border
                self.column += 2

                paid, projected = heading.get('ytd_amount'), heading.get('remaining_after_ytd')
                yearly_const = heading.get('yearly_const')
                if yearly_const:
                    paid, projected = None, yearly_const

                self.create_paid_projected(paid=paid, projected=projected)

                self.row += 1

    def create_paid_projected(self, paid, projected):
        cell = self.ws.cell(self.row, self.column, value=paid)
        cell.alignment = Alignment(horizontal="right")
        cell.fill = ROW_COLOR
        cell.border = all_side_border
        self.column += 1

        cell = self.ws.cell(self.row, self.column, value=projected)
        cell.alignment = Alignment(horizontal="right")
        cell.fill = ROW_COLOR
        cell.border = all_side_border
        self.column += 1

    def create_workbook(self):
        self.merge_cells(column_offset=3)
        tax_header = (
            "Tax calculation details from "
            f'{self.data["from_date"]:%Y-%m-%d} to '
            f'{self.data["to_date"]:%Y-%m-%d} for Fiscal Year '
            f'{self.data["fiscal_year_name"]}'
        )
        cell = self.ws.cell(self.row, self.column, value=tax_header)
        cell.border = all_side_border
        cell.alignment = Alignment(horizontal="center")
        cell.fill = LIGHT_BLUE
        self.row += 2
        self.append_employee_information()
        self.append_particulars()
        return self.workbook
