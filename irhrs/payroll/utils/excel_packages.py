from collections.abc import Mapping
import json
import time
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook, Workbook
from io import BytesIO
from datetime import datetime
from django.core.cache import cache

from irhrs.notification.utils import notify_organization
from irhrs.payroll.models.payroll import UserExperiencePackageSlot, ASSIGNED, UPDATED_PACKAGE
from irhrs.payroll.tasks import create_package_activity
from irhrs.payroll.utils.package_clone import clone_package_from_another_package
from irhrs.permission.constants.permissions.hrs_permissions import GENERATE_PAYROLL_PERMISSION

USER = get_user_model()

class ExcelDictPackage(Mapping):
    def __init__(self, file):
        wb=self._get_workbook(file)
        ws = wb.active

        # creating error book by initially copying uploaded
        # workbook, deepcopying caused error while loading
        # with openpyxl
        self.error_wb = Workbook()
        self.error_ws = self.error_wb.active
        for value in ws.values:
            self.error_ws.append(value)

        rows = ws.values
        self.header = next(rows)
        self.data = {
            row[0]: dict(zip(self.header[1:], row[1:]))
            for row in rows if row[0]
        }
        self.to_index = dict(
            zip(self.data.keys(), range(2, len(self.data) + 2))
        )
        self.error_field = len(self.header) + 1
        self.errors = {}
        self.error_ws.cell(row=1, column=self.error_field, value="Errors")

    def _get_workbook(self, file):
        if isinstance(file, Workbook):
            return file

        file=BytesIO(file.read())
        return load_workbook(filename=file)

    def __iter__(self):
        return iter(self.data)

    def __len__(self):
        return len(self.data)

    def __getitem__(self, key):
        return self.data[key]

    def __setitem__(self, key, value):
        if not value:
            return
        self.errors[key] = value
        self.error_ws.cell(
            row=self.to_index[key],
            column=self.error_field,
            value=str(value)
        )

def create_bulk_packages(assigned_date, excel_data, instance, organization, actor):
    cache.set("block_excel_package", True)
    cache.set('total_number_of_payroll_to_be_generated', len(excel_data))
    with transaction.atomic():
        for email, fields in excel_data:
            clone_and_assign_package(
                email=email,
                fields=fields,
                instance=instance,
                assigned_date=assigned_date,
                actor=actor
            )

    instance.status = "Completed"
    instance.save()
    cache.delete("failed_package_errors")
    cache.delete("block_excel_package")
    cache.delete("failed_package_errors_timestamp")
    notify_organization(
        text="Excel bulk assignment completed successfully.",
        action=instance,
        organization=organization,
        actor=actor,
        url=f"/admin/{organization.slug}/payroll/bulk-assign-package",
        permissions=[GENERATE_PAYROLL_PERMISSION]
    )
    cache.delete("payroll_generated_employee_count")
    cache.delete("total_number_of_payroll_to_be_generated")
    cache.delete("payroll_generated_employee_name")

def clone_and_assign_package(email,fields, instance, assigned_date, actor):
    user = USER.objects.get(Q(email=email) | Q(username=email))
    now =datetime.now()

    employee_count = cache.get("payroll_generated_employee_count", 0) + 1
    cache.set("payroll_generated_employee_count", employee_count)
    cache.set("payroll_generated_employee_name", user.full_name)
    name = f"{user.full_name} {user.username} {assigned_date} {time.strftime('%H:%M', time.localtime())}"
    package = clone_package_from_another_package(instance.cloned_from, name, actor)
    package.excel_package = instance
    package.save()

    fields = {k:v for k, v in fields.items() if v is not None}
    headings_to_be_updated = package.package_headings.filter(
        heading__name__in=fields
    )

    for package_heading in headings_to_be_updated:
        new_rule = str(fields[package_heading.heading.name])
        update_heading_rule(
            package_heading=package_heading,
            new_rule=new_rule
        )
    experience = user.user_experiences.first()
    user_experience_package_slot = UserExperiencePackageSlot.objects.filter(
        user_experience=experience,
        active_from_date=assigned_date
    ).first()

    if user_experience_package_slot:
        old_package = user_experience_package_slot.package
        user_experience_package_slot.package = package
        user_experience_package_slot.excel_package = instance
        user_experience_package_slot.save()
        title = f'{actor.full_name} has {ASSIGNED} a package named "{old_package.name}" to {user.full_name} by bulk package assign feature'
        create_package_activity(title=title, package=old_package, action=ASSIGNED, assigned_to=user)
        return

    UserExperiencePackageSlot.objects.create(
        package=package,
        user_experience=experience,
        active_from_date=assigned_date,
        excel_package=instance
    )
    title = f'{actor.full_name} has {ASSIGNED} a package named "{package.name}" to {user.full_name} by bulk package assign feature'
    create_package_activity(title=title, package=package, action=ASSIGNED, assigned_to=user)

def strip_rule(rules):
    rules = json.loads(rules)
    for rule in rules:
        rule['rule'] = rule['rule'].strip()
    return json.dumps(rules)

def strip_packages(package):
    package_headings = package.package_headings.all()
    for package_heading in package_headings:
        package_heading.rules = strip_rule(package_heading.rules)
        package_heading.save()

def update_rule(rule, new_rule):
    rules = json.loads(rule)
    if len(rules) != 1:
        return json.dumps(rules)

    rules[0]['rule'] = new_rule
    return json.dumps(rules)

def update_heading_rule(package_heading, new_rule):
    rule = update_rule(package_heading.rules, new_rule)
    package_heading.rules = rule
    package_heading.save()
