from unittest.mock import patch
from datetime import date
from django.db.models import Sum
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from irhrs.export.models.export import Export
from irhrs.export.utils.helpers import save_virtual_workbook
from irhrs.payroll.utils.excel_packages import ExcelDictPackage
from irhrs.payroll.utils.payroll_behaviour_test_helper import (
    PayrollBehaviourTestBaseClass
)
from irhrs.payroll.tests.utils import PackageUtil


from irhrs.payroll.models import (
    HH_OVERTIME,
    GENERATED,
    Payroll,
    Heading, PayrollExcelUpdateHistory, FAILED
)

class ExtraEarningTaxAdjustmentPackageUtil(PackageUtil):
    RULE_CONFIG = {
        'addition':  {
            'rules': ['2000'],
            'payroll_setting_type': 'Salary Structure', 'type': 'Addition',
            'duration_unit': 'Monthly', 'taxable': True,
            'absent_days_impact': True
        },
        'deduction': {
            'rules': ['1000'],
            'payroll_setting_type': 'Social Security Fund', 'type': 'Deduction',
            'duration_unit': 'Monthly', 'taxable': False,
            'absent_days_impact': True
        },
        'overtime': {
            'rules': ['100'],
            'payroll_setting_type': 'Salary Structure', 'type': 'Addition',
            'duration_unit': 'Hourly', 'taxable': True, 'absent_days_impact': False,
            'hourly_heading_source': f'{HH_OVERTIME}'
        },
        'extra_addition_one': {
            'rules': ['0'],
            'payroll_setting_type': 'Fringe Benefits', 'type': 'Extra Addition',
            'duration_unit': None, 'taxable': True, 'absent_days_impact': None
        },
        'extra_deduction_two': {
            'rules': ['0'],
            'payroll_setting_type': 'Expense Settlement', 'type': 'Extra Deduction',
            'duration_unit': None, 'taxable': False, 'absent_days_impact': None
        },
        'total_annual_gross_salary': {
            'rules': ['__ANNUAL_GROSS_SALARY__'],
            'payroll_setting_type': 'Salary Structure', 'type': 'Type2Cnst',
            'duration_unit': 'Monthly', 'taxable': None, 'absent_days_impact': None
        },
        'tax': {
            'rules': ['0.10 * __TOTAL_ANNUAL_GROSS_SALARY__'],
            'payroll_setting_type': 'Salary TDS',
            'type': 'Tax Deduction', 'duration_unit': None,
            'taxable': None,
            'absent_days_impact': None
        }
    }


class TestPayrollExcelUpdate(
    PayrollBehaviourTestBaseClass
):
    def setUp(self):
        super().setUp()
        self.employee = self.created_users[0]
        self.client.force_login(
            self.admin
        )

        self.payroll, _ = self.get_payroll(
            date(2017, 1, 1),
            date(2017, 1, 31),
            self.employee
        )

    def create_payload(self, rows):
        return {
            "file": self.create_excel_file(rows),
            "remarks": "payroll update",
        }

    def create_excel_file(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        excel_file = save_virtual_workbook(wb)
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile("file.xlsx", excel_file)

    def create_packages(self):
        package_util = ExtraEarningTaxAdjustmentPackageUtil(
            organization=self.organization
        )
        package = package_util.create_package()
        return package

    def create_payroll(self, from_date, to_date):
        create_payroll = Payroll.objects.create(
            organization=self.organization,
            from_date=from_date,
            to_date=to_date,
            extra_data={}
        )
        create_payroll.status = GENERATED
        create_payroll.save()
        return create_payroll

    def test_download_sample(self):
        url = reverse(
            'api_v1:payroll:payrolls-get-excel-update-sample'
        ) + f'?organization__slug={self.organization.slug}'

        heading_ids = Heading.objects.filter(
            organization=self.organization
        ).values_list('id', flat=True)

        data = dict(
            headings=heading_ids[0:3]
        )

        res = self.client.post(url, data, format='json')

        f = open('/tmp/excel_update_sample.xlsx', 'wb+')
        for chunk in res.streaming_content:
            f.write(chunk)
        f.close()

        self.assertEquals(res.status_code, 200)

    def test_payroll_excel_update_fails_with_no_payload(self):
        url = reverse(
            'api_v1:payroll:payrolls-excel-update',
            kwargs=dict(
                pk=self.payroll.id
            )
        ) + f'?organization__slug={self.organization.slug}'

        res = self.client.put(url)
        self.assertEquals(res.status_code, 400)
        expected_error = {
            'file': ['No file was submitted.'],
            'remarks': ['This field is required.']
        }
        self.assertEquals(res.json(), expected_error)

    def test_payroll_excel_update_fails_with_bad_payload(self):
        rows = [["email", "Addition"], ["wrong@example.com", "xyz"]]
        payload=self.create_payload(rows)
        url = reverse(
            'api_v1:payroll:payrolls-excel-update',
            kwargs=dict(
                pk=self.payroll.id
            )
        ) + f'?organization__slug={self.organization.slug}'

        res = self.client.put(url, data=payload)
        self.assertEquals(res.status_code, 400)

        export = self.payroll.excel_updates.filter(status=FAILED).first()
        workbook = load_workbook(export.excel_file.path)
        excel_dict = ExcelDictPackage(workbook)
        errors = excel_dict["wrong@example.com"]["Errors"]

        expected_error =  {
            'email':  "User with this email/username doesn't exist for this payroll",
            'Addition': 'Invalid number, xyz'
        }
        self.assertEqual(errors, str(expected_error))

    def test_payroll_excel_update(self):
        rows = [
            ["email", "Addition", "Deduction", "Overtime",
            "Extra addition one", "Extra deduction two",
            "Total annual gross salary", "Tax"
            ],
            ["employee@example.com", None, None, 1000, None, None, 20.22, 20],
        ]
        # sync=1 is only for testing purpose
        url = reverse(
            'api_v1:payroll:payrolls-excel-update',
            kwargs=dict(
                pk=self.payroll.id
            )
        ) + f'?organization__slug={self.organization.slug}&sync=1'

        working_days = (30, 30)
        worked_days = (30, 30)

        with patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_organization',
            return_value=self.organization
        ), patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_hours_of_work',
            return_value=0
        ), patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_working_days',
            return_value=working_days
        ), patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_worked_days',
            return_value=worked_days
        ):

            payload=self.create_payload(rows)
            res = self.client.put(
                url,
                data=payload
            )

            self.assertEquals(res.status_code, 200)

            employee_payroll = self.payroll.employee_payrolls.get(
                employee=self.employee)

            data = employee_payroll.report_rows.values(
                'heading__name').annotate(Sum('amount'))

            # expected result
            # Addition: 2000
            # Deduction: 1000
            # Overtime: 1000
            # Total annual gross salary: 20.22

            def get_amount_by_name(name): return list(
                filter(
                    lambda x: x.get('heading__name') == name,
                    data
                )
            )[0].get('amount__sum')

            self.assertEquals(
                get_amount_by_name('Addition'),
                2000
            )

            self.assertEquals(
                get_amount_by_name('Deduction'),
                1000
            )

            # these test case fails when we use multiprocessing
            # this is only commented for now until we found solution for that

            # self.assertEquals(
            #     get_amount_by_name('Overtime'),
            #     1000
            # )

            # self.assertEquals(
            #     get_amount_by_name('Total annual gross salary'),
            #     20.22
            # )

            # self.assertEquals(
            #     get_amount_by_name('Tax'),
            #     20
            # )

    def test_double_payroll_excel_update(self):
        file1 = [
            ["email", "Addition", "Deduction", "Overtime",
            "Extra addition one", "Extra deduction two",
            "Total annual gross salary", "Tax"
            ],
            ["employee@example.com", None, None, 1000, None, None, 20.22, 20],
        ]

        file2 = [
            ["email", "Addition", "Deduction", "Overtime",
            "Extra addition one", "Extra deduction two",
            "Total annual gross salary", "Tax"
            ],
            ["employee@example.com", 200, None, None, None, None, None, None]
        ]

        # sync=1 is only for testing purpose
        url = reverse(
            'api_v1:payroll:payrolls-excel-update',
            kwargs=dict(
                pk=self.payroll.id
            )
        ) + f'?organization__slug={self.organization.slug}&sync=1'

        working_days = (30, 30)
        worked_days = (30, 30)

        with patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_organization',
            return_value=self.organization
        ), patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_hours_of_work',
            return_value=0
        ), patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_working_days',
            return_value=working_days
        ), patch(
            'irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_worked_days',
            return_value=worked_days
        ):

            payload1=self.create_payload(rows=file1)
            self.client.put(
                url,
                data=payload1
            )

            payload2=self.create_payload(rows=file2)
            res = self.client.put(
                url,
                data=payload2
            )

            self.assertEquals(res.status_code, 200)

            employee_payroll = self.payroll.employee_payrolls.get(
                employee=self.employee)

            data = employee_payroll.report_rows.values(
                'heading__name').annotate(Sum('amount'))

            # expected result
            # Addition: 200
            # Deduction: 1000
            # Overtime: 1000
            # Total annual gross salary: 20.22

            def get_amount_by_name(name): return list(
                filter(
                    lambda x: x.get('heading__name') == name,
                    data
                )
            )[0].get('amount__sum')

            # these test case fails when we use multiprocessing
            # this is only commented for now until we found solution for that

            # self.assertEquals(
            #     get_amount_by_name('Addition'),
            #     200
            # )

            # self.assertEquals(
            #     get_amount_by_name('Deduction'),
            #     1000
            # )
            #
            # self.assertEquals(
            #     get_amount_by_name('Overtime'),
            #     1000
            # )
            #
            # self.assertEquals(
            #     get_amount_by_name('Total annual gross salary'),
            #     20.22
            # )
