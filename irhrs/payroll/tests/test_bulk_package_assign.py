from datetime import timedelta
from unittest.mock import patch
from django.test import override_settings
from openpyxl import Workbook, load_workbook
import json
from django.urls import reverse
from django.core.cache import cache

from irhrs.core.utils.common import get_today
from irhrs.common.api.tests.common import RHRSTestCaseWithExperience
from irhrs.notification.models.notification import OrganizationNotification
from irhrs.organization.api.v1.tests.factory import FiscalYearFactory
from irhrs.payroll.models.payroll import ExcelPayrollPackage, PackageHeading
from irhrs.payroll.tests.factory import OrganizationPayrollConfigFactory
from irhrs.payroll.tests.utils import PackageUtil
from irhrs.payroll.utils.excel_packages import ExcelDictPackage
from irhrs.users.models.experience import UserExperience

TEST_DIR = "test_media"


def save_virtual_workbook(workbook):
    from tempfile import NamedTemporaryFile

    with NamedTemporaryFile() as tf:
        workbook.save(tf.name)
        from io import BytesIO

        in_memory = BytesIO(tf.read())
        return in_memory.getvalue()


class TestExcelPayrollPackage(RHRSTestCaseWithExperience):
    organization_name = "Test"
    users = [
        ("admin@example.com", "password", "Male", "executive"),
        ("usera@example.com", "password", "Male", "junior assistant"),
        ("userb@example.com", "password", "Male", "sr assistant"),
    ]

    def setUp(self):
        super().setUp()
        self.default_package = PackageUtil(
            organization=self.organization
        ).create_package()
        self.start_at, self.end_at = get_today().replace(
            month=1, day=1
        ), get_today().replace(month=12, day=31)
        self.fiscal_year = FiscalYearFactory(
            organization=self.organization,
            start_at=self.start_at,
            end_at=self.end_at,
            applicable_from=self.start_at,
            applicable_to=self.end_at
        )
        self.payroll_config = OrganizationPayrollConfigFactory(
            start_fiscal_year=self.fiscal_year, organization=self.organization
        )
        self.client.force_login(self.admin)

    def tearDown(self):
        cache.clear()
        import shutil
        from django.conf import settings

        test_media = settings.PROJECT_DIR + "/" + TEST_DIR
        from pathlib import Path

        if Path(test_media).exists():
            shutil.rmtree(test_media)

    def create_excel_file(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        excel_file = save_virtual_workbook(wb)
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile("excel_file.xlsx", excel_file)

    @property
    def url(self):
        return reverse(
            "api_v1:payroll:excel-bulk-assign-list",
            kwargs={"organization_slug": "test"},
        )

    def create_payload(self, rows):
        return {
            "excel_file": self.create_excel_file(rows),
            "name": "bulk package",
            "cloned_from": self.default_package.id,
            "assigned_date": get_today() + timedelta(days=1),
        }

    def get_error_package(self):
        self.assertTrue(cache.get("failed_package_errors"))
        error_file = cache.get("failed_package_errors").split("/media/")[-1]
        from django.conf import settings

        excel_file_path = (
            settings.PROJECT_DIR + "/" + settings.MEDIA_ROOT + "/" + error_file
        )
        workbook = load_workbook(excel_file_path, data_only=True)
        return ExcelDictPackage(workbook)

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_fails_for_no_payload(self):
        response = self.client.post(self.url)
        self.assertEqual(response.status_code, 400, response.json())
        error = {
            "assigned_date": ["This field is required."],
            "cloned_from": ["This field is required."],
            "excel_file": ["No file was submitted."],
            "name": ["This field is required."],
        }
        self.assertEqual(response.json(), error)

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_fails_for_bad_heading(self):
        BAD_HEADING = "Basic Salary"
        bad_rows = [["email", BAD_HEADING], [self.admin.email, "30000"]]
        bad_payload = self.create_payload(rows=bad_rows)
        response = self.client.post(self.url, data=bad_payload)
        self.assertEqual(response.status_code, 400, response.json())

        error = {
            "non_field_errors": [
                f"Heading {BAD_HEADING} not found in {self.default_package}"
            ]
        }
        self.assertEqual(response.json(), error)

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_fails_for_assigned_date_in_no_fiscal_year(self):
        bad_rows = [["email", "Basic salary"], [self.admin.email, "30000"]]
        bad_payload = self.create_payload(rows=bad_rows)
        bad_payload["assigned_date"] = self.end_at + timedelta(days=1)
        response = self.client.post(self.url, data=bad_payload)
        self.assertEqual(response.status_code, 400, response.json())

        error = {
            "assigned_date": [
                "Can not assign package. Fiscal year doesn't exist for given date."
            ]
        }
        self.assertEqual(response.json(), error)

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_fails_for_bad_row(self):
        BAD_EMAIL = "random@email.com"
        INVALID_NUMBER = "abcd"
        bad_data = [["email", "Basic salary"], [BAD_EMAIL, INVALID_NUMBER]]
        data = {
            "excel_file": self.create_excel_file(bad_data),
            "name": "bulk package",
            "cloned_from": self.default_package.id,
            "assigned_date": get_today() + timedelta(days=1),
        }
        response = self.client.post(self.url, data=data)
        self.assertEqual(response.status_code, 201, response.json())
        excel_payroll_package = ExcelPayrollPackage.objects.get()
        self.assertEqual(excel_payroll_package.status, "Failed")
        error_dict_package = self.get_error_package()
        errors = error_dict_package[BAD_EMAIL]["Errors"]
        expected_error = {
            "Basic salary": f"Invalid number, {INVALID_NUMBER}",
            "email": "User with this email/username doesn't exist",
        }
        self.assertEqual(errors, str(expected_error))

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_fails_for_assigned_date_before_appoint_date(self):
        bad_data = [["email", "Basic salary"], [self.admin.email, "30000"]]
        data = {
            "excel_file": self.create_excel_file(bad_data),
            "name": "bulk package",
            "cloned_from": self.default_package.id,
            "assigned_date": get_today() - timedelta(days=1),
        }
        response = self.client.post(self.url, data=data)
        self.assertEqual(response.status_code, 201, response.json())
        excel_payroll_package = ExcelPayrollPackage.objects.get()
        self.assertEqual(excel_payroll_package.status, "Failed")
        error_dict_package = self.get_error_package()
        excel_error = error_dict_package[self.admin.email]["Errors"]
        expected_error = {
            "appoint_date": f"package assignment date cannot be before appoint date, {get_today()}"
        }
        self.assertEqual(excel_error, str(expected_error))

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_fails_for_end_date_before_assigned_date(self):
        bad_data = [["email", "Basic salary"], [self.admin.email, "30000"]]
        user_xp = UserExperience.objects.get(user=self.admin)
        user_xp.start_date = self.fiscal_year.end_at - timedelta(days=60)
        from datetime import date
        year = get_today().year
        self.fiscal_year.start_at = date(year, 1, 1)
        self.fiscal_year.end_at = date(year, 12, 31)
        user_xp.end_date = date(year,12,30)
        user_xp.save()
        data = {
            "excel_file": self.create_excel_file(bad_data),
            "name": "bulk package",
            "cloned_from": self.default_package.id,
            "assigned_date": date(year, 12, 31)
        }
        response = self.client.post(self.url, data=data)
        self.assertEqual(response.status_code, 201, response.json())
        excel_payroll_package = ExcelPayrollPackage.objects.get()
        self.assertEqual(excel_payroll_package.status, "Failed")
        error_dict_package = self.get_error_package()
        excel_error = error_dict_package[self.admin.email]["Errors"]
        expected_error = {
            "user_experience_end_date": (
                "Cannot assign package to user with experience end date "
                "before assigned date"
            )
        }
        self.assertEqual(excel_error, str(expected_error))

    def check_bulk_assign_package_success(self, rows, new_rule):
        with patch(
            "irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_hours_of_work",
            return_value=0,
        ), patch(
            "irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_working_days",
            return_value=30,
        ), patch(
            "irhrs.payroll.utils.calculator.EmployeeSalaryCalculator.get_worked_days",
            return_value=30,
        ):
            data = {
                "excel_file": self.create_excel_file(rows),
                "name": "bulk package",
                "cloned_from": self.default_package.id,
                "assigned_date": get_today() + timedelta(days=1),
            }
            self.assertFalse(ExcelPayrollPackage.objects.exists())
            response = self.client.post(self.url, data=data)
            excel_payroll_package = ExcelPayrollPackage.objects.get()
            self.assertEqual(response.status_code, 201, response.json())
            self.assertEqual(excel_payroll_package.status, "Completed")
            self.assertTrue(excel_payroll_package.package_slots.exists())
            basic = PackageHeading.objects.filter(
                package__excel_package=excel_payroll_package,
                heading__name="Basic salary",
            )

            self.assertTrue(basic.exists())
            rules = json.loads(basic.get().rules)
            self.assertEqual(rules[0]["rule"], new_rule)
            self.assertTrue(
                OrganizationNotification.objects.filter(
                    text="Excel bulk assignment completed successfully."
                ).exists()
            )

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_succeeds(self):
        BASIC_AMOUNT = "200000"
        rows = [["email", "Basic salary"], [self.admin.email, BASIC_AMOUNT]]
        self.check_bulk_assign_package_success(rows=rows, new_rule=BASIC_AMOUNT)

    @override_settings(MEDIA_ROOT=(TEST_DIR + "/media"))
    def test_bulk_assign_package_succeeds_with_none_value(self):
        BASIC_AMOUNT = None
        rows = [["email", "Basic salary"], [self.admin.email, BASIC_AMOUNT]]
        basic = PackageHeading.objects.get(
            package__excel_package__isnull=True,
            heading__name="Basic salary",
        )
        default_rule = json.loads(basic.rules)[0]["rule"]
        self.check_bulk_assign_package_success(rows=rows, new_rule=default_rule)
