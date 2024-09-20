from datetime import datetime
import pytz

from django.db.models import OuterRef, Subquery, Count, Q, Exists, QuerySet
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from rest_framework import mixins, status
from rest_framework.viewsets import GenericViewSet
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import action
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, MultipleChoiceFilter

from irhrs.attendance.utils.attendance import get_adjustment_request_forwarded_to
from irhrs.core.utils import subordinates
from irhrs.core.utils import nested_getattr
from irhrs.core.utils.common import get_today

from irhrs.core.utils.filters import OrderingFilterMap, SearchFilter, FilterMapBackend
from irhrs.core.utils.subordinates import set_supervisor_permissions, find_immediate_subordinates
from irhrs.export.mixins.export import BackgroundExportMixin
from irhrs.organization.models import Organization
from irhrs.task.api.v1.serializers.worklog import WorLogListSerializer, \
    WorLogToDoStartSerializer, WorLogCreateDailyTaskSerializer, WorLogToDoBulkCreateSerializer, \
    WorkLogActionRemarksSerializer, WorkLogBulkApproveSerializer, \
    WorkLogActionBulkRemarksSerializer, WorkLogHistorySerializer, WorkLogReportSerializer, \
    WorkLogBulkSendSerializer
from irhrs.task.models import WorkLog, WORKLOG_ACTION, DRAFT, REQUESTED, APPROVED, DENIED, \
    FORWARDED, CONFIRMED, CANCELED, ACKNOWLEDGED, WorkLogAction, TODO, SENT
from irhrs.task.models.settings import UserActivityProject

KTM = pytz.timezone("Asia/Kathmandu")


class WorkLogFilter(FilterSet):
    """filters WorkLog module by `status`, 'project_id`, `activity_id`, 'sender_id`"""

    status = MultipleChoiceFilter(
        choices=WORKLOG_ACTION,
        method='get_status_in',
        label='Status'
    )

    def get_status_in(self, queryset, name, value):
        return queryset.filter(**{
            'status__in': value,
        })

    class Meta:
        model = WorkLog
        fields = ['status', 'project_id', 'activity_id', 'sender_id']


class WorkLogAPIViewSet(BackgroundExportMixin, mixins.ListModelMixin, GenericViewSet):
    """WorkLog Workflow:

    :func:`~create_todo` User creates list of TO_DO
        status of WorkLog instance: TO_DO

    :func:`~start_todo` User starts TO_DO, one at a time
        status of WorkLog instance: TO_DO

    :func:`~complete_todo` User completes TO_DO, one at a time
        status of WorkLog instance: DRAFT

    :func:`~create_worklog` User/Supervisor/HR can bypass above steps and directly create WorkLog
        status of WorkLog instance: DRAFT

    :func:`~update_daily_task` Update WorkLog if any changes to be done
        status of WorkLog instance: DRAFT

    :func:`~send_worklog` User sends WorkLog to its Supervisor/HR if everything is ok
        status of WorkLog instance: REQUESTED

    :func:`~cancel_action` User cancels created WorkLog
        status of WorkLog instance: CANCELED

    :func:`~approve_create_request`  Supervisor/HR approves WorkLog sent by User/Supervisor
        status of WorkLog instance: APPROVED

    :func:`~deny_create_request`  Supervisor/HR denies WorkLog
        status of WorkLog instance: DENIED

    :func:`~forward_create_request`  Supervisor forwards WorkLog to another level of Supervisor
        status of WorkLog instance: FORWARDED

    :func:`~acknowledge_action` Once WorkLog is APPROVED, User acknowledge it
        status of WorkLog instance: ACKNOWLEDGED

    :func:`~confirm_action` Once WorkLog is ACKNOWLEDGED by user, HR confirms it
        status of WorkLog instance: CONFIRMED

    :func:`~history` History of all the action on the WorkLog

    :func:`~export_worklog` Export WorkLog detail of particular status

    """
    queryset = WorkLog.objects.all()
    filter_backends = (DjangoFilterBackend, OrderingFilterMap, SearchFilter)
    search_fields = (
        'project__name',
        'activity__name',
        'task__title',
        'core_task__title',
        'activity_description'
    )
    ordering_fields_map = {
        "modified_at": "modified_at",
        "activity_description": "activity_description"
    }
    filter_class = WorkLogFilter
    serializer_class = WorLogListSerializer
    export_fields = []

    @property
    def mode(self):
        mode = self.request.query_params.get('as')
        if mode not in ['hr', 'supervisor']:
            return 'user'

        return mode

    # Key is action request and value is current status requirement
    action_request_required_status = {
        DRAFT: [TODO],
        REQUESTED: [DRAFT],
        APPROVED: [REQUESTED, FORWARDED],
        DENIED: [REQUESTED, DENIED],
        FORWARDED: [REQUESTED, FORWARDED],
        CONFIRMED: [ACKNOWLEDGED],
        CANCELED: [DRAFT, REQUESTED],
        ACKNOWLEDGED: [APPROVED]
    }

    def validate_previous_status(self, instance: WorkLog, action: str) -> None:
        """validate previous status before taking any action(DRAFT, REQUESTED, APPROVED, etc.)

        :param instance: worklog instance
        :param action: action of worklog to be taken(DRAFT, REQUESTED, etc)
        :return: None
        """
        if instance.status not in self.action_request_required_status[action]:
            raise ValidationError(
                f'{self.action_request_required_status[action]} '
                f'is required for status to be {action}'
            )
        return

    def add_worklog_action(self, action: str) -> Response:
        """checks validation before taking action and creates history for each action taken

        :param action: action of worklog to be taken(DRAFT, REQUESTED, etc)
        :return: HTTP_200_OK
        """
        instance = self.get_object()

        self.validate_previous_status(instance, action)

        if (action in [REQUESTED, CANCELED]) and (self.request.user != instance.sender):
            return Response(
                dict(
                    non_fields_errors=[f'Only owner can perform {action} action']
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        remarks = self.request.data.get('remarks')
        if not remarks:
            raise ValidationError({"remarks": "This field may not be blank"})

        score = None
        if self.action == "approve_create_request":
            score = self.request.data.get('score')
            if not score:
                raise ValidationError({"score": "This field may not be blank"})

        WorkLogAction.objects.create(
            action_performed_by=self.request.user,
            worklog=instance,
            action=action,
            remarks=remarks,
            score=score,
            action_date=datetime.now(KTM)
        )
        return Response(status=status.HTTP_200_OK)

    def add_worklog_bulk_action(self, action: str, status: str=None) -> Response:
        """checks validation before taking action and creates history for bulk action taken

        :param action: action of worklog to be taken(DRAFT, REQUESTED, etc)
        :param status: current status of worklog instance
        :return: HTTP status code
        """
        data = self.request.data
        for req in data.get("requests"):
            instance = get_object_or_404(self.get_queryset(), id=req.get('worklog'))
            if action == FORWARDED:
                # when work log is forwarded, change receiver to next level supervisor
                forward_to = get_adjustment_request_forwarded_to(instance)
                if not forward_to:
                    raise ValidationError("Supervisor is not assigned. Please assign supervisor.")
                instance.receiver = forward_to.supervisor
                instance.save()

            if status and self.mode == 'supervisor' and not subordinates.authority_exists(
                instance.sender, self.request.user, status
            ):
                raise ValidationError({
                    f"You can not {status} this request."
                })
            self.validate_previous_status(instance, action)

        if action == DENIED and self.mode not in ["hr", "supervisor"]:
            raise ValidationError("You cannot approve this request.")

        if action == FORWARDED and self.mode != "supervisor":
            raise ValidationError("Only supervisor can forward worklog.")

        if action == CONFIRMED and self.mode != "hr":
            raise ValidationError("Only hr can confirm worklog.")

        if action == ACKNOWLEDGED and self.mode != "user":
            raise ValidationError("Only owner can acknowledge the worklog.")

        if action in [DENIED, FORWARDED, CONFIRMED, CANCELED]:
            ser = WorkLogActionBulkRemarksSerializer(
                data=self.request.data, context={
                    "action": action,
                    "user": self.request.user
                }
            )
            ser.is_valid(raise_exception=True)
            ser.save()
            return Response(ser.data)

    def get_stats_info(self, queryset: QuerySet) -> dict:
        """counts number of worklog instance in each status

        :param queryset: worklog queryset
        :return: dictionary of count of each unique status of worklog
        """
        stats_list = queryset.order_by('status').values(
            'status'
        ).annotate(
            Count('status')
        )
        stats = dict()
        for item in stats_list:
            stats[item['status']] = item['status__count']

        actions = {
            APPROVED, ACKNOWLEDGED, CANCELED, CONFIRMED, DRAFT, DENIED,
            FORWARDED, REQUESTED, TODO
        }
        differences = actions - stats.keys()
        # update value to zero which value is not stored in queryset
        # suppose if there is only one status in queryset(say DRAFT) then below code assign count 0
        # to every other actions beside the status present in queryset(in our case DRAFT)
        stats.update({
            difference: 0 for difference in differences
        })
        return stats

    def get_queryset(self) -> QuerySet:
        queryset = super().get_queryset().select_related(
            'project',
            'activity',
            'task'
        )

        if self.mode == "user":
            queryset = queryset.filter(sender=self.request.user)

        if self.mode == "supervisor":
            queryset = queryset.filter(
                receiver=self.request.user
            )

        return queryset.annotate(
            status=Subquery(
                WorkLogAction.objects.filter(
                    worklog=OuterRef('pk')
                ).order_by('-created_at').values('action')[:1]
            )
        )

    def paginate_queryset(self, queryset):
        page = super().paginate_queryset(queryset)
        if self.mode == 'supervisor' and self.action != 'history':
            return set_supervisor_permissions(page, self.request.user.id, 'sender')
        return page

    def get_paginated_response(self, data):
        response = super().get_paginated_response(data)
        response.data['stats'] = self.get_stats_info(self.get_queryset())
        return response

    def get_serializer_context(self) -> dict:
        ctx = super().get_serializer_context()
        if self.action == 'create_daily_task':
            ctx['mode'] = self.mode
        return ctx

    def create_worklog(self, request, **kwargs) -> Response:
        """creates to_dos and histories in bulk"""

        serializer = WorLogToDoBulkCreateSerializer(
            data=request.data,
            context={"user": self.request.user}
        )
        serializer.is_valid(raise_exception=True)

        if not kwargs.get('remarks'):
            raise ValidationError("This field may not be blank")

        instance = serializer.save()
        if not isinstance(instance, list):
            instance = [instance]

        WorkLogAction.objects.bulk_create(
            [
                WorkLogAction(
                    action_performed_by=self.request.user,
                    worklog=i,
                    action=kwargs.get('action'),
                    remarks=kwargs.get('remarks', 'Default remark'),
                    action_date=datetime.now(KTM)
                ) for i in instance
            ]
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(
        methods=["POST"], detail=False, url_path="create-todo",
        serializer_class=WorLogToDoBulkCreateSerializer
    )
    def create_todo(self, request, **kwargs) -> Response:
        """create to_dos and histories in bulk

        remarks and action are forcefully fed to history
        """
        kwargs['remarks'] = "Entry by user"
        kwargs['action'] = TODO
        return self.create_worklog(request, **kwargs)

    @action(
        methods=["PUT"], detail=True, url_path="start-todo",
        serializer_class=WorLogToDoStartSerializer
    )
    def start_todo(self, request, pk=None) -> Response:
        ser = self.get_serializer(
            self.get_object(), data=request.data
        )
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    @action(
        methods=["POST"], detail=False, url_path="create-daily-task",
        serializer_class=WorLogCreateDailyTaskSerializer
    )
    def create_daily_task(self, request, **kwargs) -> Response:
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        worklog = ser.save()
        WorkLogAction.objects.create(
            action_performed_by=self.request.user,
            worklog=worklog,
            action=DRAFT,
            remarks="Daily Log created by user",
            action_date=datetime.now(KTM)
        )
        return Response(ser.data)


    @action(
        methods=["PUT"], detail=True, url_path="update-daily-task",
        serializer_class=WorLogCreateDailyTaskSerializer
    )
    def update_daily_task(self, request, pk=None) ->Response:
        ser = self.get_serializer(
            self.get_object(), data=request.data
        )
        ser.is_valid(raise_exception=True)
        ser.save()
        WorkLogAction.objects.create(
            action_performed_by=self.request.user,
            worklog=self.get_object(),
            action=DRAFT,
            remarks="Daily Log updated by user",
            action_date=datetime.now(KTM)
        )
        return Response(ser.data)

    @action(
        methods=['POST'], detail=True, url_path='send-worklog',
        serializer_class=WorkLogActionRemarksSerializer
    )
    def send_worklog(self, request, pk=None, **kwargs) -> Response:
        return self.add_worklog_action(REQUESTED)

    @action(
        methods=['POST'], detail=False, url_path='approve-create-request',
        serializer_class=WorkLogBulkApproveSerializer
    )
    def approve_create_request(self, request, **kwargs) -> Response:
        if self.mode not in ["hr", "supervisor"]:
            raise ValidationError("You cannot approve this request.")
        data = request.data
        for req in data.get("requests"):
            instance = get_object_or_404(self.get_queryset(), id=req.get('worklog'))
            if self.mode == 'supervisor' and not subordinates.authority_exists(
                instance.sender, self.request.user, 'approve'
            ):
                raise ValidationError({
                    "You can not approve this request."
                })
            if instance.status not in [REQUESTED, FORWARDED]:
                raise ValidationError("Worklog must be in requested state before approval.")

        ser = WorkLogBulkApproveSerializer(
            data=data, context={'user': self.request.user}
        )
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    @action(
        methods=['POST'], detail=False, url_path='deny-create-request',
        serializer_class=WorkLogActionBulkRemarksSerializer
    )
    def deny_create_request(self, request, **kwargs) -> Response:
        return self.add_worklog_bulk_action(DENIED, 'deny')

    @action(
        methods=['POST'], detail=False, url_path='forward-create-request',
        serializer_class=WorkLogActionBulkRemarksSerializer
    )
    def forward_create_request(self, request, pk=None, **kwargs) -> Response:
        return self.add_worklog_bulk_action(FORWARDED, 'forward')

    @action(
        methods=['POST'], detail=False, url_path='confirm-action',
        serializer_class=WorkLogActionBulkRemarksSerializer
    )
    def confirm_action(self, request, pk=None, **kwargs) -> Response:
        return self.add_worklog_bulk_action(CONFIRMED)

    @action(
        methods=['POST'], detail=True, url_path='cancel-action',
        serializer_class=WorkLogActionRemarksSerializer
    )
    def cancel_action(self, request, pk=None, **kwargs) -> Response:
        return self.add_worklog_action(CANCELED)

    @action(
        methods=['POST'], detail=True, url_path='acknowledge-action',
        serializer_class=WorkLogActionRemarksSerializer
    )
    def acknowledge_action(self, request, pk=None, **kwargs) -> Response:
        return self.add_worklog_action(ACKNOWLEDGED)

    # def return_response(self, queryset):
    #     all_subordinates = self.request.user.subordinates_pks
    #     immediate_subordinates = find_immediate_subordinates(self.request.user.id)
    #     paginated_queryset = self.paginate_queryset(queryset)
    #     data = WorLogListSerializer(
    #         paginated_queryset, many=True).data
    #     response = super().get_paginated_response(data)
    #     response.data['stats'] = self.get_stats_info(
    #         self.get_queryset().filter(
    #             Q(sender_id__in=all_subordinates)
    #         )
    #     )
    #     return response
    #
    # @action(
    #     methods=['GET'], detail=False, url_path='get-supervisor-filter'
    # )
    # def get_supervisor_filter(self, request):
    #     if self.mode != "supervisor":
    #         raise PermissionDenied
    #     w_status = self.request.query_params.get('status')
    #     queryset = self.filter_queryset(self.get_queryset())
    #     immediate_subordinates = find_immediate_subordinates(self.request.user.id)
    #     all_subordinates = self.request.user.subordinates_pks
    #     if w_status in [TODO, DRAFT, REQUESTED]:
    #         queryset = queryset.filter(
    #             Q(receiver=self.request.user)
    #         )
    #     if w_status == FORWARDED:
    #         queryset = queryset.filter(
    #             Q(status=FORWARDED) & ~Q(receiver=self.request.user) &
    #             Q(sender_id__in=all_subordinates)
    #         )
    #     else:
    #         queryset = queryset.filter(Q(sender_id__in=all_subordinates))
    #     return self.return_response(queryset)

    @action(methods=['GET'], detail=True, url_name='history',
            url_path='history')
    def history(self, request, *args, **kwargs) -> Response:
        worklog = self.get_object()
        qs = worklog.worklog_actions.all()

        paginated_queryset = self.paginate_queryset(qs)
        data = WorkLogHistorySerializer(
            paginated_queryset, many=True).data
        return self.get_paginated_response(data)

    def get_organization(self) -> Organization:
        return self.request.user.detail.organization

    @action(methods=['POST'], detail=False, url_path="export-worklog")
    def export_worklog(self, response, **kwargs) -> Response:
        """exports specific status of  work log

        Note: This export doesn't is sync i.e, doesn't go in background task
        """
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="worklog.xlsx"'
        wb = Workbook()
        sheet = wb['Sheet']
        bold = Font(bold=True)
        align_center = Alignment(horizontal='center')
        strp_time_format = "%Y-%m-%d %H:%M:%S"
        queryset = self.filter_queryset(self.get_queryset())
        organization = self.get_organization()
        # logo = nested_getattr(organization, 'appearance.logo')
        # if not logo:
        #     raise ValidationError("Organization logo not found")
        # sheet.merge_cells(
        #     start_row=1,
        #     start_column=1, end_row=1,
        #     end_column=7
        # )
        # try:
        #     image_obj = Image(logo)
        # except FileNotFoundError:
        #     raise ValidationError("Organization logo not found")
        # sheet.add_image(image_obj, anchor="A1")
        # dimension = sheet.row_dimensions[1]
        # dimension.height = image_obj.height * 0.75

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        sheet.cell(row=1, column=1, value=organization.name)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        sheet.cell(row=2, column=1,
                   value=f"Downloaded at: ({get_today(with_time=True).strftime(strp_time_format)})")

        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=15)
        cell = sheet.cell(row=4, column=1, value="Work Log")
        cell.font = bold
        cell.alignment = align_center

        if self.mode in ['hr', 'supervisor']:
            cell = sheet.cell(column=1, row=7, value="Employee Name")
            cell.font = bold
            cell.alignment = align_center

        cell = sheet.cell(column=2, row=7, value="Project")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=3, row=7, value="Activity")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=4, row=7, value="Task")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=5, row=7, value="Core Task")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=6, row=7, value="Start Time")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=7, row=7, value="End Time")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=8, row=7, value="Quantity")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=9, row=7, value="Unit")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=10, row=7, value="Remarks")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=11, row=7, value="Employee Rate")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=12, row=7, value="Employee Rate Amount")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=13, row=7, value="Client Rate")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=14, row=7, value="Client Rate Amount")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=15, row=7, value="Score")
        cell.font = bold
        cell.alignment = align_center

        cell = sheet.cell(column=16, row=7, value="Feedback")
        cell.font = bold
        cell.alignment = align_center

        for column_index, worklog in enumerate(queryset):
            column = column_index + 8
            employee = nested_getattr(worklog, 'sender.full_name', default="N/A")
            project = nested_getattr(worklog, 'project.name', default="N/A")
            activity = nested_getattr(worklog, 'activity.name', default="N/A")
            task = nested_getattr(worklog, 'task.title', default="N/A")
            core_task = nested_getattr(worklog, 'core_task.title', default="N/A")
            start = worklog.start_time
            end = worklog.end_time
            start_time = start.astimezone().strftime(strp_time_format) if start else 'N/A'
            end_time = end.astimezone().strftime(strp_time_format) if end else 'N/A'
            requested = worklog.unit or 0
            measure = nested_getattr(worklog, 'activity.unit', default="N/A")
            remarks = worklog.activity_description
            w_status = worklog.status
            employee_rate = nested_getattr(worklog, 'activity.employee_rate', default=0)
            total_employee_amount = employee_rate * requested or 0
            client_rate = nested_getattr(worklog, 'activity.client_rate', default=0)
            worklog_action = worklog.worklog_actions.filter(action=w_status).first()
            total_client_amount = client_rate * requested or 0
            score = worklog_action.score if worklog_action else 0
            if w_status in [APPROVED, ACKNOWLEDGED, CONFIRMED]:
                score = worklog.worklog_actions.filter(action=APPROVED).first().score
            feedback = worklog_action.remarks if worklog_action else "N/A"

            if self.mode in ['hr', 'supervisor']:
                sheet.cell(column=1, row=column, value=employee)
            sheet.cell(column=2, row=column, value=project)
            sheet.cell(column=3, row=column, value=activity)
            sheet.cell(column=4, row=column, value=task)
            sheet.cell(column=5, row=column, value=core_task)
            sheet.cell(column=6, row=column, value=start_time)
            sheet.cell(column=7, row=column, value=end_time)
            sheet.cell(column=8, row=column, value=requested)
            sheet.cell(column=9, row=column, value=measure)
            sheet.cell(column=10, row=column, value=remarks)
            sheet.cell(column=11, row=column, value=employee_rate)
            sheet.cell(column=12, row=column, value=total_employee_amount)
            sheet.cell(column=13, row=column, value=client_rate)
            sheet.cell(column=14, row=column, value=total_client_amount)
            sheet.cell(column=15, row=column, value=score or 0)
            sheet.cell(column=16, row=column, value=feedback)

        last_row_count = sheet.max_row
        sheet.merge_cells(
            start_row=last_row_count+1, start_column=1, end_row=last_row_count+1, end_column=14
        )
        cell = sheet.cell(row=last_row_count + 1, column=1, value="Average Score")
        cell.font = bold
        cell.alignment = Alignment(horizontal='center')

        def formula(col):
            return "=AVERAGE({}:{})".format(
                f"{get_column_letter(col)}{6}",
                f"{get_column_letter(col)}{last_row_count}",
            )

        cell = sheet.cell(
            row=last_row_count + 1, column=15,
            value=formula(15)
        )
        cell.font = bold
        cell.alignment = Alignment(horizontal='right')

        wb.save(response)
        return response

    def get_frontend_redirect_url(self) -> str:
        return f'/admin/task/worklog/'


class WorkLogReportFilter(FilterSet):
    class Meta:
        model = WorkLog
        fields = ['project_id', 'sender_id', 'worklog_actions__action_date']


class WorkLogReportViewSet(mixins.ListModelMixin, GenericViewSet):
    """Detail report of work log

    Note: Only work log with billable project are seen in the report
    """
    queryset = WorkLog.objects.all()
    filter_backends = (DjangoFilterBackend, OrderingFilterMap, SearchFilter, FilterMapBackend)
    search_fields = (
        'project__name',
        'activity__name',
        'task__title',
        'core_task__title'
    )
    filter_class = WorkLogReportFilter
    filter_map = {
        'start_date': 'worklog_actions__action_date__date__gte',
        'end_date': 'worklog_actions__action_date__date__lte'
    }
    serializer_class = WorkLogReportSerializer

    @property
    def mode(self) -> str:
        mode = self.request.query_params.get('as')
        if mode not in ['hr', 'supervisor']:
            return 'user'

        return mode

    def get_queryset(self) -> QuerySet:
        queryset = super().get_queryset()
        if self.mode == "user":
            queryset = queryset.filter(sender=self.request.user)

        if self.mode == "supervisor":
            queryset = queryset.filter(receiver=self.request.user)

        queryset = queryset.filter(
            worklog_actions__action__in=[CONFIRMED, SENT]
        ).annotate(is_billable=Exists(
            UserActivityProject.objects.filter(
                user=OuterRef('sender'),
                project=OuterRef('project'),
                activity=OuterRef('activity'),
                is_billable=True
            )
        )
        ).filter(is_billable=True)
        return queryset.distinct()

    def get_serializer_context(self) -> dict:
        ctx = super().get_serializer_context()
        ctx['mode'] = self.mode
        return ctx

    @action(
        methods=['POST'], detail=False, url_path='send-worklog-to-payroll',
        serializer_class=WorkLogBulkSendSerializer
    )
    def send_worklog_to_payroll(self, request, **kwargs) -> Response:
        """sends scores of worklog to the payroll and changes worklog status to SENT"""

        ser = WorkLogBulkSendSerializer(
            data=self.request.data,
            context={
                "user": self.request.user
            }
        )
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)
