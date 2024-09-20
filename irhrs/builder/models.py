from openpyxl import Workbook
from django.contrib.contenttypes.models import ContentType
from django.db import models
from django.db.models import Avg, Min, Max, Count, Sum, Q

from irhrs.builder.constants import FIELDS_FILTER_CHOICES, \
    FIELDS_ORDERING_CHOICES, FIELDS_AGGREGATE_CHOICES
from irhrs.common.models import BaseModel


class Report(BaseModel):
    name = models.CharField(max_length=100)
    description = models.TextField(blank=True, null=True)
    app = models.CharField(max_length=100)
    model = models.CharField(max_length=100)
    model_url = models.CharField(max_length=200)

    def __str__(self):
        return self.name

    @property
    def root_class(self):
        return ContentType.objects.get(app_label=self.app,
                                       model=self.model).model_class()

    @property
    def report_filter_fields(self):
        if hasattr(self, 'reportfilter'):
            return self.reportfilter.reportfilterfield.all()
        return None

    @staticmethod
    def add_aggregates(queryset, display_fields):
        agg_funcs = {
            'Avg': Avg, 'Min': Min, 'Max': Max, 'Count': Count, 'Sum': Sum
        }

        for display_field in display_fields:
            if display_field.aggregate:
                func = agg_funcs[display_field.aggregate]
                full_name = display_field.field_path
                queryset = queryset.annotate(func(full_name))

        return queryset

    def get_report_queryset(self):
        model_class = ContentType.objects.get(app_label=self.app,
                                              model=self.model).model_class()
        base_query = model_class.objects.all().order_by()
        if hasattr(self, 'reportfilter'):
            _filter_operation = self.reportfilter.operation
            q_operation = _filter_operation.replace('(', 'Q(')
            for filter_field in self.reportfilter.reportfilterfield.all().order_by(
                    'id'):
                filter_string = filter_field.field_path + '__' + filter_field.filter_type
                if filter_field.filter_type == "isnull":
                    if filter_field.filter_value in ["0", "False", "false"]:
                        filter_value = False
                    elif filter_field.filter_value in ["1", "True", "true"]:
                        filter_value = True
                    else:
                        filter_value = None

                elif filter_field.filter_type == "in":
                    filter_value = filter_field.filter_value.split(',')
                else:
                    filter_value = filter_field.filter_value
                    if filter_field.filter_type == 'range':
                        filter_value = [filter_value,
                                        filter_field.filter_value2]

                _filter_operation = f"Q({filter_string}='{filter_value}')" \
                    if not filter_field.apply_not else f"~Q({filter_string}={filter_value})"
                q_operation = q_operation.replace(filter_field.field_path,
                                                  _filter_operation)

            if q_operation:
                base_query = base_query.filter(eval(q_operation))

        all_display_fields = self.reportdisplayfield.all().order_by('id')
        base_query = self.add_aggregates(base_query, all_display_fields)

        _query_select = []
        _display_obj_map = {}
        for d in all_display_fields:
            display_field_key = d.field_path
            if d.field_type == "Property":
                pass
            elif d.aggregate == "Avg":
                display_field_key += '__avg'
            elif d.aggregate == "Max":
                display_field_key += '__max'
            elif d.aggregate == "Min":
                display_field_key += '__min'
            elif d.aggregate == "Count":
                display_field_key += '__count'
            elif d.aggregate == "Sum":
                display_field_key += '__sum'
            if d.field_type not in ('Property',):
                _query_select.append(display_field_key)
                _display_obj_map[display_field_key] = d
        ordering = []
        for order in all_display_fields.filter(ordering__isnull=False):
            _order = order.field_path if order.ordering == 'Asc' else '-' + order.field_path
            ordering.append(_order)

        base_query = base_query.values_list(*_query_select).order_by(
            *ordering).distinct()
        return _query_select, _display_obj_map, base_query

    def export_result_wb(self):
        _select, _displays, _queryset = self.get_report_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = self.name[:30]
        columns = [_displays[i].display for i in _select]
        for x, column in enumerate(columns, start=1):
            value = column
            _ = ws.cell(row=1, column=x, value=value)

        for row, obj in enumerate(_queryset, start=4):
            for col, column in enumerate(obj, start=1):
                value = column
                _ = ws.cell(row=row, column=col, value=value)

        return wb


class ReportRelatedFieldsInfo(BaseModel):
    report = models.ForeignKey(Report, on_delete=models.CASCADE,
                               related_name="%(class)s")
    url = models.CharField(max_length=200)
    field_verbose = models.CharField(max_length=50)
    field_path = models.CharField(max_length=100)

    def __str__(self):
        return self.url


class ReportDisplayField(BaseModel):
    report = models.ForeignKey(Report, on_delete=models.CASCADE,
                               related_name="%(class)s")
    field_name = models.CharField(max_length=500)
    field_path = models.CharField(max_length=255)
    field_type = models.CharField(max_length=255)
    display = models.CharField(max_length=50)
    aggregate = models.CharField(
        max_length=5,
        choices=FIELDS_AGGREGATE_CHOICES,
        blank=True,
        null=True
    )
    group = models.BooleanField(default=False)

    ordering = models.CharField(max_length=4,
                                choices=FIELDS_ORDERING_CHOICES, blank=True,
                                null=True, db_index=True)

    def __str__(self):
        return self.display


class ReportFilter(BaseModel):
    report = models.OneToOneField(Report, on_delete=models.CASCADE,
                                  related_name="%(class)s")
    operation = models.CharField(max_length=200)

    def __str__(self):
        return self.operation + 'on' + self.report.__str__()


class ReportFilterField(BaseModel):
    report_filter = models.ForeignKey(ReportFilter, on_delete=models.CASCADE,
                                      related_name="%(class)s")
    field_name = models.CharField(max_length=500)
    field_path = models.CharField(max_length=255)
    field_type = models.CharField(max_length=255)
    display = models.CharField(max_length=50)
    filter_type = models.CharField(
        max_length=15,
        choices=FIELDS_FILTER_CHOICES,
        default='exact',
        db_index=True
    )
    filter_value = models.CharField(max_length=255)
    filter_value2 = models.CharField(max_length=255, blank=True, null=True)
    apply_not = models.BooleanField(default=False)

    def __str__(self):
        return self.display
