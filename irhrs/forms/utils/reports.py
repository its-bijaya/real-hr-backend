import os
import re

import pytz
from lxml import html
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.cell import Cell
from openpyxl import Workbook
from django.conf import settings
from django.urls import reverse
from django.db.models import Q, Count
from django.urls import reverse
from openpyxl.styles import Alignment
from irhrs.questionnaire.models.helpers import (
    CHECKBOX,
    RADIO,
    SHORT,
    LONG,
    RATING_SCALE,
    DATE,
    TIME,
    DURATION,
    DATE_TIME,
    DATE_WITHOUT_YEAR,
    DATE_TIME_WITHOUT_YEAR,
    FILE_UPLOAD,
    MULTIPLE_CHOICE_GRID,
    CHECKBOX_GRID,
)
from irhrs.core.utils import nested_getattr
from django.utils import timezone
from irhrs.core.utils.common import get_today
from irhrs.forms.models import (
    Form,
    UserForm,
    FormQuestion,
    FormQuestionSection,
    UserFormAnswerSheet,
    UserFormIndividualQuestionAnswer,
    AnonymousFormIndividualQuestionAnswer,
    AnonymousFormAnswerSheet,
)
from irhrs.users.models import UserDetail


def in_dictlist(key, value, my_dictlist):
    """
    Checks for key and value in a list of dicts.
    Returns entry if found else returns empty dict.
    """

    for entry in my_dictlist:
        if entry[key] == value:
            return entry
    return {}


def transform_aggregation(aggregated_list):
    """
    transform results like [{"answers": [3]}] to [{"answers": 3}]
    according to JSON agreement with Frontend
    """

    results = []
    for aggregated_values in aggregated_list:
        result = {}
        for key, val in aggregated_values.items():
            if key == "answers":
                try:
                    result["answer"] = val[0]
                except (KeyError, IndexError):
                    result["answer"] = val
            else:
                result[key] = aggregated_values[key]
        results.append(result)
    return results


def get_prepoulated_offset(form_question, question_answers):
    aggregated_list = question_answers.values('answers').annotate(count=Count('id')).order_by('answers')
    prepoulated_count = 5
    if aggregated_list.count() <= prepoulated_count:
        return None
    next_offset = prepoulated_count
    return next_offset


def get_aggreate_for_date_and_time_types(form_question, question_answers, context=None):
    aggregated_list = question_answers.values('answers').annotate(count=Count('id')).order_by('answers')
    pre_poulated_count = 5
    aggregated_list = aggregated_list[:pre_poulated_count]
    if aggregated_list:
        return transform_aggregation(aggregated_list)


def get_aggregate_for_rating_scale(form_question, question_answers, context=None):
    aggregated_list = question_answers.values('answers').annotate(count=Count('id')).order_by()
    result = transform_aggregation(aggregated_list)
    rating_scale_qa = question_answers.filter(
        question=form_question
    ).first()
    if rating_scale_qa:
        rating_scale_max = rating_scale_qa.question.question.rating_scale

        possible_scores = []
        for score in range(rating_scale_max + 1):
            # rating scale goes like 0, 0.5, 1, 1.5 etc
            # get list of all possible rating scale values
            possible_scores.append(score)
            possible_scores.append(score + 0.5)
        # removes last score because it is redundant
        possible_scores.pop()

        for score in possible_scores:
            if not in_dictlist("answer", score, result):
                result.append({
                    "answer": score,
                    "count": 0
                })
        result = sorted(result, key=lambda x: x["answer"] if x["answer"] else 0)
        return result
    return []


def get_aggregate_for_select_type(form_question, question_answers, context=None):
    choices = form_question.question.all_answer_choices.all().values_list('title', flat=True)
    annotate_kwargs = dict()
    for choice in choices:
        annotate_kwargs[f'{choice}__count'] = Count(
            'id',
            filter=Q(
                answers__contains=[
                    {
                        "title": choice,
                        "is_correct": True
                    }
                ]
            )
        )
    result = question_answers.aggregate(**annotate_kwargs)
    transformed_result = []
    for choice in choices:
        res = {
            "answer": choice,
            "count": result.get(f"{choice}__count") or 0
        }
        transformed_result.append(res)
    return transformed_result


def get_aggregate_for_grid_types(form_question, question_answers, context=None):
    choices = form_question.question.extra_data.get("rows", list())
    sub_choices = form_question.question.extra_data.get("columns", list())

    annotate_kwargs = dict()
    for choice in choices:
        for sub_choice in sub_choices:
            fil = {
                f"answers__{choice}__contains": [sub_choice]
            }
            annotate_kwargs[f'{choice}__{sub_choice}'] = Count(
                'id',
                filter=Q(**fil)
            )

    result = question_answers.aggregate(**annotate_kwargs)
    transformed_result = dict()
    for key, val in result.items():
        choice, sub_choice = key.split('__')
        transformed_result.setdefault(choice, dict())
        transformed_result[choice][sub_choice] = val
    return transformed_result

def get_aggregate_for_short_long_answer(form_question, question_answers, context=None):
    results = []
    for qa in question_answers:
        answer = qa.answers
        if answer:
            results.append(answer[0])
    return results


def get_aggregate_for_file_upload(form_question, question_answers, context=None):
    results = []
    question_answers = question_answers.order_by('answers').values('answers')[:4]
    for qa in question_answers:
        answer = qa['answers']
        if answer and isinstance(answer, list):
            file_url = answer[0].get('file_url')
            original_nice_file_name = answer[0].get("file_name")
            filename = os.path.basename(file_url)
            request = context.get('request')
            organization = context.get('organization')
            file_download_link = reverse(
                'api_v1:forms:forms-answer-sheets-download-form-attachment',
                kwargs={
                    'organization_slug': organization.slug,
                }
            ) + f"?file_name={original_nice_file_name}&file_uuid={filename}"
            answer[0]['file_url'] = request.build_absolute_uri(
                file_download_link
            )
            results.append(qa['answers'][0])
    return results


def get_blank_responses_for_char_types(form_question, question_answers):
    blank_response = question_answers.filter(answers__contains=[""])
    return blank_response

def get_blank_responses_for_file_upload(form_question, question_answers):
    blank_response = question_answers.filter(answers__contains=[
        {
            "file_name": "",
            "file_url": ""
        }
    ])
    return blank_response

def get_blank_responses_for_select_type(form_question, question_answers):
    choices = form_question.question.all_answer_choices.all().values_list('title', flat=True)
    blank_response_payload = [
        {"title": choice, "is_correct": False} for choice in choices
    ]
    blank_response = question_answers.filter(answers__contains=blank_response_payload)
    return blank_response


def get_blank_responses_for_grid_type(form_question, question_answers):
    return 0

blank_answer_count_map = {
    RATING_SCALE: get_blank_responses_for_char_types,
    SHORT: get_blank_responses_for_char_types,
    LONG: get_blank_responses_for_char_types,
    DATE: get_blank_responses_for_char_types,
    TIME: get_blank_responses_for_char_types,
    DURATION: get_blank_responses_for_char_types,
    DATE_TIME: get_blank_responses_for_char_types,
    DATE_WITHOUT_YEAR: get_blank_responses_for_char_types,
    DATE_TIME_WITHOUT_YEAR: get_blank_responses_for_char_types,
    FILE_UPLOAD: get_blank_responses_for_file_upload,
    CHECKBOX: get_blank_responses_for_select_type,
    RADIO: get_blank_responses_for_select_type,
    MULTIPLE_CHOICE_GRID: get_blank_responses_for_grid_type,
    CHECKBOX_GRID: get_blank_responses_for_grid_type,
}

aggregate_function_map = {
    RATING_SCALE: get_aggregate_for_rating_scale,
    SHORT: get_aggregate_for_short_long_answer,
    LONG: get_aggregate_for_short_long_answer,
    DATE: get_aggreate_for_date_and_time_types,
    TIME: get_aggreate_for_date_and_time_types,
    DURATION: get_aggreate_for_date_and_time_types,
    DATE_TIME: get_aggreate_for_date_and_time_types,
    DATE_WITHOUT_YEAR: get_aggreate_for_date_and_time_types,
    DATE_TIME_WITHOUT_YEAR: get_aggreate_for_date_and_time_types,
    FILE_UPLOAD: get_aggregate_for_file_upload,
    CHECKBOX: get_aggregate_for_select_type,
    RADIO: get_aggregate_for_select_type,
    MULTIPLE_CHOICE_GRID: get_aggregate_for_grid_types,
    CHECKBOX_GRID: get_aggregate_for_grid_types
}


def get_aggregate_for_type(form_question, question_answers, context):
    aggregate_function = aggregate_function_map.get(
        form_question.question.answer_choices
    )
    blank_response_counter = blank_answer_count_map.get(
        form_question.question.answer_choices
    )
    if not aggregate_function or not blank_response_counter:
        return []
    blank_responses = blank_response_counter(form_question, question_answers)
    if blank_responses:
        blank_response_ids = blank_responses.values('id')
        blank_count = blank_responses.count()
        question_answers = question_answers.exclude(id__in=blank_response_ids)
    else:
        blank_count = 0
    result = aggregate_function(form_question, question_answers, context)
    final_response = {
        "empty_response_count": blank_count,
        "responses": result
    }
    return final_response


def cleanhtml(raw_text):
    return html.fromstring(raw_text).text_content().strip()


def get_ordered_form_question_texts(form, request_details):
    final_questions = []
    question_description=[]
    is_user = request_details.get("is_user")
    questions = FormQuestion.objects.filter(
        question_section__question_set__forms=form
    ).order_by('question_section', 'order')
    if is_user:
        questions = questions.filter(answer_visible_to_all_users=True)
    for form_question in questions:
        question = form_question.question
        if question.answer_choices == CHECKBOX:
            choices_count = question.all_answer_choices.count()
            mcq_repeated_questions = choices_count * [question.title]
            final_questions.extend(mcq_repeated_questions)
        elif question.answer_choices in [MULTIPLE_CHOICE_GRID, CHECKBOX_GRID]:
            choices = question.extra_data.get("rows", list())
            choices_count = len(choices)
            final_questions.extend([f"{question.title}[{choice}]" for choice in choices])
        else:
            final_questions.append(question.title)
            question_description.append(question.description)
    final_question_texts = list(map(lambda x: cleanhtml(x), final_questions))
    return final_question_texts, question_description   

def get_answer_text_for_checkbox(answer, result):
    # repeat answer for MCQ multiple times
    choices = answer.answers
    for choice in choices:
        if choice.get("is_correct") is True:
            result.append(choice.get("title"))
        else:
            result.append("")


def get_answer_text_for_grid_types(individual_qa, result):
    # repeat answer for MCQ multiple times
    answer_choice = individual_qa.question.question.answer_choices
    answers = individual_qa.answers
    choices = individual_qa.question.question.extra_data.get("rows")
    if answer_choice == MULTIPLE_CHOICE_GRID:
        formatted_result = [
            "".join(answers.get(choice)) if answers.get(choice) else ""
            for choice in choices
        ]
    else:
        formatted_result = [
            ",".join(answers.get(choice)) if answers.get(choice) else ""
            for choice in choices
        ]
    result.extend(formatted_result)

def get_answer_text_for_radio(answer, result):
    choices = answer.answers
    correct_choice_found = False
    for choice in choices:
        if choice.get("is_correct") is True:
            result.append(choice.get("title"))
            correct_choice_found = True
    if not correct_choice_found:
        result.append('')

def get_answer_text_for_file_upload(answer, result):
    answer_content = answer.answers or ''
    if answer_content and isinstance(answer_content, list):
        original_nice_file_name = answer_content[0].get("file_name")
        file_url = answer_content[0].get("file_url")
        filename = os.path.basename(file_url)
        if filename:
            file_download_link = settings.BACKEND_URL + reverse(
                'api_v1:forms:forms-answer-sheets-download-form-attachment',
                kwargs={
                    'organization_slug': answer.answer_sheet.form.organization.slug,
                }
            ) + f"?file_name={original_nice_file_name}&file_uuid={filename}"
        else:
            file_download_link = ''
        result.append(file_download_link)
    else:
        result.append(str(answer_content))

def get_answer_text_default(answer, result):
    answer_content = answer.answers
    if answer_content:
        try:
            result.append(answer_content[0])
        except KeyError:
            result.append(str(answer_content))
    else:
        result.append('')

answer_choice_text_map = {
    CHECKBOX: get_answer_text_for_checkbox,
    RADIO: get_answer_text_for_radio,
    FILE_UPLOAD: get_answer_text_for_file_upload,
    MULTIPLE_CHOICE_GRID: get_answer_text_for_grid_types,
    CHECKBOX_GRID: get_answer_text_for_grid_types,
}


def append_text_for_answer_choice_to_result(answer, result):
    answer_choices = answer.question.question.answer_choices
    get_text_function = answer_choice_text_map.get(answer_choices)
    if get_text_function:
        result = get_text_function(answer, result)
    else:
        result = get_answer_text_default(answer, result)


def get_ordered_answer_texts(form, users, form_fill_date, request_details):
    final_answers = []
    individual_qa_model = (
        AnonymousFormIndividualQuestionAnswer
        if form.is_anonymously_fillable
        else UserFormIndividualQuestionAnswer
    )
    answer_sheet_model = (
        AnonymousFormAnswerSheet
        if form.is_anonymously_fillable
        else UserFormAnswerSheet
    )
    answer_sheet_filters = dict(
        form=form,
    )
    if not form.is_anonymously_fillable:
        answer_sheet_filters.update({
            'user__in': users,
            'is_approved': True
        })

    answer_sheets = answer_sheet_model.objects.filter(
        **answer_sheet_filters
    )
    is_user = request_details.get("is_user")
    for answer_sheet in answer_sheets:
        result = []
        ordered_answers = individual_qa_model.objects.filter(
            answer_sheet=answer_sheet
        ).order_by('question__question_section', 'question__order')
        if is_user:
            ordered_answers = ordered_answers.filter(question__answer_visible_to_all_users=True)
        if form_fill_date:
            ordered_answers = ordered_answers.filter(
                created_at__date=form_fill_date
            )
        first_answer = ordered_answers.first()
        if first_answer:
            answered_date = first_answer.created_at
            dt_format = "%Y-%m-%d %H:%M:%S"
            kathmandu = pytz.timezone('Asia/Kathmandu')
            result.append(
                answered_date.astimezone(kathmandu).strftime(
                    dt_format
                )
            )
                
            if not form.is_anonymously_fillable:
                user = answer_sheet.user
                user_detail = UserDetail.objects.get(user=user)   
                result.extend([
                    user.full_name,
                    user.username,
                    user.email,
                    nested_getattr(user_detail,'employment_level.title', default='Not Assigned'),
                    nested_getattr(user_detail,'job_title.title', default='Not Assigned'),
                    nested_getattr(user_detail,'branch.name', default='Not Assigned'),
                    nested_getattr(user_detail,'employment_status.title', default='Not Assigned'),
                ])                
        for answer in ordered_answers:
            append_text_for_answer_choice_to_result(answer, result)
        final_answers.append(result)
    return final_answers


def create_form_report(form, users, form_fill_date, request_details):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Form Answers"
    organization = form.organization
    
    # for organization logo in downloaded excel report:
    # logo = nested_getattr(organization, 'appearance.logo')
    # if logo:
    #     sheet.merge_cells(
    #         start_row=1,
    #         start_column=1, end_row=1,
    #         end_column=7
    #     )
    #     try:
    #         image_obj = Image(logo)
    #         sheet.add_image(image_obj, anchor="A1")
    #         dimension = sheet.row_dimensions[1]
    #         dimension.height = image_obj.height * 0.75
    #     except FileNotFoundError:
    #         sheet.cell(
    #             row=1, column=1,
    #             value=f"Logo not found"
    #         )

    bold_font = Font(bold=True)
    center_alignment = Alignment(
        vertical='center',
        horizontal='center'
    )

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    sheet.cell(row=2, column=1, value=organization.name)
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)

    strp_time_format = "%Y-%m-%d %H:%M:%S"
    sheet.cell(
        row=3, column=1,
        value=f"Generated at: ({get_today(with_time=True).strftime(strp_time_format)})"
    )

    sheet.cell(
        row=5, column=1,
        value=f"Form: {form.name}"
    )

    sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=7)
    questions, descriptions = get_ordered_form_question_texts(form, request_details)
    anonymously_fillable = form.is_anonymously_fillable

    row = sheet.max_row+1
    column = 9
    timestamp = sheet.cell(row=row,column=1,value="Timestamp")
    timestamp.font = bold_font
    timestamp.alignment = center_alignment
    if not anonymously_fillable:
        heading_cell = ("Name", "Username","Email", "Employment Level", "Job Title", "Branch", "Employment Type")

        for col, header in enumerate(heading_cell, start=2):
            header_cell = sheet.cell(row=row, column=col, value=header)
            header_cell.font = bold_font
            header_cell.alignment = center_alignment
            
        q_cell = 9
        for question, description in zip(questions,descriptions):
            question_cell= sheet.cell(row=row,column=q_cell,value=question)
            description_cell = sheet.cell(row=row+1,column=column,value=description)
            question_cell.font = bold_font
            question_cell.alignment = center_alignment
            description_cell.alignment = Alignment(
                wrap_text=True,
                vertical='top',
            )
            q_cell += 1
    else:
        for question in questions:
            cell = sheet.cell(row=row,column=column, value=question)
            cell.font = bold_font
            cell.alignment = center_alignment
            column += 1
    answers = get_ordered_answer_texts(form, users, form_fill_date, request_details)
    for answer in answers:
        sheet.append(answer)
    return workbook
