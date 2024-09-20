import openpyxl
import json

from irhrs.core.utils import get_system_admin
from irhrs.core.utils.common import DummyObject
from django.contrib.auth import get_user_model
from irhrs.users.api.v1.serializers.insurance import UserInsuranceSerializer
from irhrs.users.models import UserDetail, UserContactDetail

PATH = 'insurance.xlsx'
USER = get_user_model()


def main():
    system_bot = get_system_admin()
    wb = openpyxl.load_workbook(PATH)
    sheet = wb.active
    error_wb = openpyxl.Workbook()
    error_sheet = error_wb.active
    header_one = error_sheet.cell(row=1, column=1)
    header_one.value = 'Code'
    header_two = error_sheet.cell(row=1, column=2)
    header_two.value = 'Emails'
    header_three = error_sheet.cell(row=1, column=3)
    header_three.value = 'Errors'
    success_count = 0
    error_row_start = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user_detail = UserDetail.objects.filter(code=row[0]).first()
        if user_detail:
            user = user_detail.user
        else:
            user = USER.objects.filter(email=row[1]).first()
        if user:
            dependent_list = []
            for i in range(10, 17):
                if row[i]:
                    dependent_ids = UserContactDetail.objects.filter(
                        user=user, name=row[i], is_dependent=True
                    ).values_list('slug', flat=True)
                    if dependent_ids:
                        dependent_list.append(dependent_ids[0])

            data = {
                'user': user,
                'insured_scheme': row[2],
                'policy_name': row[3],
                'policy_provider': row[4],
                'insured_amount': row[5],
                'annual_premium': row[6],
                'policy_type': row[7].lower().replace(' ', '_'),
                'start_date': row[8].date(),
                'end_date': row[9].date(),
                'dependent': dependent_list
            }

            context = {
                "user": user,
                "request": DummyObject(method='POST', user=system_bot),
                "send_notification": False
            }
            ser = UserInsuranceSerializer(data=data, context=context)
            if ser.is_valid():
                ser.save()
                success_count += 1
            else:
                error_row_start += 1
                error_col_one = error_sheet.cell(row=error_row_start, column=1)
                error_col_one.value = row[0]
                error_col_two = error_sheet.cell(row=error_row_start, column=2)
                error_col_two.value = row[1]
                error_col_three = error_sheet.cell(
                    row=error_row_start, column=3)
                error_col_three.value = json.dumps(ser.errors)
                print('-----------INVALID------------', json.dumps(ser.errors))

    print(f"successfully created {success_count} insurance info")
    if error_row_start > 1:
        print(f"{error_row_start - 1} "
              f"Failed insurance imported into failed_insurance_list.xlsx")
        error_wb.save('failed_insurance_list.xlsx')


if __name__ == "__main__":
    main()
