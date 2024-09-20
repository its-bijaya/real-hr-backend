from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from irhrs.users.models.medical_and_legal import UserLegalInfo
from irhrs.users.models.user import UserDetail
User = get_user_model()

# file path of employee legal information in excel file format
path = "empLegalInfo.xlsx"
emp_column, pan_column, cit_column, citizenship_column, ssfid_column = 0, 1, 2, 3, 4 


def get_legal_data(path=path):
    workbook = load_workbook(path)
    workspace = workbook.active
    # get iterable emp_code to get user_id
    emp_codes = (cell.value for cell in workspace['A'][1:])
    users_dict = dict(UserDetail.objects.filter(code__in=emp_codes).values_list('code', 'user__id'))
    field_datas = workspace.iter_rows(min_row=2, values_only=True)
    return users_dict, field_datas


def main():
    UserLegalInfo.objects.all().delete()

    users_legal_datas = []
    users_dict, legal_datas = get_legal_data()
    for legal_info in legal_datas:
        user_id = users_dict[str(legal_info[0])]
        # legal_info = process_legal_info(legal_info)
        user_legal_info = UserLegalInfo(
            user_id=user_id,
            pan_number=legal_info[pan_column] or '',
            ssfid=legal_info[ssfid_column] or '',
            cit_number=legal_info[cit_column] or '',
            citizenship_number=legal_info[citizenship_column]
        )
        users_legal_datas.append(user_legal_info)

    created_objects = UserLegalInfo.objects.bulk_create(users_legal_datas)
    print(f"successfully created {len(created_objects)} legal info")


if __name__ == "__main__":
    main()
