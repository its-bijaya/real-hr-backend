import openpyxl

from irhrs.recruitment.models import JobApply


def main():
    """Exports full_name and email for the given job_slug (For MCA)

    Available job_apply_status
        - applied,
        - shortlisted,
        - interviewed,
        - selected,
        - screened,
        - rejected,
        - reference_verified
        - pre_screening_interviewed,
        - assessment_taken,
        - salary_declared
    """

    job_slug = "trainee-assistants"   # job slug here
    wb = openpyxl.Workbook()
    sheet = wb.active

    c1 = sheet.cell(row=1, column=1)
    c1.value = "Full Name"
    c2 = sheet.cell(row=1, column=2)
    c2.value = "Email"

    for index, apply in enumerate(
        JobApply.objects.filter(job__slug=job_slug), 2
    ):
        cell1 = sheet.cell(row=index, column=1)
        cell1.value = apply.applicant.user.full_name
        cell2 = sheet.cell(row=index, column=2)
        cell2.value = apply.applicant.user.email

    wb.save(f'{job_slug}.xlsx')


if __name__ == '__main__':
    main()
