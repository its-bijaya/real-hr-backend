"""Assign branch to all staff in bulk (Currently for Laxmi Bank)"""

from irhrs.organization.models import OrganizationBranch
from irhrs.users.models import UserExperience
import openpyxl
from django.contrib.auth import get_user_model

User = get_user_model()
PATH = "branch.xlsx"    # provide path to the excel file
ORGANIZATION_SLUG = "twitter"


def main():
    wb = openpyxl.load_workbook(PATH)
    sheet = wb.active
    failed_wb = openpyxl.Workbook()
    failed_sheet = failed_wb.active
    failed_sheet.title = "Failed branch"
    c1 = failed_sheet.cell(row=1, column=1)
    c1.value = sheet.cell(row=1, column=1).value
    c2 = failed_sheet.cell(row=1, column=2)
    c2.value = sheet.cell(row=1, column=2).value
    success_users_email = []
    count = 1
    for email, branch_name in sheet.iter_rows(min_row=2, values_only=True):
        user_experience = UserExperience.objects.filter(user__email=email).first()

        if not user_experience:
            count += 1
            c1 = failed_sheet.cell(row=count, column=1)
            c1.value = email
            c2 = failed_sheet.cell(row=count, column=2)
            c2.value = branch_name
            continue
        try:
            branch = OrganizationBranch.objects.filter(
                organization__slug=ORGANIZATION_SLUG
            ).get(name=branch_name)
            user_experience.branch = branch
            user_experience.save()
            success_users_email.append(email)

        except OrganizationBranch.DoesNotExist:
            count += 1
            c1 = failed_sheet.cell(row=count, column=1)
            c1.value = email
            c2 = failed_sheet.cell(row=count, column=2)
            c2.value = branch_name

    failed_wb.save("failed_branch.xlsx")
    print(f"Successfully assigned branch to {len(success_users_email)} user email.")

    if count > 1:
        print(f"\nCannot assign {count-1} email to branch.")
        print("Failed user are uploaded in same directory "
              "in excel format.")


if __name__ == "__main__":
    main()
